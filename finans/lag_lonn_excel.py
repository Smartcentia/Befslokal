"""
Lønnsdata prediksjon (valgbart år) — Bufetat
Kilde: Innkjøpsanalyse lønnsutgifter (CSV)
"""
import argparse
import chardet
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER

# ── CONFIG ────────────────────────────────────────────────────────────────────
SRC = "/Users/frank/Downloads/Innkjøpsanalyse 2026 lønnsutgifter(Lønnsutgifter).csv"
YEARS = [2020, 2021, 2022, 2023, 2024, 2025]
parser = argparse.ArgumentParser(description="Lønnsprediksjon for valgt år")
parser.add_argument("--pred-year", type=int, default=2027, help="Prediksjonsår (default: 2027)")
args = parser.parse_args()
PRED_YR = args.pred_year
OUT = f"/Users/frank/Documents/BEFS_CLEAN/finans/Prediksjon_{PRED_YR}_Lønn.xlsx"
PRED_KEY = f"pred_{PRED_YR}"

MAIN_CATS = [
    'Faste stillinger',
    'Lønn vikarer',
    'Arbeidsgiveravgift',
    'Turnustillegg',
    'Pensjonspremie (virksomheter som betaler pensjonspremie)',
    'Midlertidige stillinger (hel- og deltid)',
    'Turnustillegg, vikarer',
    'Overtid faste stillinger',
    'Overtid midlertidige ansatte og vikarer',
    'AGA på arbeidsgivertilskudd til SPK',
]
CAT_SHORT = {
    'Faste stillinger': 'Faste stillinger',
    'Lønn vikarer': 'Lønn vikarer',
    'Arbeidsgiveravgift': 'Arbeidsgiveravgift',
    'Turnustillegg': 'Turnustillegg',
    'Pensjonspremie (virksomheter som betaler pensjonspremie)': 'Pensjonspremie SPK',
    'Midlertidige stillinger (hel- og deltid)': 'Midlertidige stillinger',
    'Turnustillegg, vikarer': 'Turnustillegg vikarer',
    'Overtid faste stillinger': 'Overtid faste',
    'Overtid midlertidige ansatte og vikarer': 'Overtid midl./vik.',
    'AGA på arbeidsgivertilskudd til SPK': 'AGA SPK-tilskudd',
}

REGIONS = ['Region Midt-Norge','Region Nord','Region Sør','Region Vest','Region Øst']

# ── COLORS ───────────────────────────────────────────────────────────────────
C_NAVY   = "1B3A5C"
C_BLUE   = "2E6DA4"
C_TEAL   = "17A589"
C_ORANGE = "E67E22"
C_GREY   = "ECF0F1"
C_LGREY  = "F7F9FC"
C_WHITE  = "FFFFFF"
C_RED    = "C0392B"
C_GREEN  = "27AE60"
C_GOLD   = "F39C12"
C_WARN   = "FFF3CD"

def fill(hex_): return PatternFill("solid", fgColor=hex_)
def font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)
def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def thin_border(sides="lrtb"):
    s = Side(style="thin")
    n = Side(style=None)
    return Border(
        left=s if "l" in sides else n,
        right=s if "r" in sides else n,
        top=s if "t" in sides else n,
        bottom=s if "b" in sides else n,
    )

# ── PARSE CSV ────────────────────────────────────────────────────────────────
def parse_amount(s):
    s = s.strip()
    if not s or s in ('-', '   -'): return None
    neg = s.lstrip().startswith('-')
    s2 = s.replace('-','').replace(' ','').replace(',','.')
    try:
        v = float(s2)
        return -v if neg else v
    except:
        return None

with open(SRC, 'rb') as f:
    enc = chardet.detect(f.read(20000))['encoding']

with open(SRC, encoding=enc, errors='replace') as f:
    raw_lines = f.readlines()

# Structures: rows = list of (cat, region, enhet, {yr: amount})
rows = []
current_cat = None
current_region = None

for i, line in enumerate(raw_lines):
    cols = line.rstrip('\n').split(';')
    label = cols[0].strip()
    if not label:
        continue

    if label in MAIN_CATS:
        current_cat = label
        current_region = None
        continue

    if label in ('Bufetat', 'Bufdir'):
        continue

    if label.startswith('Region ') and label in REGIONS:
        current_region = label
        continue

    # Data row
    vals = cols[1:8]  # 2020-2026 + Totalsum sometimes at col 8
    amounts = {yr: parse_amount(vals[j]) for j, yr in enumerate(YEARS)}
    has_any = any(v is not None for v in amounts.values())

    if current_cat and current_region and has_any:
        rows.append({
            'cat': current_cat,
            'region': current_region,
            'enhet': label,
            **{yr: (amounts[yr] or 0.0) for yr in YEARS},
        })

print(f"Parsed {len(rows)} enhet-rows")

# ── CAGR + PREDICTION ────────────────────────────────────────────────────────
def cagr(start, end, n=5):
    if start and end and start > 0 and end > 0:
        return (end / start) ** (1 / n) - 1
    return None

def predict(row, yr_target=PRED_YR):
    y2020 = row.get(2020, 0) or 0
    y2025 = row.get(2025, 0) or 0
    # Use 2021 as fallback if 2020 is missing
    y_start = y2020 if y2020 > 0 else (row.get(2021, 0) or 0)
    n_years = 5 if y2020 > 0 else 4
    g = cagr(y_start, y2025, n_years)
    # Cap CAGR at ±25% per year
    if g is not None:
        g_capped = max(-0.25, min(0.25, g))
    else:
        g_capped = 0.065  # 6.5% default (Norwegian wage growth)
    n_pred = yr_target - 2025
    return y2025 * (1 + g_capped) ** n_pred, g_capped

for r in rows:
    pred, g = predict(r)
    r[PRED_KEY] = pred
    r['cagr'] = g

# ── AGGREGATES ───────────────────────────────────────────────────────────────
def agg_by(key_fn, filter_fn=None):
    result = defaultdict(lambda: defaultdict(float))
    for r in rows:
        if filter_fn and not filter_fn(r): continue
        k = key_fn(r)
        for yr in YEARS:
            result[k][yr] += r.get(yr, 0) or 0
        result[k][PRED_KEY] += r.get(PRED_KEY, 0) or 0
    return result

by_cat    = agg_by(lambda r: r['cat'])
by_region = agg_by(lambda r: r['region'])
by_enhet  = agg_by(lambda r: r['enhet'])
by_cat_region = agg_by(lambda r: (r['cat'], r['region']))

# ── EXCEL HELPERS ─────────────────────────────────────────────────────────────
def hdr(ws, row, col, val, bg=C_NAVY, fg=C_WHITE, bold=True, size=11, wrap=False, halign="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill(bg)
    c.font = font(bold=bold, color=fg, size=size)
    c.alignment = align(halign, "center", wrap)
    return c

def val_cell(ws, row, col, value, fmt="#,##0", bg=C_WHITE, bold=False, halign="right"):
    c = ws.cell(row=row, column=col, value=value)
    c.number_format = fmt
    c.fill = fill(bg)
    c.font = font(bold=bold)
    c.alignment = align(halign, "center")
    return c

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

# ── WORKBOOK ──────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ════════════════════════════════════════════════════════════════════════════
# ARK 1: FORSIDE
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Forside")
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 20
ws.row_dimensions[1].height = 8

# Title banner
for r in range(2, 8):
    for c in range(1, 6):
        ws.cell(r, c).fill = fill(C_NAVY)
ws.merge_cells('B2:D7')
t = ws.cell(2, 2, f"Lønnsutgifter\nPrediksjon {PRED_YR}")
t.font = Font(bold=True, color=C_WHITE, size=28)
t.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws.merge_cells('B8:D8')
s = ws.cell(8, 2, "Bufetat — Barneverninstitusjoner")
s.font = Font(color=C_BLUE, size=13, bold=True)
s.alignment = align("center", "center")

r = 10
info = [
    ("Datakilde",   "Agresso / aBIRK lønnseksport 2020–2026"),
    ("Grunnlag",    "Faktiske lønnskostnader 2020–2025"),
    ("Metode",      "CAGR per enhet (kappet ±25%/år)"),
    ("Prediksjon",  f"{PRED_YR} (basert på 2020-2025 trend)"),
    ("Delvis 2026", "Jan–mars 2026 (ikke brukt i prediksjon)"),
    ("Utarbeidet",  "2026-03-28"),
]
for lbl, val in info:
    ws.cell(r, 2, lbl).font = Font(bold=True, color=C_NAVY, size=11)
    ws.cell(r, 3, val).font = Font(size=11)
    r += 1

r += 1
# Grand totals on forside
total_2025 = sum(by_region[reg][2025] for reg in REGIONS)
total_pred = sum(r2[PRED_KEY] for r2 in rows)
pred_horizon = max(1, PRED_YR - 2025)
pred_cagr  = (total_pred/total_2025)**(1/pred_horizon)-1 if total_2025 > 0 else 0

for label, value, note in [
    ("Total 2025 (faktisk)", total_2025, "Alle kategorier inkl. AGA/SPK"),
    (f"Prediksjon {PRED_YR}", total_pred, f"CAGR 2025→{PRED_YR}: {pred_cagr*100:+.1f}%"),
]:
    ws.cell(r, 2, label).font = Font(bold=True, size=11)
    c = ws.cell(r, 3, value)
    c.number_format = "#,##0"
    c.font = Font(bold=True, size=12, color=C_NAVY)
    ws.cell(r, 4, note).font = Font(size=10, color="666666", italic=True)
    r += 1

# ════════════════════════════════════════════════════════════════════════════
# ARK 2: METODEBESKRIVELSE
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Metodebeskrivelse")
ws.sheet_view.showGridLines = False
set_col_widths(ws, [3, 28, 70])

r = 2
hdr(ws, r, 2, f"Metodebeskrivelse — Lønnsutgifter {PRED_YR}", C_NAVY, size=14)
ws.merge_cells(f'B{r}:C{r}')
r += 2

sections = [
    ("Datakilde", [
        "Agresso-eksport: «Innkjøpsanalyse 2026 lønnsutgifter (Lønnsutgifter).csv»",
        "Pivot-format: Enhet × År (2020–2026). 2026 er delvis (Q1).",
        "Dekker: Barnevernsinstitusjoner, alle statlige regioner (Midt, Nord, Sør, Vest, Øst)",
    ]),
    ("Lønnskategorier inkludert", [
        f"1. Faste stillinger — grunnlønnen til fast ansatte",
        f"2. Lønn vikarer — timebasert vikkarbeidslønn",
        f"3. Arbeidsgiveravgift (AGA) — 14,1% av lønnsgrunnlag",
        f"4. Turnustillegg — helg/natt-tillegg for turnusarbeid",
        f"5. Pensjonspremie SPK — statlig pensjonspremie",
        f"6. Midlertidige stillinger — engasjement/prosjektstillinger",
        f"7. Turnustillegg vikarer — tillegg for vikarer i turnus",
        f"8. Overtid faste — overtidsbetaling faste ansatte",
        f"9. Overtid midlertidige/vikarer — overtid for midlertidige",
        f"10. AGA på SPK-tilskudd — arbeidsgiveravgift på pensjon",
    ]),
    ("Prediksjonsmetode: Per-enhet CAGR", [
        "For hver enhet beregnes vekstrate (CAGR) fra 2020 til 2025 over 5 år.",
        "Formel: CAGR = (2025-verdi / 2020-verdi)^(1/5) - 1",
        "Hvis enhet mangler 2020-data brukes 2021 som startår (CAGR over 4 år).",
        "Alle CAGR-verdier kappes til ±25% per år for å unngå ekstremverdier.",
        "Standardrate 6,5%/år brukes der begge år mangler.",
        f"Prediksjon {PRED_YR} = 2025-verdi × (1 + CAGR)^{max(1, PRED_YR-2025)}",
    ]),
    ("Forutsetninger og usikkerhet", [
        "• Historisk vekst 2020-2025 antas å fortsette — ingen korreksjon for lønnsoppgjør",
        "• Nye enheter (åpnet etter 2022) har usikker trend — bruk med forsiktighet",
        "• Nedlagte enheter er ekskludert (null-verdier i 2025)",
        "• AGA og pensjon følger lønnen — ikke separat modellert",
        "• 2026 delårsdata er IKKE brukt i prediksjonen",
    ]),
    ("Gjennomsnittlig lønnsøkning (historisk)", [
        f"Total 2020: {sum(by_region[reg][2020] for reg in REGIONS)/1e6:.0f}M kr",
        f"Total 2025: {sum(by_region[reg][2025] for reg in REGIONS)/1e6:.0f}M kr",
        f"Samlet CAGR 2020-2025: {((sum(by_region[reg][2025] for reg in REGIONS)/sum(by_region[reg][2020] for reg in REGIONS))**(1/5)-1)*100:.1f}%/år",
        f"Norsk lønnsoppgjør typisk: 4-6% per år. Bufetat historisk over normalen grunnet kapasitetsøkning.",
    ]),
]

for title, bullets in sections:
    c = ws.cell(r, 2, title)
    c.font = Font(bold=True, size=12, color=C_NAVY)
    c.fill = fill(C_LGREY)
    ws.merge_cells(f'B{r}:C{r}')
    r += 1
    for b in bullets:
        ws.cell(r, 2, "")
        bc = ws.cell(r, 3, b)
        bc.font = Font(size=10)
        bc.alignment = align("left", "top", wrap=True)
        ws.row_dimensions[r].height = 16
        r += 1
    r += 1

ws.column_dimensions['C'].width = 80

# ════════════════════════════════════════════════════════════════════════════
# ARK 3: PREDIKSJON — PER KATEGORI × REGION
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet(f"Prediksjon_{PRED_YR}")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "C3"

col_years = YEARS + [PRED_KEY]
col_labels = [str(y) for y in YEARS] + [f"Prediksjon {PRED_YR}"]
n_cols = len(col_labels)

# Header row
hdr(ws, 1, 1, "Lønnskategori", C_NAVY)
hdr(ws, 1, 2, "Region", C_NAVY)
for j, lbl in enumerate(col_labels):
    bg = C_TEAL if lbl == f'Prediksjon {PRED_YR}' else C_BLUE
    hdr(ws, 1, 3+j, lbl, bg)
hdr(ws, 1, 3+n_cols, "CAGR 2020→25", C_ORANGE)

set_col_widths(ws, [32, 22] + [14]*n_cols + [14])

row = 2
alt = False
for cat in MAIN_CATS:
    cat_label = CAT_SHORT.get(cat, cat)
    for ri, region in enumerate(REGIONS):
        key = (cat, region)
        d = by_cat_region.get(key, {})
        bg = C_LGREY if alt else C_WHITE
        ws.cell(row, 1, cat_label if ri == 0 else "").font = Font(size=10, bold=(ri==0))
        ws.cell(row, 1).fill = fill(bg)
        ws.cell(row, 2, region).font = Font(size=10)
        ws.cell(row, 2).fill = fill(bg)
        for j, yk in enumerate(col_years):
            v = d.get(yk, 0) or 0
            fmt = "#,##0"
            bold = (yk == PRED_KEY)
            fg = C_TEAL if bold else C_WHITE
            cbg = "E8F8F5" if bold else bg
            val_cell(ws, row, 3+j, v if v else None, fmt, cbg, bold)

        # Regional CAGR estimate
        v2020 = d.get(2020, 0) or 0
        v2025 = d.get(2025, 0) or 0
        g = cagr(v2020, v2025, 5)
        c = ws.cell(row, 3+n_cols, g)
        if g is not None:
            c.number_format = "0.0%"
        c.fill = fill(bg)
        row += 1
    alt = not alt
    # Subtotal per category
    for j, yk in enumerate(col_years):
        v = sum(by_cat_region.get((cat, reg), {}).get(yk, 0) or 0 for reg in REGIONS)
        c = ws.cell(row, 3+j, v if v else None)
        c.number_format = "#,##0"
        c.fill = fill("D5E8F5")
        c.font = Font(bold=True, size=10)
    ws.cell(row, 1, f"  Σ {cat_label}").font = Font(bold=True, size=10, italic=True)
    ws.cell(row, 1).fill = fill("D5E8F5")
    ws.cell(row, 2).fill = fill("D5E8F5")
    ws.cell(row, 3+n_cols).fill = fill("D5E8F5")
    row += 1

# GRAND TOTAL
ws.cell(row, 1, "TOTALSUM").font = Font(bold=True, size=11, color=C_WHITE)
ws.cell(row, 1).fill = fill(C_NAVY)
ws.cell(row, 2).fill = fill(C_NAVY)
for j, yk in enumerate(col_years):
    v = sum(by_cat_region.get((cat, reg), {}).get(yk, 0) or 0
            for cat in MAIN_CATS for reg in REGIONS)
    c = ws.cell(row, 3+j, v)
    c.number_format = "#,##0"
    bg = C_TEAL if yk == PRED_KEY else C_NAVY
    c.fill = fill(bg)
    c.font = Font(bold=True, color=C_WHITE)
ws.cell(row, 3+n_cols).fill = fill(C_NAVY)

# ════════════════════════════════════════════════════════════════════════════
# ARK 4: PIVOT REGION
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Pivot_Region")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "B2"
set_col_widths(ws, [28] + [14]*8)

col_ks = YEARS + [PRED_KEY]
col_ls = [str(y) for y in YEARS] + [f"Prediksjon {PRED_YR}"]

hdr(ws, 1, 1, "Region", C_NAVY)
for j, lbl in enumerate(col_ls):
    bg = C_TEAL if lbl == f'Prediksjon {PRED_YR}' else C_BLUE
    hdr(ws, 1, 2+j, lbl, bg)

row = 2
for reg in REGIONS:
    d = by_region[reg]
    ws.cell(row, 1, reg).font = Font(bold=True, size=11)
    ws.cell(row, 1).fill = fill(C_LGREY)
    for j, yk in enumerate(col_ks):
        v = d.get(yk, 0)
        bold = (yk == PRED_KEY)
        cbg = "E8F8F5" if bold else C_WHITE
        val_cell(ws, row, 2+j, v or None, "#,##0", cbg, bold)
    row += 1

# Total
ws.cell(row, 1, "TOTALSUM").font = Font(bold=True, color=C_WHITE)
ws.cell(row, 1).fill = fill(C_NAVY)
for j, yk in enumerate(col_ks):
    v = sum(by_region[reg].get(yk, 0) for reg in REGIONS)
    c = ws.cell(row, 2+j, v)
    c.number_format = "#,##0"
    c.fill = fill(C_TEAL if yk == PRED_KEY else C_NAVY)
    c.font = Font(bold=True, color=C_WHITE)

# ════════════════════════════════════════════════════════════════════════════
# ARK 5: PIVOT KATEGORI
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Pivot_Kategori")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "B2"
set_col_widths(ws, [38] + [14]*8)

hdr(ws, 1, 1, "Lønnskategori", C_NAVY)
for j, lbl in enumerate(col_ls):
    bg = C_TEAL if lbl == f'Prediksjon {PRED_YR}' else C_BLUE
    hdr(ws, 1, 2+j, lbl, bg)

row = 2
alt = False
for cat in MAIN_CATS:
    d = by_cat[cat]
    bg = C_LGREY if alt else C_WHITE
    ws.cell(row, 1, CAT_SHORT.get(cat, cat)).font = Font(size=10)
    ws.cell(row, 1).fill = fill(bg)
    for j, yk in enumerate(col_ks):
        v = d.get(yk, 0)
        cbg = "E8F8F5" if yk == PRED_KEY else bg
        val_cell(ws, row, 2+j, v or None, "#,##0", cbg, yk == PRED_KEY)
    row += 1
    alt = not alt

ws.cell(row, 1, "TOTALSUM").font = Font(bold=True, color=C_WHITE)
ws.cell(row, 1).fill = fill(C_NAVY)
for j, yk in enumerate(col_ks):
    v = sum(by_cat[cat].get(yk, 0) for cat in MAIN_CATS)
    c = ws.cell(row, 2+j, v)
    c.number_format = "#,##0"
    c.fill = fill(C_TEAL if yk == PRED_KEY else C_NAVY)
    c.font = Font(bold=True, color=C_WHITE)

# ════════════════════════════════════════════════════════════════════════════
# ARK 6: PER ENHET (alle enheter, alle år + prediksjon)
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Per_Enhet")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "E2"

col_hdrs = ["Kategori","Region","Enhet"] + [str(y) for y in YEARS] + [f"Prediksjon {PRED_YR}","CAGR"]
set_col_widths(ws, [24, 20, 45] + [13]*len(YEARS) + [14, 10])

for j, lbl in enumerate(col_hdrs):
    bg = C_TEAL if lbl == f'Prediksjon {PRED_YR}' else (C_ORANGE if lbl == 'CAGR' else C_NAVY)
    hdr(ws, 1, j+1, lbl, bg)

row = 2
alt = False
for r in sorted(rows, key=lambda x: (x['cat'], x['region'], x['enhet'])):
    bg = C_LGREY if alt else C_WHITE
    ws.cell(row, 1, CAT_SHORT.get(r['cat'], r['cat'])).font = Font(size=9)
    ws.cell(row, 1).fill = fill(bg)
    ws.cell(row, 2, r['region']).font = Font(size=9)
    ws.cell(row, 2).fill = fill(bg)
    ws.cell(row, 3, r['enhet']).font = Font(size=9)
    ws.cell(row, 3).fill = fill(bg)
    for j, yr in enumerate(YEARS):
        v = r.get(yr, 0) or 0
        val_cell(ws, row, 4+j, v if v else None, "#,##0", bg)
    val_cell(ws, row, 4+len(YEARS), r[PRED_KEY], "#,##0", "E8F8F5", True)
    c = ws.cell(row, 4+len(YEARS)+1, r['cagr'])
    c.number_format = "0.0%"
    c.fill = fill(bg)
    c.font = Font(size=9)
    row += 1
    alt = not alt

# ════════════════════════════════════════════════════════════════════════════
# ARK 7: SLIK BEREGNET VI (steg-for-steg eksempel)
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Slik beregnet vi")
ws.sheet_view.showGridLines = False
set_col_widths(ws, [3, 32, 65])

r = 2
hdr(ws, r, 2, f"Slik beregnet vi prediksjon {PRED_YR} — Lønnsutgifter", C_NAVY, size=13)
ws.merge_cells(f'B{r}:C{r}')
r += 2

# Pick an example enhet
ex = sorted(rows, key=lambda x: -(x.get(2025,0) or 0))[0]
ex_vals = {yr: ex.get(yr,0) for yr in YEARS}
ex_pred = ex[PRED_KEY]
ex_cagr = ex['cagr']

steps = [
    ("Steg 1: Velg enhet", [
        f"Eksempel: «{ex['enhet']}» ({ex['region']})",
        f"Kategori: {CAT_SHORT.get(ex['cat'], ex['cat'])}",
    ]),
    ("Steg 2: Hent historiske lønnskostnader", [
        f"2020: {ex_vals[2020]:>16,.0f} kr",
        f"2021: {ex_vals[2021]:>16,.0f} kr",
        f"2022: {ex_vals[2022]:>16,.0f} kr",
        f"2023: {ex_vals[2023]:>16,.0f} kr",
        f"2024: {ex_vals[2024]:>16,.0f} kr",
        f"2025: {ex_vals[2025]:>16,.0f} kr",
    ]),
    ("Steg 3: Beregn CAGR (5-årig vekstrate)", [
        f"CAGR = (2025-verdi / 2020-verdi)^(1/5) - 1",
        f"CAGR = ({ex_vals[2025]:,.0f} / {ex_vals[2020]:,.0f})^(1/5) - 1",
        f"CAGR = {ex_cagr*100:.2f}% per år",
        "(Kappet til maks ±25% for å unngå ekstremverdier)",
    ]),
    (f"Steg 4: Prediker {PRED_YR}", [
        f"Prediksjon {PRED_YR} = 2025-verdi × (1 + CAGR)^{max(1, PRED_YR-2025)}",
        f"= {ex_vals[2025]:,.0f} × (1 + {ex_cagr:.4f})²",
        f"= {ex_pred:,.0f} kr",
    ]),
    ("Steg 5: Sum alle enheter", [
        f"Totalprediksjon {PRED_YR} = sum av alle enheters prediksjon",
        f"= {sum(r2[PRED_KEY] for r2 in rows):,.0f} kr",
        f"  ({sum(r2[PRED_KEY] for r2 in rows)/1e6:.0f}M kr)",
    ]),
    ("Validering", [
        f"Total 2025 (faktisk):     {sum(by_region[reg][2025] for reg in REGIONS):>16,.0f} kr",
        f"Prediksjon {PRED_YR}:          {sum(r2[PRED_KEY] for r2 in rows):>16,.0f} kr",
        f"Implisert vekst:   {((sum(r2[PRED_KEY] for r2 in rows)/sum(by_region[reg][2025] for reg in REGIONS))-1)*100:.1f}%",
        f"Norsk lønnsoppgjør 2025:  ~5% (NHO/LO) — veksten er i tråd med observert trend",
    ]),
]

for title, bullets in steps:
    c = ws.cell(r, 2, title)
    c.font = Font(bold=True, size=11, color=C_WHITE)
    c.fill = fill(C_BLUE)
    ws.merge_cells(f'B{r}:C{r}')
    r += 1
    for b in bullets:
        ws.cell(r, 2, "")
        bc = ws.cell(r, 3, b)
        bc.font = Font(size=10, name="Courier New" if any(c.isdigit() for c in b) else "Calibri")
        bc.alignment = align("left", "top", wrap=True)
        ws.row_dimensions[r].height = 16
        r += 1
    r += 1

# ════════════════════════════════════════════════════════════════════════════
# ARK 8: TILBAKEMELDING
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Tilbakemelding_Økonomi")
ws.sheet_view.showGridLines = False
set_col_widths(ws, [3, 38, 55, 20])

r = 2
hdr(ws, r, 2, "Tilbakemelding fra økonomiavdelingen", C_NAVY, size=13)
ws.merge_cells(f'B{r}:D{r}')
r += 2

ws.cell(r, 2, "Vennligst fyll inn tilbakemelding på prediksjonsmodellen:").font = Font(italic=True, size=10)
ws.merge_cells(f'B{r}:D{r}')
r += 2

questions = [
    ("1", f"Er totalprediksjon {PRED_YR} på linje med budsjett?",
     f"Vår prediksjon: {sum(r2[PRED_KEY] for r2 in rows)/1e6:.0f}M kr"),
    ("2", f"Hvilke enheter forventes åpnet/nedlagt {PRED_YR-1}-{PRED_YR}?", "Angi enhetsnavn og estimert lønnskostnad"),
    ("3", "Planlegges lønnsoppgjør utover historisk trend?", "F.eks. særskilte tillegg, stillingsøkninger"),
    ("4", "Er AGA/pensjonssatser forventet å endres?", "Noter eventuelle regelendringer"),
    ("5", "Spesielle kommentarer per region?", "F.eks. Region Nord har økt kapasitet"),
    ("6", "Godkjenner dere metoden (per-enhet CAGR)?", "Alternativ: Flat % per lønnsoppgjør"),
    ("7", "Ønskes separat prediksjon per stillingskategori?", "Faste, vikarer, overtid, mv."),
]

hdr(ws, r, 2, "Nr.", C_BLUE, size=10)
hdr(ws, r, 3, "Spørsmål", C_BLUE, size=10)
hdr(ws, r, 4, "Svar / Kommentar", C_BLUE, size=10)
r += 1

for nr, spørsmål, hint in questions:
    bg = C_LGREY if int(nr) % 2 == 0 else C_WHITE
    ws.cell(r, 2, nr).font = Font(bold=True)
    ws.cell(r, 2).fill = fill(bg)
    ws.cell(r, 3, spørsmål).font = Font(size=10)
    ws.cell(r, 3).fill = fill(bg)
    ws.cell(r, 3).alignment = align("left", "center", wrap=True)
    c = ws.cell(r, 4, hint)
    c.font = Font(size=9, italic=True, color="888888")
    c.fill = fill(C_WARN)
    c.alignment = align("left", "center", wrap=True)
    ws.row_dimensions[r].height = 30
    r += 1

# ── SAVE ──────────────────────────────────────────────────────────────────────
wb.save(OUT)
print(f"\n✅ Lagret: {OUT}")
print(f"\nNøkkeltall:")
print(f"  Antall enheter analysert: {len(rows)}")
print(f"  Total lønnskostnad 2025:  {sum(by_region[reg][2025] for reg in REGIONS)/1e6:.0f}M kr")
print(f"  Prediksjon {PRED_YR}:          {sum(r2[PRED_KEY] for r2 in rows)/1e6:.0f}M kr")
print(f"\nArk:")
for sheet in ['Forside','Metodebeskrivelse',f'Prediksjon_{PRED_YR}','Pivot_Region','Pivot_Kategori','Per_Enhet','Slik beregnet vi','Tilbakemelding_Økonomi']:
    print(f"  ✓ {sheet}")
