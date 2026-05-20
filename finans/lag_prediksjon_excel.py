"""
Lager Excel-fil for økonomisjekk av prediksjon (valgbart år).
Kjør: python3 finans/lag_prediksjon_excel.py --out-year 2026

Metodikk (CAGR, datakilde, forskjell fra backend): finans/METODE_Prediksjon_Økonomi.md
"""
import argparse
import csv
from collections import defaultdict
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.formatting.rule import FormulaRule

# ── Konstanter ──────────────────────────────────────────────────────────────
CSV_FIL  = Path(__file__).parent / "Eiendom 202001 til 202512 til Øystein(AGRESSO).csv"
parser = argparse.ArgumentParser(description="Lag økonomiprediksjon Excel")
parser.add_argument("--out-year", type=int, default=2027, help="År i filnavn (default: 2027)")
parser.add_argument(
    "--no-protect",
    action="store_true",
    help="Ikke lås formelkolonner (I/L m.m.) — kun for feilsøking; risiko for ødelagte formler",
)
args = parser.parse_args()
UT_FIL   = Path(__file__).parent / f"Prediksjon_{args.out_year}_Økonomi.xlsx"
PRED_YEAR = args.out_year
PRED_PREV_YEAR = PRED_YEAR - 1
PRED_SHEET_NAME = f"Prediksjon_{PRED_YEAR}"

RENE_BA  = {"IV", "IW", "LE", "H1", "H2", "HB", "RE"}
ÅR_HIST  = ["2022", "2023", "2024", "2025"]
ALLE_ÅR  = ["2020", "2021", "2022", "2023", "2024", "2025"]
AGGREGAT = {"204416", "500401", "300200"}

# SRS-kategorisering basert på kontokode
ANLEGG_INTERVALLER = [(1040, 1298), (3800, 3810), (4930, 4999), (6000, 6071)]
ANLEGG_ENKELT      = {6551, 7800}

def bestem_srs(konto_str):
    try:
        k = int(konto_str.strip())
    except (ValueError, AttributeError):
        return "Annet"
    if any(lo <= k <= hi for lo, hi in ANLEGG_INTERVALLER) or k in ANLEGG_ENKELT:
        return "SRS 17"
    if 5000 <= k <= 5999 or k == 7100:
        return "SRS 25"
    if k >= 3000:
        return "SRS 10"
    return "Annet"

BA_FORKLARING = {
    "MV": "MV = Motverdi / viderefakturert — sjekk om kostnaden gjelder et annet koststed",
    "MP": "MP = Motpost til MV-bilag — netter normalt mot tilhørende MV-bilag",
}

# Farger
BLÅ_TEKST   = "0000FF"
SORT_TEKST  = "000000"
GRÅ_BG      = "F2F2F2"
HEADER_BG   = "1F3864"
HEADER_FG   = "FFFFFF"
AKSENT_BG   = "D9E1F2"
GULL_BG     = "FFF2CC"
RØD_BG      = "FFE0E0"
GRØNN_BG    = "E2EFDA"
HVIT        = "FFFFFF"

# Talformat
KR_FMT      = '#,##0;(#,##0);"-"'
KR_M_FMT    = '#,##0.0;(#,##0.0);"-"'  # millioner
PCT_FMT     = "0.0%;(0.0%);-"
INT_FMT     = "#,##0"

def rens(raw):
    if not raw or not raw.strip(): return 0.0
    s = raw.strip()
    neg = s.startswith("(") and s.endswith(")")
    if neg: s = s[1:-1]
    s = s.replace(" ", "").replace(",", ".")
    try:
        v = float(s); return -v if neg else v
    except: return 0.0

def kant(thin=True):
    s = Side(style="thin" if thin else "medium")
    return Border(left=s, right=s, top=s, bottom=s)

def header_cell(ws, row, col, tekst, bg=HEADER_BG, fg=HEADER_FG, bold=True, center=True, wrap=False):
    c = ws.cell(row=row, column=col, value=tekst)
    c.font = Font(name="Arial", bold=bold, color=fg, size=10)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center" if center else "left",
                             vertical="center", wrap_text=wrap)
    c.border = kant()
    return c

def data_cell(ws, row, col, value=None, fmt=None, bold=False, bg=None,
              color=SORT_TEKST, center=False, formula=None):
    c = ws.cell(row=row, column=col, value=formula if formula else value)
    c.font = Font(name="Arial", bold=bold, color=color, size=10)
    if bg: c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center" if center else "right",
                             vertical="center")
    c.border = kant()
    if fmt: c.number_format = fmt
    return c


def _apply_high_growth_conditional(
    ws,
    first_row: int,
    last_row: int,
    col_2025: int,
    col_pred: int,
) -> None:
    """Gul bakgrunn (B:L) når predikert år / 2025-aktual > 1,25 — typisk sjekkliste."""
    if last_row < first_row:
        return
    c_base = get_column_letter(col_2025)
    c_pred = get_column_letter(col_pred)
    rng = f"B{first_row}:L{last_row}"
    r0 = first_row
    formula = f"AND(${c_base}{r0}>0,${c_pred}{r0}/${c_base}{r0}>1.25)"
    fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ws.conditional_formatting.add(rng, FormulaRule(formula=[formula], fill=fill))


# ── Les CSV ──────────────────────────────────────────────────────────────────
print("Leser CSV …")
rader        = []
mv_mp_rader  = []
alle_rader   = []   # alle år, RENE_BA (for SRS_Oversikt)
anlegg_rader = []   # anleggskonto-poster (for Anleggsregister)

with open(CSV_FIL, encoding="latin-1") as f:
    reader = csv.DictReader(f, delimiter=";")
    for row in reader:
        ar    = row.get("År","").strip()
        ba    = row.get("BA","").strip()
        konto = row.get("Konto","").strip()
        if ba in RENE_BA and ar in ÅR_HIST:
            rader.append(row)
        if ba in RENE_BA and ar in ALLE_ÅR:
            alle_rader.append(row)
        if ba in ("MV","MP") and ar == "2025":
            mv_mp_rader.append(row)
        if bestem_srs(konto) == "SRS 17" and ar in ALLE_ÅR:
            anlegg_rader.append(row)

print(f"  {len(rader)} rene rader  |  {len(mv_mp_rader)} MV/MP-rader  |  {len(anlegg_rader)} anlegg-rader")

# ── Aggregér data ────────────────────────────────────────────────────────────
def agg(rader, key_fn):
    d = defaultdict(lambda: defaultdict(float))
    for r in rader:
        ar = r.get("År","").strip()
        k  = key_fn(r)
        d[k][ar] += rens(r.get("Beløp",""))
    return d

uk_data    = agg(rader, lambda r: r.get("Underkategorier(T)","").strip())
reg_data   = agg(rader, lambda r: r.get("Region","").strip())
konto_data = agg(rader, lambda r: (r.get("Konto","").strip(), r.get("Konto(T)","").strip()))
lev_data   = agg(rader, lambda r: (r.get("Resk.nr","").strip(), r.get("Resk.nr(T)","").strip()))
dim1_data  = agg(rader, lambda r: (r.get("Dim1","").strip(), r.get("Dim1(T)","").strip()))
ba_data    = agg(rader, lambda r: r.get("BA","").strip())

UNDERKATEGORIER = ["Husleie", "Drift- og vedlikeholdskostnader",
                   "Utskiftings- og utviklingskostnader",
                   "Forsyningskostnader", "Renholdstjenester"]
REGIONENE = ["Øst","Vest","Sør","Midt","Nord","Bufdir"]

# Robust CAGR: bruker 2022→2024 som standard, men faller tilbake på
# 2023→2025 dersom 2022 er en statistisk outlier (> 1.5× medianen av 2023-2025).
# Forklaring: engangs-faktureringer (f.eks. store Statsbygg-oppgjør) i 2022
# gir kunstig negativ trend som undervurderer fremtidige kostnader.
def _er_outlier_2022(verdier_2022_2025: list, terskel: float = 1.3) -> bool:
    """Returnerer True hvis 2022-verdien er mer enn terskel × medianen av 2023–2025."""
    v22, v23, v24, v25 = verdier_2022_2025
    etterfølgende = sorted([v23, v24, v25])
    median_etter = etterfølgende[1]          # median av de tre neste
    return median_etter > 0 and v22 > terskel * median_etter


def robust_cagr(v_start, v_slutt, år_diff: int = 2) -> float:
    """CAGR over år_diff år. Returnerer 0 hvis manglende data."""
    if v_start > 0 and v_slutt > 0:
        return (v_slutt / v_start) ** (1 / år_diff) - 1
    return 0.0


# CAGR per underkategori (med outlier-korreksjon for 2022)
def cagr_22_24(uk):
    verdier = [uk_data[uk].get(ar, 0) for ar in ÅR_HIST]  # [2022,2023,2024,2025]
    if _er_outlier_2022(verdier):
        # Bruk 2023→2025 kun hvis det gir en MINDRE NEGATIV rate enn 2022→2024.
        # Forhindrer at korreksjon forverrer situasjonen (f.eks. der 2023 var en topp).
        cagr_alt = robust_cagr(verdier[1], verdier[3], år_diff=2)
        cagr_std = robust_cagr(verdier[0], verdier[2], år_diff=2)
        if cagr_alt > cagr_std:
            return cagr_alt
    return robust_cagr(verdier[0], verdier[2], år_diff=2)

# Total per år
totaler = {ar: sum(uk_data[uk].get(ar,0) for uk in UNDERKATEGORIER) for ar in ÅR_HIST}

# ── Bygg Excel ───────────────────────────────────────────────────────────────
wb = Workbook()

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 – FORSIDE
# ════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Forside"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 36
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 16
ws.column_dimensions["E"].width = 16
ws.column_dimensions["F"].width = 16
ws.column_dimensions["G"].width = 16
ws.column_dimensions["H"].width = 16
ws.column_dimensions["I"].width = 14
ws.row_dimensions[1].height = 8

# Tittel
ws.merge_cells("B2:I2")
c = ws["B2"]
c.value = f"Eiendomskostnader Bufetat — Prediksjon {PRED_YEAR}"
c.font = Font(name="Arial", bold=True, size=16, color=HEADER_BG)
c.alignment = Alignment(horizontal="left", vertical="center")

ws.merge_cells("B3:I3")
c = ws["B3"]
c.value = (
    "Grunnlag: Agresso GL, Innkjøpskategori 01 (Leie av lokaler og tilknyttede utgifter), 2020–2025.\n\n"
    f"Tabellen under er koblet til fanen «{PRED_SHEET_NAME}». Endre historikk og vekst der (og evt. «Antagelser») — "
    "ikke skriv inn tall direkte i kolonne C–F her, for da erstattes lenkene og prediksjonen oppdateres ikke som tenkt."
)
c.font = Font(name="Arial", size=10, color="595959")
c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
ws.row_dimensions[3].height = 56

ws.row_dimensions[4].height = 12

# ── Nøkkeltall-bokser (rad 5-8) ──
bokser = [
    ("B", "E", "Totalkostnad 2025\n(ren basis, ekskl. MV/MP)",
     totaler["2025"], "kr", GRØNN_BG),
    ("F", "I", f"Prediksjon {PRED_YEAR}\n(se '{PRED_SHEET_NAME}'-fane)",
     None, "Se ark →", AKSENT_BG),
]
for c1, c2, tittel, verdi, enhet, bg in bokser:
    ws.merge_cells(f"{c1}5:{c2}5")
    t = ws[f"{c1}5"]
    t.value = tittel
    t.font = Font(name="Arial", bold=True, size=10, color="404040")
    t.fill = PatternFill("solid", start_color=bg)
    t.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    t.border = kant()
    ws.row_dimensions[5].height = 32

    ws.merge_cells(f"{c1}6:{c2}6")
    v = ws[f"{c1}6"]
    if verdi is not None:
        v.value = verdi
        v.number_format = KR_FMT
        v.font = Font(name="Arial", bold=True, size=18, color=HEADER_BG)
    else:
        v.value = enhet
        v.font = Font(name="Arial", bold=True, size=14, color=HEADER_BG)
    v.fill = PatternFill("solid", start_color=bg)
    v.alignment = Alignment(horizontal="center", vertical="center")
    v.border = kant()
    ws.row_dimensions[6].height = 36

    ws.merge_cells(f"{c1}7:{c2}7")
    e = ws[f"{c1}7"]
    if verdi is not None: e.value = enhet
    e.fill = PatternFill("solid", start_color=bg)
    e.font = Font(name="Arial", size=9, color="595959")
    e.alignment = Alignment(horizontal="center")
    e.border = kant()

ws.row_dimensions[8].height = 10

# ── Historikk-tabell (rad 9–) ──
header_cell(ws, 9, 2, "Underkategori", bold=True, center=False)
for i, ar in enumerate(ÅR_HIST):
    header_cell(ws, 9, 3+i, ar)
header_cell(ws, 9, 7, "CAGR\n2022→2024", wrap=True)
header_cell(ws, 9, 8, "Andel 2025")
ws.row_dimensions[9].height = 28

total_25 = totaler["2025"]
for ri, uk in enumerate(UNDERKATEGORIER):
    row = 10 + ri
    bg = GRÅ_BG if ri % 2 == 0 else HVIT
    c = ws.cell(row=row, column=2, value=uk)
    c.font = Font(name="Arial", size=10)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = kant()
    for ci, ar in enumerate(ÅR_HIST):
        data_cell(ws, row, 3+ci, uk_data[uk].get(ar, 0), KR_FMT, bg=bg)
    v25 = uk_data[uk].get("2025", 0)
    cg = cagr_22_24(uk)
    data_cell(ws, row, 7, cg, PCT_FMT, bg=bg)
    data_cell(ws, row, 8, v25 / total_25 if total_25 else 0, PCT_FMT, bg=bg)

# Totallinje (midlertidige summer — overskrives med formler etter Prediksjon-ark + lenking)
tot_row = 10 + len(UNDERKATEGORIER)
c = ws.cell(row=tot_row, column=2, value="TOTAL")
c.font = Font(name="Arial", bold=True, size=10)
c.fill = PatternFill("solid", start_color=AKSENT_BG)
c.alignment = Alignment(horizontal="left", vertical="center")
c.border = kant()
for ci, ar in enumerate(ÅR_HIST):
    data_cell(ws, tot_row, 3 + ci, totaler[ar], KR_FMT, bold=True, bg=AKSENT_BG)
data_cell(ws, tot_row, 7, None, bg=AKSENT_BG)
data_cell(ws, tot_row, 8, 1.0, PCT_FMT, bold=True, bg=AKSENT_BG)

ws.row_dimensions[tot_row].height = 22

# Note om MV/MP
note_row = tot_row + 2
ws.merge_cells(f"B{note_row}:I{note_row}")
n = ws[f"B{note_row}"]
n.value = ("⚠️  MV/MP-posteringer 2025 (netto +18M) er IKKE inkludert i tabellen over. "
           "Se fane 'Flaggede_MV_MP' for detaljer. Scenario A (inkl. MV/MP) gir 2025-total på 567.5M.")
n.font = Font(name="Arial", size=9, color="7F0000", italic=True)
n.fill = PatternFill("solid", start_color=RØD_BG)
n.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
n.border = kant()
ws.row_dimensions[note_row].height = 28

# ════════════════════════════════════════════════════════════════════════════
# ARK «Antagelser» — globale brytere (innsettingsposisjon 1 = rett etter Forside)
# ════════════════════════════════════════════════════════════════════════════
ANTAG_SHEET = "Antagelser"
w_ant = wb.create_sheet(ANTAG_SHEET, 1)
w_ant.sheet_view.showGridLines = False
w_ant.column_dimensions["A"].width = 52
w_ant.column_dimensions["B"].width = 18
w_ant.merge_cells("A1:B1")
t = w_ant["A1"]
t.value = f"Globale justeringer (påvirker {PRED_SHEET_NAME} og Prediksjon_Per_Eiendom)"
t.font = Font(name="Arial", bold=True, size=12, color=HEADER_BG)
t.alignment = Alignment(horizontal="left", vertical="center")
w_ant["A3"] = (
    "Ekstra vekst legges inn etter rad-vekst (1+blå celle), som (1+ekstra). "
    "0 % = ingen effekt. Bruk f.eks. +2 % om du vil skru alle anslag opp likt."
)
w_ant["A3"].font = Font(name="Arial", size=9, italic=True, color="595959")
w_ant["A3"].alignment = Alignment(wrap_text=True)
w_ant.row_dimensions[3].height = 36

header_cell(w_ant, 5, 1, "Parameter", center=False)
header_cell(w_ant, 5, 2, "Verdi", center=True)
w_ant.row_dimensions[5].height = 22

def _antag_input_row(row: int, label: str, default: float) -> None:
    c = w_ant.cell(row=row, column=1, value=label)
    c.font = Font(name="Arial", size=10)
    c.fill = PatternFill("solid", start_color=GRÅ_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = kant()
    b = w_ant.cell(row=row, column=2, value=default)
    b.font = Font(name="Arial", color=BLÅ_TEKST, size=10)
    b.fill = PatternFill("solid", start_color=GULL_BG)
    b.number_format = PCT_FMT
    b.alignment = Alignment(horizontal="center", vertical="center")
    b.border = kant()


_antag_label1 = f"Ekstra vekst mot prediksjon {PRED_PREV_YEAR} (etter rad-vekst)"
_antag_label2 = f"Ekstra vekst mot prediksjon {PRED_YEAR} (etter rad-vekst)"
_antag_input_row(6, _antag_label1, 0.0)
_antag_input_row(7, _antag_label2, 0.0)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 – PREDIKSJON
# ════════════════════════════════════════════════════════════════════════════
wp = wb.create_sheet(PRED_SHEET_NAME)
wp.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGHIJKL",
                  [3, 36, 16, 16, 16, 16, 4, 16, 16, 4, 16, 16]):
    wp.column_dimensions[get_column_letter(ord(col)-64)].width = w

# Tittel
wp.merge_cells("B2:L2")
c = wp["B2"]
c.value = f"Prediksjon {PRED_YEAR} — Redigerbare forutsetninger"
c.font = Font(name="Arial", bold=True, size=14, color=HEADER_BG)
c.alignment = Alignment(horizontal="left", vertical="center")

wp.merge_cells("B3:L3")
c = wp["B3"]
c.value = (
    "Her endrer du tall som styrer prediksjonen: kolonne C–F (historikk) og H/K (vekst som %). "
    f"Globale ekstra: «{ANTAG_SHEET}». Kolonne I og L beregnes (formler) — de er låst; skriv ikke der. "
    "Betinget formatering (gul rad): prediksjon siste år (L) mer enn 25 % høyere enn 2025-faktisk (F). "
    "Forside-tabellen lenker hit. MV/MP ekskludert."
)
c.font = Font(name="Arial", size=9, color="595959", italic=True)
c.alignment = Alignment(horizontal="left")

wp.merge_cells("B4:L4")
c4 = wp["B4"]
c4.value = (
    "VIKTIG: Bruk Microsoft Excel (ikke Numbers). Under Formler → Beregningsalternativer skal «Automatisk» "
    "være på (ikke «Manuell»), ellers oppdateres ikke prediksjonene. "
    "Prosentfelt: skriv 5 % eller 0,05 — ikke «5» uten %-format."
)
c4.font = Font(name="Arial", size=9, color="7F0000", bold=True)
c4.fill = PatternFill("solid", start_color=GULL_BG)
c4.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
c4.border = kant()
wp.row_dimensions[4].height = 52

wp.row_dimensions[5].height = 8

# Header rad 6
cols = ["Underkategori", "2022\nAktual", "2023\nAktual", "2024\nAktual",
        "2025\nAktual", "", f"Veksttakt\n{PRED_PREV_YEAR} (blå)", f"Prediksjon\n{PRED_PREV_YEAR}",
        "", f"Veksttakt\n{PRED_YEAR} (blå)", f"Prediksjon\n{PRED_YEAR}"]
for ci, h in enumerate(cols):
    header_cell(wp, 6, 2+ci, h, wrap=True)
wp.row_dimensions[6].height = 32

# Veksttakt-celler lagres for referanse
vekst_26_celler = {}  # uk → excel-celleref
vekst_27_celler = {}
aktual_25_celler = {}

for ri, uk in enumerate(UNDERKATEGORIER):
    row = 7 + ri
    bg = GRÅ_BG if ri % 2 == 0 else HVIT

    c = wp.cell(row=row, column=2, value=uk)
    c.font = Font(name="Arial", size=10)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = kant()

    for ci, ar in enumerate(ÅR_HIST):
        data_cell(wp, row, 3+ci, uk_data[uk].get(ar,0), KR_FMT, bg=bg)

    # Tom skillekolonne (G=kol 7)
    wp.cell(row=row, column=7).fill = PatternFill("solid", start_color=GRÅ_BG)

    # Veksttakt for år før prediksjonsår (blå, redigerbar) – kolonne H = 8
    cg = cagr_22_24(uk)
    vc26 = wp.cell(row=row, column=8, value=round(cg, 4))
    vc26.font = Font(name="Arial", color=BLÅ_TEKST, size=10)
    vc26.fill = PatternFill("solid", start_color=GULL_BG)
    vc26.number_format = PCT_FMT
    vc26.alignment = Alignment(horizontal="center", vertical="center")
    vc26.border = kant()
    vekst_26_celler[uk] = f"H{row}"
    aktual_25_celler[uk] = f"F{row}"

    # Prediksjon for år før prediksjonsår = aktual_2025 * (1 + vekst) * (1 + global ekstra)
    pred26 = wp.cell(row=row, column=9,
                     value=f"=F{row}*(1+H{row})*(1+'{ANTAG_SHEET}'!$B$6)")
    pred26.font = Font(name="Arial", size=10)
    pred26.fill = PatternFill("solid", start_color=AKSENT_BG)
    pred26.number_format = KR_FMT
    pred26.alignment = Alignment(horizontal="right", vertical="center")
    pred26.border = kant()

    # Tom skille kol 10
    wp.cell(row=row, column=10).fill = PatternFill("solid", start_color=GRÅ_BG)

    # Veksttakt for prediksjonsår (blå) – kolonne K = 11
    vc27 = wp.cell(row=row, column=11, value=round(cg, 4))
    vc27.font = Font(name="Arial", color=BLÅ_TEKST, size=10)
    vc27.fill = PatternFill("solid", start_color=GULL_BG)
    vc27.number_format = PCT_FMT
    vc27.alignment = Alignment(horizontal="center", vertical="center")
    vc27.border = kant()
    vekst_27_celler[uk] = f"K{row}"

    # Prediksjon for prediksjonsår = forrige pred * (1 + vekst) * (1 + global ekstra)
    pred27 = wp.cell(row=row, column=12,
                     value=f"=I{row}*(1+K{row})*(1+'{ANTAG_SHEET}'!$B$7)")
    pred27.font = Font(name="Arial", size=10)
    pred27.fill = PatternFill("solid", start_color=GRØNN_BG)
    pred27.number_format = KR_FMT
    pred27.alignment = Alignment(horizontal="right", vertical="center")
    pred27.border = kant()

# Totalrader
tot_r = 7 + len(UNDERKATEGORIER)
wp.cell(row=tot_r, column=2, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
wp.cell(row=tot_r, column=2).fill = PatternFill("solid", start_color=AKSENT_BG)
wp.cell(row=tot_r, column=2).border = kant()
wp.cell(row=tot_r, column=2).alignment = Alignment(horizontal="left")

for ci, ar in enumerate(ÅR_HIST):
    data_cell(wp, tot_r, 3+ci, totaler[ar], KR_FMT, bold=True, bg=AKSENT_BG)
wp.cell(row=tot_r, column=7).fill = PatternFill("solid", start_color=GRÅ_BG)
wp.cell(row=tot_r, column=10).fill = PatternFill("solid", start_color=GRÅ_BG)

# Sum-formler for prediksjon totalt
first_data_row = 7
last_data_row  = 7 + len(UNDERKATEGORIER) - 1

tc26 = wp.cell(row=tot_r, column=9,
               value=f"=SUM(I{first_data_row}:I{last_data_row})")
tc26.font = Font(name="Arial", bold=True, size=10)
tc26.fill = PatternFill("solid", start_color=AKSENT_BG)
tc26.number_format = KR_FMT
tc26.border = kant()
tc26.alignment = Alignment(horizontal="right", vertical="center")

tc27 = wp.cell(row=tot_r, column=12,
               value=f"=SUM(L{first_data_row}:L{last_data_row})")
tc27.font = Font(name="Arial", bold=True, size=10)
tc27.fill = PatternFill("solid", start_color=GRØNN_BG)
tc27.number_format = KR_FMT
tc27.border = kant()
tc27.alignment = Alignment(horizontal="right", vertical="center")
wp.row_dimensions[tot_r].height = 22

pred_uk_tot_row = tot_r  # rad med SUM på prediksjonsark (brukes i Forside og «Slik beregnet vi»)

_apply_high_growth_conditional(wp, first_data_row, last_data_row, col_2025=6, col_pred=12)

# Forside: koble nøkkeltall til formler som oppdateres når bruker endrer arkene
ws["B6"].value = f"=SUM('{PRED_SHEET_NAME}'!F{first_data_row}:F{last_data_row})"
ws["B6"].number_format = KR_FMT
ws["B6"].font = Font(name="Arial", bold=True, size=18, color=HEADER_BG)
ws["F6"].value = f"='{PRED_SHEET_NAME}'!L{pred_uk_tot_row}"
ws["F6"].number_format = KR_FMT
ws["F6"].font = Font(name="Arial", bold=True, size=18, color=HEADER_BG)

# Forside tabell: lenke historikk til dette arket + dynamiske CAGR/andel + SUM i totalrad
_fs_first = 10
_fs_last = 10 + len(UNDERKATEGORIER) - 1
_fs_tot = _fs_last + 1
for ri in range(len(UNDERKATEGORIER)):
    fs_row = 10 + ri
    pr_row = 7 + ri
    bg = GRÅ_BG if ri % 2 == 0 else HVIT
    for ci in range(4):
        col = 3 + ci
        letter = get_column_letter(col)
        c = ws.cell(row=fs_row, column=col, value=f"='{PRED_SHEET_NAME}'!{letter}{pr_row}")
        c.number_format = KR_FMT
        c.font = Font(name="Arial", size=10)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = kant()
    data_cell(
        ws,
        fs_row,
        7,
        formula=f"=IF(AND(C{fs_row}>0,E{fs_row}>0),(E{fs_row}/C{fs_row})^(1/2)-1,0)",
        fmt=PCT_FMT,
        bg=bg,
    )
    data_cell(
        ws,
        fs_row,
        8,
        formula=(
            f"=IF(SUM($F${_fs_first}:$F${_fs_last})>0,"
            f"F{fs_row}/SUM($F${_fs_first}:$F${_fs_last}),0)"
        ),
        fmt=PCT_FMT,
        bg=bg,
    )

for ci in range(4):
    col = 3 + ci
    letter = get_column_letter(col)
    data_cell(
        ws,
        _fs_tot,
        col,
        formula=f"=SUM({letter}{_fs_first}:{letter}{_fs_last})",
        fmt=KR_FMT,
        bold=True,
        bg=AKSENT_BG,
    )
zg = ws.cell(row=_fs_tot, column=7)
zg.fill = PatternFill("solid", start_color=AKSENT_BG)
zg.border = kant()
data_cell(
    ws,
    _fs_tot,
    8,
    formula=f"=SUM(H{_fs_first}:H{_fs_last})",
    fmt=PCT_FMT,
    bold=True,
    bg=AKSENT_BG,
)

# Prediksjon-ark: totalrad for 2022–2025 = SUM av datarader (ikke faste tall)
for ci in range(4):
    col = 3 + ci
    letter = get_column_letter(col)
    data_cell(
        wp,
        pred_uk_tot_row,
        col,
        formula=f"=SUM({letter}{first_data_row}:{letter}{last_data_row})",
        fmt=KR_FMT,
        bold=True,
        bg=AKSENT_BG,
    )

# KPI-note
note_r = tot_r + 2
wp.merge_cells(f"B{note_r}:L{note_r}")
n = wp[f"B{note_r}"]
n.value = ("Statsbygg-andel 2025: ~228M kr (41% av total). "
           f"Hvert prosentpoeng KPI-økning = +2.3M kr på {PRED_YEAR}-anslaget fra Statsbygg alene. "
           "Juster veksttakt for 'Husleie' opp/ned tilsvarende.")
n.font = Font(name="Arial", size=9, italic=True, color="404040")
n.fill = PatternFill("solid", start_color=GULL_BG)
n.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
n.border = kant()
wp.row_dimensions[note_r].height = 28

# Forklaring farger
note_r2 = note_r + 2
for ci, (txt, col) in enumerate([
    ("Blå = justerbar vekst % (endre → prediksjon følger)", BLÅ_TEKST),
    ("Sort = beregnet formel", SORT_TEKST),
    ("Gul bakgrunn = input", SORT_TEKST),
    (f"Blå bakgrunn = {PRED_PREV_YEAR}-prediksjon", SORT_TEKST),
    (f"Grønn bakgrunn = {PRED_YEAR}-prediksjon", SORT_TEKST),
]):
    c = wp.cell(row=note_r2, column=2+ci*2, value=txt)
    c.font = Font(name="Arial", size=9, color=col, italic=True)
    wp.merge_cells(start_row=note_r2, start_column=2+ci*2,
                   end_row=note_r2, end_column=3+ci*2)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 – PIVOT UNDERKATEGORI × ÅR
# ════════════════════════════════════════════════════════════════════════════
wu = wb.create_sheet("Pivot_Underkategori")
wu.sheet_view.showGridLines = False
wu.column_dimensions["A"].width = 3
wu.column_dimensions["B"].width = 46
for i, ar in enumerate(ÅR_HIST):
    wu.column_dimensions[get_column_letter(3+i)].width = 16
wu.column_dimensions[get_column_letter(3+len(ÅR_HIST))].width = 16
wu.column_dimensions[get_column_letter(4+len(ÅR_HIST))].width = 12

wu.merge_cells("B2:H2")
wu["B2"].value = "Pivot: Underkategori × År (ekskl. MV/MP)"
wu["B2"].font = Font(name="Arial", bold=True, size=13, color=HEADER_BG)

header_cell(wu, 4, 2, "Underkategori", center=False)
for i, ar in enumerate(ÅR_HIST):
    header_cell(wu, 4, 3+i, ar)
header_cell(wu, 4, 3+len(ÅR_HIST), "Endring\n22→25", wrap=True)
header_cell(wu, 4, 4+len(ÅR_HIST), "Andel\n2025", wrap=True)
wu.row_dimensions[4].height = 28

all_uk = sorted(uk_data.keys(), key=lambda k: -uk_data[k].get("2025",0))
grand_25 = sum(uk_data[k].get("2025",0) for k in all_uk)

for ri, uk in enumerate(all_uk):
    row = 5 + ri
    bg = GRÅ_BG if ri % 2 == 0 else HVIT
    c = wu.cell(row=row, column=2, value=uk or "(tom)")
    c.font = Font(name="Arial", size=10)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = kant()
    for ci, ar in enumerate(ÅR_HIST):
        data_cell(wu, row, 3+ci, uk_data[uk].get(ar,0), KR_FMT, bg=bg)
    v22 = uk_data[uk].get("2022",0)
    v25 = uk_data[uk].get("2025",0)
    endring = (v25-v22)/v22 if v22 else 0
    data_cell(wu, row, 3+len(ÅR_HIST), endring, PCT_FMT, bg=bg)
    data_cell(wu, row, 4+len(ÅR_HIST), v25/grand_25 if grand_25 else 0, PCT_FMT, bg=bg)

tot_r = 5 + len(all_uk)
c = wu.cell(row=tot_r, column=2, value="TOTAL")
c.font = Font(name="Arial", bold=True, size=10)
c.fill = PatternFill("solid", start_color=AKSENT_BG)
c.border = kant()
c.alignment = Alignment(horizontal="left")
for ci, ar in enumerate(ÅR_HIST):
    data_cell(wu, tot_r, 3+ci, sum(uk_data[uk].get(ar,0) for uk in all_uk),
              KR_FMT, bold=True, bg=AKSENT_BG)
data_cell(wu, tot_r, 3+len(ÅR_HIST), None, bg=AKSENT_BG)
data_cell(wu, tot_r, 4+len(ÅR_HIST), 1.0, PCT_FMT, bold=True, bg=AKSENT_BG)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 4 – PIVOT REGION × ÅR
# ════════════════════════════════════════════════════════════════════════════
wr = wb.create_sheet("Pivot_Region")
wr.sheet_view.showGridLines = False
wr.column_dimensions["A"].width = 3
wr.column_dimensions["B"].width = 22
for i in range(len(ÅR_HIST)+2):
    wr.column_dimensions[get_column_letter(3+i)].width = 16

wr.merge_cells("B2:H2")
wr["B2"].value = "Pivot: Region × År (ekskl. MV/MP)"
wr["B2"].font = Font(name="Arial", bold=True, size=13, color=HEADER_BG)

header_cell(wr, 4, 2, "Region", center=False)
for i, ar in enumerate(ÅR_HIST):
    header_cell(wr, 4, 3+i, ar)
header_cell(wr, 4, 3+len(ÅR_HIST), "Endring\n22→25", wrap=True)
header_cell(wr, 4, 4+len(ÅR_HIST), "Andel\n2025", wrap=True)
wr.row_dimensions[4].height = 28

reg_total_25 = sum(reg_data[r].get("2025",0) for r in reg_data)
for ri, reg in enumerate(sorted(reg_data.keys(),
                                 key=lambda k: -reg_data[k].get("2025",0))):
    row = 5 + ri
    bg = GRÅ_BG if ri % 2 == 0 else HVIT
    c = wr.cell(row=row, column=2, value=reg or "(tom)")
    c.font = Font(name="Arial", size=10)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = kant()
    for ci, ar in enumerate(ÅR_HIST):
        data_cell(wr, row, 3+ci, reg_data[reg].get(ar,0), KR_FMT, bg=bg)
    v22 = reg_data[reg].get("2022",0)
    v25 = reg_data[reg].get("2025",0)
    endring = (v25-v22)/v22 if v22 else 0
    data_cell(wr, row, 3+len(ÅR_HIST), endring, PCT_FMT, bg=bg)
    data_cell(wr, row, 4+len(ÅR_HIST), v25/reg_total_25 if reg_total_25 else 0, PCT_FMT, bg=bg)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 5 – TOPP LEVERANDØRER
# ════════════════════════════════════════════════════════════════════════════
wl = wb.create_sheet("Pivot_Leverandør")
wl.sheet_view.showGridLines = False
wl.column_dimensions["A"].width = 3
wl.column_dimensions["B"].width = 10
wl.column_dimensions["C"].width = 46
for i in range(len(ÅR_HIST)):
    wl.column_dimensions[get_column_letter(4+i)].width = 16
wl.column_dimensions[get_column_letter(4+len(ÅR_HIST))].width = 16   # Total
wl.column_dimensions[get_column_letter(5+len(ÅR_HIST))].width = 12   # Andel
wl.column_dimensions[get_column_letter(6+len(ÅR_HIST))].width = 8    # Trend

wl.merge_cells(f"B2:{get_column_letter(6+len(ÅR_HIST))}2")
wl["B2"].value = "Alle leverandører — kostnad per år (ekskl. MV/MP)"
wl["B2"].font = Font(name="Arial", bold=True, size=13, color=HEADER_BG)

header_cell(wl, 4, 2, "Resk.nr", center=True)
header_cell(wl, 4, 3, "Leverandørnavn", center=False)
for i, ar in enumerate(ÅR_HIST):
    header_cell(wl, 4, 4+i, ar)
header_cell(wl, 4, 4+len(ÅR_HIST), "Total\n2022–2025", wrap=True)
header_cell(wl, 4, 5+len(ÅR_HIST), "Andel\n2025", wrap=True)
header_cell(wl, 4, 6+len(ÅR_HIST), "Trend", center=True)
wl.row_dimensions[4].height = 28

alle_lev = sorted(lev_data.keys(), key=lambda k: -lev_data[k].get("2025",0))
lev_total_25 = sum(lev_data[k].get("2025",0) for k in lev_data)
lev_data_start_row = 5

for ri, (lid, lnavn) in enumerate(alle_lev):
    row = lev_data_start_row + ri
    bg = GRÅ_BG if ri % 2 == 0 else HVIT
    c = wl.cell(row=row, column=2, value=lid)
    c.font = Font(name="Arial", size=10)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = kant()

    c2 = wl.cell(row=row, column=3, value=lnavn)
    c2.font = Font(name="Arial", size=10)
    c2.fill = PatternFill("solid", start_color=bg)
    c2.alignment = Alignment(horizontal="left", vertical="center")
    c2.border = kant()

    for ci, ar in enumerate(ÅR_HIST):
        data_cell(wl, row, 4+ci, lev_data[(lid,lnavn)].get(ar,0), KR_FMT, bg=bg)
    total_lev = sum(lev_data[(lid,lnavn)].get(ar,0) for ar in ÅR_HIST)
    data_cell(wl, row, 4+len(ÅR_HIST), total_lev, KR_FMT, bg=bg, bold=True)
    v25 = lev_data[(lid,lnavn)].get("2025",0)
    data_cell(wl, row, 5+len(ÅR_HIST), v25/lev_total_25 if lev_total_25 else 0, PCT_FMT, bg=bg)
    v22 = lev_data[(lid,lnavn)].get("2022",0)
    trend = "▲" if v25 > v22 else ("▼" if v25 < v22 else "→")
    ct = wl.cell(row=row, column=6+len(ÅR_HIST), value=trend)
    ct.font = Font(name="Arial", size=11, bold=True,
                   color=("006600" if trend == "▲" else ("CC0000" if trend == "▼" else "404040")))
    ct.fill = PatternFill("solid", start_color=bg)
    ct.alignment = Alignment(horizontal="center", vertical="center")
    ct.border = kant()

# Autofilter
lev_last_row = lev_data_start_row + len(alle_lev) - 1
wl.auto_filter.ref = f"B4:{get_column_letter(6+len(ÅR_HIST))}{lev_last_row}"
wl.freeze_panes = "B5"
print(f"  → {len(alle_lev)} leverandører lagt til")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 6 – AGGREGAT-KOSTSEDER (204416 m.fl.)
# ════════════════════════════════════════════════════════════════════════════
wa = wb.create_sheet("Aggregat_Kostseder")
wa.sheet_view.showGridLines = False
wa.column_dimensions["A"].width = 3
wa.column_dimensions["B"].width = 14
wa.column_dimensions["C"].width = 50
for i in range(len(ÅR_HIST)+1):
    wa.column_dimensions[get_column_letter(4+i)].width = 16

wa.merge_cells("B2:I2")
wa["B2"].value = "⚠️  Aggregat-kostseder — kostnader IKKE fordelt til enkelteiendommer"
wa["B2"].font = Font(name="Arial", bold=True, size=13, color="7F0000")
wa.merge_cells("B3:I3")
wa["B3"].value = ("Disse kosstedene aggregerer kostnader for MANGE bygg i Agresso. "
                  "~150M av 2025-summen er 'parkert' her uten kobling til eiendom. "
                  "Bør sjekkes mot Agresso-koststedshierarkiet.")
wa["B3"].font = Font(name="Arial", size=9, italic=True, color="595959")
wa["B3"].alignment = Alignment(wrap_text=True)
wa.row_dimensions[3].height = 28

header_cell(wa, 5, 2, "Dim1-kode")
header_cell(wa, 5, 3, "Koststed-navn", center=False)
for i, ar in enumerate(ÅR_HIST):
    header_cell(wa, 5, 4+i, ar)
header_cell(wa, 5, 4+len(ÅR_HIST), "Andel\n2025", wrap=True)
wa.row_dimensions[5].height = 28

all_dim1 = sorted(dim1_data.keys(), key=lambda k: -dim1_data[k].get("2025",0))
total_25_dim1 = sum(dim1_data[k].get("2025",0) for k in all_dim1)

for ri, (kode, navn) in enumerate(all_dim1[:40]):
    row = 6 + ri
    is_agg = kode in AGGREGAT
    bg = RØD_BG if is_agg else (GRÅ_BG if ri % 2 == 0 else HVIT)
    c = wa.cell(row=row, column=2, value=kode)
    c.font = Font(name="Arial", size=10, bold=is_agg)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = kant()
    c2 = wa.cell(row=row, column=3, value=(navn or "(tom)")[:60])
    c2.font = Font(name="Arial", size=10, bold=is_agg)
    c2.fill = PatternFill("solid", start_color=bg)
    c2.alignment = Alignment(horizontal="left", vertical="center")
    c2.border = kant()
    for ci, ar in enumerate(ÅR_HIST):
        data_cell(wa, row, 4+ci, dim1_data[(kode,navn)].get(ar,0), KR_FMT, bg=bg, bold=is_agg)
    v25 = dim1_data[(kode,navn)].get("2025",0)
    data_cell(wa, row, 4+len(ÅR_HIST),
              v25/total_25_dim1 if total_25_dim1 else 0, PCT_FMT, bg=bg, bold=is_agg)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 7 – FLAGGEDE MV/MP-POSTERINGER (utvidet)
# ════════════════════════════════════════════════════════════════════════════

# Forberegn bilag-netto og linjektel per bilagsnr
_bilag_netto  = defaultdict(float)
_bilag_linjer = defaultdict(int)
for _r in mv_mp_rader:
    _nr = _r.get("Bilagsnr","").strip()
    _bilag_netto[_nr]  += rens(_r.get("Beløp",""))
    _bilag_linjer[_nr] += 1

# Sorter: etter Bilagsnr (gruppert), deretter størst beløp i gruppen
_bilag_maks = {nr: max(abs(rens(r.get("Beløp",""))) for r in mv_mp_rader
               if r.get("Bilagsnr","").strip() == nr)
               for nr in _bilag_netto}
mv_mp_sortert = sorted(mv_mp_rader,
                        key=lambda r: (-_bilag_maks.get(r.get("Bilagsnr","").strip(), 0),
                                       r.get("Bilagsnr",""),
                                       -abs(rens(r.get("Beløp","")))))

wf = wb.create_sheet("Flaggede_MV_MP")
wf.sheet_view.showGridLines = False
# 18 kolonner: A(marg) B(BA) C(type-forkl) D(Bilagsnr) E(dato) F(periode) G(region)
#              H(Dim1-kode) I(Dim1-navn) J(Konto) K(Konto(T)) L(Underkategori)
#              M(Lev.nr) N(Lev.navn) O(Beløp) P(Tekst) Q(Bilag-netto) R(Linjer)
col_widths_f = [3, 5, 34, 14, 12, 8, 8, 10, 36, 8, 26, 26, 10, 34, 14, 52, 14, 6]
for ci, w in enumerate(col_widths_f, 1):
    wf.column_dimensions[get_column_letter(ci)].width = w

wf.merge_cells("B2:R2")
wf["B2"].value = "Flaggede MV/MP-posteringer 2025 — alle poster med full kontekst"
wf["B2"].font = Font(name="Arial", bold=True, size=13, color="7F0000")
wf.merge_cells("B3:R3")
wf["B3"].value = (
    "MV = Motverdi/Viderefakturert  |  MP = Motpost.  Netto 2025: +18M — IKKE i prediksjonsbasis (se sluttresultat).  "
    "Bilag med Bilag-netto ≠ 0 krever manuell avklaring.  "
    "Sjekk Tekst-feltet for kontekst — Agresso-narrativ forklarer hva posteringen gjelder.  "
    "MV uten tilhørende MP på samme bilagsnr = viderefakturering til tredjepart."
)
wf["B3"].font = Font(name="Arial", size=9, italic=True, color="595959")
wf["B3"].alignment = Alignment(wrap_text=True)
wf.row_dimensions[3].height = 40

_mvmp_hdrs = ["BA", "Type-forklaring", "Bilagsnr", "Bilagsdato", "Periode", "Region",
              "Dim1-kode", "Dim1-navn", "Konto", "Konto(T)", "Underkategori",
              "Leverandørnr", "Leverandørnavn", "Beløp (kr)", "Tekst (full)",
              "Bilag-nettossum", "Linjer\ni bilag"]
for ci, h in enumerate(_mvmp_hdrs):
    header_cell(wf, 5, 2+ci, h, wrap=True)
wf.row_dimensions[5].height = 32

_prev_bilag = None
_alt_bg_idx = 0
for ri, row in enumerate(mv_mp_sortert):
    r = 6 + ri
    ba  = row.get("BA","")
    nr  = row.get("Bilagsnr","").strip()
    # Alternerende farge per bilagsgruppe
    if nr != _prev_bilag:
        _alt_bg_idx += 1
        _prev_bilag = nr
    base_bg = RØD_BG if ba == "MV" else ("E8E8E8" if _alt_bg_idx % 2 == 0 else GRÅ_BG)
    bg = base_bg

    netto = _bilag_netto.get(nr, 0.0)
    netto_bg = "FFE0B2" if abs(netto) > 1 else bg   # oransje hvis ubalansert

    vals = [
        ba,
        BA_FORKLARING.get(ba, ba),
        nr,
        row.get("Bilagsdato",""),
        row.get("Periode",""),
        row.get("Region",""),
        row.get("Dim1",""),
        row.get("Dim1(T)",""),
        row.get("Konto",""),
        row.get("Konto(T)",""),
        row.get("Underkategorier(T)",""),
        row.get("Resk.nr",""),
        row.get("Resk.nr(T)",""),
        rens(row.get("Beløp","")),
        row.get("Tekst",""),          # full Tekst, ikke trunkert
        netto,
        _bilag_linjer.get(nr, 1),
    ]
    beløp_col  = 15   # O = col index 15 (1-based from B=2)
    netto_col  = 16   # Q
    linjer_col = 17   # R

    for ci, v in enumerate(vals):
        col_idx = 2 + ci
        use_bg = netto_bg if ci in (beløp_col-2, netto_col-2) else bg
        c = wf.cell(row=r, column=col_idx, value=v)
        c.font = Font(name="Arial", size=9)
        c.fill = PatternFill("solid", start_color=use_bg)
        c.border = kant()
        is_amt = ci in (13, 15)   # Beløp og Bilag-netto
        is_int = ci == 16          # Linjer
        c.alignment = Alignment(
            horizontal="right" if is_amt else ("center" if is_int else "left"),
            vertical="center", wrap_text=(ci in (1, 14)))
        if is_amt:
            c.number_format = KR_FMT

wf.freeze_panes = "D6"
wf.auto_filter.ref = f"B5:R{5 + len(mv_mp_sortert)}"
print(f"  → {len(mv_mp_sortert)} MV/MP-rader (alle) lagt til med {len(_bilag_netto)} unike bilag")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 8 – PREDIKSJON PER EIENDOM (Dim1/Koststed)
# ════════════════════════════════════════════════════════════════════════════
we = wb.create_sheet("Prediksjon_Per_Eiendom")
we.sheet_view.showGridLines = False

# Kolonnebredder: A(marg) B(kode) C(navn) D-G(år) H(cagr) I(vekst26) J(pred26) K(vekst27) L(pred27)
col_w = [3, 12, 46, 16, 16, 16, 16, 10, 14, 16, 14, 16]
for i, w in enumerate(col_w, 1):
    we.column_dimensions[get_column_letter(i)].width = w

we.merge_cells("B2:L2")
c = we["B2"]
c.value = f"Prediksjon {PRED_YEAR} per eiendom (Dim1/Koststed) — redigerbare veksttakter"
c.font = Font(name="Arial", bold=True, size=14, color=HEADER_BG)
c.alignment = Alignment(horizontal="left", vertical="center")

we.merge_cells("B3:L3")
c = we["B3"]
c.value = (
    "Blå celler = justerbar veksttakt per eiendom. "
    f"Ark «{ANTAG_SHEET}» gir ekstra vekst på alle linjer. "
    "Rød rad = aggregat-koststed (dekker mange bygg, ikke én eiendom). "
    "Betinget formatering (gul rad): prediksjon siste år (L) mer enn 25 % høyere enn 2025-faktisk (G). "
    "Basis = 2022–2025 faktisk, ekskl. MV/MP."
)
c.font = Font(name="Arial", size=9, italic=True, color="595959")
c.alignment = Alignment(horizontal="left", wrap_text=True)
we.row_dimensions[3].height = 20
we.row_dimensions[4].height = 8

# Header rad 5
hdrs5 = ["Dim1-kode", "Eiendom / Koststed-navn",
         "2022\nAktual", "2023\nAktual", "2024\nAktual", "2025\nAktual",
         "CAGR\n22→24", f"Veksttakt\n{PRED_PREV_YEAR}", f"Prediksjon\n{PRED_PREV_YEAR}",
         f"Veksttakt\n{PRED_YEAR}", f"Prediksjon\n{PRED_YEAR}"]
for ci, h in enumerate(hdrs5):
    header_cell(we, 5, 2+ci, h, wrap=True)
we.row_dimensions[5].height = 32

# Sorter alle Dim1 etter 2025-faktisk, høyest først
alle_dim1 = sorted(dim1_data.keys(), key=lambda k: -dim1_data[k].get("2025", 0))
total_25_alle = sum(dim1_data[k].get("2025", 0) for k in alle_dim1)

for ri, (kode, navn) in enumerate(alle_dim1):
    row = 6 + ri
    is_agg = kode in AGGREGAT
    bg = RØD_BG if is_agg else (GRÅ_BG if ri % 2 == 0 else HVIT)

    # Dim1-kode
    c = we.cell(row=row, column=2, value=kode)
    c.font = Font(name="Arial", size=10, bold=is_agg)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = kant()

    # Navn
    c2 = we.cell(row=row, column=3, value=(navn or "(tom)")[:60])
    c2.font = Font(name="Arial", size=10, bold=is_agg)
    c2.fill = PatternFill("solid", start_color=bg)
    c2.alignment = Alignment(horizontal="left", vertical="center")
    c2.border = kant()

    # Historiske år
    for ci, ar in enumerate(ÅR_HIST):
        data_cell(we, row, 4+ci, dim1_data[(kode, navn)].get(ar, 0), KR_FMT, bg=bg, bold=is_agg)

    # CAGR per eiendom (med outlier-korreksjon for 2022)
    verdier_e = [dim1_data[(kode, navn)].get(ar, 0) for ar in ÅR_HIST]
    cg_std = robust_cagr(verdier_e[0], verdier_e[2], år_diff=2)
    if _er_outlier_2022(verdier_e):
        cg_alt = robust_cagr(verdier_e[1], verdier_e[3], år_diff=2)
        cg = cg_alt if cg_alt > cg_std else cg_std
    else:
        cg = cg_std
    # Begrens til -50% / +50% for å unngå ekstreme utslag på enkelteiendommer
    cg = max(-0.50, min(0.50, cg))

    c_cagr = we.cell(row=row, column=8, value=round(cg, 4))
    c_cagr.font = Font(name="Arial", size=10, color="595959", bold=is_agg)
    c_cagr.fill = PatternFill("solid", start_color=bg)
    c_cagr.number_format = PCT_FMT
    c_cagr.alignment = Alignment(horizontal="center", vertical="center")
    c_cagr.border = kant()

    # Veksttakt for år før prediksjonsår (blå, redigerbar) – kolonne I=9
    vc26 = we.cell(row=row, column=9, value=round(cg, 4))
    vc26.font = Font(name="Arial", color=BLÅ_TEKST, size=10)
    vc26.fill = PatternFill("solid", start_color=GULL_BG)
    vc26.number_format = PCT_FMT
    vc26.alignment = Alignment(horizontal="center", vertical="center")
    vc26.border = kant()

    # Prediksjon for år før prediksjonsår (samme globale ekstra som underkategori-ark)
    pred26 = we.cell(row=row, column=10, value=f"=G{row}*(1+I{row})*(1+'{ANTAG_SHEET}'!$B$6)")
    pred26.font = Font(name="Arial", size=10, bold=is_agg)
    pred26.fill = PatternFill("solid", start_color=AKSENT_BG)
    pred26.number_format = KR_FMT
    pred26.alignment = Alignment(horizontal="right", vertical="center")
    pred26.border = kant()

    # Veksttakt for prediksjonsår (blå, redigerbar) – kolonne K=11
    vc27 = we.cell(row=row, column=11, value=round(cg, 4))
    vc27.font = Font(name="Arial", color=BLÅ_TEKST, size=10)
    vc27.fill = PatternFill("solid", start_color=GULL_BG)
    vc27.number_format = PCT_FMT
    vc27.alignment = Alignment(horizontal="center", vertical="center")
    vc27.border = kant()

    # Prediksjon for prediksjonsår
    pred27 = we.cell(row=row, column=12, value=f"=J{row}*(1+K{row})*(1+'{ANTAG_SHEET}'!$B$7)")
    pred27.font = Font(name="Arial", size=10, bold=is_agg)
    pred27.fill = PatternFill("solid", start_color=GRØNN_BG)
    pred27.number_format = KR_FMT
    pred27.alignment = Alignment(horizontal="right", vertical="center")
    pred27.border = kant()

# Totallinje
n_eiend = len(alle_dim1)
tot_r = 6 + n_eiend
c = we.cell(row=tot_r, column=2, value="TOTAL")
c.font = Font(name="Arial", bold=True, size=10)
c.fill = PatternFill("solid", start_color=AKSENT_BG)
c.border = kant()
c.alignment = Alignment(horizontal="center")

c2 = we.cell(row=tot_r, column=3, value=f"Alle {n_eiend} kostseder")
c2.font = Font(name="Arial", bold=True, size=10)
c2.fill = PatternFill("solid", start_color=AKSENT_BG)
c2.border = kant()
c2.alignment = Alignment(horizontal="left")

for ci, ar in enumerate(ÅR_HIST):
    col_l = get_column_letter(4 + ci)
    data_cell(we, tot_r, 4+ci,
              formula=f"=SUM({col_l}6:{col_l}{tot_r-1})",
              fmt=KR_FMT, bold=True, bg=AKSENT_BG)

# CAGR-kolonnen blank
c_cg = we.cell(row=tot_r, column=8)
c_cg.fill = PatternFill("solid", start_color=AKSENT_BG)
c_cg.border = kant()

# Veksttakt-kolonner blank
for col in [9, 11]:
    cx = we.cell(row=tot_r, column=col)
    cx.fill = PatternFill("solid", start_color=AKSENT_BG)
    cx.border = kant()

# Sum prediksjon for år før prediksjonsår og prediksjonsår
data_cell(we, tot_r, 10,
          formula=f"=SUM(J6:J{tot_r-1})",
          fmt=KR_FMT, bold=True, bg=AKSENT_BG)
data_cell(we, tot_r, 12,
          formula=f"=SUM(L6:L{tot_r-1})",
          fmt=KR_FMT, bold=True, bg=GRØNN_BG)
we.row_dimensions[tot_r].height = 22

_apply_high_growth_conditional(we, 6, tot_r - 1, col_2025=7, col_pred=12)

# Forklaringsnotat under tabellen
note_e = tot_r + 2
we.merge_cells(f"B{note_e}:L{note_e}")
n = we[f"B{note_e}"]
n.value = ("⚠️  Røde rader = aggregat-kostseder (204416, 500401, 300200) som dekker mange bygg. "
           "Disse bør bekreftes mot Agresso-hierarkiet. "
           "CAGR er begrenset til ±50% per eiendom for å unngå enkeltårs-utslag. "
           f"Totalen her vil avvike noe fra {PRED_SHEET_NAME}-fanen (som bruker underkategori-CAGR).")
n.font = Font(name="Arial", size=9, italic=True, color="7F0000")
n.fill = PatternFill("solid", start_color=RØD_BG)
n.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
n.border = kant()
we.row_dimensions[note_e].height = 32

# Fryse topprad
we.freeze_panes = "D6"

print(f"  → {n_eiend} eiendommer/kostseder lagt til")

PER_EIENDOM_FIRST_ROW = 6
PER_EIENDOM_TOT_ROW = tot_r
PER_EIENDOM_LAST_DATA_ROW = tot_r - 1

# ════════════════════════════════════════════════════════════════════════════
# SHEET 9 – TILBAKEMELDING FRA ØKONOMI
# ════════════════════════════════════════════════════════════════════════════
wtb = wb.create_sheet("Tilbakemelding_Økonomi")
wtb.sheet_view.showGridLines = False
wtb.column_dimensions["A"].width = 3
wtb.column_dimensions["B"].width = 60
wtb.column_dimensions["C"].width = 36
wtb.column_dimensions["D"].width = 14
wtb.column_dimensions["E"].width = 12

wtb.merge_cells("B2:E2")
c = wtb["B2"]
c.value = f"Tilbakemelding fra økonomi — Prediksjon {PRED_YEAR}"
c.font = Font(name="Arial", bold=True, size=15, color=HEADER_BG)
c.alignment = Alignment(horizontal="left", vertical="center")

wtb.merge_cells("B3:E3")
c = wtb["B3"]
c.value = "Fyll inn de gule cellene og send tilbake til eiendomsteamet."
c.font = Font(name="Arial", size=10, italic=True, color="595959")
c.alignment = Alignment(horizontal="left")
wtb.row_dimensions[3].height = 18
wtb.row_dimensions[4].height = 10

# Seksjonstittel
wtb.merge_cells("B5:E5")
c = wtb["B5"]
c.value = f"DEL 1 — Veksttakter per kostnadskategori  (se fanen «{PRED_SHEET_NAME}»)"
c.font = Font(name="Arial", bold=True, size=11, color=HVIT)
c.fill = PatternFill("solid", start_color=HEADER_BG)
c.alignment = Alignment(horizontal="left", vertical="center")
c.border = kant()
wtb.row_dimensions[5].height = 22

# Spørsmål-header
for ci, h in enumerate(["Kostnadskategori / spørsmål", "Svar / vurdering", "Forslag til %", "Avklart?"]):
    header_cell(wtb, 6, 2+ci, h, center=(ci > 0))
wtb.row_dimensions[6].height = 20

spørsmål = [
    ("Husleie  (+2.6% CAGR)  —  Statsbygg = 41% / 228M kr\n\n"
     f"→ KPI-regulerte kontrakter. Realistisk med 2–3%? Porteføljeendringer {PRED_PREV_YEAR}–{PRED_YEAR}?",
     "Skriv inn …"),
    ("Drift og vedlikehold  (−17% CAGR, 92M i 2024)\n\n"
     "→ Stor nedgang. Fortsetter, eller er tiltak planlagt som snur trenden?",
     "Skriv inn …"),
    ("Utskifting og utvikling  (+86% CAGR, 42M i 2024)\n\n"
     f"→ Tredoblet 2022→2024. Planlagte investeringer? Budsjett {PRED_PREV_YEAR}–{PRED_YEAR}?\n"
     "   (NB: satt til 0% i prediksjon — økonomi må bekrefte forventet nivå)",
     "Skriv inn …"),
    ("Forsyning / strøm  (−16% CAGR, 28M i 2024)\n\n"
     "→ Strømprisfall. Hva er forventning fremover?",
     "Skriv inn …"),
    ("Renholdstjenester  (+5.2% CAGR, 20M i 2024)\n\n"
     "→ Stabil vekst. Er 5% riktig, eller er det kontraktsendringer?",
     "Skriv inn …"),
]
for ri, (sp, pl) in enumerate(spørsmål):
    row = 7 + ri
    bg = GRÅ_BG if ri % 2 == 0 else HVIT
    c = wtb.cell(row=row, column=2, value=sp)
    c.font = Font(name="Arial", size=10)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.border = kant()
    wtb.row_dimensions[row].height = 52
    # Svar-celle (gul, redigerbar)
    cs = wtb.cell(row=row, column=3, value=pl)
    cs.font = Font(name="Arial", size=10, italic=True, color="595959")
    cs.fill = PatternFill("solid", start_color=GULL_BG)
    cs.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cs.border = kant()
    # % forslag
    cp = wtb.cell(row=row, column=4)
    cp.fill = PatternFill("solid", start_color=GULL_BG)
    cp.border = kant()
    cp.number_format = PCT_FMT
    cp.alignment = Alignment(horizontal="center")
    # Avklart
    ca = wtb.cell(row=row, column=5)
    ca.fill = PatternFill("solid", start_color=GULL_BG)
    ca.border = kant()
    ca.alignment = Alignment(horizontal="center")

# DEL 2
r_del2 = 7 + len(spørsmål) + 1
wtb.merge_cells(f"B{r_del2}:E{r_del2}")
c = wtb[f"B{r_del2}"]
c.value = "DEL 2 — Aggregat-kostseder og mulig dobbeltføring"
c.font = Font(name="Arial", bold=True, size=11, color=HVIT)
c.fill = PatternFill("solid", start_color="7F0000")
c.alignment = Alignment(horizontal="left", vertical="center")
c.border = kant()
wtb.row_dimensions[r_del2].height = 22

del2_sp = [
    ("Koststed 204416 «Regionkontor ØV»: 134.8M i 2025 — er dette reelle eiendomskostnader\n"
     "eller et aggregeringspunkt for mange bygg?",
     "Skriv inn …"),
    ("Koststed 500401 «Regionkontor felleskostnader»: 12.8M i 2025 — bekreftes dette\n"
     "som reelle kostnader, ikke viderefakturering?",
     "Skriv inn …"),
    ("MV/MP-posteringer (netto +18M i 2025): er disse engangsposteringer\n"
     "eller gjentakende? Skal de inn i prediksjonsbasis?",
     "Skriv inn …"),
]
for ri, (sp, pl) in enumerate(del2_sp):
    row = r_del2 + 1 + ri
    bg = RØD_BG if ri % 2 == 0 else HVIT
    c = wtb.cell(row=row, column=2, value=sp)
    c.font = Font(name="Arial", size=10)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.border = kant()
    wtb.row_dimensions[row].height = 46
    cs = wtb.cell(row=row, column=3, value=pl)
    cs.font = Font(name="Arial", size=10, italic=True, color="595959")
    cs.fill = PatternFill("solid", start_color=GULL_BG)
    cs.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cs.border = kant()
    for col in [4, 5]:
        cx = wtb.cell(row=row, column=col)
        cx.fill = PatternFill("solid", start_color=GULL_BG)
        cx.border = kant()
        cx.alignment = Alignment(horizontal="center")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 10 – METODEBESKRIVELSE
# ════════════════════════════════════════════════════════════════════════════
wm = wb.create_sheet("Metodebeskrivelse")
wm.sheet_view.showGridLines = False
wm.column_dimensions["A"].width = 3
wm.column_dimensions["B"].width = 22
wm.column_dimensions["C"].width = 54
wm.column_dimensions["D"].width = 22

wm.merge_cells("B2:D2")
c = wm["B2"]
c.value = f"Metodebeskrivelse — Prediksjon {PRED_YEAR}"
c.font = Font(name="Arial", bold=True, size=14, color=HEADER_BG)
c.alignment = Alignment(horizontal="left", vertical="center")
wm.row_dimensions[2].height = 28

seksjoner = [
    ("Datagrunnlag",
     "Agresso GL-eksport, Innkjøpskategori 01 «Leie av lokaler og tilknyttede utgifter».\n"
     "136 055 transaksjonslinjer for perioden 2020–2025.\n"
     "Kilde: Eiendom 202001 til 202512 til Øystein (AGRESSO).csv"),
    ("Filtrering",
     "Bare bilagsarter IV, IW, LE, H1, H2, HB, RE er inkludert.\n"
     "MV (Motverdi/viderefakturert) og MP (Motpost) er UTELATT fra prediksjonsbasis\n"
     "— disse er regnskapsmessige motposter, ikke reelle kostnader.\n"
     "Prediksjonsperiode: 2022–2025 (2020–2021 utelatt grunnet COVID-avvik)."),
    ("Metode",
     "CAGR (Compound Annual Growth Rate) 2022→2024 per underkategori.\n"
     "Formel: CAGR = (Aktual_2024 / Aktual_2022)^(1/2) − 1\n"
     f"Prediksjon {PRED_PREV_YEAR} = Aktual_2025 × (1 + CAGR) × (1 + ekstra fra «{ANTAG_SHEET}»)\n"
     f"Prediksjon {PRED_YEAR} = Prediksjon_{PRED_PREV_YEAR} × (1 + CAGR) × (1 + ekstra fra «{ANTAG_SHEET}»)\n\n"
     "Unntak: «Utskiftings- og utviklingskostnader» er holdt flat på 2025-nivå (42.3M).\n"
     "Historisk CAGR +86% skyldtes engangsøkning 2022→2023, ikke en reell trend."),
    ("Per eiendom",
     "Fanen «Prediksjon_Per_Eiendom» bruker samme CAGR-metode per Dim1-koststed.\n"
     "CAGR er begrenset til ±50% per eiendom for å unngå enkeltårs-utslag.\n"
     "Aggregat-kostseder (204416, 500401, 300200) er markert rødt."),
    ("Usikkerhet",
     "Modellen er mekanisk (historisk trend).\n\n"
     "Merk: KPI-klausul-justeringer på Statsbygg-kontrakter er INKLUDERT —\n"
     "Husleie-CAGR (+2,6%) gjenspeiler historiske KPI-reguleringer.\n\n"
     "Faktorer som IKKE er inkludert:\n"
     "  • Nye eller avviklede leieobjekter\n"
     "  • Politiske beslutninger om institusjonsnedleggelser\n"
     "  • Energiprisutvikling\n"
     f"Økonomi bør justere veksttaktene i «{PRED_SHEET_NAME}» og evt. globale ekstra i «{ANTAG_SHEET}»."),
    ("Statsbygg",
     "Statsbygg utgjør ~41% av totalkostnad 2025 (ca. 228M kr).\n"
     "KPI-regulerte kontrakter: hvert prosentpoeng KPI-økning = +2.3M kr.\n"
     "Statsbygg-identifisering: leverandørnavn ILIKE '%statsbygg%'."),
]

for ri, (tittel, tekst) in enumerate(seksjoner):
    row = 4 + ri * 3
    bg = AKSENT_BG if ri % 2 == 0 else GRÅ_BG
    c = wm.cell(row=row, column=2, value=tittel)
    c.font = Font(name="Arial", bold=True, size=11, color=HEADER_BG)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = kant()
    wm.row_dimensions[row].height = 20
    c2 = wm.cell(row=row+1, column=2, value=tekst)
    c2.font = Font(name="Arial", size=10)
    c2.fill = PatternFill("solid", start_color=HVIT)
    c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c2.border = kant()
    wm.merge_cells(start_row=row+1, start_column=2, end_row=row+1, end_column=4)
    wm.row_dimensions[row+1].height = 80
    wm.row_dimensions[row+2].height = 6

# ════════════════════════════════════════════════════════════════════════════
# SHEET 11 – SLIK BEREGNET VI PREDIKSJON
# ════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Slik beregnet vi")
ws5.sheet_view.showGridLines = False
ws5.column_dimensions["A"].width = 3
ws5.column_dimensions["B"].width = 6
ws5.column_dimensions["C"].width = 36
ws5.column_dimensions["D"].width = 36
ws5.column_dimensions["E"].width = 20
ws5.column_dimensions["F"].width = 36

ws5.merge_cells("B2:F2")
c = ws5["B2"]
c.value = f"Slik beregnet vi prediksjon {PRED_YEAR} — steg for steg"
c.font = Font(name="Arial", bold=True, size=14, color=HEADER_BG)
c.alignment = Alignment(horizontal="left", vertical="center")

ws5.merge_cells("B3:F3")
c = ws5["B3"]
c.value = ("Dette arket viser nøyaktig hvilke tall vi hentet ut, hvilke vi filtrerte bort, "
           f"og hvordan vi kom frem til prediksjon for {PRED_YEAR}.")
c.font = Font(name="Arial", size=10, italic=True, color="595959")
c.alignment = Alignment(horizontal="left", wrap_text=True)
ws5.row_dimensions[3].height = 20

ws5.merge_cells("B4:F4")
c = ws5["B4"]
c.value = ("✅  Korrigert anslag: «Utskifting og utvikling» settes til 0% vekst "
           "(ikke +86,4% historisk CAGR). Den historiske økningen skyldtes en engangsøkning i 2022–2023, "
           "ikke en vedvarende trend. Utskifting holdes på 2025-nivå (42,3M) inntil økonomi bekrefter investeringsplan.")
c.font = Font(name="Arial", size=10, color="006600", bold=True)
c.fill = PatternFill("solid", start_color=GRØNN_BG)
c.alignment = Alignment(horizontal="left", wrap_text=True)
c.border = kant()
ws5.row_dimensions[4].height = 44

def faseoverskrift(ws, row, tekst):
    ws.merge_cells(f"B{row}:F{row}")
    c = ws[f"B{row}"]
    c.value = tekst
    c.font = Font(name="Arial", bold=True, size=11, color=HVIT)
    c.fill = PatternFill("solid", start_color=HEADER_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = kant()
    ws.row_dimensions[row].height = 22

def steg_rad(ws, row, nr, hva, hvordan, resultat, merknad, bg=HVIT):
    vals = [nr, hva, hvordan, resultat, merknad]
    for ci, v in enumerate(vals):
        c = ws.cell(row=row, column=2+ci, value=v)
        c.font = Font(name="Arial", size=10)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="left" if ci != 0 else "center",
                                 vertical="top", wrap_text=True)
        c.border = kant()
    ws.row_dimensions[row].height = 50

# Header
for ci, h in enumerate(["Steg", "Hva", "Hvordan", "Resultat", "Merknad"]):
    header_cell(ws5, 6, 2+ci, h, center=(ci == 0))
ws5.row_dimensions[6].height = 20

faseoverskrift(ws5, 7, "  FASE 1 — Hente ut data fra Agresso")
steg_rad(ws5, 8, "1",
         "Agresso GL-eksport\nAlle transaksjoner i kategori 01",
         "Innkjøpskategorier = «01»\nPeriode: 2020–2025",
         "136 055 linjer",
         "Totalsum: 2 956M kr\n(6 år alle bilagsarter)", GRÅ_BG)
steg_rad(ws5, 9, "2",
         "Behold kun gyldige bilagsarter\n(«rene» poster)",
         "BA ∈ {IV, IW, LE, H1, H2, HB, RE}",
         "→ 80 545 linjer (2022–2025)",
         "IV = vanlig faktura\nH1/H2 = omposteringer\nRE = reverseringer")

faseoverskrift(ws5, 10, "  FASE 2 — Velge prediksjonsbasis (hvilke år)")
steg_rad(ws5, 11, "3",
         "Bruk kun 2022–2025\n(stabil periode)",
         "År ∈ {2022, 2023, 2024, 2025}",
         "80 545 linjer\n(2022–2025)",
         "2020–2021 utelatt:\nCOVID-avvik og\nufullstendige data", GRÅ_BG)

# Faktiske summer per underkategori
faseoverskrift(ws5, 12, "    Faktiske summer per underkategori (kr, ekskl. MV/MP)")
for ci2, h in enumerate(["Underkategori", "2022", "2023", "2024", "2025"]):
    header_cell(ws5, 13, 2+ci2, h, center=(ci2 > 0))
ws5.row_dimensions[13].height = 18

uk_korte = {
    "Husleie": "Husleie",
    "Drift- og vedlikeholdskostnader": "Drift og vedlikehold",
    "Utskiftings- og utviklingskostnader": "Utskifting og utvikling",
    "Forsyningskostnader": "Forsyning / strøm",
    "Renholdstjenester": "Renholdstjenester",
}
for ri2, uk in enumerate(UNDERKATEGORIER):
    row = 14 + ri2
    bg = GRÅ_BG if ri2 % 2 == 0 else HVIT
    c = ws5.cell(row=row, column=2, value=uk_korte.get(uk, uk))
    c.font = Font(name="Arial", size=10); c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="left"); c.border = kant()
    for ci2, ar in enumerate(ÅR_HIST):
        data_cell(ws5, row, 3+ci2, round(uk_data[uk].get(ar,0)/1e6)*1000000, KR_FMT, bg=bg)
    ws5.row_dimensions[row].height = 18

# Totallinje i tabellen
tot_r5 = 14 + len(UNDERKATEGORIER)
c = ws5.cell(row=tot_r5, column=2, value="TOTAL")
c.font = Font(name="Arial", bold=True, size=10)
c.fill = PatternFill("solid", start_color=AKSENT_BG); c.border = kant()
c.alignment = Alignment(horizontal="left")
for ci2, ar in enumerate(ÅR_HIST):
    data_cell(ws5, tot_r5, 3+ci2,
              formula=f"={get_column_letter(3+ci2)}14:{get_column_letter(3+ci2)}{tot_r5-1}",
              fmt=KR_FMT, bold=True, bg=AKSENT_BG)
ws5.row_dimensions[tot_r5].height = 20

faseoverskrift(ws5, tot_r5+1, "  FASE 3 — Beregne veksttakt (CAGR 2022 → 2024)")
for ci2, h in enumerate(["Underkategori", "2022", "2024", "CAGR", "Tolkning"]):
    header_cell(ws5, tot_r5+2, 2+ci2, h, center=(ci2 in [1,2,3]))
ws5.row_dimensions[tot_r5+2].height = 18

tolkninger = {
    "Husleie": "Stabil KPI-vekst",
    "Drift- og vedlikeholdskostnader": "Effektivisering eller nedprioritering",
    "Utskiftings- og utviklingskostnader": "Strukturell økning — USIKKER",
    "Forsyningskostnader": "Strømprisfall",
    "Renholdstjenester": "Stabil vekst",
}
for ri2, uk in enumerate(UNDERKATEGORIER):
    row = tot_r5 + 3 + ri2
    bg = GRÅ_BG if ri2 % 2 == 0 else HVIT
    is_utskifting = "Utskifting" in uk
    cg = cagr_22_24(uk)
    cg_vis = "0,0% ← korrigert" if is_utskifting else f"{cg*100:+.1f}%"
    c = ws5.cell(row=row, column=2, value=uk_korte.get(uk, uk))
    c.font = Font(name="Arial", size=10, bold=is_utskifting)
    c.fill = PatternFill("solid", start_color=GRØNN_BG if is_utskifting else bg)
    c.border = kant(); c.alignment = Alignment(horizontal="left")
    data_cell(ws5, row, 3, uk_data[uk].get("2022",0), KR_FMT,
              bg=GRØNN_BG if is_utskifting else bg)
    data_cell(ws5, row, 4, uk_data[uk].get("2024",0), KR_FMT,
              bg=GRØNN_BG if is_utskifting else bg)
    c4 = ws5.cell(row=row, column=5, value=cg_vis)
    c4.font = Font(name="Arial", size=10, bold=is_utskifting,
                   color="006600" if is_utskifting else SORT_TEKST)
    c4.fill = PatternFill("solid", start_color=GRØNN_BG if is_utskifting else bg)
    c4.border = kant(); c4.alignment = Alignment(horizontal="center")
    c5 = ws5.cell(row=row, column=6, value=tolkninger.get(uk,""))
    c5.font = Font(name="Arial", size=10, italic=True)
    c5.fill = PatternFill("solid", start_color=GRØNN_BG if is_utskifting else bg)
    c5.border = kant(); c5.alignment = Alignment(horizontal="left")
    ws5.row_dimensions[row].height = 18

last_cagr_row = tot_r5 + 3 + len(UNDERKATEGORIER) - 1
faseoverskrift(ws5, last_cagr_row+1, f"  FASE 4 — Beregne prediksjon {PRED_PREV_YEAR} og {PRED_YEAR}")
steg_rad(ws5, last_cagr_row+2, "4",
         f"Prediksjon {PRED_PREV_YEAR}\n= Faktisk 2025 × (1 + CAGR)",
         "Husleie: 352,6M × 1,026 = 361,8M\n(osv. per underkategori)",
         f"≈ anslag for {PRED_PREV_YEAR}",
         "CAGR fra Fase 3\nappliseres på faktisk 2025-tall.", GRÅ_BG)
steg_rad(ws5, last_cagr_row+3, "5",
         f"Prediksjon {PRED_YEAR}\n= Prediksjon {PRED_PREV_YEAR} × (1 + CAGR)",
         "Husleie: forrige prediksjon × (1 + vekst)\n(osv. per underkategori)",
         f"≈ anslag for {PRED_YEAR}",
         "Utskifting holdes flat\npå 2025-nivå (42,3M).")

# Sluttresultat-boks
sr_row = last_cagr_row + 4
faseoverskrift(ws5, sr_row, "  Sluttresultat")
ws5.merge_cells(f"B{sr_row+1}:C{sr_row+1}")
c = ws5[f"B{sr_row+1}"]
c.value = f"Prediksjon {PRED_YEAR} — anbefalt anslag"
c.font = Font(name="Arial", bold=True, size=12, color=HEADER_BG)
c.fill = PatternFill("solid", start_color=GRØNN_BG)
c.alignment = Alignment(horizontal="left", vertical="center")
c.border = kant()
ws5.row_dimensions[sr_row+1].height = 28

c2 = ws5.cell(row=sr_row+1, column=4, value=f"='{PRED_SHEET_NAME}'!L{pred_uk_tot_row}")
c2.font = Font(name="Arial", bold=True, size=14, color="006600")
c2.fill = PatternFill("solid", start_color=GRØNN_BG)
c2.number_format = KR_FMT
c2.alignment = Alignment(horizontal="center", vertical="center")
c2.border = kant()

c3 = ws5.cell(row=sr_row+1, column=5, value=f"← justerbart i {PRED_SHEET_NAME}, {ANTAG_SHEET} og per rad")
c3.font = Font(name="Arial", size=10, italic=True, color="595959")
c3.fill = PatternFill("solid", start_color=GRØNN_BG)
c3.alignment = Alignment(horizontal="left", vertical="center")
c3.border = kant()

# ════════════════════════════════════════════════════════════════════════════
# SHEET 12 – SRS OVERSIKT (beregnet fra CSV, alle år 2020–2025)
# ════════════════════════════════════════════════════════════════════════════
wsov = wb.create_sheet("SRS_Oversikt")
wsov.sheet_view.showGridLines = False

# Bygg srs_konto_data: {srs_gruppe: {(konto, konto_t): {år: beløp}}}
srs_konto_data = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
srs_konto_ant  = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
for _r in alle_rader:
    _konto   = _r.get("Konto","").strip()
    _konto_t = _r.get("Konto(T)","").strip()
    _ar      = _r.get("År","").strip()
    _belop   = rens(_r.get("Beløp",""))
    _srs     = bestem_srs(_konto)
    if _ar in ALLE_ÅR:
        srs_konto_data[_srs][(_konto, _konto_t)][_ar] += _belop
        srs_konto_ant[_srs][(_konto, _konto_t)][_ar]  += 1

SRS_GRUPPER = [
    ("SRS 17", "Anleggsmidler og avskrivninger (konto 1040–1298, 3800–3810, 4930–4999, 6000–6071, 6551, 7800)", "DAEEF3"),
    ("SRS 25", "Lønn og reise (konto 5000–5999, 7100)", "EBF1DE"),
    ("SRS 10", "Driftskostnader (alle andre kostnadskontoer ≥3000)", "FFF2CC"),
    ("Annet",  "Øvrige (balansekontoer, inntektskontoer)", "F2F2F2"),
]

# Kolonnebredder
wsov.column_dimensions["A"].width = 3
wsov.column_dimensions["B"].width = 10
wsov.column_dimensions["C"].width = 38
for i, ar in enumerate(ALLE_ÅR):
    wsov.column_dimensions[get_column_letter(4+i)].width = 16
wsov.column_dimensions[get_column_letter(4+len(ALLE_ÅR))].width = 16   # Endring

wsov.merge_cells(f"B2:{get_column_letter(4+len(ALLE_ÅR))}2")
c = wsov["B2"]
c.value = "SRS-oversikt — Kostnader per konto og år (beregnet fra Agresso-data)"
c.font = Font(name="Arial", bold=True, size=13, color=HEADER_BG)
c.alignment = Alignment(horizontal="left", vertical="center")

wsov.merge_cells(f"B3:{get_column_letter(4+len(ALLE_ÅR))}3")
c = wsov["B3"]
c.value = "Kilde: Agresso GL, Innkjøpskategori 01, 2020–2025. Bilagsarter: IV, IW, LE, H1, H2, HB, RE."
c.font = Font(name="Arial", size=9, italic=True, color="595959")
c.alignment = Alignment(horizontal="left")
wsov.row_dimensions[4].height = 8

current_row = 5
for srs_navn, srs_beskr, srs_fargebg in SRS_GRUPPER:
    kontoer = srs_konto_data.get(srs_navn, {})
    if not kontoer:
        continue

    # Gruppeoverskrift
    wsov.merge_cells(f"B{current_row}:{get_column_letter(4+len(ALLE_ÅR))}{current_row}")
    c = wsov[f"B{current_row}"]
    c.value = f"{srs_navn}  —  {srs_beskr}"
    c.font = Font(name="Arial", bold=True, size=11, color=HVIT)
    c.fill = PatternFill("solid", start_color=HEADER_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = kant()
    wsov.row_dimensions[current_row].height = 22
    current_row += 1

    # Header
    header_cell(wsov, current_row, 2, "Konto", center=True)
    header_cell(wsov, current_row, 3, "Kontonavn", center=False)
    for i, ar in enumerate(ALLE_ÅR):
        header_cell(wsov, current_row, 4+i, ar)
    header_cell(wsov, current_row, 4+len(ALLE_ÅR), "Endring\n2020→2025", wrap=True)
    wsov.row_dimensions[current_row].height = 28
    current_row += 1

    # Kontolinjer
    sortert_kontoer = sorted(kontoer.keys(),
                              key=lambda k: -kontoer[k].get("2025", kontoer[k].get("2024", 0)))
    subtotaler = defaultdict(float)
    for ri2, (konto, konto_t) in enumerate(sortert_kontoer):
        bg = srs_fargebg if ri2 % 2 == 0 else HVIT
        c = wsov.cell(row=current_row, column=2, value=konto)
        c.font = Font(name="Arial", size=10)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = kant()
        c2 = wsov.cell(row=current_row, column=3, value=konto_t)
        c2.font = Font(name="Arial", size=10)
        c2.fill = PatternFill("solid", start_color=bg)
        c2.alignment = Alignment(horizontal="left", vertical="center")
        c2.border = kant()
        for i, ar in enumerate(ALLE_ÅR):
            v = kontoer[(konto, konto_t)].get(ar, 0)
            subtotaler[ar] += v
            data_cell(wsov, current_row, 4+i, v, KR_FMT, bg=bg)
        v20 = kontoer[(konto, konto_t)].get("2020", 0)
        v25 = kontoer[(konto, konto_t)].get("2025", 0)
        endring = (v25-v20)/abs(v20) if v20 else 0
        data_cell(wsov, current_row, 4+len(ALLE_ÅR), endring, PCT_FMT, bg=bg)
        wsov.row_dimensions[current_row].height = 18
        current_row += 1

    # Subtotallinje
    c = wsov.cell(row=current_row, column=2, value=srs_navn)
    c.font = Font(name="Arial", bold=True, size=10)
    c.fill = PatternFill("solid", start_color=AKSENT_BG)
    c.border = kant()
    c.alignment = Alignment(horizontal="center")
    c2 = wsov.cell(row=current_row, column=3, value="SUBTOTAL")
    c2.font = Font(name="Arial", bold=True, size=10)
    c2.fill = PatternFill("solid", start_color=AKSENT_BG)
    c2.border = kant()
    c2.alignment = Alignment(horizontal="left")
    for i, ar in enumerate(ALLE_ÅR):
        data_cell(wsov, current_row, 4+i, subtotaler[ar], KR_FMT, bold=True, bg=AKSENT_BG)
    st20 = subtotaler.get("2020", 0)
    st25 = subtotaler.get("2025", 0)
    endring_tot = (st25-st20)/abs(st20) if st20 else 0
    data_cell(wsov, current_row, 4+len(ALLE_ÅR), endring_tot, PCT_FMT, bold=True, bg=AKSENT_BG)
    wsov.row_dimensions[current_row].height = 22
    current_row += 2

wsov.freeze_panes = "D6"
wsov.auto_filter.ref = f"B{5}:{get_column_letter(4+len(ALLE_ÅR))}{current_row}"
print(f"  → SRS_Oversikt: {sum(len(srs_konto_data[g]) for g in srs_konto_data)} kontoer fordelt på {len(SRS_GRUPPER)} SRS-grupper")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 13 – ANLEGGSREGISTER FRA CSV (Konto-basert, alle år)
# ════════════════════════════════════════════════════════════════════════════
wanl = wb.create_sheet("Anleggsregister_CSV")
wanl.sheet_view.showGridLines = False
# Kolonner: A(marg) B(Dim6/anleggsnr) C(Konto) D(Konto(T)) E(Dim1) F(Dim1(T)) G(År) H(Periode) I(Bilagsnr) J(Tekst) K(Beløp)
for ci, w in enumerate([3, 14, 8, 28, 10, 40, 6, 8, 14, 50, 14], 1):
    wanl.column_dimensions[get_column_letter(ci)].width = w

wanl.merge_cells("B2:K2")
c = wanl["B2"]
c.value = "Anleggsregister — SRS 17 anleggsposteringer fra Agresso (beregnet fra CSV)"
c.font = Font(name="Arial", bold=True, size=14, color=HEADER_BG)
c.alignment = Alignment(horizontal="left", vertical="center")

wanl.merge_cells("B3:K3")
c = wanl["B3"]
c.value = (
    "Filtert på anleggskontoer: 1040–1298, 3800–3810, 4930–4999, 6000–6071, 6551, 7800.  "
    "Dim6 = Anleggsnummer (påkrevd for disse kontiene iht. SRS 17).  "
    "Tekst-feltet inneholder Agresso-narrativ om hva anleggsmidlet gjelder."
)
c.font = Font(name="Arial", size=9, italic=True, color="595959")
c.alignment = Alignment(horizontal="left", wrap_text=True)
wanl.row_dimensions[3].height = 32
wanl.row_dimensions[4].height = 8

_anl_hdrs = ["Anleggsnr (Dim6)", "Konto", "Konto(T)", "Dim1-kode", "Dim1-navn",
             "År", "Periode", "Bilagsnr", "Tekst (full)", "Beløp (kr)"]
for ci, h in enumerate(_anl_hdrs):
    header_cell(wanl, 5, 2+ci, h, wrap=True)
wanl.row_dimensions[5].height = 28

# Sorter etter Dim6 (anleggsnr), deretter År, Periode
anlegg_sortert = sorted(anlegg_rader,
                         key=lambda r: (r.get("Dim6","") or "ZZZZ",
                                        r.get("År",""),
                                        r.get("Periode","")))

_prev_dim6 = None
_anl_alt = 0
for ri, row in enumerate(anlegg_sortert):
    r = 6 + ri
    dim6 = row.get("Dim6","").strip()
    if dim6 != _prev_dim6:
        _anl_alt += 1
        _prev_dim6 = dim6
    bg = AKSENT_BG if _anl_alt % 2 == 0 else HVIT
    vals = [
        dim6 or "(mangler)",
        row.get("Konto",""),
        row.get("Konto(T)",""),
        row.get("Dim1",""),
        row.get("Dim1(T)",""),
        row.get("År",""),
        row.get("Periode",""),
        row.get("Bilagsnr",""),
        row.get("Tekst",""),      # full Tekst
        rens(row.get("Beløp","")),
    ]
    for ci, v in enumerate(vals):
        c = wanl.cell(row=r, column=2+ci, value=v)
        c.font = Font(name="Arial", size=9)
        c.fill = PatternFill("solid", start_color=bg)
        c.border = kant()
        is_amt = ci == 9
        c.alignment = Alignment(
            horizontal="right" if is_amt else ("center" if ci in (0,1,3,5,6) else "left"),
            vertical="center", wrap_text=(ci == 8))
        if is_amt:
            c.number_format = KR_FMT

wanl.freeze_panes = "C6"
wanl.auto_filter.ref = f"B5:K{5 + len(anlegg_sortert)}"
print(f"  → Anleggsregister_CSV: {len(anlegg_sortert)} anleggsposteringer")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 14 – SRS SAMSVARRAPPORT (hentet fra BEFS 2025-03-27)
# ════════════════════════════════════════════════════════════════════════════
wsrs = wb.create_sheet("SRS_Samsvarrapport")
wsrs.sheet_view.showGridLines = False
for col, w in zip(range(1, 9), [3, 28, 16, 16, 16, 16, 16, 16]):
    wsrs.column_dimensions[get_column_letter(col)].width = w

wsrs.merge_cells("B2:H2")
c = wsrs["B2"]
c.value = "SRS-samsvarrapport — Statlig Regnskapsstandard"
c.font = Font(name="Arial", bold=True, size=14, color=HEADER_BG)
c.alignment = Alignment(horizontal="left", vertical="center")

wsrs.merge_cells("B3:H3")
c = wsrs["B3"]
c.value = "Kilde: knowme-frontend-amber.vercel.app/financials/srs  |  Alle tilgjengelige år (2020–2025)"
c.font = Font(name="Arial", size=9, italic=True, color="595959")
c.alignment = Alignment(horizontal="left")
wsrs.row_dimensions[4].height = 8

# ── Del 1: SRS-kategorisering per år ──────────────────────────────────────
wsrs.merge_cells("B5:H5")
c = wsrs["B5"]
c.value = "SRS-kategorisering — Drift / Investering / Gjennomstrømning per år"
c.font = Font(name="Arial", bold=True, size=11, color=HVIT)
c.fill = PatternFill("solid", start_color=HEADER_BG)
c.alignment = Alignment(horizontal="left", vertical="center")
c.border = kant()
wsrs.row_dimensions[5].height = 22

# Header
for ci, h in enumerate(["År", "Totalt ant. transaksj.", "Drift — antall",
                         "Drift — beløp", "Gjennomstrøm. — antall",
                         "Gjennomstrøm. — beløp", "Investering — beløp"]):
    header_cell(wsrs, 6, 2+ci, h, wrap=True)
wsrs.row_dimensions[6].height = 32

srs_år_data = [
    # (år, tot_antall, drift_ant, drift_belop, gj_ant, gj_belop, inv_belop)
    ("2025", 29135, 24878, 516806292, 4050,   -70167,   50786376),
    ("2024", 22037, 19452, 499435693, 2585,  5539467,   27662099),
    ("2023", 19439, 16927, 508195023, 2512,  4810862,   27878970),
    ("2022", 21842, 19967, 520216469, 1875,  2304445,   10015873),
    ("2021", 20412, 19077, 366427162, 1335,  -224147,          0),
    ("2020", 23011, 19952, 415581494, 3059,   259829,          0),
]
for ri, (ar, tot, da, db, ga, gb, ib) in enumerate(srs_år_data):
    row = 7 + ri
    bg = AKSENT_BG if ar == "2025" else (GRÅ_BG if ri % 2 == 0 else HVIT)
    bold = ar == "2025"
    vals = [ar, tot, da, db, ga, gb, ib]
    fmts = [None, INT_FMT, INT_FMT, KR_FMT, INT_FMT, KR_FMT, KR_FMT]
    for ci, (v, fmt) in enumerate(zip(vals, fmts)):
        c = wsrs.cell(row=row, column=2+ci, value=v)
        c.font = Font(name="Arial", size=10, bold=bold)
        c.fill = PatternFill("solid", start_color=bg)
        c.border = kant()
        c.alignment = Alignment(horizontal="center" if ci == 0 else "right",
                                 vertical="center")
        if fmt: c.number_format = fmt
    wsrs.row_dimensions[row].height = 20

# ── Del 2: Samsvarsstatus ──────────────────────────────────────────────────
r2 = 7 + len(srs_år_data) + 2
wsrs.merge_cells(f"B{r2}:H{r2}")
c = wsrs[f"B{r2}"]
c.value = "SRS-samsvarsstatus 2025 (fra BEFS-systemet)"
c.font = Font(name="Arial", bold=True, size=11, color=HVIT)
c.fill = PatternFill("solid", start_color=HEADER_BG)
c.alignment = Alignment(horizontal="left", vertical="center")
c.border = kant()
wsrs.row_dimensions[r2].height = 22

for ci, h in enumerate(["Krav", "Status", "Detaljer"]):
    header_cell(wsrs, r2+1, 2+ci, h, center=False)
wsrs.row_dimensions[r2+1].height = 20

srs_status = [
    ("SRS-kategorisering", "✅ Komplett",
     "29 135 transaksjoner kategorisert Drift / Investering / Gjennomstrømning (2025)"),
    ("Koststed-kobling (Dim1 → eiendom)", "⚠️  Delvis",
     "292 av 572 kosteder koblet til BEFS-eiendom (51%). 280 ikke koblet."),
    ("SRS 13 – Leie av lokaler", "✅ Komplett",
     "Konto 6300 (privat): 214 265 140 kr  |  Konto 6310 (Statsbygg): 145 953 567 kr"),
    ("SRS 17 – Anleggsmidler", "✅ Komplett",
     "348 aktive anleggsmidler med avskrivningsplan. Bokført verdi: 260 973 243 kr"),
    ("SRS 10 – Nøytralisering", "📋 Planlagt",
     "Motbilag (33xx/39xx) generert for avskrivninger. Avventer fase 3 (avskrivningsmotor)."),
]
statsbg = {
    "✅ Komplett": GRØNN_BG,
    "⚠️  Delvis":  GULL_BG,
    "📋 Planlagt": GRÅ_BG,
}
for ri, (krav, status, detaljer) in enumerate(srs_status):
    row = r2 + 2 + ri
    bg = statsbg.get(status, HVIT)
    for ci, v in enumerate([krav, status, detaljer]):
        c = wsrs.cell(row=row, column=2+ci, value=v)
        c.font = Font(name="Arial", size=10)
        c.fill = PatternFill("solid", start_color=bg)
        c.border = kant()
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    # Slå sammen detalj-kolonner B+C (col 4-8)
    try:
        wsrs.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
    except Exception:
        pass
    wsrs.row_dimensions[row].height = 28

# ── Del 3: SRS 13 – Leieavtale-kontoer ────────────────────────────────────
r3 = r2 + 2 + len(srs_status) + 2
wsrs.merge_cells(f"B{r3}:H{r3}")
c = wsrs[f"B{r3}"]
c.value = "SRS 13 — Leieavtaler 2025 per konto"
c.font = Font(name="Arial", bold=True, size=11, color=HVIT)
c.fill = PatternFill("solid", start_color=HEADER_BG)
c.alignment = Alignment(horizontal="left", vertical="center")
c.border = kant()
wsrs.row_dimensions[r3].height = 22

for ci, h in enumerate(["Konto", "Beskrivelse", "Type", "Antall bilag", "Total beløp"]):
    header_cell(wsrs, r3+1, 2+ci, h, center=(ci >= 3))
wsrs.row_dimensions[r3+1].height = 20

leie_data = [
    ("6300", "Leie lokaler andre utleiere", "Privat",     4569, 214265140),
    ("6310", "Leie lokaler fra Statsbygg",  "Statsbygg",  2427, 145953567),
]
for ri, (konto, besk, type_, ant, belop) in enumerate(leie_data):
    row = r3 + 2 + ri
    bg = GRÅ_BG if ri % 2 == 0 else HVIT
    for ci, v in enumerate([konto, besk, type_, ant, belop]):
        c = wsrs.cell(row=row, column=2+ci, value=v)
        c.font = Font(name="Arial", size=10)
        c.fill = PatternFill("solid", start_color=bg)
        c.border = kant()
        c.alignment = Alignment(horizontal="center" if ci in [0,2,3] else "left",
                                 vertical="center")
        if ci == 3: c.number_format = INT_FMT
        if ci == 4: c.number_format = KR_FMT
    wsrs.row_dimensions[row].height = 20

# Totallinje
r3_tot = r3 + 2 + len(leie_data)
for ci, v in enumerate(["", "TOTAL", "", sum(r[3] for r in leie_data),
                         sum(r[4] for r in leie_data)]):
    c = wsrs.cell(row=r3_tot, column=2+ci, value=v)
    c.font = Font(name="Arial", bold=True, size=10)
    c.fill = PatternFill("solid", start_color=AKSENT_BG)
    c.border = kant()
    c.alignment = Alignment(horizontal="center" if ci in [0,2,3] else "left", vertical="center")
    if ci == 3: c.number_format = INT_FMT
    if ci == 4: c.number_format = KR_FMT
wsrs.row_dimensions[r3_tot].height = 20

# ════════════════════════════════════════════════════════════════════════════
# SHEET 13 – ANLEGGSREGISTER SRS 17
# ════════════════════════════════════════════════════════════════════════════
wan = wb.create_sheet("Anleggsregister_SRS17")
wan.sheet_view.showGridLines = False
for col, w in zip(range(1, 10), [3, 8, 14, 16, 16, 16, 14, 14, 14]):
    wan.column_dimensions[get_column_letter(col)].width = w

wan.merge_cells("B2:I2")
c = wan["B2"]
c.value = "Anleggsregister — SRS 17 Balanseførte anleggsmidler"
c.font = Font(name="Arial", bold=True, size=14, color=HEADER_BG)
c.alignment = Alignment(horizontal="left", vertical="center")

wan.merge_cells("B3:I3")
c = wan["B3"]
c.value = "Kilde: knowme-frontend-amber.vercel.app/financials/anlegg  |  Lineær avskrivning over gjenværende leieperiode"
c.font = Font(name="Arial", size=9, italic=True, color="595959")
c.alignment = Alignment(horizontal="left")
wan.row_dimensions[4].height = 8

# ── Nøkkeltall-bokser ───────────────────────────────────────────────────────
nøkkel = [
    ("B", "D", "Anleggsmidler\n(aktive poster)",  348,           None),
    ("E", "G", "Bokført verdi\n(01.01.2025)",      260973243,     KR_FMT),
    ("H", "I", "Årlig avskrivning\n(konto 6010)",  5369354,       KR_FMT),
]
for c1, c2, tittel, verdi, fmt in nøkkel:
    wan.merge_cells(f"{c1}5:{c2}5")
    t = wan[f"{c1}5"]
    t.value = tittel
    t.font = Font(name="Arial", bold=True, size=10, color="404040")
    t.fill = PatternFill("solid", start_color=AKSENT_BG)
    t.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    t.border = kant()
    wan.row_dimensions[5].height = 30

    wan.merge_cells(f"{c1}6:{c2}6")
    v = wan[f"{c1}6"]
    v.value = verdi
    v.font = Font(name="Arial", bold=True, size=16, color=HEADER_BG)
    v.fill = PatternFill("solid", start_color=AKSENT_BG)
    v.alignment = Alignment(horizontal="center", vertical="center")
    v.border = kant()
    if fmt: v.number_format = fmt
    wan.row_dimensions[6].height = 32

wan.row_dimensions[7].height = 10

# ── Avskrivningsplan 2025–2033 ─────────────────────────────────────────────
wan.merge_cells("B8:I8")
c = wan["B8"]
c.value = "Avskrivningsplan SRS 17 — 2025 til 2033"
c.font = Font(name="Arial", bold=True, size=11, color=HVIT)
c.fill = PatternFill("solid", start_color=HEADER_BG)
c.alignment = Alignment(horizontal="left", vertical="center")
c.border = kant()
wan.row_dimensions[8].height = 22

for ci, h in enumerate(["År", "Restverdi inng.", "Avskrivning", "Restverdi utg."]):
    header_cell(wan, 9, 2+ci, h)
wan.row_dimensions[9].height = 20

avskr_plan = [
    (2025, 260973243,  -5369354, 255603889),
    (2026, 255603889,  -5281068, 250322821),
    (2027, 250322821,  -4734124, 245588697),
    (2028, 245588697,  -4675174, 240913523),
    (2029, 240913523,  -4617887, 236295636),
    (2030, 236295636,  -4481877, 231813759),
    (2031, 231813759,  -3353635, 228460124),
    (2032, 228460124,  -3345481, 225114643),
    (2033, 225114643,  -3193674,          0),
]
for ri, (ar, inn, avskr, ut) in enumerate(avskr_plan):
    row = 10 + ri
    is_target_year = ar == PRED_YEAR
    bg = GULL_BG if is_target_year else (GRØNN_BG if ar == 2025 else (GRÅ_BG if ri % 2 == 0 else HVIT))
    for ci, v in enumerate([ar, inn, avskr, ut]):
        c = wan.cell(row=row, column=2+ci, value=v)
        c.font = Font(name="Arial", size=10, bold=is_target_year)
        c.fill = PatternFill("solid", start_color=bg)
        c.border = kant()
        c.alignment = Alignment(horizontal="center" if ci == 0 else "right", vertical="center")
        if ci > 0: c.number_format = KR_FMT
    wan.row_dimensions[row].height = 20
    if is_target_year:
        wan.cell(row=row, column=6, value="← Prediksjonshorisonten").font = Font(
            name="Arial", size=9, italic=True, color="595959")

wan.row_dimensions[10 + len(avskr_plan)].height = 8

# ── Topp anleggsmidler ──────────────────────────────────────────────────────
r_anl = 10 + len(avskr_plan) + 2
wan.merge_cells(f"B{r_anl}:I{r_anl}")
c = wan[f"B{r_anl}"]
c.value = "Topp anleggsmidler etter anskaffelseskost (348 aktive poster totalt)"
c.font = Font(name="Arial", bold=True, size=11, color=HVIT)
c.fill = PatternFill("solid", start_color=HEADER_BG)
c.alignment = Alignment(horizontal="left", vertical="center")
c.border = kant()
wan.row_dimensions[r_anl].height = 22

for ci, h in enumerate(["Dim6", "Koststed", "Anleggsnavn", "Anskaffelse", "Bokført", "Status"]):
    header_cell(wan, r_anl+1, 2+ci, h, center=(ci in [0,1,5]))
wan.row_dimensions[r_anl+1].height = 20
wan.column_dimensions["D"].width = 44  # Navn-kolonne bredere

anlegg_topp = [
    ("499990", "Div.", "Påkostninger Eikelund Ungdomssenter",          33022236, 33022236),
    ("299990", "Div.", "Påkostninger Innlandet barnevernssenter",        5700000,  5700000),
    ("403208", "Driftsavd.", "Eikelund ombygging jf. b. 800312760",    5179831,  5179831),
    ("403208", "Driftsavd.", "Eikelund ombygging, ref. B800312760",    5179831,  5179831),
    ("284416", "Div.", "Nordsetvegen 27, investeringer i ombygging",   5000000,  5000000),
    ("403208", "Driftsavd.", "Ombygging på Eikelund bygg F",            4658025,  4658025),
    ("403208", "Driftsavd.", "Eikelund ombygging - September/Oktober", 4638314,  4638314),
    ("403208", "Driftsavd.", "Avtale 2024/50584 Eikelund Bygg F",      4476233,  4476233),
    ("403208", "Driftsavd.", "Ombygging på Eikelund",                   4402219,  4402219),
    ("599990", "Div.", "Påkostninger Ranheim 2025",                    4200000,  4200000),
    ("299990", "Div.", "Historisk anlegg - overf. anlegg 1000+",       3980791,  3980791),
    ("321204", "Div.", "Ombygging hovedhuset til 6-13",                3219438,  3219438),
]
for ri, (dim6, kst, navn, ansk, bokf) in enumerate(anlegg_topp):
    row = r_anl + 2 + ri
    bg = GRÅ_BG if ri % 2 == 0 else HVIT
    for ci, v in enumerate([dim6, kst, navn, ansk, bokf, "AKTIV"]):
        c = wan.cell(row=row, column=2+ci, value=v)
        c.font = Font(name="Arial", size=9)
        c.fill = PatternFill("solid", start_color=bg)
        c.border = kant()
        c.alignment = Alignment(
            horizontal="center" if ci in [0,1,5] else ("left" if ci == 2 else "right"),
            vertical="center", wrap_text=(ci == 2))
        if ci in [3, 4]: c.number_format = KR_FMT
        if ci == 5:
            c.font = Font(name="Arial", size=9, color="006600", bold=True)
    wan.row_dimensions[row].height = 18

# Note
r_note_a = r_anl + 2 + len(anlegg_topp) + 1
wan.merge_cells(f"B{r_note_a}:I{r_note_a}")
n = wan[f"B{r_note_a}"]
n_target = next((x for x in avskr_plan if x[0] == PRED_YEAR), None)
if n_target:
    _, _, avskr_target, ut_target = n_target
    note_year_text = f"Avskrivning {PRED_YEAR}: {avskr_target:,.0f} kr  →  Restverdi utg. {PRED_YEAR}: {ut_target:,.0f} kr."
else:
    note_year_text = f"Prediksjonsår {PRED_YEAR} finnes ikke i avskrivningsplanen."
n.value = ("Viser topp 12 av 348 aktive anleggsmidler sortert etter anskaffelseskost. "
           "Total bokført verdi per 01.01.2025: 260 973 243 kr. "
           f"{note_year_text}")
n.font = Font(name="Arial", size=9, italic=True, color="595959")
n.fill = PatternFill("solid", start_color=GRÅ_BG)
n.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
n.border = kant()
wan.row_dimensions[r_note_a].height = 32

# Frys topprader
wan.freeze_panes = "B10"

# ════════════════════════════════════════════════════════════════════════════
# LAGRE
# ════════════════════════════════════════════════════════════════════════════
# Automatisk beregning + beskyttelse av formelceller (bruker redigerer C–F, H, K)
try:
    calc = getattr(wb, "calculation", None)
    if calc is not None:
        calc.calcMode = "auto"
        calc.fullCalcOnLoad = True
except Exception:
    pass

_sp = dict(
    sheet=True,
    selectLockedCells=True,
    selectUnlockedCells=True,
    formatCells=False,
    formatRows=False,
    formatColumns=False,
)

def _apply_pred_uk_protection() -> None:
    wpp = wb[PRED_SHEET_NAME]
    fr, lr = first_data_row, last_data_row
    tr = pred_uk_tot_row
    for row in range(fr, lr + 1):
        for col in (2, 3, 4, 5, 6, 8, 11):
            wpp.cell(row=row, column=col).protection = Protection(locked=False)
        for col in (9, 12):
            wpp.cell(row=row, column=col).protection = Protection(locked=True)
    for col in range(2, 13):
        wpp.cell(row=tr, column=col).protection = Protection(locked=True)
    wpp.protection = SheetProtection(**_sp)


def _apply_antagelser_protection() -> None:
    wa = wb[ANTAG_SHEET]
    wa.cell(row=6, column=2).protection = Protection(locked=False)
    wa.cell(row=7, column=2).protection = Protection(locked=False)
    wa.protection = SheetProtection(**_sp)


def _apply_per_eiendom_protection() -> None:
    wep = wb["Prediksjon_Per_Eiendom"]
    for row in range(PER_EIENDOM_FIRST_ROW, PER_EIENDOM_LAST_DATA_ROW + 1):
        for col in (2, 3, 4, 5, 6, 7, 8, 9, 11):
            wep.cell(row=row, column=col).protection = Protection(locked=False)
        for col in (10, 12):
            wep.cell(row=row, column=col).protection = Protection(locked=True)
    for col in range(2, 13):
        wep.cell(row=PER_EIENDOM_TOT_ROW, column=col).protection = Protection(locked=True)
    wep.protection = SheetProtection(**_sp)


if not args.no_protect:
    _apply_pred_uk_protection()
    _apply_antagelser_protection()
    _apply_per_eiendom_protection()

wb.save(str(UT_FIL))
print(f"\n✅  Lagret: {UT_FIL}")
print(f"   Sheets: {', '.join(wb.sheetnames)}")
