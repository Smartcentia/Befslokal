#!/usr/bin/env python3
"""
Sammenlign Master_enheter_register.xlsx med BEFS-databasen (lønnskostnader).

Samme eiendoms-oppløsning som import_salary_from_master_xlsx.py:
  Dim1 → property via unit_id_erp eller koststed_kode (første treff per nøkkel),
  deretter fuzzy navn (≥ 0,88) med global «én eiendom per navnetreff»-begrensning.

Lønn i DB sammenlignes mot summen for den **oppløste** property_id (ikke bare koststed=Dim1).

Kjøring:
  cd backend && python3 scripts/compare_master_enheter_db_costs.py --master /path/Master_enheter_register.xlsx
  railway run -- bash -c 'cd backend && python3 scripts/compare_master_enheter_db_costs.py --master ... --csv rapport.csv'

Krever DATABASE_URL (f.eks. railway run eller backend/.env).
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import re
import sys
from collections import defaultdict
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Optional

from dotenv import load_dotenv

load_dotenv(Path(__file__).resolve().parent.parent / ".env")

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from sqlalchemy import text

import app.db.base  # noqa: F401
from app.db.session import SessionLocal


def _norm_dim(v: Any) -> str | None:
    if v is None or v == "":
        return None
    s = str(v).strip()
    if not s:
        return None
    try:
        return str(int(float(s)))
    except (TypeError, ValueError):
        return s


def _f(x: Any) -> float | None:
    if x is None or x == "":
        return None
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def _fuzzy_ratio(a: str, b: str) -> float:
    return SequenceMatcher(None, a.lower().strip(), b.lower().strip()).ratio()


def _find_property_by_name(
    navn: str,
    name_pairs: list[tuple[str, str]],
    used: set[str],
    cutoff: float = 0.88,
) -> Optional[str]:
    if not navn or not navn.strip():
        return None
    best_id: Optional[str] = None
    best_sc = 0.0
    nl = navn.lower().strip()
    for pid, pname in name_pairs:
        if pid in used or not pname:
            continue
        pl = pname.lower().strip()
        if nl == pl:
            return pid
        sc = _fuzzy_ratio(navn, pname)
        if len(pl) >= 6 and len(nl) >= 6 and (nl in pl or pl in nl):
            sc = max(sc, 0.9)
        if sc > best_sc:
            best_sc = sc
            best_id = pid
    if best_id is not None and best_sc >= cutoff:
        return best_id
    return None


def _load_master_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Master_enheter" not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"Fant ikke arket «Master_enheter» i {path}")
    ws = wb["Master_enheter"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    header = [str(h) if h is not None else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(header) if h}
    ycols: list[str] = []
    for h in header:
        m = re.match(r"^Lønn_(\d{4})_INNJKOP$", h)
        if m:
            ycols.append(m.group(1))
    ycols = sorted(ycols)
    out: list[dict[str, Any]] = []
    for r in rows[1:]:
        if not r:
            continue
        d1 = r[idx.get("Dim1", -1)] if "Dim1" in idx else None
        nk = _norm_dim(d1)
        if not nk:
            continue
        poster = None
        if "Poster_AGRESSO" in idx:
            pv = r[idx["Poster_AGRESSO"]]
            try:
                poster = int(float(pv)) if pv is not None and str(pv).strip() != "" else None
            except (TypeError, ValueError):
                poster = None
        rowd: dict[str, Any] = {
            "dim1": nk,
            "enhet_agresso": r[idx["Enhet_navn_AGRESSO"]] if "Enhet_navn_AGRESSO" in idx else "",
            "kryssjekk": r[idx["Kryssjekk"]] if "Kryssjekk" in idx else "",
            "poster_agresso": poster,
            "sum_agresso": _f(r[idx["Sum_beløp_AGRESSO"]]) if "Sum_beløp_AGRESSO" in idx else None,
        }
        for y in ycols:
            col = f"Lønn_{y}_INNJKOP"
            if col in idx:
                rowd[f"excel_{y}"] = _f(r[idx[col]])
            else:
                rowd[f"excel_{y}"] = None
        out.append(rowd)
    return ycols, out


async def _map_dim_to_property_db(db) -> dict[str, str]:
    sql = text(
        """
        SELECT property_id::text AS property_id, unit_id_erp, koststed_kode
        FROM properties
        WHERE (unit_id_erp IS NOT NULL AND TRIM(unit_id_erp) <> '')
           OR (koststed_kode IS NOT NULL AND TRIM(koststed_kode) <> '')
        """
    )
    by_dim: dict[str, str] = {}
    result = await db.execute(sql)
    for row in result.mappings().all():
        pid = row["property_id"]
        for key in ("unit_id_erp", "koststed_kode"):
            d = _norm_dim(row[key])
            if d and d not in by_dim:
                by_dim[d] = pid
    return by_dim


async def _fetch_property_names_db(db) -> list[tuple[str, str]]:
    r = await db.execute(
        text(
            """
            SELECT property_id::text, COALESCE(NULLIF(TRIM(name), ''), '') AS n
            FROM properties
            """
        )
    )
    return [(str(row[0]), str(row[1])) for row in r.fetchall() if row[1]]


async def _load_db_salary_by_property_db(db) -> dict[str, dict[int, float]]:
    sql = text(
        """
        SELECT
            sc.property_id::text AS pid,
            sc.year,
            COALESCE(sc.faste_stillinger, 0)::float + COALESCE(sc.vikarer, 0)::float
            + COALESCE(sc.arbeidsgiveravgift, 0)::float AS total
        FROM salary_costs sc
        WHERE sc.property_id IS NOT NULL
        """
    )
    by_prop: dict[str, dict[int, float]] = defaultdict(lambda: defaultdict(float))
    result = await db.execute(sql)
    for row in result.mappings().all():
        pid = row["pid"]
        by_prop[pid][int(row["year"])] += float(row["total"] or 0)
    return {k: dict(v) for k, v in by_prop.items()}


async def _load_db_salary_rowcount_by_property_db(db) -> dict[str, int]:
    sql = text(
        """
        SELECT sc.property_id::text AS pid, COUNT(*)::int AS n
        FROM salary_costs sc
        WHERE sc.property_id IS NOT NULL
        GROUP BY sc.property_id
        """
    )
    out: dict[str, int] = {}
    result = await db.execute(sql)
    for row in result.mappings().all():
        out[row["pid"]] = int(row["n"])
    return out


async def _load_db_properties_by_dim_both_db(db) -> dict[str, list[dict[str, Any]]]:
    sql = text(
        """
        SELECT
            TRIM(p.koststed_kode) AS ks,
            TRIM(p.unit_id_erp) AS ue,
            p.property_id::text AS property_id,
            COALESCE(NULLIF(TRIM(p.name), ''), NULLIF(TRIM(p.address), ''), '') AS visningsnavn
        FROM properties p
        WHERE (p.koststed_kode IS NOT NULL AND TRIM(p.koststed_kode) <> '')
           OR (p.unit_id_erp IS NOT NULL AND TRIM(p.unit_id_erp) <> '')
        """
    )
    by_dim: dict[str, list[dict[str, Any]]] = defaultdict(list)
    result = await db.execute(sql)
    for row in result.mappings().all():
        pid = row["property_id"]
        vis = row["visningsnavn"]
        for keycol in ("ks", "ue"):
            nk = _norm_dim(row[keycol])
            if not nk:
                continue
            lst = by_dim[nk]
            if not any(x["property_id"] == pid for x in lst):
                lst.append({"property_id": pid, "visningsnavn": vis})
    return dict(by_dim)


async def _db_salary_orphans_summary() -> dict[str, int]:
    """Lønndata i DB som ikke kan kobles til Dim1 via koststed_kode."""
    sql = text(
        """
        SELECT
            COUNT(*) FILTER (
                WHERE p.koststed_kode IS NULL OR TRIM(p.koststed_kode) = ''
            )::int AS salary_rows_uten_koststed,
            COUNT(*)::int AS salary_rows_totalt
        FROM salary_costs sc
        JOIN properties p ON p.property_id = sc.property_id
        WHERE sc.property_id IS NOT NULL
        """
    )
    async with SessionLocal() as db:
        row = (await db.execute(sql)).mappings().first()
        return {
            "salary_rows_uten_koststed": int(row["salary_rows_uten_koststed"] or 0),
            "salary_rows_totalt": int(row["salary_rows_totalt"] or 0),
        }


def _row_status(
    excel_vals: dict[int, float | None],
    db_vals: dict[int, float],
    years: list[str],
) -> str:
    any_excel = any(excel_vals.get(int(y)) is not None for y in years)
    any_db = any(db_vals.get(int(y), 0) != 0 for y in years)
    if not any_excel and not any_db:
        return "ingen_lonn_hverken"
    if any_excel and not any_db:
        return "master_har_lonn_mangler_i_db"
    if not any_excel and any_db:
        return "db_har_lonn_mangler_i_master"
    return "har_lonn_i_begge"


def _close_ratio(a: float, b: float, tol: float) -> bool:
    if a == 0 and b == 0:
        return True
    if a == 0 or b == 0:
        return False
    return abs(a - b) / max(abs(a), abs(b)) <= tol


def _master_forventer_innkjop_lonn(m: dict[str, Any], years: list[str]) -> bool:
    if (m.get("kryssjekk") or "").strip() == "begge_kilder":
        return True
    return any((m.get(f"excel_{y}") or 0) != 0 for y in years)


def _master_har_agresso_data(m: dict[str, Any]) -> bool:
    if (m.get("poster_agresso") or 0) > 0:
        return True
    s = m.get("sum_agresso")
    return s is not None and s != 0


def _gap_beskrivelse(
    mangler_eiendom: bool,
    forventer_innkjop: bool,
    salary_rows: int,
    har_avvik: bool,
) -> str:
    if mangler_eiendom:
        if forventer_innkjop:
            return "Ingen eiendom (import-logikk: Dim1→unit/koststed, deretter fuzzy navn); lønn i DB ikke koblet"
        return "Ingen eiendom for Dim1 i BEFS (sjekk unit_id_erp/koststed_kode og navn)"
    if forventer_innkjop and salary_rows == 0:
        return "Eiendom funnet, men ingen salary_costs-rader (mangler import?)"
    if forventer_innkjop and har_avvik:
        return "Har salary_costs, men avvik mot master (innkjøp) for minst ett år"
    if forventer_innkjop:
        return "Innkjøp-lønn i master og data i DB (sjekk detaljkolonner)"
    return "Master forventer ikke innkjøp-lønn for raden; Dim1-kobling evt. kun AGRESSO"


async def run(master_path: Path, csv_path: Path | None, tol: float) -> None:
    years, master = _load_master_rows(master_path)
    if not master:
        print("Ingen rader med Dim1 i master.")
        return

    print("=" * 72)
    print("Master enheter ↔ salary_costs (samme oppløsning som import fra master-xlsx)")
    print(f"Master-fil: {master_path}")
    print(f"Start: {datetime.now().isoformat(timespec='seconds')}")
    print(f"År fra master: {', '.join(years)}")
    print("=" * 72)

    async with SessionLocal() as db:
        dim_map = await _map_dim_to_property_db(db)
        name_pairs = await _fetch_property_names_db(db)
        db_sal_by_prop = await _load_db_salary_by_property_db(db)
        db_salary_rows_by_prop = await _load_db_salary_rowcount_by_property_db(db)
        db_props = await _load_db_properties_by_dim_both_db(db)

    orphans = await _db_salary_orphans_summary()

    used_name_match: set[str] = set()
    resolution_order: list[tuple[str, Optional[str], str]] = []
    for m in master:
        dim = m["dim1"]
        navn = (m.get("enhet_agresso") or "").strip() or str(dim)
        pid: Optional[str] = dim_map.get(dim)
        how = "dim1_kart"
        if not pid:
            pid = _find_property_by_name(navn, name_pairs, used_name_match, cutoff=0.88)
            if pid:
                how = "fuzzy_navn"
                used_name_match.add(pid)
            else:
                how = "ingen"
        resolution_order.append((dim, pid, how))

    rows_out: list[dict[str, Any]] = []
    stats = defaultdict(int)
    gap_stats = defaultdict(int)
    res_stats = defaultdict(int)

    for m, (dim, resolved_pid, kobling) in zip(master, resolution_order):
        res_stats[kobling] += 1
        props = db_props.get(dim, [])
        sy = db_sal_by_prop.get(resolved_pid, {}) if resolved_pid else {}
        excel_by_y = {int(y): m.get(f"excel_{y}") for y in years}
        st = _row_status(excel_by_y, sy, years)
        stats[st] += 1

        if resolved_pid:
            prop_ids = resolved_pid[:8]
            extra = [p for p in props if p["property_id"] != resolved_pid][:2]
            if extra:
                prop_ids += ";" + ";".join(p["property_id"][:8] for p in extra)
        else:
            prop_ids = ";".join(p["property_id"][:8] for p in props[:3])
            if len(props) > 3:
                prop_ids += f";+{len(props) - 3}"

        n_props = len(props)
        mangler_eiendom = resolved_pid is None
        forventer_innkjop = _master_forventer_innkjop_lonn(m, years)
        sal_rows = db_salary_rows_by_prop.get(resolved_pid, 0) if resolved_pid else 0
        har_agresso = _master_har_agresso_data(m)

        mism = 0
        close = 0
        for y in years:
            yi = int(y)
            ex = excel_by_y.get(yi)
            dbv = sy.get(yi, 0.0)
            if ex is None:
                continue
            if _close_ratio(ex, dbv, tol):
                close += 1
            else:
                mism += 1
        har_avvik = mism > 0
        gap_ingen_eiendom = mangler_eiendom
        gap_ingen_lonn_import = (not mangler_eiendom) and forventer_innkjop and sal_rows == 0
        gap_besk = _gap_beskrivelse(mangler_eiendom, forventer_innkjop, sal_rows, har_avvik)

        if gap_ingen_eiendom and har_agresso:
            gap_stats["dim1_med_agresso_i_master_uten_eiendom_i_db"] += 1
        if gap_ingen_eiendom:
            gap_stats["dim1_uten_import_oppløsning"] += 1
        if gap_ingen_lonn_import:
            gap_stats["dim1_med_eiendom_men_ingen_salary_costs"] += 1
        if forventer_innkjop and (mangler_eiendom or sal_rows == 0):
            gap_stats["master_forventer_innkjop_men_db_mangler_kobling_eller_lonn"] += 1
        if forventer_innkjop and har_avvik and not gap_ingen_lonn_import and not mangler_eiendom:
            gap_stats["har_lonn_men_avvik_mot_master"] += 1

        row: dict[str, Any] = {
            "kobling_som_import": kobling,
            "resolved_property_id": resolved_pid or "",
            "gap_ingen_eiendom_for_dim1": gap_ingen_eiendom,
            "gap_ingen_salary_costs_import": gap_ingen_lonn_import,
            "gap_master_forventer_innkjop_lonn": forventer_innkjop,
            "gap_master_har_agresso_posteringer": har_agresso,
            "db_antall_salary_costs_rader": sal_rows,
            "gap_forklaring": gap_besk,
            "dim1": dim,
            "enhet_agresso": m.get("enhet_agresso") or "",
            "kryssjekk": m.get("kryssjekk") or "",
            "poster_agresso": m.get("poster_agresso"),
            "sum_belop_agresso": m.get("sum_agresso"),
            "antall_properties_med_dim1": n_props,
            "property_ids_kort": prop_ids,
            "status_lønn": st,
            "år_med_avvik_utenfor_tol": mism,
            "år_match_tol": close,
        }
        for y in years:
            yi = int(y)
            ex = excel_by_y.get(yi)
            dbv = sy.get(yi, 0.0)
            row[f"excel_lønn_{y}"] = ex
            row[f"db_lønn_total_{y}"] = round(dbv, 2)
            if ex is None:
                row[f"diff_{y}"] = None
                row[f"avvik_pct_{y}"] = None
                continue
            diff = ex - dbv
            row[f"diff_{y}"] = round(diff, 2)
            if ex == 0 and dbv == 0:
                row[f"avvik_pct_{y}"] = 0.0
            elif max(abs(ex), abs(dbv)) > 0:
                row[f"avvik_pct_{y}"] = round(100.0 * diff / max(abs(ex), abs(dbv)), 2)
            else:
                row[f"avvik_pct_{y}"] = None
        rows_out.append(row)

    dim_master = {m["dim1"] for m in master}
    dim_db = set(db_props.keys())
    print(f"\nRader i master (med Dim1): {len(master)}")
    print(f"Unike Dim1 i master: {len(dim_master)}")
    print(f"Unike Dim1-nøkler i DB (koststed eller unit_id_erp): {len(dim_db)}")
    print(f"Dim1 i master uten eiendom med samme koststed/EnhetID: {len(dim_master - dim_db)}")
    print(f"Dim1 i DB (koststed/EnhetID) som ikke finnes i master-fil: {len(dim_db - dim_master)}")
    print("\n--- Oppløsning (som import_salary_from_master_xlsx) ---")
    for k in sorted(res_stats.keys()):
        print(f"  {k}: {res_stats[k]}")
    print("\n--- DB-orphans (lønn uten koststed_kode på eiendom) ---")
    print(
        f"  salary_costs-rader uten koststed_kode: {orphans['salary_rows_uten_koststed']} "
        f"/ totalt {orphans['salary_rows_totalt']}"
    )
    print("\n--- Mangler i BEFS sett fra master (gap) ---")
    for k, v in sorted(gap_stats.items(), key=lambda x: -x[1]):
        print(f"  {k}: {v}")
    print("\nStatus (lønn master vs DB):")
    for k in sorted(stats.keys()):
        print(f"  {k}: {stats[k]}")

    tol_mismatch = sum(1 for r in rows_out if r["år_med_avvik_utenfor_tol"] > 0)
    print(f"\nRader med minst ett år utenfor {tol:.0%} relativt avvik (der begge har tall): {tol_mismatch}")

    if csv_path:
        fieldnames = list(rows_out[0].keys()) if rows_out else []
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            w.writerows(rows_out)
        print(f"\nCSV: {csv_path}")
    else:
        bad = [
            r
            for r in rows_out
            if r["gap_ingen_eiendom_for_dim1"] or r["gap_ingen_salary_costs_import"]
        ]
        bad.sort(
            key=lambda x: (
                0 if x["gap_ingen_eiendom_for_dim1"] else 1,
                -(x.get("sum_belop_agresso") or 0),
            )
        )
        print("\n--- Opptil 15 rader med data-mangel (gap) ---")
        for r in bad[:15]:
            g = r["gap_forklaring"]
            if len(g) > 55:
                g = g[:55] + "…"
            print(f"  Dim1={r['dim1']} {str(r['enhet_agresso'])[:36]} | {g}")


def main() -> None:
    ap = argparse.ArgumentParser(description="Sammenlign Master_enheter med salary_costs")
    ap.add_argument("--master", type=Path, required=True, help="Master_enheter_register.xlsx")
    ap.add_argument("--csv", type=Path, help="Skriv full sammenligning til CSV")
    ap.add_argument(
        "--tol",
        type=float,
        default=0.02,
        help="Relativ toleranse for «match» per år (default 2%%)",
    )
    args = ap.parse_args()
    asyncio.run(run(args.master, args.csv, args.tol))


if __name__ == "__main__":
    main()
