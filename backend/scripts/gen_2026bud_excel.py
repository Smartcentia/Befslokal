"""
Regenerate 2026bud.xlsx with corrected regional distribution.
Run via: railway run --service BEFS1 python3 -m app.scripts.gen_2026bud_excel
Or: DATABASE_URL=... python3 backend/scripts/gen_2026bud_excel.py
"""
import os
import sys
import zipfile
from datetime import date
from collections import defaultdict

import psycopg2
import psycopg2.extras
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

DATABASE_URL = os.environ.get("DATABASE_URL", "")
if not DATABASE_URL:
    print("ERROR: DATABASE_URL not set", file=sys.stderr)
    sys.exit(1)

# Regionale vekstrater brukt i BEFS-prediksjonen (reverse-engineered fra økonomibudsjettet)
REGION_GROWTH = {
    "Øst":        (1.1701, "+17,0%"),
    "Sør":        (0.8922, "−10,8%"),
    "Nord":       (1.2662, "+26,6%"),
    "Vest":       (0.8508, "−14,9%"),
    "Midt-Norge": (1.0213, " +2,1%"),
    "Bufdir":     (1.1423, "+14,2%"),
    None:         (1.0350, " +3,5%"),   # Nasjonal fallback
}

# Convert asyncpg URL to psycopg2
db_url = DATABASE_URL.replace("postgresql+asyncpg://", "postgresql://")

print("Connecting to DB...")
conn = psycopg2.connect(db_url)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

# ── 1. Fetch GL 2025 netto per property (riktig konto-klassifisering) ──
# Kilde: gl_transactions med korrigert srs_kategori (oppdatert 2026-05-09)
# Erstatter finance_budget.kontant_2025 som hadde feil kategori på 9 kontoer
cur.execute("""
    SELECT
        p.property_id,
        p.name,
        p.region,
        SUM(CASE WHEN gt.belop > 0 THEN gt.belop ELSE 0 END) AS regn25,
        SUM(CASE WHEN gt.srs_kategori = 'Lokaler'     THEN gt.belop ELSE 0 END) AS lokaler_25,
        SUM(CASE WHEN gt.srs_kategori = 'Vedlikehold' THEN gt.belop ELSE 0 END) AS vedlikehold_25,
        SUM(CASE WHEN gt.srs_kategori = 'Drift'       THEN gt.belop ELSE 0 END) AS drift_25
    FROM gl_transactions gt
    JOIN properties p ON gt.property_id::text = p.property_id::text
    WHERE gt.ar = 2025
    GROUP BY p.property_id, p.name, p.region
    HAVING SUM(gt.belop) > 0
    ORDER BY p.region NULLS LAST, SUM(gt.belop) DESC
""")
rows_regn25 = {r["property_id"]: r for r in cur.fetchall()}
print(f"  GL 2025 netto: {len(rows_regn25)} properties")

# ── 1b. Fetch GL 2025 konto-detaljer per property (for kontostruktur-fane) ──
cur.execute("""
    SELECT
        p.property_id,
        p.name,
        p.region,
        gt.konto,
        gt.konto_navn,
        gt.srs_kategori,
        SUM(gt.belop) AS netto
    FROM gl_transactions gt
    JOIN properties p ON gt.property_id::text = p.property_id::text
    WHERE gt.ar = 2025
      AND gt.konto IN ('1268','4960','6300','6310','6320','6340','6360','6364',
                       '6365','6390','6391','6395','6396','6398','6630','6632','6662')
    GROUP BY p.property_id, p.name, p.region, gt.konto, gt.konto_navn, gt.srs_kategori
    HAVING SUM(gt.belop) > 0
    ORDER BY p.region NULLS LAST, gt.srs_kategori, gt.konto
""")
rows_konto_detalj = cur.fetchall()
print(f"  Konto-detaljer: {len(rows_konto_detalj)} rader")

# ── 2. Fetch finance_dept_2026 budget per property ──
cur.execute("""
    SELECT
        p.property_id,
        p.name,
        p.region,
        SUM(fb.amount) AS bud26
    FROM finance_budget fb
    JOIN properties p ON p.property_id::text = fb.property_id::text
    WHERE fb.data_source = 'finance_dept_2026'
      AND fb.year = 2026
    GROUP BY p.property_id, p.name, p.region
    HAVING SUM(fb.amount) > 0
    ORDER BY p.region NULLS LAST, SUM(fb.amount) DESC
""")
rows_bud26 = {r["property_id"]: r for r in cur.fetchall()}
print(f"  finance_dept_2026: {len(rows_bud26)} properties")

# ── 3. Fetch BEFS prediction 2026 per property ──
cur.execute("""
    SELECT
        p.property_id,
        p.name,
        p.region,
        SUM(b.amount) AS pred26
    FROM budget b
    JOIN properties p ON p.property_id::text = b.property_id::text
    WHERE b.year = 2026
      AND b.data_source = 'okonomi_regional_2026'
    GROUP BY p.property_id, p.name, p.region
    HAVING SUM(b.amount) > 0
""")
rows_pred26 = {r["property_id"]: r for r in cur.fetchall()}
print(f"  BEFS pred 2026: {len(rows_pred26)} properties")

conn.close()

# ── 4. Build merged per-property rows ──
all_property_ids = set(rows_regn25) | set(rows_bud26) | set(rows_pred26)

def get_name_region(pid):
    for d in [rows_regn25, rows_bud26, rows_pred26]:
        if pid in d:
            return d[pid]["name"], d[pid]["region"]
    return pid, None

REGION_ORDER = ["Øst", "Sør", "Vest", "Midt-Norge", "Nord", "Bufdir", None]
REGION_DISPLAY = {
    "Øst": "Øst",
    "Sør": "Sør",
    "Vest": "Vest",
    "Midt-Norge": "Midt-Norge",
    "Nord": "Nord",
    "Bufdir": "Bufdir",
    None: "Nasjonal",
}

prop_rows = []
for pid in all_property_ids:
    name, region = get_name_region(pid)
    r25 = rows_regn25.get(pid, {})
    regn25       = float(r25.get("regn25")       or 0)
    lokaler_25   = float(r25.get("lokaler_25")   or 0)
    vedlikehold_25 = float(r25.get("vedlikehold_25") or 0)
    drift_25     = float(r25.get("drift_25")     or 0)
    pred26 = float(rows_pred26.get(pid, {}).get("pred26") or 0)
    bud26  = float(rows_bud26.get(pid, {}).get("bud26") or 0)

    # Flagg ufullstendig økonomibudsjett: bud26 < 20% av regn25 og bud26 < 500 000
    bud26_ufullstendig = (
        bud26 > 0 and regn25 > 0
        and bud26 < regn25 * 0.20
        and bud26 < 500_000
    )
    if bud26_ufullstendig:
        bud26_display = 0.0
    else:
        bud26_display = bud26

    prop_rows.append({
        "pid": pid,
        "name": name,
        "region": region,
        "region_display": REGION_DISPLAY.get(region, region or "Nasjonal"),
        "regn25": regn25,
        "lokaler_25": lokaler_25,
        "vedlikehold_25": vedlikehold_25,
        "drift_25": drift_25,
        "pred26": pred26,
        "bud26": bud26_display,
        "bud26_raw": bud26,
        "bud26_ufullstendig": bud26_ufullstendig,
    })

# Sort: region order, then by bud26 desc
def sort_key(r):
    try:
        ri = REGION_ORDER.index(r["region"])
    except ValueError:
        ri = 99
    return (ri, -r["bud26"])

prop_rows.sort(key=sort_key)

# ── 5. Build regional aggregates ──
reg_agg = defaultdict(lambda: {"regn25": 0.0, "pred26": 0.0, "bud26": 0.0, "n": 0})
for r in prop_rows:
    key = r["region_display"]
    reg_agg[key]["regn25"] += r["regn25"]
    reg_agg[key]["pred26"] += r["pred26"]
    reg_agg[key]["bud26"] += r["bud26"]
    reg_agg[key]["n"] += 1

# ── 6. Create workbook ──
wb = openpyxl.Workbook()

# ─── Helper styles ───
HEADER_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Arial", size=10)
SUBHDR_FILL = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
SUBHDR_FONT = Font(color="FFFFFF", bold=True, name="Arial", size=10)
NASJONAL_FILL = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
NASJONAL_FONT = Font(color="808080", italic=True, name="Arial", size=9)
NORMAL_FONT = Font(name="Arial", size=9)
BOLD_FONT = Font(name="Arial", size=9, bold=True)
NOK_FMT = '#,##0'
PCT_FMT = '0.0%'
thin = Side(style='thin', color="CCCCCC")
THIN_BORDER = Border(bottom=thin)

def hdr(ws, cell, val, fill=HEADER_FILL, font=HEADER_FONT, align="center"):
    c = ws[cell]
    c.value = val
    c.fill = fill
    c.font = font
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

def nok(ws, cell, val, bold=False, italic=False, color=None):
    c = ws[cell]
    c.value = val
    c.number_format = NOK_FMT
    f = Font(name="Arial", size=9, bold=bold, italic=italic,
             color=color if color else "000000")
    c.font = f
    c.alignment = Alignment(horizontal="right")

# ═══════════════════════════════════════════════════════════
# Sheet 1: Sammenligning 2026
# ═══════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Sammenligning 2026"

# Row 1 – note
ws1.merge_cells("A1:H1")
ws1["A1"].value = (
    "✅  Regnskap 2025 (Øk.) = Kontant 2025-data (faktiske regnskapsdata fra økonomi)"
)
ws1["A1"].font = Font(name="Arial", size=9, italic=True, color="1F5C2B")
ws1["A1"].fill = PatternFill("solid", start_color="E8F5E9", end_color="E8F5E9")
ws1.row_dimensions[1].height = 22

# Row 2 – headers
cols = ["Eiendom", "Region", "Regn. 2025 (Øk.)", "BEFS Prediksjon 2026",
        "Budsjett 2026 (Økonomi)", "Avvik (BEFS − Øk.)", "Avvik %", "Merknad"]
for ci, h in enumerate(cols, 1):
    cell = ws1.cell(row=2, column=ci)
    cell.value = h
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[2].height = 30

col_widths = [42, 12, 18, 22, 22, 20, 10, 72]
for ci, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(ci)].width = w

# Data rows
row = 3
for p in prop_rows:
    is_nasjonal = p["region"] is None
    is_sekkepost = p["name"] in ("Statlig", "Ideell")

    # Del 1 — BEFS-metode
    if is_sekkepost:
        befs_del = "Sekkepost — nasjonale konsolideringsobjekter"
    elif p["regn25"] > 0:
        faktor, pct = REGION_GROWTH.get(p["region"], (1.0, "?"))
        befs_del = f"BEFS: Kontant-2025 × {faktor:.4f} ({p['region'] or 'Nasjonal'} {pct})"
    else:
        befs_del = "BEFS: Hentet fra økonomibudsjett 2026 (ingen 2025-historikk)"

    # Del 2 — Økonomi-status
    if is_sekkepost:
        ok_del = ""
    elif p.get("bud26_ufullstendig"):
        raw_k = int((p.get("bud26_raw") or 0) / 1000)
        ok_del = f"Øk.: ufullstendig ({raw_k}k NOK, < 20% av regn25)"
    elif p["bud26"] == 0:
        ok_del = "Øk.: ikke budsjettert 2026"
    else:
        ok_del = "Øk.: fullstendig budsjett"

    merknad = befs_del if not ok_del else f"{befs_del}  |  {ok_del}"

    fill = NASJONAL_FILL if is_nasjonal else None
    font_kw = dict(italic=is_nasjonal, color="808080" if is_nasjonal else "000000")

    # Name
    c = ws1.cell(row=row, column=1, value=p["name"])
    c.font = Font(name="Arial", size=9, **font_kw)
    c.alignment = Alignment(horizontal="left")
    if fill:
        c.fill = fill

    # Region
    c = ws1.cell(row=row, column=2, value=p["region_display"])
    c.font = Font(name="Arial", size=9, **font_kw)
    c.alignment = Alignment(horizontal="center")
    if fill:
        c.fill = fill

    # Regn 2025
    c = ws1.cell(row=row, column=3, value=p["regn25"] if p["regn25"] else None)
    c.number_format = NOK_FMT
    c.font = Font(name="Arial", size=9, **font_kw)
    c.alignment = Alignment(horizontal="right")
    if fill:
        c.fill = fill

    # BEFS pred 2026
    c = ws1.cell(row=row, column=4, value=p["pred26"] if p["pred26"] else None)
    c.number_format = NOK_FMT
    c.font = Font(name="Arial", size=9, **font_kw)
    c.alignment = Alignment(horizontal="right")
    if fill:
        c.fill = fill

    # Budget 2026
    c = ws1.cell(row=row, column=5, value=p["bud26"] if p["bud26"] else None)
    c.number_format = NOK_FMT
    c.font = Font(name="Arial", size=9, **font_kw)
    c.alignment = Alignment(horizontal="right")
    if fill:
        c.fill = fill

    # Avvik NOK — blank hvis økonomi = 0 (ikke budsjettert) eller < 50 000 (ikke meningsfull)
    r_col = get_column_letter(4)
    ok_col = get_column_letter(5)
    avvik_formula = f"=IF(OR({ok_col}{row}=0,{ok_col}{row}<50000),\"\",{r_col}{row}-{ok_col}{row})"
    c = ws1.cell(row=row, column=6, value=avvik_formula)
    c.number_format = '#,##0;(#,##0);-'
    c.font = Font(name="Arial", size=9, **font_kw)
    c.alignment = Alignment(horizontal="right")
    if fill:
        c.fill = fill

    # Avvik % — blank hvis økonomi < 50 000 (ikke meningsfull sammenligning)
    avvik_pct_formula = f"=IF(AND({ok_col}{row}<>0,{ok_col}{row}>=50000),F{row}/{ok_col}{row},\"\")"
    c = ws1.cell(row=row, column=7, value=avvik_pct_formula)
    c.number_format = '0.0%;(0.0%);-'
    c.font = Font(name="Arial", size=9, **font_kw)
    c.alignment = Alignment(horizontal="right")
    if fill:
        c.fill = fill

    # Merknad
    c = ws1.cell(row=row, column=8, value=merknad)
    c.font = Font(name="Arial", size=9, italic=True, color="606060")
    c.alignment = Alignment(horizontal="left", wrap_text=True, vertical="top")
    if fill:
        c.fill = fill

    row += 1

# Total row
total_row = row
ws1.cell(row=total_row, column=1, value="TOTAL").font = BOLD_FONT
ws1.cell(row=total_row, column=1).fill = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
for ci in range(2, 9):
    ws1.cell(row=total_row, column=ci).fill = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")

for col_i, col_letter in [(3, "C"), (4, "D"), (5, "E")]:
    c = ws1.cell(row=total_row, column=col_i)
    c.value = f"=SUM({col_letter}3:{col_letter}{total_row-1})"
    c.number_format = NOK_FMT
    c.font = BOLD_FONT
    c.alignment = Alignment(horizontal="right")
    c.fill = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")

c = ws1.cell(row=total_row, column=6)
c.value = f"=D{total_row}-E{total_row}"
c.number_format = '#,##0;(#,##0);-'
c.font = BOLD_FONT
c.alignment = Alignment(horizontal="right")
c.fill = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")

c = ws1.cell(row=total_row, column=7)
c.value = f"=IF(E{total_row}<>0,F{total_row}/E{total_row},\"\")"
c.number_format = '0.0%;(0.0%);-'
c.font = BOLD_FONT
c.alignment = Alignment(horizontal="right")
c.fill = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")

ws1.freeze_panes = "A3"

# ═══════════════════════════════════════════════════════════
# Sheet 2: Regionsoversikt
# ═══════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Regionsoversikt")

ws2.merge_cells("A1:G1")
ws2["A1"].value = f"Regionsoversikt 2026 — generert {date.today().isoformat()}"
ws2["A1"].font = Font(name="Arial", size=12, bold=True, color="1F4E79")
ws2["A1"].alignment = Alignment(horizontal="left")
ws2.row_dimensions[1].height = 25

hdrs2 = ["Region", "Regn. 2025 (Øk.)", "BEFS Pred. 2026",
          "Budsj. 2026 (Øk.)", "Avvik (NOK)", "Avvik (%)", "Eiendommer"]
for ci, h in enumerate(hdrs2, 1):
    cell = ws2.cell(row=2, column=ci)
    cell.value = h
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws2.row_dimensions[2].height = 30

region_display_order = ["Øst", "Sør", "Vest", "Midt-Norge", "Nord", "Bufdir", "Nasjonal"]
col_widths2 = [16, 20, 20, 20, 18, 12, 12]
for ci, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w

row2 = 3
for rname in region_display_order:
    agg = reg_agg.get(rname)
    if not agg:
        continue
    is_nasjonal = rname == "Nasjonal"
    fill = NASJONAL_FILL if is_nasjonal else None
    fkw = dict(italic=is_nasjonal, color="808080" if is_nasjonal else "000000")

    c = ws2.cell(row=row2, column=1, value=rname + (" (sekkepost)" if is_nasjonal else ""))
    c.font = Font(name="Arial", size=10, bold=not is_nasjonal, **fkw)
    c.alignment = Alignment(horizontal="left")
    if fill:
        c.fill = fill

    for ci, key, fmt in [
        (2, "regn25", NOK_FMT),
        (3, "pred26", NOK_FMT),
        (4, "bud26", NOK_FMT),
    ]:
        c = ws2.cell(row=row2, column=ci, value=agg[key])
        c.number_format = fmt
        c.font = Font(name="Arial", size=10, **fkw)
        c.alignment = Alignment(horizontal="right")
        if fill:
            c.fill = fill

    # Avvik NOK (pred - bud)
    c = ws2.cell(row=row2, column=5)
    c.value = f"=C{row2}-D{row2}"
    c.number_format = '#,##0;(#,##0);-'
    c.font = Font(name="Arial", size=10, **fkw)
    c.alignment = Alignment(horizontal="right")
    if fill:
        c.fill = fill

    # Avvik %
    c = ws2.cell(row=row2, column=6)
    c.value = f"=IF(D{row2}<>0,E{row2}/D{row2},\"\")"
    c.number_format = '0.0%;(0.0%);-'
    c.font = Font(name="Arial", size=10, **fkw)
    c.alignment = Alignment(horizontal="right")
    if fill:
        c.fill = fill

    # Eiendommer count
    c = ws2.cell(row=row2, column=7, value=agg["n"])
    c.font = Font(name="Arial", size=10, **fkw)
    c.alignment = Alignment(horizontal="center")
    if fill:
        c.fill = fill

    row2 += 1

# Total row for regions
total_row2 = row2
ws2.cell(row=total_row2, column=1, value="TOTAL").font = BOLD_FONT
for ci in range(1, 8):
    ws2.cell(row=total_row2, column=ci).fill = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")

for ci, ltr in [(2, "B"), (3, "C"), (4, "D")]:
    c = ws2.cell(row=total_row2, column=ci)
    c.value = f"=SUM({ltr}3:{ltr}{total_row2-1})"
    c.number_format = NOK_FMT
    c.font = BOLD_FONT
    c.alignment = Alignment(horizontal="right")
    c.fill = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")

c = ws2.cell(row=total_row2, column=5)
c.value = f"=C{total_row2}-D{total_row2}"
c.number_format = '#,##0;(#,##0);-'
c.font = BOLD_FONT
c.alignment = Alignment(horizontal="right")
c.fill = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")

c = ws2.cell(row=total_row2, column=6)
c.value = f"=IF(D{total_row2}<>0,E{total_row2}/D{total_row2},\"\")"
c.number_format = '0.0%;(0.0%);-'
c.font = BOLD_FONT
c.alignment = Alignment(horizontal="right")
c.fill = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")

ws2.freeze_panes = "A3"

# ═══════════════════════════════════════════════════════════
# Sheet 3: Kontostruktur 2025
# ═══════════════════════════════════════════════════════════
ws3k = wb.create_sheet("Kontostruktur 2025")

# Konto-til-kategori og beskrivelse
KONTO_META = {
    "6300": ("Lokaler",     "Leie lokaler andre utleiere"),
    "6310": ("Lokaler",     "Leie lokaler fra Statsbygg"),
    "6391": ("Lokaler",     "Leie parkeringsplass"),
    "6395": ("Lokaler",     "Fellesutgifter andre utleiere"),
    "6396": ("Lokaler",     "Fellesutgifter (BAD) Statsbygg"),
    "6320": ("Drift",       "Renovasjon, vann, avløp o.l."),
    "6340": ("Drift",       "Strøm og oppvarming"),
    "6360": ("Drift",       "Renhold lokaler"),
    "6364": ("Drift",       "Vakthold lokaler"),
    "6365": ("Drift",       "Vaktmestertjenester"),
    "6390": ("Drift",       "Annen kostnad lokaler"),
    "1268": ("Vedlikehold", "Fast bygningsinventar og påkostning, leide bygg >50k"),
    "4960": ("Vedlikehold", "Fast bygningsinventar >50k"),
    "6398": ("Vedlikehold", "Fellesutgifter Statsbygg – indre vedlikehold"),
    "6630": ("Vedlikehold", "Reparasjon og vedlikehold leide lokaler"),
    "6632": ("Vedlikehold", "Oppgradering og påkostning leide lokaler <50k"),
    "6662": ("Vedlikehold", "Reparasjon og vedlikehold av anlegg, serviceavtaler"),
}

KAT_FILL = {
    "Lokaler":     PatternFill("solid", start_color="DDEEFF", end_color="DDEEFF"),
    "Drift":       PatternFill("solid", start_color="DDFFF0", end_color="DDFFF0"),
    "Vedlikehold": PatternFill("solid", start_color="FFF6DD", end_color="FFF6DD"),
}
KAT_HDR_FILL = {
    "Lokaler":     PatternFill("solid", start_color="2E75B6", end_color="2E75B6"),
    "Drift":       PatternFill("solid", start_color="375623", end_color="375623"),
    "Vedlikehold": PatternFill("solid", start_color="C55A11", end_color="C55A11"),
}

ws3k.column_dimensions["A"].width = 42
ws3k.column_dimensions["B"].width = 12
ws3k.column_dimensions["C"].width = 8
ws3k.column_dimensions["D"].width = 40
ws3k.column_dimensions["E"].width = 16
ws3k.column_dimensions["F"].width = 16
ws3k.column_dimensions["G"].width = 16
ws3k.column_dimensions["H"].width = 16

# Note row
ws3k.merge_cells("A1:H1")
ws3k["A1"].value = (
    "Kontostruktur 2025 — GL-regnskap netto per eiendom · "
    "Konto-kategorisering oppdatert 2026-05-09 (17 kontoer, 29 178 rader reklassifisert)"
)
ws3k["A1"].font = Font(name="Arial", size=9, italic=True, color="1F5C2B")
ws3k["A1"].fill = PatternFill("solid", start_color="E8F5E9", end_color="E8F5E9")
ws3k.row_dimensions[1].height = 22

# Header
k_headers = ["Eiendom", "Region", "Konto", "Beskrivelse",
             "Kategori", "Netto 2025 (kr)", "% av eiendom", ""]
for ci, h in enumerate(k_headers, 1):
    c = ws3k.cell(row=2, column=ci, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws3k.row_dimensions[2].height = 28

# Build per-property total for % calc
prop_totals_k: dict = {}
for rd in rows_konto_detalj:
    pid = rd["property_id"]
    prop_totals_k[pid] = prop_totals_k.get(pid, 0) + float(rd["netto"] or 0)

# Aggregate: Totalt per konto for alle eiendommer
konto_totals: dict = {}
for rd in rows_konto_detalj:
    k = rd["konto"]
    konto_totals[k] = konto_totals.get(k, 0) + float(rd["netto"] or 0)

# Write rows grouped by property
row3k = 3
prev_region = None
prev_pid = None

for rd in sorted(rows_konto_detalj, key=lambda x: (
    REGION_ORDER.index(x["region"]) if x["region"] in REGION_ORDER else 99,
    x["name"] or "",
    x["srs_kategori"] or "",
    x["konto"] or ""
)):
    pid = rd["property_id"]
    region = rd["region"]
    konto = rd["konto"] or ""
    kat = rd["srs_kategori"] or KONTO_META.get(konto, ("",))[0]
    netto = float(rd["netto"] or 0)
    prop_total = prop_totals_k.get(pid, 1)

    # Region separator
    if region != prev_region:
        ws3k.merge_cells(f"A{row3k}:H{row3k}")
        c = ws3k.cell(row=row3k, column=1, value=f"── {REGION_DISPLAY.get(region, region or 'Nasjonal')} ──")
        c.fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
        c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        ws3k.row_dimensions[row3k].height = 18
        row3k += 1
        prev_region = region
        prev_pid = None

    fill = KAT_FILL.get(kat, PatternFill("solid", start_color="F9F9F9", end_color="F9F9F9"))

    ws3k.cell(row=row3k, column=1, value=rd["name"] if pid != prev_pid else "").font = Font(name="Arial", size=9)
    ws3k.cell(row=row3k, column=2, value=REGION_DISPLAY.get(region, region or "Nasjonal")).font = Font(name="Arial", size=9, color="888888")
    ws3k.cell(row=row3k, column=3, value=konto).font = Font(name="Arial", size=9, bold=True)
    ws3k.cell(row=row3k, column=4, value=KONTO_META.get(konto, ("", rd.get("konto_navn","")))[1] or rd.get("konto_navn","")).font = Font(name="Arial", size=9)
    c_kat = ws3k.cell(row=row3k, column=5, value=kat)
    c_kat.fill = fill
    c_kat.font = Font(name="Arial", size=9, bold=True, color={
        "Lokaler": "1565C0", "Drift": "1B5E20", "Vedlikehold": "BF360C"
    }.get(kat, "000000"))
    c_nok = ws3k.cell(row=row3k, column=6, value=netto)
    c_nok.number_format = NOK_FMT
    c_nok.font = Font(name="Arial", size=9)
    c_nok.alignment = Alignment(horizontal="right")
    c_pct = ws3k.cell(row=row3k, column=7, value=netto / prop_total if prop_total else 0)
    c_pct.number_format = "0.0%"
    c_pct.font = Font(name="Arial", size=9, color="888888")
    c_pct.alignment = Alignment(horizontal="right")
    for ci in range(1, 8):
        ws3k.cell(row=row3k, column=ci).fill = fill

    prev_pid = pid
    row3k += 1

# Totals by konto
row3k += 1
ws3k.merge_cells(f"A{row3k}:D{row3k}")
c = ws3k.cell(row=row3k, column=1, value="TOTALT PER KONTO — alle eiendommer 2025")
c.fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
ws3k.row_dimensions[row3k].height = 22
row3k += 1

current_kat = None
kat_sum = 0.0
grand_total = 0.0
for konto in sorted(konto_totals.keys()):
    kat = KONTO_META.get(konto, ("",))[0]
    if kat != current_kat:
        if current_kat:
            # Print kat subtotal
            c = ws3k.cell(row=row3k, column=4, value=f"Delsum {current_kat}")
            c.font = Font(name="Arial", size=9, bold=True)
            c = ws3k.cell(row=row3k, column=6, value=kat_sum)
            c.number_format = NOK_FMT
            c.font = Font(name="Arial", size=9, bold=True)
            c.alignment = Alignment(horizontal="right")
            for ci in range(1, 8):
                ws3k.cell(row=row3k, column=ci).fill = KAT_HDR_FILL.get(current_kat, HEADER_FILL)
                ws3k.cell(row=row3k, column=ci).font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
            row3k += 1
        kat_sum = 0.0
        current_kat = kat
    v = konto_totals[konto]
    kat_sum += v
    grand_total += v
    fill = KAT_FILL.get(kat, PatternFill())
    desc = KONTO_META.get(konto, ("", konto))[1]
    ws3k.cell(row=row3k, column=3, value=konto).font = Font(name="Arial", size=9, bold=True)
    ws3k.cell(row=row3k, column=4, value=desc).font = Font(name="Arial", size=9)
    ws3k.cell(row=row3k, column=5, value=kat).font = Font(name="Arial", size=9, color={
        "Lokaler": "1565C0", "Drift": "1B5E20", "Vedlikehold": "BF360C"}.get(kat, "000000"))
    c = ws3k.cell(row=row3k, column=6, value=v)
    c.number_format = NOK_FMT
    c.alignment = Alignment(horizontal="right")
    for ci in range(1, 8):
        ws3k.cell(row=row3k, column=ci).fill = fill
    row3k += 1

# Last kat subtotal
if current_kat:
    c = ws3k.cell(row=row3k, column=4, value=f"Delsum {current_kat}")
    c = ws3k.cell(row=row3k, column=6, value=kat_sum)
    c.number_format = NOK_FMT
    c.font = Font(name="Arial", size=9, bold=True)
    c.alignment = Alignment(horizontal="right")
    for ci in range(1, 8):
        ws3k.cell(row=row3k, column=ci).fill = KAT_HDR_FILL.get(current_kat, HEADER_FILL)
        ws3k.cell(row=row3k, column=ci).font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    row3k += 1

# Grand total
ws3k.cell(row=row3k, column=1, value="TOTAL ALLE KONTOER").font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
c = ws3k.cell(row=row3k, column=6, value=grand_total)
c.number_format = NOK_FMT
c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
c.alignment = Alignment(horizontal="right")
for ci in range(1, 8):
    ws3k.cell(row=row3k, column=ci).fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    ws3k.cell(row=row3k, column=ci).font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

ws3k.freeze_panes = "A3"

# ═══════════════════════════════════════════════════════════
# Sheet 4: Metodikk og Kilder
# ═══════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Metodikk og Kilder")
ws3.column_dimensions["A"].width = 40
ws3.column_dimensions["B"].width = 70

metodikk_rows = [
    ("BEFS 2026 — Metodikk og Datakilder", None),
    (None, None),
    ("Formål", "Sammenligne BEFS sin prediksjon for 2026 mot økonomiavdelingens vedtatte budsjett"),
    (None, None),
    ("── KOLONNE C: Regnskap 2025 (Øk.) ──", None),
    ("Datakilde", "Kontant 2025 (faktiske regnskapsdata fra økonomiavdelingen)"),
    ("Tabell", "finance_budget WHERE data_source='kontant_2025' AND year=2025"),
    ("Notat", "Koststed 204416 redistribuert til 59 eiendom via adressematching mot Dim 2(T)"),
    ("Notat", "'Statlig' og 'Ideell' er nasjonale sekkepost-objekter (region=NULL i DB)"),
    (None, None),
    ("── KOLONNE D: BEFS Prediksjon 2026 ──", None),
    ("Datakilde", "BEFS-prediksjon basert på økonomiavdelingens regnskap 2025 med regionale vekstrater"),
    ("Tabell", "budget WHERE year=2026 AND data_source='okonomi_regional_2026'"),
    ("Metode", "Kontant 2025 per eiendom × regionens vekstrate (reverse-engineered fra økonomibudsjettet)"),
    ("Bakgrunn", (
        "Økonomi bruker institusjonsspesifikke vekstrater per region, IKKE flat prisjustering. "
        "Ratene er beregnet ved å sammenligne økonomiavd. sitt vedtatte budsjett 2026 (finance_dept_2026) "
        "mot faktisk regnskap 2025 (kontant_2025) per region via koststed_mapping."
    )),
    (None, None),
    ("── REGIONALE VEKSTRATER (beregnet 2026-05-07) ──", None),
    ("Øst",        "+17.01%  →  Kontant 2025 × 1.1701"),
    ("Sør",        "−10.78%  →  Kontant 2025 × 0.8922"),
    ("Nord",       "+26.62%  →  Kontant 2025 × 1.2662"),
    ("Vest",       "−14.92%  →  Kontant 2025 × 0.8508"),
    ("Midt-Norge", " +2.13%  →  Kontant 2025 × 1.0213"),
    ("Bufdir",     "+14.23%  →  Kontant 2025 × 1.1423"),
    ("Nasjonal",   " +3.50%  →  Kontant 2025 × 1.0350  (fallback for region=NULL)"),
    (None, None),
    ("── KOLONNE E: Budsjett 2026 (Økonomi) ──", None),
    ("Datakilde", "Økonomiavdelingens vedtatte budsjett for 2026"),
    ("Tabell", "finance_budget WHERE data_source='finance_dept_2026' AND year=2026"),
    (None, None),
    ("── ØKONOMIAVDELINGENS BUDSJETTMETODIKK ──", None),
    ("Metode", "Regional totalramme fastsettes sentralt av økonomiavdelingen, deretter fordelt per institusjon/koststed"),
    ("Resultat 2026", "189 eiendommer budsjettert. Total ~567 MNOK  (BEFS prediksjon: ~566 MNOK — avvik −0,1%)"),
    ("Ufullstendige poster", "11 eiendommer har økonomibudsjett < 20% av 2025-regnskap — mulig delvis registrering i kildesystemet:"),
    ("  Bufetathus Drammen",        "293k budsjettert vs. 5 318k regnskap 2025 (5,5 %)"),
    ("  Familievernkontoret Homansbyen", "1k budsjettert vs. 395k regnskap 2025 (0,3 %)"),
    ("  Familievernkontoret Asker og Bærum", "1k budsjettert vs. 258k regnskap 2025 (0,5 %)"),
    ("  + 8 andre", "Se merknad-kolonne (H) i Sammenligning 2026-fanen for fullstendig liste"),
    ("Ikke budsjettert", "22 eiendommer mangler økonomibudsjett helt (inkl. AVVIKLES-enheter og småobjekter)"),
    (None, None),
    ("── BEFS PREDIKSJON — UNNTAK FRA STANDARDMETODE ──", None),
    ("Standardmetode", "Kontant-2025 per eiendom × regionens vekstrate (se tabell over)"),
    ("Ingen 2025-historikk", (
        "5 eiendommer manglet Kontant-2025-data og fikk prediksjon hentet direkte fra "
        "økonomiavdelingens budsjett 2026: Fana og Ytrebygda avd. Eikelund, Fana og Ytrebygda avd. Skjold, "
        "Bodø behandlingssenter avd. 3, Nye Kvæfjord ungdomssenter avd. 3, Stjørdal Ungdomssenter — Boenhet"
    )),
    ("Individuell dokumentasjon", "Se Merknad-kolonne (H) i Sammenligning 2026-fanen — hver rad viser BEFS-metode og økonomi-status"),
    (None, None),
    ("── REGIONER ──", None),
    ("Nasjonal (sekkepost)", "Eiendommer med region=NULL i properties-tabellen"),
    ("Inkluderer", "'Statlig' (~147M i 2025) og 'Ideell' (~22M i 2025)"),
    (None, None),
    ("── HISTORIKK ──", None),
    ("v1 (forkastet)", "Holt-Winters ML-modell — forkastet pga. for lite historikk"),
    ("v2 (forkastet)", "Flat 3.5% prisjustering for alle regioner (data_source='kontant_2025_plus_3.5pct')"),
    ("v3 (gjeldende)", "Regionale vekstrater reverse-engineered fra økonomibudsjettet (data_source='okonomi_regional_2026')"),
    ("Kilde", f"Generert {date.today().isoformat()} av BEFS-systemet"),
]

for i, (a, b) in enumerate(metodikk_rows, 1):
    ca = ws3.cell(row=i, column=1, value=a)
    cb = ws3.cell(row=i, column=2, value=b)
    if i == 1:
        ca.font = Font(name="Arial", size=12, bold=True, color="1F4E79")
    elif a and a.startswith("──"):
        ca.font = Font(name="Arial", size=9, bold=True, color="2E75B6")
    else:
        ca.font = Font(name="Arial", size=9)
        cb.font = Font(name="Arial", size=9)

# ── 7. Save ──
out_path = os.environ.get("OUTPUT", "/tmp/2026bud.xlsx")
wb.save(out_path)
print(f"\nSaved: {out_path}")

# Print summary
print("\n=== REGIONAL SUMMARY ===")
for rname in region_display_order:
    agg = reg_agg.get(rname)
    if agg:
        print(f"  {rname:15s}: Regn25={agg['regn25']/1e6:7.1f}M  Pred26={agg['pred26']/1e6:7.1f}M  Bud26={agg['bud26']/1e6:7.1f}M  n={agg['n']}")
