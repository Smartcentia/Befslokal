"""
Generer fullstendig prediksjon-2027 Excel uten database (OFFLINE DEMO).

Struktur og ark matcher API: _build_prediksjon_2027_excel_workbook.
Tall er syntetiske — bytt til export_prediksjon_2027_xlsx.py nar DATABASE_URL virker.

Kjor:
  cd backend && python3 scripts/generate_prediksjon_2027_demo_xlsx.py
"""
from __future__ import annotations

import argparse
import os
import sys

if __name__ == "__main__":
    _backend = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if _backend not in sys.path:
        sys.path.insert(0, _backend)

# Samme forklaringstekst som i app/api/v1/financials.py (linje ~1295)
FORKLARING: list[tuple[str, str | None]] = [
    ("", ""),
    ("OM PREDIKSJONEN", None),
    (
        "Datagrunnlag",
        "Agresso GL-transaksjoner 2021–2025, gruppert per eiendom og SRS-kategori (Drift / Investering / Gjennomstrømning). Kun poster med beløp > 0 og kobling til BEFS-eiendom er inkludert.",
    ),
    ("Historikk brukt", "5 år (2021–2025). Eiendommer uten GL-data i denne perioden er ikke inkludert i prediksjonen."),
    ("Antall eiendommer", "190 eiendommer med prediksjon, fordelt på alle regioner."),
    ("", ""),
    ("ALGORITME — STEG 1: HOLT-WINTERS", None),
    ("Metode", "Holt-Winters dobbel eksponensiell utjevning med dempet trend. Kjøres per eiendom og per SRS-kategori separat."),
    ("α = 0,70", "Vekting av nivå — nyere år gir størst bidrag til estimatet."),
    ("β = 0,30", "Trendglattning — demper brå svingninger i trenden."),
    ("φ = 0,85", "Dempingsfaktor — hindrer ukontrollert eksponensiell trend-ekstrapolasjon."),
    ("Inflasjonsfallback", "3,5 % per år brukes som vekstrate for eiendommer med < 2 år historikk."),
    (
        "Minimumsgulv",
        "Hvis HW-prediksjon faller under 10 % av historisk gjennomsnitt, brukes gjennomsnitt × inflasjon i stedet. Forhindrer urealistisk nedgang (eks. Åsen, Tune).",
    ),
    ("", ""),
    ("ALGORITME — STEG 2: XGBOOST-GULV", None),
    ("Metode", "Maskinlæringsmodell (XGBoost Regressor) trent på tvers av alle eiendommer. Brukes som nedre grense — ikke som erstatning for HW."),
    (
        "Treningsdata",
        "232 (eiendom, kategori)-par med ≥ 4 år historikk. 127 unike eiendommer. Fordeling: Drift 146, Gjennomstrømning 80, Investering 6.",
    ),
    (
        "Features",
        "log(historisk gjennomsnitt), log(siste kjente år), antall år data, relativ trend (siste/tidlige år), variasjonskoeffisient, SRS-kategori (encoded).",
    ),
    ("Formel", "Endelig verdi = max(HW-prediksjon,  XGB-prediksjon × gulv-faktor)"),
    ("", ""),
    ("SCENARIER", None),
    (
        "XGB Gulv 70%",
        "Konservativt scenario — anbefalt for budsjettplanlegging og ressursallokering. XGBoost-gulvet settes til 70 % av tverreiendommsprediksjon. 179 av 190 eiendommer fikk gulvet aktivert.",
    ),
    (
        "XGB Gulv 50%",
        "Optimistisk scenario — viser nedre grense ved god effektivisering. XGBoost-gulvet settes til 50 % av tverreiendommsprediksjon. 130 av 190 eiendommer fikk gulvet aktivert.",
    ),
    ("", ""),
    ("TOLKNING AV RESULTATER", None),
    (
        "Diff 70-50 (NOK)",
        "Differansen mellom de to scenariene per eiendom. Stor differanse = HW-modellen gir lav prediksjon og gulvet slår inn.",
    ),
    (
        "Endring 70% vs 2025",
        "Prosentvis endring fra faktisk GL 2025 til XGB Gulv 70%-estimat. Verdier over 30 % kan indikere kraftig kostnadsvekst — vurder årsak.",
    ),
    (
        "is_synthetic = true",
        "Alle predikterte budsjettlinjer lagres med dette flagget i budget-tabellen, slik at de skilles fra vedtatte budsjett.",
    ),
    (
        "data_source",
        "holt_winters_2027_xgb70 (Ark 2 XGB 70%) / holt_winters_2027_xgb50 (Ark 2 XGB 50%). Brukes til å filtrere i avviksanalysen.",
    ),
    ("", ""),
    ("DRILL-DOWN ARK", None),
    (
        "Eiendom_kategori",
        "Faktisk GL 2025 og 2027-prediksjon (XGB 70/50) per eiendom og SRS-kategori. Bruk property_id som nøkkel mot andre ark.",
    ),
    ("GL_konto", "Aggregerte kontoposter for 2025 (kun beløp > 0 i grunnlagslinjer), alle eiendommer."),
    (
        "GL_bilag",
        "Enkeltbilag/linjer for 2025; maks ca. 80 000 rader, sortert etter eiendom og beløp. For full historikk: bruk BEFS eller database.",
    ),
    ("", ""),
    ("BEGRENSNINGER", None),
    ("Ikke inkludert", "Nybygg uten historikk. Eiendommer avviklet etter 2025. Strukturelle endringer etter 2025."),
    (
        "Investering-kategorien",
        "Kun 6 treningspar — XGB-estimater for Investering er mindre pålitelige enn Drift/Gjennomstrømning.",
    ),
    (
        "Ansvarsfraskrivelse",
        "Tallene er estimater for planleggingsformål — ikke godkjente budsjetttall. Endelig budsjett vedtas av ledelsen.",
    ),
    (
        "MERKNAD DEMO-FIL",
        "Denne arbeidsboken er generert offline med syntetiske eiendommer og beløp (generate_prediksjon_2027_demo_xlsx.py). Erstatt med eksport fra BEFS for produksjonsdata.",
    ),
]


def _demo_portfolio():
    """Returnerer prop_info, gl_2025, scenario_data xgb70/xgb50 per property_id."""
    regions_cycle = ["Nord", "Sør", "Vest", "Midt", "Øst", "Bufdir"]
    prop_info: dict[str, dict] = {}
    gl_2025: dict[str, float] = {}
    xgb70: dict[str, float] = {}
    xgb50: dict[str, float] = {}

    for i in range(18):
        pid = f"{i+1:08d}-0000-4000-8000-{i+1:012d}"
        reg = regions_cycle[i % len(regions_cycle)]
        name = f"Demo eiendom {reg} {i+1:02d}"
        prop_info[pid] = {"name": name, "region": reg}
        base = 3_500_000 + (i * 437_000) % 9_000_000
        g = float(base)
        gl_2025[pid] = g
        xgb70[pid] = round(g * 1.042 + 120_000, 0)
        xgb50[pid] = round(g * 1.028 + 80_000, 0)
        if xgb50[pid] > xgb70[pid]:
            xgb50[pid] = round(xgb70[pid] * 0.98, 0)

    return prop_info, gl_2025, {"xgb70": xgb70, "xgb50": xgb50}


def _split_category(gl: float, p70: float, p50: float):
    """Tre kategorier med andeler som summerer til 1."""
    parts = [("Drift", 0.62), ("Gjennomstrømning", 0.28), ("Investering", 0.10)]
    out = []
    for kat, share in parts:
        out.append(
            (
                kat,
                round(gl * share, 0),
                round(p70 * share, 0),
                round(p50 * share, 0),
            )
        )
    return out


def main() -> int:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    p = argparse.ArgumentParser()
    p.add_argument(
        "-o",
        "--output",
        default=os.path.join(
            os.path.dirname(__file__),
            "..",
            "..",
            "tools",
            "prediksjon-excel-viewer-pakke",
            "prediksjon_2027_export.xlsx",
        ),
    )
    args = p.parse_args()

    prop_info, gl_2025, sc = _demo_portfolio()
    scenario_data = sc
    all_pids = sorted(
        set(gl_2025.keys()) | set(scenario_data["xgb70"].keys()) | set(scenario_data["xgb50"].keys())
    )

    wb = openpyxl.Workbook()
    fill_blue = PatternFill("solid", fgColor="1D4ED8")
    fill_gray = PatternFill("solid", fgColor="374151")
    fill_green = PatternFill("solid", fgColor="065F46")
    fill_dark = PatternFill("solid", fgColor="1E3A5F")
    fill_teal = PatternFill("solid", fgColor="0F766E")

    def cell(ws, row, col, value, bold=False, fill=None, num_format=None):
        c = ws.cell(row=row, column=col, value=value)
        if bold:
            c.font = Font(bold=True)
        if fill:
            c.fill = fill
            c.font = Font(bold=True, color="FFFFFF")
        if num_format:
            c.number_format = num_format
        c.alignment = Alignment(horizontal="right" if isinstance(value, (int, float)) else "left")
        return c

    # --- Sammendrag ---
    ws1 = wb.active
    ws1.title = "Sammendrag"
    ws1.merge_cells("A1:D1")
    ws1["A1"] = "Budsjettprediksjon 2027 — Sammenligning av scenarier (OFFLINE DEMO)"
    ws1["A1"].font = Font(bold=True, size=14)
    headers = ["Metrikk", "2025 Faktisk (GL)", "XGB Gulv 70%", "XGB Gulv 50%"]
    for col, h in enumerate(headers, 1):
        cell(ws1, 2, col, h, fill=fill_blue)

    total_70 = sum(scenario_data["xgb70"].values())
    total_50 = sum(scenario_data["xgb50"].values())
    total_25 = sum(gl_2025.values())

    def pst(a, b):
        return round((a - b) / b * 100, 1) if b else None

    rows_kpi = [
        ("Total kostnadsbudsjett (NOK)", total_25, total_70, total_50),
        ("Endring vs 2025 (%)", None, pst(total_70, total_25), pst(total_50, total_25)),
        ("Antall eiendommer", len(gl_2025), len(scenario_data["xgb70"]), len(scenario_data["xgb50"])),
    ]
    for i, (label, v25, v70, v50) in enumerate(rows_kpi, 3):
        ws1.cell(row=i, column=1, value=label).font = Font(bold=True)
        for col, val in [(2, v25), (3, v70), (4, v50)]:
            c = ws1.cell(row=i, column=col, value=val)
            if isinstance(val, float) and abs(val) > 1000:
                c.number_format = "#,##0"

    ws1.cell(row=7, column=1, value="Region-fordeling").font = Font(bold=True, size=12)
    for col, h in enumerate(["Region", "2025 Faktisk", "XGB 70%", "XGB 50%", "Diff 70 vs 50"], 1):
        cell(ws1, 8, col, h, fill=fill_gray)

    regions: dict[str, dict] = {}
    for pid in all_pids:
        reg = prop_info.get(pid, {}).get("region", "Ukjent")
        if reg not in regions:
            regions[reg] = {"gl": 0.0, "xgb70": 0.0, "xgb50": 0.0}
        regions[reg]["gl"] += gl_2025.get(pid, 0)
        regions[reg]["xgb70"] += scenario_data["xgb70"].get(pid, 0)
        regions[reg]["xgb50"] += scenario_data["xgb50"].get(pid, 0)

    for i, (reg, vals) in enumerate(sorted(regions.items()), 9):
        ws1.cell(row=i, column=1, value=reg)
        for col, val in [
            (2, vals["gl"]),
            (3, vals["xgb70"]),
            (4, vals["xgb50"]),
            (5, vals["xgb70"] - vals["xgb50"]),
        ]:
            c = ws1.cell(row=i, column=col, value=round(val, 0))
            c.number_format = "#,##0"

    for col in range(1, 6):
        ws1.column_dimensions[get_column_letter(col)].width = 22

    # --- Alle eiendommer ---
    ws2 = wb.create_sheet("Alle eiendommer")
    headers2 = [
        "Eiendom",
        "Region",
        "2025 Faktisk (NOK)",
        "XGB 70% (NOK)",
        "XGB 50% (NOK)",
        "Diff 70-50 (NOK)",
        "Endring 70% vs 2025 (%)",
    ]
    for col, h in enumerate(headers2, 1):
        cell(ws2, 1, col, h, fill=fill_green)

    for i, pid in enumerate(sorted(all_pids, key=lambda p: -scenario_data["xgb70"].get(p, 0)), 2):
        name = prop_info.get(pid, {}).get("name", pid)
        reg = prop_info.get(pid, {}).get("region", "Ukjent")
        v25 = gl_2025.get(pid, 0)
        v70 = scenario_data["xgb70"].get(pid, 0)
        v50 = scenario_data["xgb50"].get(pid, 0)
        diff = v70 - v50
        endring = round((v70 - v25) / v25 * 100, 1) if v25 > 0 else None
        ws2.cell(row=i, column=1, value=name)
        ws2.cell(row=i, column=2, value=reg)
        for col, val in [(3, v25), (4, v70), (5, v50), (6, diff)]:
            c = ws2.cell(row=i, column=col, value=round(val, 0))
            c.number_format = "#,##0"
        if endring is not None:
            c = ws2.cell(row=i, column=7, value=endring)
            c.number_format = '0.0"%"'

    for col, w in [(1, 40), (2, 16), (3, 20), (4, 20), (5, 20), (6, 20), (7, 22)]:
        ws2.column_dimensions[get_column_letter(col)].width = w

    # --- Forklaring ---
    ws3 = wb.create_sheet("Forklaring")

    def txt(ws, row, col, value, bold=False, italic=False, size=11, fill=None, color="000000"):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, italic=italic, size=size, color=color if not fill else color)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if fill:
            c.fill = fill
            c.font = Font(bold=bold, size=size, color="FFFFFF")
        return c

    ws3.merge_cells("A1:C1")
    txt(ws3, 1, 1, "Budsjettprediksjon 2027 — Metodebeskrivelse", bold=True, size=14, fill=fill_dark)
    txt(ws3, 2, 1, "OFFLINE DEMO — Generert uten database (BEFS-mal)", italic=True, size=10, color="555555")

    row = 4
    for label, tekst in FORKLARING:
        if tekst is None:
            ws3.merge_cells(f"A{row}:C{row}")
            txt(ws3, row, 1, label, bold=True, size=11, fill=fill_dark)
        elif label == "":
            row += 1
            continue
        else:
            txt(ws3, row, 1, label, bold=True, size=10)
            ws3.merge_cells(f"B{row}:C{row}")
            txt(ws3, row, 2, tekst, size=10)
            ws3.row_dimensions[row].height = max(15, len(tekst) // 6)
        row += 1

    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 55
    ws3.column_dimensions["C"].width = 20

    # --- Eiendom_kategori ---
    ws4 = wb.create_sheet("Eiendom_kategori")
    hk = [
        "property_id",
        "eiendom",
        "region",
        "srs_kategori",
        "gl_2025",
        "pred_2027_xgb70",
        "pred_2027_xgb50",
        "endring_pct_70_vs_gl",
    ]
    for col, title in enumerate(hk, 1):
        c = ws4.cell(1, col, title)
        c.fill = fill_teal
        c.font = Font(bold=True, color="FFFFFF")
    r4 = 2
    for pid in sorted(all_pids):
        info = prop_info[pid]
        gtot = gl_2025[pid]
        p70t = scenario_data["xgb70"][pid]
        p50t = scenario_data["xgb50"][pid]
        for kat, g, p70, p50 in _split_category(gtot, p70t, p50t):
            endr = round((p70 - g) / g * 100, 1) if g > 0 else None
            rowv = [pid, info["name"], info["region"], kat, g, p70, p50, endr]
            for c, val in enumerate(rowv, 1):
                cell = ws4.cell(r4, c, val)
                if c >= 5 and c <= 7 and isinstance(val, (int, float)):
                    cell.number_format = "#,##0"
                if c == 8 and val is not None:
                    cell.number_format = '0.0"%"'
            r4 += 1

    for col, w in enumerate([38, 36, 14, 18, 16, 16, 16, 18], 1):
        ws4.column_dimensions[get_column_letter(col)].width = w

    # --- GL_konto ---
    ws5 = wb.create_sheet("GL_konto")
    kk = [
        "property_id",
        "eiendom",
        "region",
        "ar",
        "srs_kategori",
        "konto",
        "konto_navn",
        "belop",
        "antall_transaksjoner",
    ]
    for col, title in enumerate(kk, 1):
        c = ws5.cell(1, col, title)
        c.fill = fill_teal
        c.font = Font(bold=True, color="FFFFFF")

    kontoplan = [
        ("6300", "Leie lokaler", "Drift", 0.45),
        ("7100", "Annen drift", "Drift", 0.35),
        ("6900", "Kontor og forbruk", "Drift", 0.20),
        ("8500", "Gjennomføring", "Gjennomstrømning", 0.85),
        ("8510", "Annen gjennomstrømning", "Gjennomstrømning", 0.15),
        ("1280", "Investering bygg", "Investering", 0.70),
        ("4960", "Investering utstyr", "Investering", 0.30),
    ]

    r5 = 2
    for pid in sorted(all_pids):
        info = prop_info[pid]
        gtot = gl_2025[pid]
        for kat, share_gl, share_p70, share_p50 in [
            ("Drift", 0.62, 0.62, 0.62),
            ("Gjennomstrømning", 0.28, 0.28, 0.28),
            ("Investering", 0.10, 0.10, 0.10),
        ]:
            kat_gl = gtot * share_gl
            for konto, knavn, kkat, weight in kontoplan:
                if kkat != kat:
                    continue
                bel = round(kat_gl * weight, 0)
                if bel < 1:
                    continue
                ws5.cell(r5, 1, pid)
                ws5.cell(r5, 2, info["name"])
                ws5.cell(r5, 3, info["region"])
                ws5.cell(r5, 4, 2025)
                ws5.cell(r5, 5, kat)
                ws5.cell(r5, 6, konto)
                ws5.cell(r5, 7, knavn)
                c = ws5.cell(r5, 8, bel)
                c.number_format = "#,##0"
                ws5.cell(r5, 9, max(1, int(bel / 250_000)))
                r5 += 1

    for col, w in enumerate([38, 34, 12, 6, 16, 10, 28, 14, 8], 1):
        ws5.column_dimensions[get_column_letter(col)].width = min(w, 50)

    # --- GL_bilag ---
    ws6 = wb.create_sheet("GL_bilag")
    ws6.merge_cells("A1:J1")
    ws6["A1"] = (
        "Merk: OFFLINE DEMO — et utvalg bilagslinjer. Ekte eksport: inntil 80 000 linjer, sortert etter eiendom og beløp."
    )
    ws6["A1"].font = Font(italic=True, size=10)
    bh = [
        "property_id",
        "eiendom",
        "region",
        "bilagsnr",
        "bilagsdato",
        "konto",
        "konto_navn",
        "srs_kategori",
        "tekst",
        "belop",
    ]
    for col, title in enumerate(bh, 1):
        c = ws6.cell(2, col, title)
        c.fill = fill_teal
        c.font = Font(bold=True, color="FFFFFF")

    rb = 3
    bilnr = 9000000
    for pid in sorted(all_pids)[:12]:
        info = prop_info[pid]
        for konto, knavn, kat, _w in kontoplan[:4]:
            for q, (dato, tekst) in enumerate(
                [
                    ("2025-03-15", f"Faktura {konto} Q1"),
                    ("2025-09-10", f"Oppfølging {konto} H2"),
                ]
            ):
                bel = round(gl_2025[pid] * 0.04 / (q + 1), 0) + 10_000
                ws6.cell(rb, 1, pid)
                ws6.cell(rb, 2, info["name"])
                ws6.cell(rb, 3, info["region"])
                ws6.cell(rb, 4, f"DEMO-{bilnr}")
                bilnr += 1
                ws6.cell(rb, 5, dato)
                ws6.cell(rb, 6, konto)
                ws6.cell(rb, 7, knavn)
                ws6.cell(rb, 8, kat)
                ws6.cell(rb, 9, tekst)
                c = ws6.cell(rb, 10, bel)
                c.number_format = "#,##0.00"
                rb += 1

    for col, w in enumerate([38, 32, 12, 14, 12, 10, 24, 14, 36, 14], 1):
        ws6.column_dimensions[get_column_letter(col)].width = min(w, 50)

    out = os.path.abspath(args.output)
    d = os.path.dirname(out)
    if d:
        os.makedirs(d, exist_ok=True)
    wb.save(out)
    print(f"Skrev full DEMO-workbook ({len(all_pids)} eiendommer): {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
