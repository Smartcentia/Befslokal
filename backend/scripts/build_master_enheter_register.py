#!/usr/bin/env python3
"""
Bygger et felles masterregister for enheter ved å kryssjekke:
- Eiendom/AGRESSO-eksport (Dim1 + Dim1(T) + region + økonomidata)
- Innkjøpsanalyse «Lønnsutgifter»-ark (enhetsnavn + årlige lønnstall)

Kjøring:
  python scripts/build_master_enheter_register.py \\
    --eiendom /path/Eiendom....xlsx \\
    --innkjop /path/Innkjøpsanalyse....xlsx \\
    --out /path/Master_enheter_register.xlsx
"""

from __future__ import annotations

import argparse
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz, process

SKIP_INNJKOP_LABELS = frozenset({"Faste stillinger", "Bufetat"})
REGION_PREFIX = "Region "


def is_number(x: Any) -> bool:
    return isinstance(x, (int, float)) and not isinstance(x, bool)


def normalize_name(s: str) -> str:
    if not s:
        return ""
    t = unicodedata.normalize("NFKC", str(s)).strip().lower()
    t = re.sub(r"\s+", " ", t)
    t = t.replace(".", "").replace(",", "")
    return t


def parse_innkjop_lonn(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Lønnsutgifter"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_idx = next(i for i, r in enumerate(rows) if r and r[0] == "Radetiketter")
    years: list[str] = [str(c) for c in rows[header_idx][1:] if c is not None]

    current_region: str | None = None
    agg: dict[str, dict[str, Any]] = {}

    for i in range(header_idx + 1, len(rows)):
        r = rows[i]
        if not r or r[0] is None:
            continue
        label = str(r[0]).strip()
        if label in SKIP_INNJKOP_LABELS:
            continue
        if label.startswith(REGION_PREFIX):
            current_region = label
            continue
        vals = list(r[1 : 1 + len(years)])
        if not any(is_number(v) for v in vals):
            continue

        row_vals: dict[str, Any] = {}
        for j, y in enumerate(years):
            v = vals[j] if j < len(vals) else None
            row_vals[y] = float(v) if is_number(v) else None

        if label not in agg:
            agg[label] = {
                "enhet_navn_innkjop": label,
                "region_innkjop_sist_sett": current_region,
                **{y: 0.0 for y in years},
            }
        for y in years:
            v = row_vals.get(y)
            if v is not None:
                agg[label][y] = agg[label][y] + v
        if current_region:
            agg[label]["region_innkjop_sist_sett"] = current_region

    # Slå sammen rader som bare skiller seg på tegnsetting/mellomrom (samme normaliserte nøkkel)
    by_norm: dict[str, dict[str, Any]] = {}
    for label, d in agg.items():
        nk = normalize_name(label)
        if nk not in by_norm:
            by_norm[nk] = {**d, "enhet_navn_innkjop": label}
        else:
            t = by_norm[nk]
            t["enhet_navn_innkjop"] = min(t["enhet_navn_innkjop"], label, key=len)
            for y in years:
                if y == "Totalsum":
                    continue
                a, b = t.get(y) or 0.0, d.get(y) or 0.0
                if a or b:
                    t[y] = (a or 0.0) + (b or 0.0)
            if d.get("region_innkjop_sist_sett"):
                t["region_innkjop_sist_sett"] = d["region_innkjop_sist_sett"]

    merged = list(by_norm.values())
    for d in merged:
        tot = 0.0
        for y in years:
            if y == "Totalsum":
                continue
            v = d.get(y)
            if v is not None:
                tot += float(v)
        d["Totalsum_beregnet"] = round(tot, 2)

    return years, merged


@dataclass
class AgressoUnit:
    dim1: Any
    dim1_t: str
    regions: set[str] = field(default_factory=set)
    row_count: int = 0
    sum_belop: float = 0.0
    min_periode: int | None = None
    max_periode: int | None = None
    min_dato: datetime | None = None
    max_dato: datetime | None = None


def parse_agresso(path: Path) -> dict[Any, AgressoUnit]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["AGRESSO"]
    units: dict[Any, AgressoUnit] = {}
    header = None
    idx: dict[str, int] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = row
            idx = {str(h): j for j, h in enumerate(header) if h is not None}
            continue
        if not row:
            continue
        d1 = row[idx["Dim1"]]
        if d1 is None or d1 == "":
            continue
        d1t = row[idx["Dim1(T)"]] or ""
        reg = row[idx["Region"]]
        belop = row[idx["Beløp"]]
        periode = row[idx["Periode"]]
        bilagsdato = row[idx["Bilagsdato"]]

        u = units.setdefault(d1, AgressoUnit(dim1=d1, dim1_t=str(d1t).strip()))
        u.regions.add(str(reg) if reg is not None else "")
        u.row_count += 1
        if is_number(belop):
            u.sum_belop += float(belop)
        if periode is not None:
            try:
                p = int(periode)
            except (TypeError, ValueError):
                p = None
            if p is not None:
                u.min_periode = p if u.min_periode is None else min(u.min_periode, p)
                u.max_periode = p if u.max_periode is None else max(u.max_periode, p)
        if isinstance(bilagsdato, datetime):
            u.min_dato = bilagsdato if u.min_dato is None else min(u.min_dato, bilagsdato)
            u.max_dato = bilagsdato if u.max_dato is None else max(u.max_dato, bilagsdato)

    wb.close()
    return units


def autosize_columns(ws: Any) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 60)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--eiendom", type=Path, required=True)
    ap.add_argument("--innkjop", type=Path, required=True)
    ap.add_argument("--out", type=Path, required=True)
    ap.add_argument("--fuzzy-min", type=int, default=88, help="Min score for fuzzy-match (0-100)")
    args = ap.parse_args()

    years, innkjop_rows = parse_innkjop_lonn(args.innkjop)
    agresso = parse_agresso(args.eiendom)

    innkjop_by_norm = {normalize_name(r["enhet_navn_innkjop"]): r for r in innkjop_rows}
    innkjop_names = [r["enhet_navn_innkjop"] for r in innkjop_rows]

    master_rows: list[dict[str, Any]] = []
    matched_innkjop_norm: set[str] = set()

    for dim1 in sorted(agresso.keys(), key=lambda x: (str(x))):
        u = agresso[dim1]
        norm_ag = normalize_name(u.dim1_t)
        inn_row = None
        match_type = ""
        score = None

        if norm_ag and norm_ag in innkjop_by_norm:
            inn_row = innkjop_by_norm[norm_ag]
            match_type = "eksakt_navn"
            score = 100
        elif norm_ag:
            best = process.extractOne(
                u.dim1_t,
                innkjop_names,
                scorer=fuzz.token_sort_ratio,
            )
            if best and best[1] >= args.fuzzy_min:
                inn_row = innkjop_by_norm[normalize_name(best[0])]
                match_type = "fuzzy"
                score = int(best[1])

        if inn_row:
            matched_innkjop_norm.add(normalize_name(inn_row["enhet_navn_innkjop"]))

        reg_str = ", ".join(sorted(x for x in u.regions if x))

        kryss = "kun_AGRESSO"
        if inn_row:
            kryss = "begge_kilder"

        row_out: dict[str, Any] = {
            "Dim1": dim1,
            "Enhet_navn_AGRESSO": u.dim1_t,
            "Region(er)_AGRESSO": reg_str,
            "Poster_AGRESSO": u.row_count,
            "Sum_beløp_AGRESSO": round(u.sum_belop, 2),
            "Periode_fra": u.min_periode,
            "Periode_til": u.max_periode,
            "Bilagsdato_fra": u.min_dato.date().isoformat() if u.min_dato else "",
            "Bilagsdato_til": u.max_dato.date().isoformat() if u.max_dato else "",
            "Match_type": match_type,
            "Match_score": score if score is not None else "",
            "Enhet_navn_INNJKOP": inn_row["enhet_navn_innkjop"] if inn_row else "",
            "Region_sist_INNJKOP": inn_row.get("region_innkjop_sist_sett") if inn_row else "",
            "Kryssjekk": kryss,
        }
        for y in years:
            if inn_row and y in inn_row:
                v = inn_row.get(y)
                row_out[f"Lønn_{y}_INNJKOP"] = round(v, 2) if is_number(v) else ""
            else:
                row_out[f"Lønn_{y}_INNJKOP"] = ""

        master_rows.append(row_out)

    only_innkjop = []
    for r in innkjop_rows:
        n = normalize_name(r["enhet_navn_innkjop"])
        if n not in matched_innkjop_norm:
            only_innkjop.append(r)

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb_out = openpyxl.Workbook()
    ws_m = wb_out.active
    ws_m.title = "Master_enheter"

    headers = list(master_rows[0].keys()) if master_rows else []
    ws_m.append(headers)
    for row in master_rows:
        ws_m.append([row.get(h) for h in headers])
    autosize_columns(ws_m)

    ws_o = wb_out.create_sheet("Kun_Innkjøp_uten_AGRESSO")
    oy = ["enhet_navn_innkjop", "region_innkjop_sist_sett"] + [y for y in years if y != "Totalsum"] + [
        "Totalsum_beregnet"
    ]
    ws_o.append(oy)
    for r in sorted(only_innkjop, key=lambda x: normalize_name(x["enhet_navn_innkjop"])):
        ws_o.append([r.get(c, "") for c in oy])
    autosize_columns(ws_o)

    by_innkjop_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in master_rows:
        inn = (r.get("Enhet_navn_INNJKOP") or "").strip()
        if inn:
            by_innkjop_name[inn].append(r)
    dim1_kollisjoner = {k: v for k, v in by_innkjop_name.items() if len(v) > 1}

    ws_c = wb_out.create_sheet("Flere_Dim1_samme_lønnsnavn")
    coll_headers = [
        "Enhet_navn_INNJKOP",
        "Antall_Dim1",
        "Dim1",
        "Enhet_navn_AGRESSO",
        "Region(er)_AGRESSO",
        "Match_type",
        "Match_score",
        "Poster_AGRESSO",
        "Sum_beløp_AGRESSO",
    ]
    ws_c.append(coll_headers)
    for inn in sorted(dim1_kollisjoner.keys(), key=lambda x: normalize_name(x)):
        gruppe = sorted(dim1_kollisjoner[inn], key=lambda x: (str(x.get("Dim1"))))
        n = len(gruppe)
        for r in gruppe:
            ws_c.append(
                [
                    inn,
                    n,
                    r.get("Dim1"),
                    r.get("Enhet_navn_AGRESSO"),
                    r.get("Region(er)_AGRESSO"),
                    r.get("Match_type"),
                    r.get("Match_score"),
                    r.get("Poster_AGRESSO"),
                    r.get("Sum_beløp_AGRESSO"),
                ]
            )
    autosize_columns(ws_c)

    ws_s = wb_out.create_sheet("Sammendrag")
    n_ag = len(agresso)
    n_inn = len(innkjop_rows)
    n_begge = sum(1 for r in master_rows if r["Kryssjekk"] == "begge_kilder")
    n_kun_ag = sum(1 for r in master_rows if r["Kryssjekk"] == "kun_AGRESSO")
    n_fuzzy = sum(1 for r in master_rows if r["Match_type"] == "fuzzy")
    ws_s.append(["Kilde", "Antall unike enheter/rader"])
    ws_s.append(["AGRESSO (unike Dim1)", n_ag])
    ws_s.append(["Innkjøpsanalyse Lønnsutgifter (aggregert navn)", n_inn])
    ws_s.append([])
    ws_s.append(["Kryssjekk", "Antall"])
    ws_s.append(["Treff i begge kilder", n_begge])
    ws_s.append(["Kun i AGRESSO (Dim1 uten lønn-treff)", n_kun_ag])
    ws_s.append(["Kun i Innkjøp (ikke matchet til Dim1)", len(only_innkjop)])
    ws_s.append([])
    ws_s.append(
        [
            "Lønnsnavn med mer enn én Dim1 (manuell avklaring)",
            len(dim1_kollisjoner),
        ]
    )
    ws_s.append(
        [
            "Dim1-rader i kollisjonsarket",
            sum(len(v) for v in dim1_kollisjoner.values()),
        ]
    )
    ws_s.append([])
    ws_s.append(["Fuzzy-match (score >= " + str(args.fuzzy_min) + ")", n_fuzzy])
    ws_s.append(["Generert", datetime.now().isoformat(timespec="seconds")])
    autosize_columns(ws_s)

    wb_out.save(args.out)
    print(f"Skrev {args.out} ({n_ag} Dim1, {n_inn} innkjøp-navn, {n_begge} dobbelttreff)")


if __name__ == "__main__":
    main()
