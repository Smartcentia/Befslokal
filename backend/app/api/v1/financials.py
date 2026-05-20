import logging
from io import BytesIO
from fastapi import APIRouter, BackgroundTasks, Depends, Query, HTTPException, status
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, desc, text
from typing import List, Dict, Any, Optional

from app.api.deps import get_db, get_current_user
from app.domains.core.models.user import User, UserRole
from app.services.analytics.financial_analysis_service import FinancialAnalysisService

logger = logging.getLogger(__name__)
router = APIRouter()

@router.get("/suppliers", response_model=Dict[str, Any])
async def get_supplier_overview(
    year: Optional[int] = Query(None, description="Filter by year (e.g. 2024, 2025)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get global supplier statistics across all properties.
    Returns aggregated costs, property counts, and details per supplier.
    Requires REGIONAL_MANAGER or ADMIN role.
    """
    # REGIONAL_MANAGER kan se sin region, ADMIN kan se alt
    if current_user.role not in [UserRole.ADMIN, UserRole.REGIONAL_MANAGER]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only regional managers and administrators can access supplier overview"
        )

    return await FinancialAnalysisService.get_global_supplier_stats(db, year=year)


@router.get("/supplier-catalog", response_model=List[Dict[str, Any]])
async def get_supplier_catalog(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Get the supplier catalog aggregated from GL transactions.
    Returns each unique supplier with their service categories.
    """
    try:
        from app.models.financial_models import GLTransaction
        from sqlalchemy import func, select

        stmt = (
            select(
                GLTransaction.leverandor_navn,
                func.string_agg(
                    GLTransaction.konto_navn.distinct(),
                    ", "
                ).label("tjenester"),
            )
            .where(
                GLTransaction.leverandor_navn.isnot(None),
                GLTransaction.leverandor_navn != "",
                GLTransaction.belop > 0,
            )
            .group_by(GLTransaction.leverandor_navn)
            .order_by(GLTransaction.leverandor_navn)
        )
        rows = (await db.execute(stmt)).fetchall()
        return [{"Leverandør": r.supplier_name, "Tjenester": r.tjenester or ""} for r in rows]
    except Exception as e:
        return []


@router.get("/patterns", response_model=Dict[str, Any])
async def get_common_patterns(
    year: Optional[int] = Query(None, description="Filter by year (e.g. 2024, 2025)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get common cost patterns across all properties.

    Returns:
    - Most common cost categories
    - Most common providers
    - Statistics

    Requires REGIONAL_MANAGER or ADMIN role.
    """
    # REGIONAL_MANAGER kan se sin region, ADMIN kan se alt
    if current_user.role not in [UserRole.ADMIN, UserRole.REGIONAL_MANAGER]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only regional managers and administrators can access cost patterns"
        )

    return await FinancialAnalysisService.get_common_patterns(db, year=year)


@router.get("/rent-reconciliation", response_model=Dict[str, Any])
async def get_rent_reconciliation(
    year: int = Query(2024, description="År for avstemming"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Grundig avstemming: kontraktsfestet vs bokført husleie.
    Viser alle GL-kontonavn med beløp, hvilke som regnes som husleie, og gap-analyse.
    """
    if current_user.role not in [UserRole.ADMIN, UserRole.REGIONAL_MANAGER]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Krever ADMIN eller REGIONAL_MANAGER")
    try:
        from app.models.financial_models import GLTransaction
        from app.domains.core.models.property import Property
        from app.domains.core.models.contract import Contract
        from app.domains.core.models.unit import Unit
        from app.models.gl_constants import is_lease_account

        # 1. Alle GL-kontonavn for året med sum (inkl. orphan)
        acct_stmt = (
            select(
                GLTransaction.konto_navn,
                GLTransaction.property_id,
                func.sum(GLTransaction.belop).label("total"),
            )
            .where(GLTransaction.ar == year, GLTransaction.belop > 0)
            .group_by(GLTransaction.konto_navn, GLTransaction.property_id)
        )
        acct_rows = (await db.execute(acct_stmt)).fetchall()

        by_account: Dict[str, Dict[str, Any]] = {}
        gl_lease_with_prop = 0.0
        gl_lease_orphan = 0.0
        gl_other_with_prop = 0.0
        gl_other_orphan = 0.0

        for row in acct_rows:
            name = row.account_name or "(tom)"
            if name not in by_account:
                by_account[name] = {"total": 0.0, "with_property": 0.0, "orphan": 0.0, "is_lease": is_lease_account(row.account_name)}
            amt = float(row.total or 0)
            by_account[name]["total"] += amt
            if row.property_id:
                by_account[name]["with_property"] += amt
                if is_lease_account(row.account_name):
                    gl_lease_with_prop += amt
                else:
                    gl_other_with_prop += amt
            else:
                by_account[name]["orphan"] += amt
                if is_lease_account(row.account_name):
                    gl_lease_orphan += amt
                else:
                    gl_other_orphan += amt

        # 2. Kontraktsfestet husleie (aktive kontrakter)
        contract_rows = (await db.execute(
            select(Contract.amount, Unit.property_id)
            .join(Unit, Contract.unit_id == Unit.unit_id)
            .where(Contract.status == "active")
        )).fetchall()

        contracted_total = 0.0
        contracted_by_prop: Dict[Any, float] = {}
        for amount_val, prop_id in contract_rows:
            amt = float((amount_val or {}).get("amount_per_year", 0) or 0) if isinstance(amount_val, dict) else float(amount_val or 0)
            contracted_total += amt
            if prop_id:
                contracted_by_prop[prop_id] = contracted_by_prop.get(prop_id, 0.0) + amt

        gl_lease_total = gl_lease_with_prop + gl_lease_orphan
        gap = contracted_total - gl_lease_total
        gap_pct = round((gap / contracted_total * 100), 1) if contracted_total > 0 else None

        # Sorter kontonavn etter total (størst først)
        account_list = [
            {"account_name": k, **v}
            for k, v in sorted(by_account.items(), key=lambda x: -x[1]["total"])
        ]

        return {
            "year": year,
            "contracted_rent_total": round(contracted_total, 0),
            "gl_lease_total": round(gl_lease_total, 0),
            "gl_lease_with_property": round(gl_lease_with_prop, 0),
            "gl_lease_orphan": round(gl_lease_orphan, 0),
            "gl_other_total": round(gl_other_with_prop + gl_other_orphan, 0),
            "gap": round(gap, 0),
            "gap_pct": gap_pct,
            "accounts": account_list[:50],
            "contract_count": len(contract_rows),
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/rent-gap", response_model=List[Dict[str, Any]])
async def get_rent_gap(
    year: int = Query(2025, description="GL-år for husleie-bokføring"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Sammenlign kontraktsfestet husleie (fra aktive kontrakter) med GL-bokført husleie
    per eiendom. Returnerer gap-beløp og gap-% sortert på størst avvik.
    """
    if current_user.role not in [UserRole.ADMIN, UserRole.REGIONAL_MANAGER]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Krever ADMIN eller REGIONAL_MANAGER")
    try:
        from app.models.financial_models import GLTransaction
        from app.domains.core.models.property import Property
        from app.domains.core.models.contract import Contract
        from app.domains.core.models.unit import Unit

        # 1. GL-husleie per property – match på property_id ELLER department_code = unit_id_erp (som financial-summary)
        from app.models.gl_constants import gl_lease_filter
        gl_stmt = (
            select(GLTransaction.property_id, GLTransaction.dim1_kode, func.sum(GLTransaction.belop).label("gl_rent"))
            .where(
                GLTransaction.ar == year,
                GLTransaction.belop > 0,
                gl_lease_filter(GLTransaction.konto_navn),
            )
            .group_by(GLTransaction.property_id, GLTransaction.dim1_kode)
        )
        gl_rows = (await db.execute(gl_stmt)).fetchall()

        # Map department_code → property_id (unit_id_erp), normalisert for "1234.0" vs "1234"
        def _norm_id(v):
            if v is None: return None
            s = str(v).strip()
            if not s: return None
            if s.endswith(".0") and s[:-2].isdigit():
                return s[:-2]
            return s

        dept_to_prop = (await db.execute(
            select(Property.unit_id_erp, Property.property_id).where(Property.unit_id_erp.isnot(None))
        )).fetchall()
        dept_to_prop_map = {}
        for r in dept_to_prop:
            if r[0]:
                k = _norm_id(r[0])
                if k:
                    dept_to_prop_map[k] = r[1]

        gl_by_prop: Dict[Any, float] = {}
        for r in gl_rows:
            amt = float(r.gl_rent or 0)
            pid = r.property_id
            dept = _norm_id(r.department_code) if r.department_code else None
            if pid:
                gl_by_prop[pid] = gl_by_prop.get(pid, 0.0) + amt
            elif dept and dept in dept_to_prop_map:
                pid_via_dept = dept_to_prop_map[dept]
                gl_by_prop[pid_via_dept] = gl_by_prop.get(pid_via_dept, 0.0) + amt

        # 2. Kontraktsfestet husleie: aktive kontrakter → unit → property
        contract_rows = (await db.execute(
            select(Contract.amount, Unit.property_id)
            .join(Unit, Contract.unit_id == Unit.unit_id)
            .where(Contract.status == "active")
        )).fetchall()

        contracted_by_prop: Dict[Any, float] = {}
        for amount_val, prop_id in contract_rows:
            if isinstance(amount_val, dict):
                amt = float(amount_val.get("amount_per_year") or 0)
            elif isinstance(amount_val, (int, float)):
                amt = float(amount_val)
            else:
                amt = 0.0
            contracted_by_prop[prop_id] = contracted_by_prop.get(prop_id, 0.0) + amt

        # 3. Property info
        prop_rows = (await db.execute(
            select(Property.property_id, Property.name, Property.address, Property.region)
        )).fetchall()
        prop_map = {r.property_id: {"name": r.name or r.address or str(r.property_id), "region": r.region or ""} for r in prop_rows}

        # 4. Merge alle properties som har enten kontraktsfestet eller GL-husleie
        all_prop_ids = {p for p in (set(gl_by_prop.keys()) | set(contracted_by_prop.keys())) if p is not None}
        result = []
        for pid in all_prop_ids:
            contracted = contracted_by_prop.get(pid, 0.0)
            gl_rent = gl_by_prop.get(pid, 0.0)
            gap = contracted - gl_rent
            gap_pct = round((gap / contracted * 100), 1) if contracted > 0 else None
            info = prop_map.get(pid, {"name": str(pid), "region": ""})
            result.append({
                "property_id": str(pid),
                "name": info["name"],
                "region": info["region"],
                "contracted_rent": round(contracted, 0),
                "gl_rent": round(gl_rent, 0),
                "gap": round(gap, 0),
                "gap_pct": gap_pct,
            })

        result.sort(key=lambda x: abs(x["gap"]), reverse=True)
        return result
    except Exception as e:
        return []


@router.get("/yoy-comparison", response_model=List[Dict[str, Any]])
async def get_yoy_comparison(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    År-over-år kostnadsvekst per kostnadskateorii: 2024 vs 2025.
    Returnerer endring i NOK og % per kategori, sortert på størst absolutt endring.
    """
    if current_user.role not in [UserRole.ADMIN, UserRole.REGIONAL_MANAGER]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Krever ADMIN eller REGIONAL_MANAGER")
    try:
        from app.models.financial_models import GLTransaction

        stmt = (
            select(
                GLTransaction.konto_navn,
                GLTransaction.ar,
                func.sum(GLTransaction.belop).label("total"),
            )
            .where(
                GLTransaction.ar.in_([2024, 2025]),
                GLTransaction.belop > 0,
                GLTransaction.konto_navn.isnot(None),
                GLTransaction.konto_navn != "",
            )
            .group_by(GLTransaction.konto_navn, GLTransaction.ar)
        )
        rows = (await db.execute(stmt)).fetchall()

        # Pivot: {category: {2024: X, 2025: Y}}
        pivot: Dict[str, Dict[int, float]] = {}
        for r in rows:
            pivot.setdefault(r.account_name, {})[r.year] = float(r.total)

        result = []
        for category, years in pivot.items():
            amt_2024 = years.get(2024, 0.0)
            amt_2025 = years.get(2025, 0.0)
            change_nok = amt_2025 - amt_2024
            change_pct = round((change_nok / amt_2024 * 100), 1) if amt_2024 > 0 else None
            result.append({
                "category": category,
                "amount_2024": round(amt_2024, 0),
                "amount_2025": round(amt_2025, 0),
                "change_nok": round(change_nok, 0),
                "change_pct": change_pct,
            })

        result.sort(key=lambda x: abs(x["change_nok"]), reverse=True)
        return result
    except Exception as e:
        return []


@router.get("/monthly-budget-actual", response_model=List[Dict[str, Any]])
async def get_monthly_budget_actual(
    year: int = Query(2025, description="År for månedlig budsjett vs faktisk"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    12 måneder med budsjett og GL-faktisk kostnad for valgt år.
    Returnerer varians (budsjett - faktisk) og varians-% per måned.
    """
    if current_user.role not in [UserRole.ADMIN, UserRole.REGIONAL_MANAGER]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Krever ADMIN eller REGIONAL_MANAGER")
    try:
        from app.models.financial_models import GLTransaction, Budget

        budget_stmt = (
            select(Budget.month, func.sum(Budget.amount).label("budget"))
            .where(Budget.year == year)
            .group_by(Budget.month)
        )
        budget_rows = (await db.execute(budget_stmt)).fetchall()
        budget_by_month = {r.month: float(r.budget) for r in budget_rows}

        actual_stmt = (
            select(GLTransaction.maaned, func.sum(GLTransaction.belop).label("actual"))
            .where(GLTransaction.ar == year, GLTransaction.belop > 0)
            .group_by(GLTransaction.maaned)
        )
        actual_rows = (await db.execute(actual_stmt)).fetchall()
        actual_by_month = {r.month: float(r.actual) for r in actual_rows}

        MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "Mai", "Jun", "Jul", "Aug", "Sep", "Okt", "Nov", "Des"]

        result = []
        for m in range(1, 13):
            budget = budget_by_month.get(m, 0.0)
            actual = actual_by_month.get(m, 0.0)
            variance = budget - actual
            variance_pct = round((variance / budget * 100), 1) if budget > 0 else None
            result.append({
                "month": m,
                "month_name": MONTH_NAMES[m - 1],
                "budget": round(budget, 0),
                "actual": round(actual, 0),
                "variance": round(variance, 0),
                "variance_pct": variance_pct,
            })
        return result
    except Exception as e:
        return []


@router.get("/srs-rapport", response_model=Dict[str, Any])
async def get_srs_rapport(
    ar: int = Query(2025, description="År for SRS-rapport"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    SRS-samsvarrapport for regnskapsavdelingen.
    Viser kategorisering (Drift/Investering/Gjennomstrømning), koststed-dekning,
    SRS 13 leieavtaler, SRS 17 anleggsmidler og compliance-status.
    """
    if current_user.role not in [UserRole.ADMIN, UserRole.REGIONAL_MANAGER]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Krever ADMIN eller REGIONAL_MANAGER")

    try:
        from app.models.financial_models import GLTransaction, KoststedMapping, FixedAsset

        # ── 1. SRS-kategorier per år ──────────────────────────────────────────
        kat_stmt = (
            select(
                GLTransaction.srs_kategori,
                GLTransaction.ar,
                func.count().label("antall"),
                func.sum(GLTransaction.belop).label("total_belop"),
            )
            .where(GLTransaction.ar.isnot(None))
            .group_by(GLTransaction.srs_kategori, GLTransaction.ar)
            .order_by(GLTransaction.ar.desc(), GLTransaction.srs_kategori)
        )
        kat_rows = (await db.execute(kat_stmt)).fetchall()
        srs_kategorier = [
            {
                "kategori": r.srs_kategori or "Ukjent",
                "ar": r.ar,
                "antall": r.antall,
                "total_belop": float(r.total_belop or 0),
            }
            for r in kat_rows
        ]

        # ── 2. Koststed-dekning ───────────────────────────────────────────────
        totalt = (await db.execute(select(func.count()).select_from(KoststedMapping))).scalar() or 0
        koblet = (await db.execute(
            select(func.count()).select_from(KoststedMapping).where(KoststedMapping.property_id.isnot(None))
        )).scalar() or 0
        koststed_dekning = {
            "totalt": totalt,
            "koblet": koblet,
            "ukoblet": totalt - koblet,
            "prosent": round(koblet / totalt * 100, 1) if totalt > 0 else 0.0,
        }

        # ── 3. SRS 13 – Leieavtaler (konto 6300 og 6310) ─────────────────────
        leie_stmt = (
            select(
                GLTransaction.konto,
                GLTransaction.konto_navn,
                func.sum(GLTransaction.belop).label("belop"),
                func.count().label("antall"),
            )
            .where(
                GLTransaction.ar == ar,
                GLTransaction.konto.in_(["6300", "6310"]),
            )
            .group_by(GLTransaction.konto, GLTransaction.konto_navn)
            .order_by(GLTransaction.konto)
        )
        leie_rows = (await db.execute(leie_stmt)).fetchall()
        srs_13_leie = [
            {
                "konto": r.konto,
                "konto_navn": r.konto_navn or ("Statsbygg-leie" if r.konto == "6310" else "Privat leie"),
                "belop": float(r.belop or 0),
                "antall": r.antall,
                "type": "Statsbygg" if r.konto == "6310" else "Privat",
            }
            for r in leie_rows
        ]

        # ── 4. SRS 17 – Anleggsmidler ─────────────────────────────────────────
        try:
            anlegg_totalt = (await db.execute(select(func.count()).select_from(FixedAsset))).scalar() or 0
            anlegg_aktiv = (await db.execute(
                select(func.count()).select_from(FixedAsset).where(FixedAsset.srs_status == "Aktiv")
            )).scalar() or 0
            anlegg_bokfort = (await db.execute(
                select(func.sum(FixedAsset.opening_balance_value)).select_from(FixedAsset)
                .where(FixedAsset.srs_status == "Aktiv")
            )).scalar()
            srs_17_anlegg = {
                "totalt": anlegg_totalt,
                "aktive": anlegg_aktiv,
                "total_bokfort": float(anlegg_bokfort or 0),
                "status": "ok" if anlegg_totalt > 0 else "ikke_befolket",
            }
        except Exception:
            srs_17_anlegg = {"totalt": 0, "aktive": 0, "total_bokfort": 0.0, "status": "ikke_befolket"}

        # ── 5. GL-dekning for valgt år ────────────────────────────────────────
        gl_ar_stmt = (
            select(
                GLTransaction.srs_kategori,
                func.count().label("antall"),
                func.sum(GLTransaction.belop).label("belop"),
            )
            .where(GLTransaction.ar == ar)
            .group_by(GLTransaction.srs_kategori)
        )
        gl_ar_rows = (await db.execute(gl_ar_stmt)).fetchall()
        gl_dette_ar = {
            r.srs_kategori or "Ukjent": {"antall": r.antall, "belop": float(r.belop or 0)}
            for r in gl_ar_rows
        }
        totalt_gl = sum(v["antall"] for v in gl_dette_ar.values())

        # ── 6. Compliance-sjekkliste ──────────────────────────────────────────
        har_kategorisering = totalt_gl > 0
        # Full kobling: alle koststed-rader i koststed_mapping skal ha property_id (100 %).
        ks_tot = koststed_dekning["totalt"]
        ks_koblet = koststed_dekning["koblet"]
        ks_ukoblet = koststed_dekning["ukoblet"]
        if ks_tot == 0:
            koststed_compliance_status = "mangler"
        elif ks_ukoblet == 0:
            koststed_compliance_status = "ok"
        elif ks_koblet == 0:
            koststed_compliance_status = "mangler"
        else:
            koststed_compliance_status = "delvis"
        har_leie_klassifisering = len(srs_13_leie) > 0
        compliance = [
            {
                "kode": "SRS-kategorisering",
                "beskrivelse": "GL-transaksjoner er kategorisert som Drift / Investering / Gjennomstrømning",
                "status": "ok" if har_kategorisering else "mangler",
                "detalj": f"{totalt_gl:,} transaksjoner for {ar}",
            },
            {
                "kode": "Koststed-kobling",
                "beskrivelse": "Koststed (Dim1) er koblet til BEFS-eiendommer",
                "status": koststed_compliance_status,
                "detalj": f"{koststed_dekning['koblet']} av {koststed_dekning['totalt']} koststed koblet ({koststed_dekning['prosent']}%)",
            },
            {
                "kode": "SRS 13 – Leie",
                "beskrivelse": "Leieavtaler klassifisert på konto 6300 (privat) og 6310 (Statsbygg)",
                "status": "ok" if har_leie_klassifisering else "mangler",
                "detalj": f"{len(srs_13_leie)} leieavtale-kontoer funnet for {ar}",
            },
            {
                "kode": "SRS 17 – Anleggsmidler",
                "beskrivelse": "Anleggsregister med avskrivningsplan (lineær over levetid)",
                "status": "ok" if srs_17_anlegg["totalt"] > 0 else "pending",
                "detalj": (
                    f"{srs_17_anlegg['aktive']} aktive anleggsmidler"
                    if srs_17_anlegg["totalt"] > 0
                    else "Fase 3 ikke startet – anleggsregisteret er tomt"
                ),
            },
            {
                "kode": "SRS 10 – Nøytralisering",
                "beskrivelse": "Motbilag (33xx/39xx) generert for alle avskrivninger",
                "status": "pending",
                "detalj": "Avventer SRS 17 avskrivningsmotor (fase 3)",
            },
        ]

        return {
            "ar": ar,
            "srs_kategorier": srs_kategorier,
            "koststed_dekning": koststed_dekning,
            "srs_13_leie": srs_13_leie,
            "srs_17_anlegg": srs_17_anlegg,
            "gl_dette_ar": gl_dette_ar,
            "compliance": compliance,
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"SRS-rapport feil: {str(e)}")


@router.get("/koststed-ukoblet", response_model=Dict[str, Any])
async def get_koststed_ukoblet(
    ar: int = Query(2025, description="År for GL-aggregat (antall og beløp)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Lister alle rader i koststed_mapping uten property_id, med GL-aktivitet for valgt år og totalt.
    Brukes til SRS-oppfølging og manuell kobling.
    """
    if current_user.role not in [UserRole.ADMIN, UserRole.REGIONAL_MANAGER]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Krever ADMIN eller REGIONAL_MANAGER")

    try:
        stmt = text("""
            SELECT km.koststed_kode, km.koststed_navn, km.region, km.eksempel_adresse,
                   COUNT(g.transaction_id) FILTER (WHERE g.ar = :ar) AS gl_antall_ar,
                   COALESCE(SUM(g.belop) FILTER (WHERE g.ar = :ar), 0) AS gl_belop_ar,
                   COUNT(g.transaction_id) AS gl_antall_alle
            FROM koststed_mapping km
            LEFT JOIN gl_transactions g ON g.dim1_kode = km.koststed_kode
            WHERE km.property_id IS NULL
            GROUP BY km.koststed_kode, km.koststed_navn, km.region, km.eksempel_adresse
            ORDER BY gl_belop_ar DESC NULLS LAST, km.koststed_kode
        """)
        rows = (await db.execute(stmt, {"ar": ar})).fetchall()
        rader = [
            {
                "koststed_kode": r[0],
                "koststed_navn": r[1] or "",
                "region": r[2] or "",
                "eksempel_adresse": r[3] or "",
                "gl_antall_ar": int(r[4] or 0),
                "gl_belop_ar": float(r[5] or 0),
                "gl_antall_alle": int(r[6] or 0),
            }
            for r in rows
        ]
        return {"ar": ar, "antall": len(rader), "rader": rader}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"koststed-ukoblet feil: {str(e)}")


# ─────────────────────────────────────────────────────────────────────────────
# SRS 17 – Anleggsmidler: populate + hent
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/srs17/populate", response_model=Dict[str, Any])
async def srs17_populate_assets(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Populerer fixed_assets fra GL-transaksjoner der konto IN ('1268', '4960').

    Logikk:
    - Gruppper på dim6_anlegg_id (anleggsnummer fra Agresso).
      Hvis dim6 mangler, gruppér på dim1_kode + tekst.
    - Anskaffelseskost = sum(belop) per gruppe, terskel >= 50 000 NOK.
    - Kjøpsdato = min(bilagsdato) for gruppen.
    - Kobler via koststed_mapping -> property_id -> aktiv Leiekontrakt -> end_date.
    - Beregner: opening_balance_value, monthly_depreciation_amount, remaining_months.
    - Hopper over anleggsmidler som allerede finnes (basert på dim6 eller navn+koststed).
    """
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Kun ADMIN")

    from app.models.financial_models import GLTransaction, KoststedMapping, FixedAsset
    from app.domains.core.models.contract import Contract
    from app.domains.core.models.unit import Unit
    from sqlalchemy import text
    from datetime import date, datetime
    import math

    REF_DATE = date(2025, 1, 1)
    TERSKEL = 50_000  # NOK

    try:
        # 1. Hent GL-rader for konto 1268 / 4960
        gl_stmt = select(
            GLTransaction.dim6_anlegg_id,
            GLTransaction.dim1_kode,
            GLTransaction.tekst,
            GLTransaction.konto,
            GLTransaction.bilagsdato,
            GLTransaction.belop,
        ).where(
            GLTransaction.konto.in_(["1268", "4960"]),
            GLTransaction.belop.isnot(None),
        )
        gl_rows = (await db.execute(gl_stmt)).fetchall()

        # 2. Grupper i Python
        groups: Dict[str, Dict] = {}
        for row in gl_rows:
            key = (row.dim6_anlegg_id or "").strip()
            if not key:
                # Fallback: dim1 + forkortet tekst (maks 80 tegn)
                tekst_short = (row.tekst or "")[:80].strip()
                key = f"{row.dim1_kode or 'X'}|{tekst_short}"
                is_grouped = True
            else:
                is_grouped = False

            if key not in groups:
                groups[key] = {
                    "dim6": (row.dim6_anlegg_id or "").strip() or None,
                    "dim1_kode": row.dim1_kode,
                    "tekst": row.tekst or "",
                    "konto": row.konto,
                    "belop_sum": 0.0,
                    "purchase_date": None,
                    "is_grouped": is_grouped,
                }
            g = groups[key]
            g["belop_sum"] += float(row.belop or 0)
            if row.bilagsdato:
                if g["purchase_date"] is None or row.bilagsdato < g["purchase_date"]:
                    g["purchase_date"] = row.bilagsdato

        # 3. Hent koststed -> property_id mapping
        km_rows = (await db.execute(
            select(KoststedMapping.koststed_kode, KoststedMapping.property_id)
            .where(KoststedMapping.property_id.isnot(None))
        )).fetchall()
        koststed_to_prop = {r.koststed_kode: r.property_id for r in km_rows}

        # 4. Hent aktive kontrakter: property_id -> tidligste end_date
        contract_stmt = (
            select(Unit.property_id, Contract.end_date)
            .join(Contract, Contract.unit_id == Unit.unit_id)
            .where(
                Contract.status == "active",
                Contract.end_date.isnot(None),
            )
            .order_by(Unit.property_id, Contract.end_date.asc())
        )
        contract_rows = (await db.execute(contract_stmt)).fetchall()
        prop_to_lease_end: Dict = {}
        for r in contract_rows:
            pid = str(r.property_id)
            if pid not in prop_to_lease_end or r.end_date < prop_to_lease_end[pid]:
                prop_to_lease_end[pid] = r.end_date

        # 5. Finn eksisterende anleggsmidler
        existing_dim6 = set()
        existing_name_koststed = set()
        existing_rows = (await db.execute(select(FixedAsset.agresso_dim6_id, FixedAsset.asset_name, FixedAsset.koststed_kode))).fetchall()
        for r in existing_rows:
            if r.agresso_dim6_id:
                existing_dim6.add(r.agresso_dim6_id)
            existing_name_koststed.add((r.asset_name[:80], r.koststed_kode or ""))

        # 6. Bygg og insert nye FixedAsset-poster
        inserted = 0
        skipped_terskel = 0
        skipped_exists = 0
        now = datetime.utcnow()

        new_assets = []
        for key, g in groups.items():
            cost = g["belop_sum"]
            if cost < TERSKEL:
                skipped_terskel += 1
                continue

            # Duplikat-sjekk
            dim6 = g["dim6"]
            if dim6 and dim6 in existing_dim6:
                skipped_exists += 1
                continue
            name_key = (g["tekst"][:80], g["dim1_kode"] or "")
            if name_key in existing_name_koststed:
                skipped_exists += 1
                continue

            # Koble til eiendom og leieavtale
            prop_id = koststed_to_prop.get(g["dim1_kode"])
            lease_end = prop_to_lease_end.get(str(prop_id)) if prop_id else None

            # Beregn avskrivning
            if lease_end and lease_end > REF_DATE:
                months_remaining = (
                    (lease_end.year - REF_DATE.year) * 12
                    + (lease_end.month - REF_DATE.month)
                )
                months_remaining = max(months_remaining, 1)
            else:
                months_remaining = None

            monthly_depr = (cost / months_remaining) if months_remaining else None

            asset = FixedAsset(
                asset_name=g["tekst"][:500] or f"Anlegg {dim6 or key[:40]}",
                property_id=prop_id,
                koststed_kode=g["dim1_kode"] or "UKJENT",
                agresso_dim6_id=dim6,
                original_account=g["konto"],
                purchase_date=g["purchase_date"],
                acquisition_cost=cost,
                opening_balance_value=cost,
                monthly_depreciation_amount=monthly_depr,
                remaining_months_at_start=months_remaining,
                lease_end_date=lease_end,
                is_grouped=g["is_grouped"],
                srs_status="Aktiv",
                created_at=now,
                updated_at=now,
            )
            new_assets.append(asset)

        db.add_all(new_assets)
        await db.commit()
        inserted = len(new_assets)

        return {
            "status": "ok",
            "inserted": inserted,
            "skipped_terskel": skipped_terskel,
            "skipped_exists": skipped_exists,
            "total_groups": len(groups),
            "message": f"Lagt inn {inserted} anleggsmidler fra GL (terskel ≥ {TERSKEL:,} NOK)",
        }

    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"SRS17 populate feil: {str(e)}")


@router.get("/srs17/anlegg", response_model=Dict[str, Any])
async def srs17_get_anlegg(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Henter alle anleggsmidler fra fixed_assets med avskrivningsplan.
    Returnerer: liste over aktiva, totalsummer, avskrivningsplan per år 2025-2030.
    """
    if current_user.role not in [UserRole.ADMIN, UserRole.REGIONAL_MANAGER]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Krever ADMIN eller REGIONAL_MANAGER")

    from app.models.financial_models import FixedAsset
    from datetime import date

    try:
        rows = (await db.execute(
            select(FixedAsset).order_by(FixedAsset.acquisition_cost.desc())
        )).scalars().all()

        assets = []
        total_bokfort = 0.0
        total_arlig_avskrivning = 0.0

        for a in rows:
            arlig = float(a.monthly_depreciation_amount or 0) * 12
            total_bokfort += float(a.opening_balance_value or 0)
            total_arlig_avskrivning += arlig
            assets.append({
                "id": str(a.id),
                "asset_name": a.asset_name,
                "koststed_kode": a.koststed_kode,
                "agresso_dim6_id": a.agresso_dim6_id,
                "original_account": a.original_account,
                "purchase_date": str(a.purchase_date) if a.purchase_date else None,
                "acquisition_cost": float(a.acquisition_cost or 0),
                "opening_balance_value": float(a.opening_balance_value or 0),
                "monthly_depreciation_amount": float(a.monthly_depreciation_amount or 0),
                "remaining_months_at_start": a.remaining_months_at_start,
                "lease_end_date": str(a.lease_end_date) if a.lease_end_date else None,
                "srs_status": a.srs_status,
                "is_grouped": a.is_grouped,
            })

        # Avskrivningsplan per år (2025–2033)
        avskrivningsplan = []
        ref = date(2025, 1, 1)
        for year in range(2025, 2034):
            ar_avskr = 0.0
            ar_restverdi = 0.0
            for a in rows:
                if not a.monthly_depreciation_amount or not a.remaining_months_at_start:
                    ar_restverdi += float(a.opening_balance_value or 0)
                    continue
                months_into = (year - 2025) * 12
                months_left = a.remaining_months_at_start - months_into
                if months_left <= 0:
                    continue
                avskr_this_year = float(a.monthly_depreciation_amount) * min(12, months_left)
                ar_avskr += avskr_this_year
                restverdi = float(a.opening_balance_value or 0) - float(a.monthly_depreciation_amount) * months_into
                ar_restverdi += max(restverdi, 0)
            avskrivningsplan.append({
                "ar": year,
                "avskrivning": round(ar_avskr, 2),
                "restverdi_inngaende": round(ar_restverdi, 2),
            })

        return {
            "totalt_antall": len(assets),
            "aktive": sum(1 for a in assets if a["srs_status"] == "Aktiv"),
            "total_bokfort_verdi": round(total_bokfort, 2),
            "total_arlig_avskrivning": round(total_arlig_avskrivning, 2),
            "assets": assets,
            "avskrivningsplan": avskrivningsplan,
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"SRS17 anlegg feil: {str(e)}")


# ─────────────────────────────────────────────────────────────────────────────
# Prediksjon 2027 – Holt-Winters sammendrag
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/prediksjon-2027", response_model=Dict[str, Any])
async def get_prediksjon_2027(
    scenario: str = "xgb70",
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Sammendrag av Holt-Winters 2027-prediksjoner på totalt, region og eiendomsnivå.

    ?scenario=xgb70  → data_source LIKE 'holt_winters_2027_xgb70'  (default)
    ?scenario=xgb50  → data_source LIKE 'holt_winters_2027_xgb50'
    ?scenario=       → alle is_synthetic rader (bakoverkompatibel)

    Returnerer:
    - generert: om budget-tabellen har is_synthetic 2027-data
    - total_2027 / total_2025_gl for direkte sammenligning
    - per_region: region-rollup med endring %
    - per_eiendom_topp20: de 20 dyreste eiendommene i 2027
    - per_kategori: Drift / Investering / Gjennomstrømning
    - sanity: advarsler om urealistiske prediksjoner
    """
    if current_user.role not in [UserRole.ADMIN, UserRole.REGIONAL_MANAGER]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Krever ADMIN eller REGIONAL_MANAGER")

    try:
        from app.models.financial_models import GLTransaction, Budget
        from app.domains.core.models.property import Property

        # Mapping: budget-tabellen bruker engelske kategorier, GL bruker norske
        BUDGET_KAT_MAP = {
            "operations": "Drift",
            "investment": "Investering",
            "property": "Gjennomstrømning",
            "other": "Annet",
        }

        # ── 1. Hent 2027-budsjett fra budget-tabellen ──────────────────────
        data_source_filter = f"holt_winters_2027_{scenario}" if scenario else None
        budget_where = [Budget.year == 2027, Budget.is_synthetic == True]
        if data_source_filter:
            budget_where.append(Budget.data_source == data_source_filter)

        budget_stmt = (
            select(
                Budget.property_id,
                Budget.category,
                func.sum(Budget.amount).label("belop_2027"),
            )
            .where(*budget_where)
            .group_by(Budget.property_id, Budget.category)
        )
        budget_rows = (await db.execute(budget_stmt)).fetchall()

        generert = len(budget_rows) > 0

        # Summér per property_id og per kategori (oversett til norske navn)
        prop_2027: Dict[str, float] = {}
        kat_2027: Dict[str, float] = {}
        for r in budget_rows:
            pid = str(r.property_id)
            prop_2027[pid] = prop_2027.get(pid, 0.0) + float(r.belop_2027 or 0)
            kat = BUDGET_KAT_MAP.get(r.category or "", r.category or "Ukjent")
            kat_2027[kat] = kat_2027.get(kat, 0.0) + float(r.belop_2027 or 0)

        total_2027 = sum(prop_2027.values())

        # ── 2. Hent 2025 GL-faktisk — netto per (eiendom, kategori) ──────────
        gl_stmt = (
            select(
                GLTransaction.property_id,
                GLTransaction.srs_kategori,
                func.sum(GLTransaction.belop).label("belop_2025"),
            )
            .where(
                GLTransaction.ar == 2025,
                GLTransaction.property_id.isnot(None),
            )
            .group_by(GLTransaction.property_id, GLTransaction.srs_kategori)
            .having(func.sum(GLTransaction.belop) > 0)
        )
        gl_rows = (await db.execute(gl_stmt)).fetchall()

        prop_2025: Dict[str, float] = {}
        kat_2025: Dict[str, float] = {}
        for r in gl_rows:
            pid = str(r.property_id)
            prop_2025[pid] = prop_2025.get(pid, 0.0) + float(r.belop_2025 or 0)
            kat = r.srs_kategori or "Ukjent"
            kat_2025[kat] = kat_2025.get(kat, 0.0) + float(r.belop_2025 or 0)

        total_2025 = sum(prop_2025.values())

        # ── 3. Property-info (navn + region) ──────────────────────────────
        prop_stmt = select(Property.property_id, Property.name, Property.address, Property.region)
        prop_rows = (await db.execute(prop_stmt)).fetchall()
        prop_info: Dict[str, Dict] = {
            str(r.property_id): {
                "name": r.name or r.address or str(r.property_id),
                "region": r.region or "Ukjent",
            }
            for r in prop_rows
        }

        # ── 4. Region-rollup ───────────────────────────────────────────────
        region_2027: Dict[str, float] = {}
        region_2025: Dict[str, float] = {}

        for pid, belop in prop_2027.items():
            reg = prop_info.get(pid, {}).get("region", "Ukjent")
            region_2027[reg] = region_2027.get(reg, 0.0) + belop

        for pid, belop in prop_2025.items():
            reg = prop_info.get(pid, {}).get("region", "Ukjent")
            region_2025[reg] = region_2025.get(reg, 0.0) + belop

        alle_regioner = sorted(set(region_2027.keys()) | set(region_2025.keys()))
        per_region = []
        for reg in alle_regioner:
            b27 = region_2027.get(reg, 0.0)
            b25 = region_2025.get(reg, 0.0)
            endring = round((b27 - b25) / b25 * 100, 1) if b25 > 0 else None
            per_region.append({
                "region": reg,
                "belop_2027": round(b27, 0),
                "belop_2025": round(b25, 0),
                "endring_pst": endring,
            })
        per_region.sort(key=lambda x: -x["belop_2027"])

        # ── 5. Topp 20 eiendommer ──────────────────────────────────────────
        alle_pids = set(prop_2027.keys()) | set(prop_2025.keys())
        per_eiendom = []
        for pid in alle_pids:
            b27 = prop_2027.get(pid, 0.0)
            b25 = prop_2025.get(pid, 0.0)
            info = prop_info.get(pid, {"name": pid, "region": "Ukjent"})
            endring = round((b27 - b25) / b25 * 100, 1) if b25 > 0 else None
            per_eiendom.append({
                "property_id": pid,
                "name": info["name"],
                "region": info["region"],
                "belop_2027": round(b27, 0),
                "belop_2025": round(b25, 0),
                "endring_pst": endring,
            })
        per_eiendom.sort(key=lambda x: -x["belop_2027"])
        per_eiendom_topp20 = per_eiendom[:20]

        # ── 6. Per kategori ────────────────────────────────────────────────
        alle_kats = sorted(set(kat_2027.keys()) | set(kat_2025.keys()))
        per_kategori = []
        for kat in alle_kats:
            b27 = kat_2027.get(kat, 0.0)
            b25 = kat_2025.get(kat, 0.0)
            endring = round((b27 - b25) / b25 * 100, 1) if b25 > 0 else None
            per_kategori.append({
                "kategori": kat,
                "belop_2027": round(b27, 0),
                "belop_2025": round(b25, 0),
                "endring_pst": endring,
            })

        # ── 7. Sanity-sjekk ────────────────────────────────────────────────
        advarsler = []
        eiendommer_uten_prediksjon = 0

        # Eiendommer i GL 2025 uten 2027-prediksjon
        for pid in prop_2025:
            if pid not in prop_2027:
                eiendommer_uten_prediksjon += 1

        # Eiendommer der 2027 > 2× 2025
        overdrevet = []
        for pid in prop_2027:
            b27 = prop_2027[pid]
            b25 = prop_2025.get(pid, 0.0)
            if b25 > 0 and b27 > 2 * b25:
                name = prop_info.get(pid, {}).get("name", pid)
                overdrevet.append(name)
        if overdrevet:
            advarsler.append(f"{len(overdrevet)} eiendommer har 2027 > 2×2025: {', '.join(overdrevet[:3])}{'...' if len(overdrevet) > 3 else ''}")

        if eiendommer_uten_prediksjon > 0:
            advarsler.append(f"{eiendommer_uten_prediksjon} eiendommer har GL 2025-data men mangler 2027-prediksjon")

        endring_total = round((total_2027 - total_2025) / total_2025 * 100, 1) if total_2025 > 0 else None

        # ── 8. Lønnsprediksjon 2027 vs 2025 ───────────────────────────────
        total_salary_2027 = 0.0
        total_salary_2025 = 0.0
        try:
            sal_2027 = await db.execute(text("""
                SELECT COALESCE(SUM(faste_stillinger + vikarer + arbeidsgiveravgift), 0)
                FROM salary_costs WHERE year = 2027 AND property_id IS NOT NULL
                  AND import_batch_id = 'holt_winters_2027'
            """))
            total_salary_2027 = float(sal_2027.scalar() or 0)
            sal_2025 = await db.execute(text("""
                SELECT COALESCE(SUM(faste_stillinger + vikarer + arbeidsgiveravgift), 0)
                FROM salary_costs WHERE year = 2025 AND property_id IS NOT NULL
            """))
            total_salary_2025 = float(sal_2025.scalar() or 0)
        except Exception as exc:
            logger.debug("salary_costs query feil i prediksjon-2027: %s", exc)

        lonn_generert = total_salary_2027 > 0
        lonn_endring_pst = (
            round((total_salary_2027 - total_salary_2025) / total_salary_2025 * 100, 1)
            if total_salary_2025 > 0 and lonn_generert
            else None
        )

        return {
            "ar": 2027,
            "scenario": scenario,
            "generert": generert,
            "antall_eiendommer": len(prop_2027),
            "total_2027": round(total_2027, 0),
            "total_2025_gl": round(total_2025, 0),
            "endring_pst": endring_total,
            "per_region": per_region,
            "per_eiendom_topp20": per_eiendom_topp20,
            "per_kategori": per_kategori,
            "sanity": {
                "ok": len(advarsler) == 0,
                "advarsler": advarsler,
                "eiendommer_uten_prediksjon": eiendommer_uten_prediksjon,
            },
            "lonn_2027": round(total_salary_2027, 0),
            "lonn_2025": round(total_salary_2025, 0),
            "lonn_generert": lonn_generert,
            "lonn_endring_pst": lonn_endring_pst,
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Prediksjon 2027 feil: {str(e)}")


async def _build_prediksjon_2027_excel_workbook(db: AsyncSession) -> BytesIO:
    """Bygger prediksjon-2027 Excel inkl. drill-ark, historikkgrunnlag og detaljert kostnadsgrunnlag."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from app.domains.core.models.property import Property
    from app.models.financial_models import Budget, GLTransaction, FinanceBudget
    from app.services.financials.prediksjon_drill_sheets import append_prediksjon_drill_sheets

    SCENARIOS = [
        ("xgb70", "Scenario xgb70 (XGB-gulv 70 %)"),
        ("xgb50", "Scenario xgb50 (XGB-gulv 50 %)"),
    ]
    BUDGET_KAT_MAP = {"operations": "Drift", "investment": "Investering", "property": "Gjennomstrømning", "other": "Annet"}

    # Fetch both scenarios
    scenario_data: dict[str, dict] = {}
    for tag, _ in SCENARIOS:
        rows = (await db.execute(
            select(Budget.property_id, Budget.category, func.sum(Budget.amount).label("belop"))
            .where(Budget.year == 2027, Budget.is_synthetic == True, Budget.data_source == f"holt_winters_2027_{tag}")
            .group_by(Budget.property_id, Budget.category)
        )).fetchall()
        scenario_data[tag] = {str(r.property_id): scenario_data.get(tag, {}).get(str(r.property_id), 0) + float(r.belop or 0) for r in rows}
        # Rebuild properly
        prop_map: dict[str, float] = {}
        for r in rows:
            pid = str(r.property_id)
            prop_map[pid] = prop_map.get(pid, 0.0) + float(r.belop or 0)
        scenario_data[tag] = prop_map

    # 2025 GL actuals — netto per eiendom (HAVING SUM > 0 utelukker negativt-nettede grupper)
    gl_rows = (await db.execute(
        select(GLTransaction.property_id, func.sum(GLTransaction.belop).label("belop"))
        .where(GLTransaction.ar == 2025, GLTransaction.property_id.isnot(None))
        .group_by(GLTransaction.property_id)
        .having(func.sum(GLTransaction.belop) > 0)
    )).fetchall()
    gl_2025: dict[str, float] = {str(r.property_id): float(r.belop or 0) for r in gl_rows}

    # BEFS 2026 estimat: kontant_2025 × regionale vekstrater (reverse-engineered fra økonomibudsjettet)
    BEFS_2026_SOURCES = ['okonomi_regional_2026']
    befs_2026_rows = (await db.execute(
        select(Budget.property_id, func.sum(Budget.amount).label("belop"))
        .where(Budget.year == 2026, Budget.is_synthetic == True, Budget.data_source.in_(BEFS_2026_SOURCES))
        .group_by(Budget.property_id)
    )).fetchall()
    befs_2026: dict[str, float] = {str(r.property_id): float(r.belop or 0) for r in befs_2026_rows}

    # Økonomi vedtatt budsjett 2026 (finance_budget-tabellen, Beløp DA)
    oko_2026_rows = (await db.execute(
        select(FinanceBudget.property_id, func.sum(FinanceBudget.amount).label("belop"))
        .where(
            FinanceBudget.year == 2026,
            FinanceBudget.data_source == "finance_dept_2026",
            FinanceBudget.property_id.isnot(None),
        )
        .group_by(FinanceBudget.property_id)
    )).fetchall()
    oko_2026: dict[str, float] = {str(r.property_id): float(r.belop or 0) for r in oko_2026_rows}

    # Property info
    prop_rows = (await db.execute(select(Property.property_id, Property.name, Property.address, Property.region))).fetchall()
    prop_info = {str(r.property_id): {"name": r.name or r.address or str(r.property_id), "region": r.region or "Ukjent"} for r in prop_rows}

    all_pids = sorted(set(list(scenario_data["xgb70"].keys()) + list(scenario_data["xgb50"].keys()) + list(gl_2025.keys())))

    wb = openpyxl.Workbook()

    # ── Ark 1: Sammendrag ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Sammendrag"
    hdr = Font(bold=True, color="FFFFFF")
    fill_blue = PatternFill("solid", fgColor="1D4ED8")
    fill_gray = PatternFill("solid", fgColor="374151")
    fill_green = PatternFill("solid", fgColor="065F46")

    def cell(ws, row, col, value, bold=False, fill=None, num_format=None):
        c = ws.cell(row=row, column=col, value=value)
        if bold: c.font = Font(bold=True)
        if fill: c.fill = fill; c.font = Font(bold=True, color="FFFFFF")
        if num_format: c.number_format = num_format
        c.alignment = Alignment(horizontal="right" if isinstance(value, (int, float)) else "left")
        return c

    # KPI-tabell
    ws1.merge_cells("A1:D1")
    ws1["A1"] = "Budsjettprediksjon 2027 — Sammenligning av scenarier"
    ws1["A1"].font = Font(bold=True, size=14)
    headers = ["Metrikk", "2025 Faktisk (GL)", "Scenario xgb70 (gulv 70 %)", "Scenario xgb50 (gulv 50 %)"]
    for col, h in enumerate(headers, 1):
        cell(ws1, 2, col, h, fill=fill_blue)

    total_70 = sum(scenario_data["xgb70"].values())
    total_50 = sum(scenario_data["xgb50"].values())
    total_25 = sum(gl_2025.values())

    def pst(a, b): return round((a - b) / b * 100, 1) if b else None

    rows_kpi = [
        ("Total kostnadsbudsjett (NOK)", total_25, total_70, total_50),
        ("Endring vs 2025 (%)", None, pst(total_70, total_25), pst(total_50, total_25)),
        ("Antall eiendommer", len(gl_2025), len(scenario_data["xgb70"]), len(scenario_data["xgb50"])),
    ]
    for i, (label, v25, v70, v50) in enumerate(rows_kpi, 3):
        ws1.cell(row=i, column=1, value=label).font = Font(bold=True)
        for col, val in [(2, v25), (3, v70), (4, v50)]:
            c = ws1.cell(row=i, column=col, value=val)
            if isinstance(val, float) and abs(val) > 1000: c.number_format = '#,##0'

    # Region-tabell
    ws1.cell(row=7, column=1, value="Region-fordeling").font = Font(bold=True, size=12)
    for col, h in enumerate(["Region", "2025 Faktisk", "XGB 70%", "XGB 50%", "Diff 70 vs 50"], 1):
        cell(ws1, 8, col, h, fill=fill_gray)

    regions: dict[str, dict] = {}
    for pid in all_pids:
        reg = prop_info.get(pid, {}).get("region", "Ukjent")
        if reg not in regions:
            regions[reg] = {"gl": 0.0, "xgb70": 0.0, "xgb50": 0.0}
        regions[reg]["gl"] += gl_2025.get(pid, 0)
        regions[reg]["xgb70"] += scenario_data["xgb70"].get(pid, 0)
        regions[reg]["xgb50"] += scenario_data["xgb50"].get(pid, 0)

    for i, (reg, vals) in enumerate(sorted(regions.items()), 9):
        ws1.cell(row=i, column=1, value=reg)
        for col, val in [(2, vals["gl"]), (3, vals["xgb70"]), (4, vals["xgb50"]), (5, vals["xgb70"] - vals["xgb50"])]:
            c = ws1.cell(row=i, column=col, value=round(val, 0))
            c.number_format = '#,##0'

    for col in range(1, 6):
        ws1.column_dimensions[get_column_letter(col)].width = 22

    # ── Ark 2: Alle eiendommer ────────────────────────────────────────────
    ws2 = wb.create_sheet("Alle eiendommer")
    headers2 = ["Eiendom", "Region", "2025 Faktisk (NOK)", "XGB 70% (NOK)", "XGB 50% (NOK)", "Diff 70-50 (NOK)", "Endring 70% vs 2025 (%)"]
    for col, h in enumerate(headers2, 1):
        cell(ws2, 1, col, h, fill=fill_green)

    for i, pid in enumerate(sorted(all_pids, key=lambda p: -scenario_data["xgb70"].get(p, 0)), 2):
        name = prop_info.get(pid, {}).get("name", pid)
        reg = prop_info.get(pid, {}).get("region", "Ukjent")
        v25 = gl_2025.get(pid, 0)
        v70 = scenario_data["xgb70"].get(pid, 0)
        v50 = scenario_data["xgb50"].get(pid, 0)
        diff = v70 - v50
        endring = round((v70 - v25) / v25 * 100, 1) if v25 > 0 else None
        ws2.cell(row=i, column=1, value=name)
        ws2.cell(row=i, column=2, value=reg)
        for col, val in [(3, v25), (4, v70), (5, v50), (6, diff)]:
            c = ws2.cell(row=i, column=col, value=round(val, 0))
            c.number_format = '#,##0'
        if endring is not None:
            c = ws2.cell(row=i, column=7, value=endring)
            c.number_format = '0.0"%"'

    for col, w in [(1, 40), (2, 16), (3, 20), (4, 20), (5, 20), (6, 20), (7, 22)]:
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ── Ark 3: Forklaring / Metodebeskrivelse ─────────────────────────────
    ws3 = wb.create_sheet("Forklaring")
    fill_dark = PatternFill("solid", fgColor="1E3A5F")
    fill_light = PatternFill("solid", fgColor="EFF6FF")

    def txt(ws, row, col, value, bold=False, italic=False, size=11, fill=None, color="000000"):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, italic=italic, size=size, color=color if fill else color)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if fill:
            c.fill = fill
            c.font = Font(bold=bold, size=size, color="FFFFFF")
        return c

    ws3.merge_cells("A1:C1")
    txt(ws3, 1, 1, "Budsjettprediksjon 2027 — Metodebeskrivelse", bold=True, size=14, fill=fill_dark)
    txt(ws3, 2, 1, "Generert av BEFS (Bufetat Eiendomsforvaltningssystem)", italic=True, size=10, color="555555")

    forklaring = [
        ("", ""),
        (
            "VIKTIG — les dette først (prosentene i oversikten)",
            "Kolonnene «Scenario xgb70 (gulv 70 %)» og «Scenario xgb50 (gulv 50 %)» er ikke «70 % eller 50 % internt forbruk», aktivitetsnivå eller Holt-Winters α. "
            "De er to modellscenarier for hvor streng den nedre grensen (gulvet) fra XGBoost skal være i forhold til maskinlæringsprediksjon på tvers av porteføljen, etter at Holt-Winters er kjørt. "
            "«2025 Faktisk (GL)» er bokført kostnad i regnskapet (alle relevante linjer summert). «Pred. 2027» er syntetiske budsjettposter — ikke samme definisjon som rå GL-sum, og eiendommer uten generert prediksjon inngår ikke i 2027-totalen men kan fortsatt ha GL i 2025. "
            "Derfor kan total pred. 2027 være lavere eller høyere enn GL 2025 uten at det alene betyr feil.",
        ),
        ("", ""),
        ("OM PREDIKSJONEN", None),
        ("Datagrunnlag", "Agresso GL-transaksjoner 2021–2025, gruppert per eiendom og SRS-kategori (Drift / Investering / Gjennomstrømning). Kun poster med beløp > 0 og kobling til BEFS-eiendom er inkludert."),
        ("Historikk brukt", "5 år (2021–2025). Eiendommer uten GL-data i denne perioden er ikke inkludert i prediksjonen."),
        ("Antall eiendommer", "190 eiendommer med prediksjon, fordelt på alle regioner."),
        ("", ""),
        ("ALGORITME — STEG 1: HOLT-WINTERS", None),
        ("Metode", "Holt-Winters dobbel eksponensiell utjevning med dempet trend. Kjøres per eiendom og per SRS-kategori separat."),
        ("α = 0,70", "Vekting av nivå — nyere år gir størst bidrag til estimatet."),
        ("β = 0,30", "Trendglattning — demper brå svingninger i trenden."),
        ("φ = 0,85", "Dempingsfaktor — hindrer ukontrollert eksponensiell trend-ekstrapolasjon."),
        ("Inflasjonsfallback", "3,5 % per år brukes som vekstrate for eiendommer med < 2 år historikk."),
        ("Minimumsgulv", "Hvis HW-prediksjon faller under 10 % av historisk gjennomsnitt, brukes gjennomsnitt × inflasjon i stedet. Forhindrer urealistisk nedgang (eks. Åsen, Tune)."),
        ("", ""),
        ("ALGORITME — STEG 2: XGBOOST-GULV", None),
        ("Metode", "Maskinlæringsmodell (XGBoost Regressor) trent på tvers av alle eiendommer. Brukes som nedre grense — ikke som erstatning for HW."),
        ("Treningsdata", "232 (eiendom, kategori)-par med ≥ 4 år historikk. 127 unike eiendommer. Fordeling: Drift 146, Gjennomstrømning 80, Investering 6."),
        ("Features", "log(historisk gjennomsnitt), log(siste kjente år), antall år data, relativ trend (siste/tidlige år), variasjonskoeffisient, SRS-kategori (encoded)."),
        ("Formel", "Endelig verdi = max(HW-prediksjon,  XGB-prediksjon × gulv-faktor)"),
        ("", ""),
        ("SCENARIER (xgb70 / xgb50)", None),
        ("Scenario xgb70 (gulv 70 %)", "Konservativt scenario — anbefalt for budsjettplanlegging. XGBoost-gulvet settes til 70 % av tverreiendomms-XGB-prediksjon der gulvet aktiveres."),
        ("Scenario xgb50 (gulv 50 %)", "Optimistisk scenario — lavere gulv (50 % av tverreiendomms-XGB-prediksjon) der gulvet aktiveres."),
        ("", ""),
        ("TOLKNING AV RESULTATER", None),
        ("Diff xgb70 − xgb50 (NOK)", "Differansen mellom de to scenariene per eiendom. Stor differanse betyr ofte at HW ga lavt tall og XGB-gulvet skiller scenariene."),
        ("Endring xgb70 vs 2025 GL", "Prosentvis endring fra faktisk GL 2025 til xgb70-estimat. Høy prosent kan bety kraftig vekst — vurder årsak og datakvalitet."),
        ("is_synthetic = true", "Alle predikterte budsjettlinjer lagres med dette flagget i budget-tabellen, slik at de skilles fra vedtatte budsjett."),
        ("data_source", "holt_winters_2027_xgb70 (Ark 2 XGB 70%) / holt_winters_2027_xgb50 (Ark 2 XGB 50%). Brukes til å filtrere i avviksanalysen."),
        ("", ""),
        ("DRILL-DOWN ARK", None),
        ("Eiendom_kategori", "Faktisk GL 2025 og 2027-prediksjon (XGB 70/50) per eiendom og SRS-kategori. Klikk på property_id eller eiendomsnavn for å hoppe videre."),
        ("GL_konto", "Aggregerte kontoposter for 2025 (kun beløp > 0 i grunnlagslinjer), alle eiendommer."),
        ("GL_bilag", "Enkeltbilag/linjer for 2025; maks ca. 80 000 rader, sortert etter eiendom og beløp for rask drill-down."),
        ("Historikk_grunnlag", "Samlet historisk kostnadsgrunnlag 2021–2025 per eiendom og SRS-kategori. Dette er basisen som prediksjonen bygger på."),
        ("Kostnadsgrunnlag_GL", "Detaljert kostnadsliste for hele historikkgrunnlaget, med år, periode, bilag, konto, leverandør, tekst og beløp."),
        ("", ""),
        ("BEGRENSNINGER", None),
        ("Ikke inkludert", "Nybygg uten historikk. Eiendommer avviklet etter 2025. Strukturelle endringer etter 2025."),
        ("Investering-kategorien", "Kun 6 treningspar — XGB-estimater for Investering er mindre pålitelige enn Drift/Gjennomstrømning."),
        ("Ansvarsfraskrivelse", "Tallene er estimater for planleggingsformål — ikke godkjente budsjetttall. Endelig budsjett vedtas av ledelsen."),
    ]

    row = 4
    for label, tekst in forklaring:
        if tekst is None:
            # Seksjonsoverskrift
            ws3.merge_cells(f"A{row}:C{row}")
            txt(ws3, row, 1, label, bold=True, size=11, fill=fill_dark)
        elif label == "":
            row += 1
            continue
        else:
            txt(ws3, row, 1, label, bold=True, size=10)
            ws3.merge_cells(f"B{row}:C{row}")
            txt(ws3, row, 2, tekst, size=10)
            ws3.row_dimensions[row].height = max(15, len(tekst) // 6)
        row += 1

    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 55
    ws3.column_dimensions["C"].width = 20

    await append_prediksjon_drill_sheets(db, wb, prop_info=prop_info)

    # ── Ark: Sammenligning 2026 ───────────────────────────────────────────────
    ws_s = wb.create_sheet("Sammenligning 2026")

    REGIONS_ORDER = ["Bufdir", "Midt-Norge", "Nord", "Sør", "Vest", "Øst"]
    fill_teal   = PatternFill("solid", fgColor="0F766E")   # header økonomi
    fill_violet = PatternFill("solid", fgColor="6D28D9")   # header BEFS
    fill_dark   = PatternFill("solid", fgColor="1F2937")   # section header
    fill_red    = PatternFill("solid", fgColor="FEE2E2")   # over budsjett
    fill_grn    = PatternFill("solid", fgColor="DCFCE7")   # under budsjett
    fill_tot    = PatternFill("solid", fgColor="F3F4F6")   # totalrad
    thin        = Side(style="thin", color="D1D5DB")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    def sc(ws, row, col, value, bold=False, fill=None, italic=False,
           num_fmt=None, align="left", color=None):
        c = ws.cell(row=row, column=col, value=value)
        f = Font(bold=bold, italic=italic, color=color or "000000")
        c.font = f
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
        c.border = border
        if fill:
            c.fill = fill
            c.font = Font(bold=bold, color="FFFFFF", italic=italic)
        if num_fmt:
            c.number_format = num_fmt
        return c

    def nok_fmt(v: float) -> str:
        """Tall-format vises via num_format — verdien sendes som float."""
        return v

    # ── Tittel ────────────────────────────────────────────────────────────────
    ws_s.merge_cells("A1:G1")
    t = ws_s["A1"]
    t.value = "Sammenligning 2026: BEFS-estimat vs. Vedtatt økonomibudsjett (Beløp DA)"
    t.font = Font(bold=True, size=13)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws_s.row_dimensions[1].height = 22

    # ── KPI-rad ───────────────────────────────────────────────────────────────
    all_s26 = set(befs_2026) | set(oko_2026)
    tot_befs = sum(befs_2026.values())
    tot_oko  = sum(oko_2026.values())
    tot_avvik = tot_befs - tot_oko
    tot_pst   = round(tot_avvik / tot_oko * 100, 1) if tot_oko else None

    kpi_labels = [
        ("BEFS-estimat 2026 (total)",   tot_befs,  fill_violet),
        ("Vedtatt budsjett 2026 (Øk.)", tot_oko,   fill_teal),
        ("Avvik (BEFS − Økonomi)",       tot_avvik, PatternFill("solid", fgColor="7C3AED" if tot_avvik > 0 else "065F46")),
        ("Avvik %",                       tot_pst,   PatternFill("solid", fgColor="374151")),
    ]
    for col, (label, val, pfill) in enumerate(kpi_labels, 1):
        sc(ws_s, 3, col, label, bold=True, fill=pfill)
        fmt = '+#,##0.0"%";-#,##0.0"%"' if label == "Avvik %" else '#,##0'
        c = ws_s.cell(row=4, column=col, value=val)
        c.number_format = fmt
        c.font = Font(bold=True, size=12)
        c.alignment = Alignment(horizontal="right")
        c.border = border
    ws_s.row_dimensions[3].height = 18
    ws_s.row_dimensions[4].height = 22

    # ── Regional tabell ───────────────────────────────────────────────────────
    ws_s.cell(row=6, column=1).value = "Regional oversikt"
    ws_s.cell(row=6, column=1).font = Font(bold=True, size=11)
    ws_s.row_dimensions[6].height = 18

    reg_hdrs = ["Region", "BEFS-estimat 2026", "Vedtatt budsjett (Øk.)", "Avvik (NOK)", "Avvik (%)"]
    for col, h in enumerate(reg_hdrs, 1):
        sc(ws_s, 7, col, h, bold=True, fill=fill_dark)
    ws_s.row_dimensions[7].height = 16

    reg_befs: dict[str, float] = {}
    reg_oko:  dict[str, float] = {}
    for pid in all_s26:
        reg = prop_info.get(pid, {}).get("region", "Ukjent")
        reg_befs[reg] = reg_befs.get(reg, 0.0) + befs_2026.get(pid, 0.0)
        reg_oko[reg]  = reg_oko.get(reg, 0.0)  + oko_2026.get(pid, 0.0)

    row = 8
    for reg in REGIONS_ORDER:
        b = reg_befs.get(reg, 0.0)
        o = reg_oko.get(reg, 0.0)
        if b == 0 and o == 0:
            continue
        avvik = b - o
        pst_v = round(avvik / o * 100, 1) if o else None
        abs_pst = abs(pst_v) if pst_v is not None else 0
        row_fill = fill_red if avvik > 0 and abs_pst > 5 else (fill_grn if avvik < 0 and abs_pst > 5 else None)
        sc(ws_s, row, 1, reg, bold=False, fill=row_fill)
        sc(ws_s, row, 2, round(b, 0), num_fmt='#,##0', align="right", fill=row_fill)
        sc(ws_s, row, 3, round(o, 0), num_fmt='#,##0', align="right", fill=row_fill)
        sc(ws_s, row, 4, round(avvik, 0), num_fmt='+#,##0;-#,##0;0', align="right", fill=row_fill)
        pst_cell = sc(ws_s, row, 5, pst_v, num_fmt='+0.0%;-0.0%;0%', align="right", fill=row_fill)
        if row_fill is None:
            pst_cell.font = Font(bold=True, color="166534" if avvik < 0 else ("991B1B" if avvik > 0 else "374151"))
        row += 1

    # Totalrad
    for col, val in enumerate([
        "TOTAL", round(tot_befs, 0), round(tot_oko, 0),
        round(tot_avvik, 0), tot_pst
    ], 1):
        fmts = [None, '#,##0', '#,##0', '+#,##0;-#,##0;0', '+0.0%;-0.0%;0%']
        c = sc(ws_s, row, col, val, bold=True, fill=fill_tot)
        c.font = Font(bold=True, color="000000")
        if fmts[col - 1]:
            c.number_format = fmts[col - 1]
        c.alignment = Alignment(horizontal="right" if col > 1 else "left")
    ws_s.row_dimensions[row].height = 16
    row += 2

    # ── Eiendommstabell ───────────────────────────────────────────────────────
    ws_s.cell(row=row, column=1).value = "Per eiendom — sortert etter absolutt avvik"
    ws_s.cell(row=row, column=1).font = Font(bold=True, size=11)
    ws_s.row_dimensions[row].height = 18
    row += 1

    prop_hdrs = ["Eiendom", "Region", "BEFS-estimat 2026", "Vedtatt budsjett (Øk.)", "Avvik (NOK)", "Avvik (%)", "Status"]
    for col, h in enumerate(prop_hdrs, 1):
        sc(ws_s, row, col, h, bold=True, fill=fill_dark)
    ws_s.row_dimensions[row].height = 16
    row += 1

    prop_rows_sorted = sorted(
        all_s26,
        key=lambda p: -abs(befs_2026.get(p, 0.0) - oko_2026.get(p, 0.0)),
    )

    for pid in prop_rows_sorted:
        b = befs_2026.get(pid, 0.0)
        o = oko_2026.get(pid, 0.0)
        if b == 0 and o == 0:
            continue
        name = prop_info.get(pid, {}).get("name", pid)
        reg  = prop_info.get(pid, {}).get("region", "Ukjent")
        avvik = b - o
        pst_v = round(avvik / o * 100, 1) if o else None
        abs_pst = abs(pst_v) if pst_v is not None else 0
        har_begge = b > 0 and o > 0
        row_fill = (fill_red if avvik > 0 else fill_grn) if har_begge else None

        is_statlig = "statlig" in name.lower()
        status = "⚠ Statlig sekkepost" if is_statlig else ("Kun BEFS" if o == 0 else ("Kun Øk." if b == 0 else ""))

        sc(ws_s, row, 1, name, fill=row_fill)
        sc(ws_s, row, 2, reg, fill=row_fill)
        sc(ws_s, row, 3, round(b, 0) if b else None, num_fmt='#,##0', align="right", fill=row_fill)
        sc(ws_s, row, 4, round(o, 0) if o else None, num_fmt='#,##0', align="right", fill=row_fill)
        sc(ws_s, row, 5, round(avvik, 0) if har_begge else None, num_fmt='+#,##0;-#,##0;0', align="right", fill=row_fill)
        sc(ws_s, row, 6, pst_v if har_begge else None, num_fmt='+0.0%;-0.0%;0%', align="right", fill=row_fill)
        sc(ws_s, row, 7, status, italic=is_statlig, fill=row_fill)
        row += 1

    # ── Kildeforklaring ───────────────────────────────────────────────────────
    row += 1
    note_lines = [
        "Kilder og metodikk:",
        "• BEFS-estimat 2026: budget-tabellen — kontant_2025 per eiendom × regional vekstrate (okonomi_regional_2026).",
        "  GL 2025 husleie + 4,7 % og GL 2025 drift + 10,0 % inflasjonsjustering. Sum ≈ 604 M kr.",
        "• Vedtatt budsjett (Økonomi): finance_budget-tabellen, data_source='finance_dept_2026'.",
        "  Beløp DA = vedtatt fulltårsramme fra Agresso — IKKE Kontantbeløp (kassabevegelse).",
        "• Avvik = BEFS − Økonomi. Rødt = BEFS over budsjett, grønt = BEFS under budsjett.",
        "• ⚠ Statlig-eiendommer er nasjonale sekkepostobjekter — ikke geografiske eiendommer.",
        "  Se _compute_statlig_split() for korrekt regional fordeling.",
    ]
    for line in note_lines:
        ws_s.merge_cells(f"A{row}:G{row}")
        c = ws_s.cell(row=row, column=1, value=line)
        c.font = Font(size=9, italic=line.startswith("•"), color="374151")
        c.alignment = Alignment(horizontal="left")
        row += 1

    # Kolonnebredder
    col_widths = [38, 14, 20, 22, 18, 12, 22]
    for i, w in enumerate(col_widths, 1):
        ws_s.column_dimensions[get_column_letter(i)].width = w

    # Frys topp-rad i eiendommstabellen
    ws_s.freeze_panes = "A8"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/prediksjon-2027/excel")
async def get_prediksjon_excel(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Laster ned Excel-fil med begge scenarier (XGB 70% og XGB 50%) side om side.
    Ark: Sammendrag, Alle eiendommer, Forklaring, Eiendom_kategori, GL_konto, GL_bilag.
    """
    from fastapi.responses import StreamingResponse

    if current_user.role not in [UserRole.ADMIN, UserRole.REGIONAL_MANAGER]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Krever ADMIN eller REGIONAL_MANAGER")

    buf = await _build_prediksjon_2027_excel_workbook(db)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=prediksjon_2027.xlsx"},
    )


@router.get("/prediksjon-2027/export.xlsx")
async def get_prediksjon_2027_export_xlsx(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    scenario: str = Query("xgb70", description="Reservert; arbeidsboken inneholder begge XGB-scenarioer."),
):
    """Samme innhold som /prediksjon-2027/excel; filnavn prediksjon_2027_export.xlsx."""
    from fastapi.responses import StreamingResponse

    if current_user.role not in [UserRole.ADMIN, UserRole.REGIONAL_MANAGER]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Krever ADMIN eller REGIONAL_MANAGER")

    buf = await _build_prediksjon_2027_excel_workbook(db)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=prediksjon_2027_export.xlsx"},
    )


@router.get("/prediksjon-2027/region/{region}", response_model=List[Dict[str, Any]])
async def get_prediksjon_by_region(
    region: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Alle eiendommer i én region med 2025 GL og 2027-prediksjon."""
    if current_user.role not in [UserRole.ADMIN, UserRole.REGIONAL_MANAGER]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Krever ADMIN eller REGIONAL_MANAGER")

    try:
        from app.models.financial_models import GLTransaction, Budget
        from app.domains.core.models.property import Property

        # Hent property_id-er i regionen
        prop_stmt = select(Property.property_id, Property.name, Property.address).where(
            Property.region == region
        )
        prop_rows = (await db.execute(prop_stmt)).fetchall()
        prop_ids = [r.property_id for r in prop_rows]
        prop_info = {
            str(r.property_id): r.name or r.address or str(r.property_id)
            for r in prop_rows
        }

        if not prop_ids:
            return []

        # 2027-budsjett
        budget_stmt = (
            select(Budget.property_id, func.sum(Budget.amount).label("belop_2027"))
            .where(Budget.year == 2027, Budget.is_synthetic == True, Budget.property_id.in_(prop_ids))
            .group_by(Budget.property_id)
        )
        budget_rows = (await db.execute(budget_stmt)).fetchall()
        prop_2027 = {str(r.property_id): float(r.belop_2027 or 0) for r in budget_rows}

        # 2025 GL — netto (regel 7 CLAUDE.md)
        gl_stmt = (
            select(GLTransaction.property_id, func.sum(GLTransaction.belop).label("belop_2025"))
            .where(GLTransaction.ar == 2025, GLTransaction.property_id.in_(prop_ids))
            .group_by(GLTransaction.property_id)
            .having(func.sum(GLTransaction.belop) > 0)
        )
        gl_rows = (await db.execute(gl_stmt)).fetchall()
        prop_2025 = {str(r.property_id): float(r.belop_2025 or 0) for r in gl_rows}

        alle_pids = set(prop_2027.keys()) | set(prop_2025.keys())
        result = []
        for pid in alle_pids:
            b27 = prop_2027.get(pid, 0.0)
            b25 = prop_2025.get(pid, 0.0)
            endring = round((b27 - b25) / b25 * 100, 1) if b25 > 0 else None
            result.append({
                "property_id": pid,
                "name": prop_info.get(pid, pid),
                "region": region,
                "belop_2025": round(b25, 0),
                "belop_2027": round(b27, 0),
                "endring_pst": endring,
            })
        result.sort(key=lambda x: -x["belop_2027"])
        return result

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Region drill-down feil: {str(e)}")


@router.get("/prediksjon-2027/kategori/{kategori}", response_model=List[Dict[str, Any]])
async def get_prediksjon_by_kategori(
    kategori: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Alle eiendommer sortert etter beløp for én SRS-kategori (Drift/Investering/Gjennomstrømning)."""
    if current_user.role not in [UserRole.ADMIN, UserRole.REGIONAL_MANAGER]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Krever ADMIN eller REGIONAL_MANAGER")

    KAT_MAP = {"Drift": "operations", "Investering": "investment", "Gjennomstrømning": "property", "Annet": "other"}
    budget_kat = KAT_MAP.get(kategori, kategori.lower())

    try:
        from app.models.financial_models import GLTransaction, Budget
        from app.domains.core.models.property import Property

        # 2027-budsjett for kategorien
        budget_stmt = (
            select(Budget.property_id, func.sum(Budget.amount).label("belop_2027"))
            .where(Budget.year == 2027, Budget.is_synthetic == True, Budget.category == budget_kat)
            .group_by(Budget.property_id)
        )
        budget_rows = (await db.execute(budget_stmt)).fetchall()
        prop_2027 = {str(r.property_id): float(r.belop_2027 or 0) for r in budget_rows}

        # 2025 GL for kategorien — netto per eiendom (regel 7 CLAUDE.md)
        gl_stmt = (
            select(GLTransaction.property_id, func.sum(GLTransaction.belop).label("belop_2025"))
            .where(
                GLTransaction.ar == 2025,
                GLTransaction.property_id.isnot(None),
                GLTransaction.srs_kategori == kategori,
            )
            .group_by(GLTransaction.property_id)
            .having(func.sum(GLTransaction.belop) > 0)
        )
        gl_rows = (await db.execute(gl_stmt)).fetchall()
        prop_2025 = {str(r.property_id): float(r.belop_2025 or 0) for r in gl_rows}

        # Property-info
        alle_pids_uuid = list(set(
            [r.property_id for r in budget_rows] + [r.property_id for r in gl_rows]
        ))
        prop_info: Dict[str, Dict] = {}
        if alle_pids_uuid:
            prop_stmt = select(Property.property_id, Property.name, Property.address, Property.region).where(
                Property.property_id.in_(alle_pids_uuid)
            )
            p_rows = (await db.execute(prop_stmt)).fetchall()
            prop_info = {
                str(r.property_id): {"name": r.name or r.address or str(r.property_id), "region": r.region or "Ukjent"}
                for r in p_rows
            }

        alle_pids = set(prop_2027.keys()) | set(prop_2025.keys())
        result = []
        for pid in alle_pids:
            b27 = prop_2027.get(pid, 0.0)
            b25 = prop_2025.get(pid, 0.0)
            info = prop_info.get(pid, {"name": pid, "region": "Ukjent"})
            endring = round((b27 - b25) / b25 * 100, 1) if b25 > 0 else None
            result.append({
                "property_id": pid,
                "name": info["name"],
                "region": info["region"],
                "belop_2025": round(b25, 0),
                "belop_2027": round(b27, 0),
                "endring_pst": endring,
            })
        result.sort(key=lambda x: -x["belop_2027"])
        return result

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Kategori drill-down feil: {str(e)}")


@router.get("/prediksjon-2027/eiendom/{property_id}", response_model=Dict[str, Any])
async def get_prediksjon_by_eiendom(
    property_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Full prediksjon-detalj for én eiendom: historikk 2021–2025 + 2027-prediksjon per kategori."""
    if current_user.role not in [UserRole.ADMIN, UserRole.REGIONAL_MANAGER]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Krever ADMIN eller REGIONAL_MANAGER")

    try:
        from app.models.financial_models import GLTransaction, Budget
        from app.domains.core.models.property import Property
        import uuid as uuid_lib

        try:
            pid_uuid = uuid_lib.UUID(property_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Ugyldig property_id format")

        # Property-info
        prop_row = (await db.execute(
            select(Property.property_id, Property.name, Property.address, Property.region)
            .where(Property.property_id == pid_uuid)
        )).fetchone()
        if not prop_row:
            raise HTTPException(status_code=404, detail="Eiendom ikke funnet")

        name = prop_row.name or prop_row.address or property_id
        region = prop_row.region or "Ukjent"

        # Historikk 2021–2025 fra GL — netto per år (regel 7 CLAUDE.md)
        hist_stmt = (
            select(GLTransaction.ar, func.sum(GLTransaction.belop).label("belop"))
            .where(
                GLTransaction.property_id == pid_uuid,
                GLTransaction.ar.between(2021, 2025),
            )
            .group_by(GLTransaction.ar)
            .having(func.sum(GLTransaction.belop) > 0)
            .order_by(GLTransaction.ar)
        )
        hist_rows = (await db.execute(hist_stmt)).fetchall()
        historikk = {str(r.ar): round(float(r.belop or 0), 0) for r in hist_rows}

        # 2027-prediksjon total
        pred_total_stmt = (
            select(func.sum(Budget.amount).label("belop_2027"))
            .where(Budget.property_id == pid_uuid, Budget.year == 2027, Budget.is_synthetic == True)
        )
        pred_total = float((await db.execute(pred_total_stmt)).scalar() or 0)

        # Per-kategori: 2025 GL + 2027 budget
        BUDGET_KAT_MAP = {
            "operations": "Drift",
            "investment": "Investering",
            "property": "Gjennomstrømning",
            "other": "Annet",
        }

        kat_2027_stmt = (
            select(Budget.category, func.sum(Budget.amount).label("belop_2027"))
            .where(Budget.property_id == pid_uuid, Budget.year == 2027, Budget.is_synthetic == True)
            .group_by(Budget.category)
        )
        kat_2027_rows = (await db.execute(kat_2027_stmt)).fetchall()
        kat_2027 = {BUDGET_KAT_MAP.get(r.category, r.category): float(r.belop_2027 or 0) for r in kat_2027_rows}

        # 2025 per kategori — netto (HAVING SUM > 0 ekskluderer negativ-nettede kategorier)
        kat_2025_stmt = (
            select(GLTransaction.srs_kategori, func.sum(GLTransaction.belop).label("belop_2025"))
            .where(
                GLTransaction.property_id == pid_uuid,
                GLTransaction.ar == 2025,
            )
            .group_by(GLTransaction.srs_kategori)
            .having(func.sum(GLTransaction.belop) > 0)
        )
        kat_2025_rows = (await db.execute(kat_2025_stmt)).fetchall()
        kat_2025 = {(r.srs_kategori or "Ukjent"): float(r.belop_2025 or 0) for r in kat_2025_rows}

        alle_kats = sorted(set(kat_2027.keys()) | set(kat_2025.keys()))
        per_kategori = []
        for kat in alle_kats:
            b27 = kat_2027.get(kat, 0.0)
            b25 = kat_2025.get(kat, 0.0)
            endring = round((b27 - b25) / b25 * 100, 1) if b25 > 0 else None
            per_kategori.append({
                "kategori": kat,
                "belop_2025": round(b25, 0),
                "belop_2027": round(b27, 0),
                "endring_pst": endring,
            })

        return {
            "property_id": property_id,
            "name": name,
            "region": region,
            "historikk": historikk,
            "prediksjon_2027": round(pred_total, 0),
            "per_kategori": per_kategori,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Eiendom drill-down feil: {str(e)}")


@router.get("/prediksjon-2027/eiendom/{property_id}/konto", response_model=List[Dict[str, Any]])
async def get_prediksjon_konto_for_eiendom(
    property_id: str,
    srs_kategori: Optional[str] = None,
    year: int = 2025,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Kontodetaljer (konto-aggregat) for én eiendom og ett år, valgfritt filtrert på srs_kategori."""
    if current_user.role not in [UserRole.ADMIN, UserRole.REGIONAL_MANAGER]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Krever ADMIN eller REGIONAL_MANAGER")

    try:
        from app.models.financial_models import GLTransaction
        import uuid as uuid_lib

        try:
            pid_uuid = uuid_lib.UUID(property_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Ugyldig property_id format")

        stmt = (
            select(
                GLTransaction.konto,
                GLTransaction.konto_navn,
                func.sum(GLTransaction.belop).label("belop"),
                func.count(GLTransaction.transaction_id).label("antall_transaksjoner"),
            )
            .where(
                GLTransaction.property_id == pid_uuid,
                GLTransaction.ar == year,
            )
        )
        if srs_kategori:
            stmt = stmt.where(GLTransaction.srs_kategori == srs_kategori)
        # Netto per konto — HAVING SUM > 0 ekskluderer kontoer der reverseringer overstiger original
        stmt = stmt.group_by(GLTransaction.konto, GLTransaction.konto_navn).having(func.sum(GLTransaction.belop) > 0).order_by(desc(func.sum(GLTransaction.belop)))

        rows = (await db.execute(stmt)).fetchall()
        return [
            {
                "konto": r.konto,
                "konto_navn": r.konto_navn,
                "belop": round(float(r.belop or 0), 0),
                "antall_transaksjoner": r.antall_transaksjoner,
            }
            for r in rows
        ]

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Konto drill-down feil: {str(e)}")


# ---------------------------------------------------------------------------
# Lønnskostnader
# ---------------------------------------------------------------------------

@router.get("/salary-costs/years", response_model=Dict[str, Any])
async def get_salary_cost_years(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Returnerer tilgjengelige år i salary_costs-tabellen."""
    if current_user.role not in [UserRole.ADMIN, UserRole.REGIONAL_MANAGER]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Krever ADMIN eller REGIONAL_MANAGER")

    try:
        result = await db.execute(
            text("SELECT DISTINCT year FROM salary_costs WHERE property_id IS NOT NULL ORDER BY year")
        )
        years = [r[0] for r in result.fetchall()]
    except Exception:
        years = []

    return {"available_years": years}


@router.get("/salary-costs", response_model=Dict[str, Any])
async def get_salary_costs(
    year: int = Query(2025, description="År for lønnskostnader"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Returnerer lønnskostnader for alle eiendommer for ett år.
    { year, total, by_property: [{property_id, property_name, faste_stillinger, vikarer, arbeidsgiveravgift, total}] }
    """
    if current_user.role not in [UserRole.ADMIN, UserRole.REGIONAL_MANAGER]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Krever ADMIN eller REGIONAL_MANAGER")

    try:
        rows = (await db.execute(text("""
            SELECT
                sc.property_id::text,
                p.name AS property_name,
                sc.faste_stillinger,
                sc.vikarer,
                sc.arbeidsgiveravgift
            FROM salary_costs sc
            LEFT JOIN properties p ON p.property_id = sc.property_id
            WHERE sc.year = :year
              AND sc.property_id IS NOT NULL
            ORDER BY p.name
        """), {"year": year})).fetchall()
    except Exception as e:
        logger.debug("salary_costs les feil: %s", e)
        return {"year": year, "total": 0, "by_property": []}

    by_property = []
    total = 0.0
    for r in rows:
        faste = float(r[2] or 0)
        vikarer = float(r[3] or 0)
        aga = float(r[4] or 0)
        row_total = faste + vikarer + aga
        total += row_total
        by_property.append({
            "property_id": r[0],
            "property_name": r[1] or "",
            "faste_stillinger": round(faste, 4),
            "vikarer": round(vikarer, 4),
            "arbeidsgiveravgift": round(aga, 4),
            "total": round(row_total, 4),
        })

    return {
        "year": year,
        "total": round(total, 4),
        "by_property": by_property,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Lønnsprediksjon 2027 – Holt-Winters
# ─────────────────────────────────────────────────────────────────────────────

async def _run_salary_prediction(year: int, history_from: int, inflation: float) -> None:
    """Background task som kjører lønnsprediksjon og lagrer til DB."""
    from app.db.session import SessionLocal
    from app.services.salary_prediction_service import SalaryPredictionService
    try:
        async with SessionLocal() as db:
            result = await SalaryPredictionService.predict_all_properties(
                db,
                target_year=year,
                history_from=history_from,
                inflation=inflation,
            )
            logger.debug("Lønnsprediksjon ferdig: %s", result)
    except Exception as exc:
        logger.debug("Lønnsprediksjon feilet: %s", exc)


@router.post("/salary-costs/predict", response_model=Dict[str, Any])
async def predict_salary_costs(
    background_tasks: BackgroundTasks,
    year: int = Query(default=2027, description="Målår for prediksjon"),
    history_from: int = Query(default=2020, description="Første år i historikken"),
    inflation: float = Query(default=0.045, description="Lønnsvekst-fallback (NHO/SSB)"),
    current_user: User = Depends(get_current_user),
):
    """
    Starter Holt-Winters lønnsprediksjon for alle eiendommer som bakgrunnsoppgave.
    Resultater skrives til salary_costs med year=target_year og import_batch_id='holt_winters_XXXX'.
    Krever ADMIN-rolle.
    """
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Kun ADMIN kan generere lønnsprediksjoner",
        )

    background_tasks.add_task(_run_salary_prediction, year, history_from, inflation)
    return {
        "status": "started",
        "year": year,
        "history_from": history_from,
        "inflation": inflation,
        "message": f"Lønnsprediksjon for {year} er startet. Hent resultater fra GET /salary-costs?year={year} om ~15 sek.",
    }


# ---------------------------------------------------------------------------
# Presentasjon — regional summary (faktisk regnskap og budsjett)
# ---------------------------------------------------------------------------

@router.get("/regional-summary", response_model=List[Dict[str, Any]])
async def get_regional_summary(
    year: int = Query(..., ge=2023, le=2030, description="Regnskapsår"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Returnerer faktiske kostnader per region for et år.
    Kilde: finance_budget med data_source='kontant_{year}'.
    Brukes av /admin/presentasjon.
    """
    data_source = f"kontant_{year}"
    try:
        rows = await db.execute(text("""
            SELECT
                COALESCE(p.region, km.region, 'Ukjent') AS region,
                SUM(fb.amount)                           AS total,
                COUNT(DISTINCT fb.property_id)
                    FILTER (WHERE fb.property_id IS NOT NULL
                              AND p.region IS NOT NULL) AS antall_eiendommer
            FROM finance_budget fb
            LEFT JOIN properties p
                ON p.property_id::text = fb.property_id::text
            LEFT JOIN koststed_mapping km
                ON km.koststed_kode = fb.koststed_kode
            WHERE fb.data_source = :ds
            GROUP BY COALESCE(p.region, km.region, 'Ukjent')
            ORDER BY total DESC
        """), {"ds": data_source})
        return [
            {
                "region": r.region,
                "total": float(r.total or 0),
                "antall_eiendommer": int(r.antall_eiendommer or 0),
            }
            for r in rows.all()
        ]
    except Exception as exc:
        logger.debug("regional-summary feilet: %s", exc)
        return []


@router.get("/budget-regional-summary", response_model=List[Dict[str, Any]])
async def get_budget_regional_summary(
    year: int = Query(..., ge=2024, le=2030, description="Budsjettår"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Returnerer budsjetterte kostnader per region for et år.
    Kilde: finance_budget med data_source='finance_dept_{year}'.
    Brukes av /admin/presentasjon.
    """
    data_source = f"finance_dept_{year}"
    try:
        rows = await db.execute(text("""
            SELECT
                COALESCE(p.region, km.region, 'Ukjent') AS region,
                SUM(fb.amount)                           AS total,
                COUNT(DISTINCT fb.property_id)
                    FILTER (WHERE fb.property_id IS NOT NULL
                              AND p.region IS NOT NULL) AS antall_eiendommer
            FROM finance_budget fb
            LEFT JOIN properties p
                ON p.property_id::text = fb.property_id::text
            LEFT JOIN koststed_mapping km
                ON km.koststed_kode = fb.koststed_kode
            WHERE fb.data_source = :ds
            GROUP BY COALESCE(p.region, km.region, 'Ukjent')
            ORDER BY total DESC
        """), {"ds": data_source})
        return [
            {
                "region": r.region,
                "total": float(r.total or 0),
                "antall_eiendommer": int(r.antall_eiendommer or 0),
            }
            for r in rows.all()
        ]
    except Exception as exc:
        logger.debug("budget-regional-summary feilet: %s", exc)
        return []


@router.get("/husleie-2026-regional", response_model=List[Dict[str, Any]])
async def get_husleie_2026_regional(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Returnerer KPI-justert husleie 2026 per region (Alternativ A, beregnet fra SSB KPI).
    Kilde: properties.husleie_2026 (satt av update_husleie_2026.py).
    Brukes av /admin/presentasjon og /financials.
    """
    try:
        rows = await db.execute(text("""
            SELECT
                COALESCE(region, 'Ukjent')   AS region,
                SUM(husleie_2026)            AS total_husleie_2026,
                COUNT(*)
                    FILTER (WHERE husleie_2026 IS NOT NULL) AS antall_eiendommer
            FROM properties
            WHERE husleie_2026 IS NOT NULL
            GROUP BY region
            ORDER BY total_husleie_2026 DESC
        """))
        return [
            {
                "region": r.region,
                "total": float(r.total_husleie_2026 or 0),
                "antall_eiendommer": int(r.antall_eiendommer or 0),
            }
            for r in rows.all()
        ]
    except Exception as exc:
        logger.debug("husleie-2026-regional feilet: %s", exc)
        return []


@router.get("/budsjett-sammenligning-regional", response_model=List[Dict[str, Any]])
async def get_budsjett_sammenligning_regional(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Returnerer autoriserte tall fra økonomiavdelingen per region.
    Kilde: befs_budsjett_sammenligning (importert fra budsjettt2026ver04).
    Kolonnene: regn_2025_ok (kontant 2025), befs_pred_2026, budsjett_2026_ok.
    """
    try:
        rows = await db.execute(text("""
            SELECT
                COALESCE(region, 'Ukjent') AS region,
                SUM(regn_2025_ok)          AS regn_2025,
                SUM(befs_pred_2026)        AS befs_2026,
                SUM(budsjett_2026_ok)      AS budsjett_2026,
                COUNT(*)                   AS antall_eiendommer
            FROM befs_budsjett_sammenligning
            GROUP BY region
            ORDER BY regn_2025 DESC NULLS LAST
        """))
        return [
            {
                "region": r.region,
                "regn_2025": float(r.regn_2025 or 0),
                "befs_2026": float(r.befs_2026 or 0),
                "budsjett_2026": float(r.budsjett_2026 or 0),
                "antall_eiendommer": int(r.antall_eiendommer or 0),
            }
            for r in rows.all()
        ]
    except Exception as exc:
        logger.error("budsjett-sammenligning-regional feilet: %s", exc)
        return []


@router.get("/budsjett-sammenligning-eiendommer", response_model=List[Dict[str, Any]])
async def get_budsjett_sammenligning_eiendommer(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Returnerer alle eiendommer fra økonomiavdelingens sammenligning.
    Kilde: befs_budsjett_sammenligning (importert fra budsjettt2026ver04).
    """
    try:
        rows = await db.execute(text("""
            SELECT
                eiendom,
                region,
                regn_2025_ok,
                befs_pred_2026,
                budsjett_2026_ok,
                merknad
            FROM befs_budsjett_sammenligning
            ORDER BY region NULLS LAST, eiendom
        """))
        return [
            {
                "eiendom": r.eiendom,
                "region": r.region or "Ukjent",
                "regn_2025": float(r.regn_2025_ok or 0),
                "befs_2026": float(r.befs_pred_2026 or 0),
                "budsjett_2026": float(r.budsjett_2026_ok or 0),
                "merknad": r.merknad,
            }
            for r in rows.all()
        ]
    except Exception as exc:
        logger.error("budsjett-sammenligning-eiendommer feilet: %s", exc)
        return []


@router.get("/konto-fordeling-per-eiendom", response_model=List[Dict[str, Any]])
async def get_konto_fordeling_per_eiendom(
    year: int = Query(2025, ge=2022, le=2026, description="Regnskapsår"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Konto-fordeling per eiendom fra GL-transaksjoner.
    Kun eiendommer som matcher økonomiavdelingens CSV (befs_budsjett_sammenligning).
    Returnerer: eiendom, region, konto, konto_navn, belop
    """
    try:
        rows = await db.execute(text("""
            SELECT
                p.name        AS eiendom,
                p.region      AS region,
                g.konto       AS konto,
                g.konto_navn  AS konto_navn,
                SUM(g.belop)  AS belop
            FROM gl_transactions g
            JOIN properties p ON g.property_id = p.property_id
            WHERE g.ar = :year
              AND p.property_id IN (
                  SELECT DISTINCT p2.property_id
                  FROM properties p2
                  JOIN befs_budsjett_sammenligning b ON
                      lower(p2.name) = lower(b.eiendom)
                      OR lower(regexp_replace(p2.name, '^\\d{4} - ', '')) = lower(b.eiendom)
              )
            GROUP BY p.name, p.region, g.konto, g.konto_navn
            HAVING SUM(g.belop) != 0
            ORDER BY p.region NULLS LAST, p.name, SUM(g.belop) DESC
        """), {"year": year})
        return [
            {
                "eiendom": r.eiendom,
                "region": r.region or "Ukjent",
                "konto": r.konto,
                "konto_navn": r.konto_navn,
                "belop": float(r.belop or 0),
            }
            for r in rows.all()
        ]
    except Exception as exc:
        logger.debug("konto-fordeling-per-eiendom feilet: %s", exc)
        return []


@router.get("/befs-synthetic-budget-regional", response_model=List[Dict[str, Any]])
async def get_befs_synthetic_budget_regional(
    year: int = Query(2026, ge=2024, le=2030, description="Budsjettår"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Returnerer BEFS eget syntetisk budsjett per region.
    Kilde: budget-tabellen med data_source='okonomi_regional_{year}'.
    Brukes av /admin/presentasjon for sammenligning mot Økonomiavd. vedtatt budsjett.
    """
    data_source = f"okonomi_regional_{year}"
    try:
        rows = await db.execute(text("""
            SELECT
                COALESCE(p.region, 'Ukjent') AS region,
                SUM(b.amount)                AS total,
                COUNT(DISTINCT b.property_id)
                    FILTER (WHERE b.property_id IS NOT NULL) AS antall_eiendommer
            FROM budget b
            LEFT JOIN properties p
                ON p.property_id::text = b.property_id::text
            WHERE b.data_source = :ds
              AND b.year = :yr
            GROUP BY COALESCE(p.region, 'Ukjent')
            ORDER BY total DESC
        """), {"ds": data_source, "yr": year})
        return [
            {
                "region": r.region,
                "total": float(r.total or 0),
                "antall_eiendommer": int(r.antall_eiendommer or 0),
            }
            for r in rows.all()
        ]
    except Exception as exc:
        logger.debug("befs-synthetic-budget-regional feilet: %s", exc)
        return []
