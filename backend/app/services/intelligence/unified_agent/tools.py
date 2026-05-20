"""
BEFS tools for Unified Agent - wrappers around ki_kollega_service methods.

Tools are created with db and service via closure so they have access to database session.
Uses query_normalizer for robust understanding (fvk -> familievernkontor, etc.).
"""

import json
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

from langchain_core.tools import tool, StructuredTool
from pydantic import create_model
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession
from app.services.mcp.handler import mcp_handler

from app.services.intelligence.ki_kollega.query_normalizer import (
    expand_query_terms,
    get_search_terms_for_property_lookup,
    normalize_query,
)
from app.services.intelligence.ki_kollega.service import KIKollegaService
from app.services.variance_service import VarianceService
from app.domains.barnevern.services.prediction_service import run_simulation

logger = logging.getLogger(__name__)

MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "Mai", "Jun", "Jul", "Aug", "Sep", "Okt", "Nov", "Des"]

# backend/data – samme kilder som /api/v1/barnevern-docs/*
def _backend_data_dir() -> Path:
    return Path(__file__).resolve().parents[4] / "data"


def _format_barnevern_reference_context(section: str = "all", max_items: int = 12) -> str:
    """
    Leser ferdig genererte JSON-filer (ikke nettverkskall).
    section: all | stprp | annual | ssb
    """
    data_dir = _backend_data_dir()
    max_items = max(1, min(int(max_items or 12), 40))
    parts: List[str] = []

    def load(path: Path) -> Dict[str, Any]:
        if not path.exists():
            return {}
        try:
            with open(path, encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            logger.warning("Kunne ikke lese %s: %s", path.name, e)
            return {}

    if section in ("all", "stprp"):
        stprp = load(data_dir / "stprp_bufdir.json")
        items = stprp.get("items") or []
        lines = [
            "=== ST.PRP. / PROP (Bufdir/Bufetat-relevant, Stortingets data) ===",
            f"Generert: {stprp.get('generated_at', '?')}  Kilde: {stprp.get('source', '?')}",
        ]
        for it in items[:max_items]:
            ref = it.get("reference") or ""
            title = (it.get("title") or it.get("short_title") or "")[:200]
            pu = it.get("prop_url") or ""
            su = it.get("storting_sak_url") or ""
            terms = ", ".join(it.get("match_terms") or [])[:120]
            lines.append(
                f"- {ref} | {title}\n  Treff: {terms}\n  Prop: {pu}\n  Sak: {su}"
            )
        if not items:
            lines.append("(ingen data – kjør fetch_stprp_bufdir.py og deploy fil)")
        parts.append("\n".join(lines))

    if section in ("all", "annual"):
        annual = load(data_dir / "bufdir_arsrapporter.json")
        items = annual.get("items") or []
        lines = [
            "=== BUFDIR ÅRSRAPPORT (lenker, siste år) ===",
            f"Generert: {annual.get('generated_at', '?')}  Kilde: {annual.get('source', '?')}",
        ]
        for it in items[:max_items]:
            y = it.get("year")
            title = it.get("title") or ""
            pdf = it.get("pdf_url") or ""
            st = it.get("status") or ""
            page = it.get("page_url") or ""
            lines.append(f"- {y}: {title}  [{st}]\n  Side: {page}\n  PDF: {pdf or '—'}")
        if not items:
            lines.append("(ingen data – kjør fetch_bufdir_annual_reports.py)")
        parts.append("\n".join(lines))

    if section in ("all", "ssb"):
        ssb = load(data_dir / "ssb_bufetat_bufdir_tables.json")
        items = ssb.get("items") or []
        lines = [
            "=== SSB TABELL-KORTLISTE (barnevern/fosterhjem/familievern) ===",
            f"Generert: {ssb.get('generated_at', '?')}  Kilde: {ssb.get('source', '?')}",
            f"Søk brukt ved bygging: {', '.join(ssb.get('queries') or [])}",
        ]
        for it in items[:max_items]:
            tid = it.get("id") or ""
            lab = (it.get("label") or "")[:160]
            du = it.get("dataUrl") or ""
            mu = it.get("metadataUrl") or ""
            lines.append(f"- Tabell {tid}: {lab}\n  Data: {du}\n  Metadata: {mu}")
        if not items:
            lines.append("(ingen data – kjør ssb_bufdir_relevant_tables.py)")
        parts.append("\n".join(lines))

    if not parts:
        return "Ukjent section (bruk all, stprp, annual eller ssb)."
    return "\n\n".join(parts)


def _finans_dir() -> Path:
    """Repo-roten / finans (samme som GET /barnevern-docs/prediction-excel)."""
    return Path(__file__).resolve().parents[5] / "finans"


def _list_prediksjon_excel_paths(kind: str = "okonomi") -> List[Path]:
    d = _finans_dir()
    if not d.is_dir():
        return []
    paths: List[Path] = []
    k = (kind or "okonomi").lower()
    if k in ("okonomi", "økonomi", "any", "all"):
        paths.extend(d.glob("Prediksjon_*_Økonomi.xlsx"))
    if k in ("lonn", "lønn", "any", "all"):
        paths.extend(d.glob("Prediksjon_*_Lønn.xlsx"))
    paths.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return paths[:20]


def _safe_prediksjon_path(filename: str) -> Optional[Path]:
    base = Path(filename).name
    if ".." in base or "/" in filename.replace(base, ""):
        return None
    if not base.startswith("Prediksjon_") or not base.endswith(".xlsx"):
        return None
    p = _finans_dir() / base
    return p if p.is_file() else None


def _read_prediksjon_excel_preview(
    path: Path,
    max_rows_per_sheet: int = 28,
    max_cols: int = 14,
    max_sheets: int = 14,
    max_total_chars: int = 14000,
) -> str:
    """Tekstlig forhåndsvisning av Excel (for KI-kontekst, ikke full eksport)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return "openpyxl mangler – kan ikke lese Excel."

    if not path.exists():
        return f"Fil finnes ikke: {path.name}"

    size = path.stat().st_size
    if size > 30 * 1024 * 1024:
        return (
            f"Fil {path.name} er for stor ({size / (1024 * 1024):.1f} MB) for automatisk uttrekk. "
            "Bruk nedlasting fra Barnevern-dokumenter eller spør om konkrete tall via SQL."
        )

    lines: List[str] = [
        f"=== Prediksjon-Excel: {path.name} ({size / 1024:.0f} KB) ===",
        f"Mapper: Antagelser, regionark og nøkkeltall ligger typisk i egne ark (se oversikt nedenfor).",
    ]
    total = 0
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return f"Kunne ikke åpne {path.name}: {e}"

    try:
        lines.append(f"Ark: {', '.join(wb.sheetnames)}")
        for sheet_name in wb.sheetnames[:max_sheets]:
            ws = wb[sheet_name]
            chunk_lines = [f"--- Ark: {sheet_name} (første {max_rows_per_sheet} rader, {max_cols} kolonner) ---"]
            for row in ws.iter_rows(
                min_row=1,
                max_row=max_rows_per_sheet,
                min_col=1,
                max_col=max_cols,
                values_only=True,
            ):
                cells: List[str] = []
                for c in row:
                    if c is None:
                        cells.append("")
                    elif isinstance(c, float):
                        cells.append(f"{c:.6g}" if abs(c) < 1e15 else str(c))
                    else:
                        s = str(c).replace("\n", " ").strip()
                        if len(s) > 90:
                            s = s[:87] + "…"
                        cells.append(s)
                chunk_lines.append(" | ".join(cells))
            block = "\n".join(chunk_lines)
            if total + len(block) > max_total_chars:
                lines.append("(… avkortet: flere ark finnes – spør mer spesifikt om ett ark eller bruk SQL for DB-prediksjon.)")
                break
            lines.append(block)
            total += len(block)
    finally:
        wb.close()

    return "\n".join(lines)


def create_befs_tools(db: AsyncSession, service: KIKollegaService, user=None) -> list:
    """
    Create LangChain tools that wrap ki_kollega_service methods.
    Uses closure to capture db, service and user (for RBAC in combine_ssb_befs).
    """

    @tool
    async def run_sql_query(question: str) -> str:
        """Kjør database-analyse mot PostgreSQL. Bruk dette for:
- Husleie/kostnader: 'hvilke eiendommer har husleie over X', 'total husleie 2025', 'kostnader per eiendom' → bruk gl_transactions med account_name
- Statistikk og telling: antall eiendommer, kontrakter, størst/minst, gjennomsnitt
- Kontraktsinformasjon: utløpende kontrakter, aktive kontrakter, leiebeløp per år
- Budsjett vs. faktisk: sammenlign budget med gl_transactions
- Avvik/HMS-saker: åpne avvik, kritiske saker
Viktig: For regnskap/kostnader 2025 brukes tabellen gl_transactions (ikke contracts). Kolonnen er account_name (ikke category).
RETRY: Hvis resultatet inneholder FEIL: eller feilmelding (f.eks. column X does not exist), reformuler spørsmålet og kall run_sql_query igjen."""
        try:
            normalized_q = expand_query_terms(normalize_query(question)) or question
            r = await service._tool_run_sql(db, normalized_q)
            if isinstance(r, dict) and "formatted" in r:
                return r
            return {"formatted": str(r), "structured_sources": []}
        except Exception as e:
            logger.error(f"run_sql_query failed: {e}")
            return {"formatted": f"Feil ved database-analyse: {str(e)}", "structured_sources": []}

    @tool
    async def lookup_properties(search_term: str):
        """Søk etter eiendommer ved navn, adresse eller bruk (f.eks. familievernkontor, fvk). Bruk for 'alle X', 'hvilke eiendommer er X', 'finn eiendom Y'."""
        try:
            terms = get_search_terms_for_property_lookup(normalize_query(search_term))
            term_to_use = terms[0] if terms else normalize_query(search_term) or search_term
            result = await service._tool_lookup_properties(db, term_to_use)
            if isinstance(result, dict):
                return result  # Full dict with formatted + structured_sources for custom ToolNode
            return {"formatted": str(result), "structured_sources": []}
        except Exception as e:
            logger.error(f"lookup_properties failed: {e}")
            return {"formatted": f"Feil ved eiendomssøk: {str(e)}", "structured_sources": []}

    @tool
    async def lookup_parties(search_term: str):
        """Søk etter parter (leietakere/leverandører) ved navn eller orgnr. Bruk for 'har vi kontrakt med X', 'leietaker Y', 'part Z'."""
        try:
            normalized = expand_query_terms(normalize_query(search_term)) or search_term.strip()
            result = await service._tool_lookup_parties(db, normalized)
            if isinstance(result, dict):
                return result  # Full dict with formatted + structured_sources for custom ToolNode
            return {"formatted": str(result), "structured_sources": []}
        except Exception as e:
            logger.error(f"lookup_parties failed: {e}")
            return {"formatted": f"Feil ved partsøk: {str(e)}", "structured_sources": []}

    @tool
    async def lookup_familievernkontor_bufdir(search_term: str):
        """Nasjonal oversikt fra Bufdir.no om familievernkontor (offisielt navn, telefon, e-post, region). Bruk når brukeren spør om familievernkontor/familievern utover eller i tillegg til BEFS-eiendommer, eller om kontaktinfo fra Bufdir."""
        try:
            return service._tool_lookup_familievernkontor_bufdir(search_term)
        except Exception as e:
            logger.error(f"lookup_familievernkontor_bufdir failed: {e}")
            return {"formatted": f"Feil ved Bufdir-oppslag: {str(e)}", "structured_sources": []}

    @tool
    async def search_documents(query: str) -> str:
        """Søk i dokumenter om rutiner, krav, instrukser. Bruk for 'hvordan'-spørsmål, HMS, prosedyrer."""
        try:
            return await service._tool_search_documents(db, query)
        except Exception as e:
            logger.error(f"search_documents failed: {e}")
            return f"Feil ved dokumentsøk: {str(e)}"

    @tool
    async def search_lovdata(query: str, limit: int = 5):
        """Søk i lover og forskrifter fra Lovdata. Bruk for juridiske spørsmål om husleie, HMS, kontraktsrett."""
        try:
            result = await service._tool_search_lovdata(query, limit=limit)
            if isinstance(result, dict):
                return result  # Full dict with formatted + structured_sources for custom ToolNode
            return {"formatted": str(result), "structured_sources": []}
        except Exception as e:
            logger.error(f"search_lovdata failed: {e}")
            return {"formatted": f"Feil ved Lovdata-søk: {str(e)}", "structured_sources": []}

    @tool
    async def assess_property_risk(property_id: str, risk_types: List[str] = None) -> str:
        """Vurder risiko for eiendom (flomfare, grunnforhold, miljø). Bruk for spørsmål om risikonivå for en spesifikk eiendom. property_id kan være UUID, navn eller adresse."""
        try:
            return await service._tool_assess_property_risk(db, property_id, risk_types)
        except Exception as e:
            logger.error(f"assess_property_risk failed: {e}")
            return f"Feil ved risikovurdering: {str(e)}"

    @tool
    async def create_jira_issue(summary: str, description: str, project_key: str = "KAN", issue_type: str = "Task") -> str:
        """Opprett en ny Jira-sak (to-do/oppgave). Bruk dette når brukeren vil registrere noe, lage en oppgave/sak eller huskeliste."""
        try:
            return await service._tool_create_jira_issue(summary, description, project_key, issue_type)
        except Exception as e:
            logger.error(f"create_jira_issue failed: {e}")
            return f"Feil ved opprettelse av Jira-sak: {str(e)}"

    @tool
    async def fetch_ssb_statistics(query: str, table_id: str = "", value_codes: dict = None) -> str:
        """Hent offisiell statistikk fra SSB (KPI, konsumprisindeks, boligpriser, befolkning). Bruk for spørsmål om nasjonal statistikk, inflasjon, eller for å sammenligne våre tall mot markedet."""
        try:
            result = await service._tool_fetch_ssb_statistics(
                query=query,
                table_id=table_id or None,
                value_codes=value_codes,
            )
            if isinstance(result, dict) and "formatted" in result:
                return result
            return {"formatted": str(result), "structured_sources": []}
        except Exception as e:
            logger.error(f"fetch_ssb_statistics failed: {e}")
            return {
                "formatted": f"Feil ved henting av SSB-statistikk: {str(e)}",
                "structured_sources": [],
            }

    @tool
    async def combine_ssb_befs_data(
        table_id: str,
        befs_dataset: str,
        join_key: str,
        year: int = 2025,
    ) -> str:
        """Kombiner SSB-statistikk med BEFS interne data. Bruk for 'sammenlign våre kostnader med KPI', 'kostnadsvekst vs inflasjon'. befs_dataset: region_costs, properties eller contracts. join_key: region, kommune eller year."""
        try:
            result = await service._tool_combine_ssb_befs_data(
                db=db,
                table_id=table_id,
                befs_dataset=befs_dataset,
                join_key=join_key,
                year=year,
                user=user,
            )
            if isinstance(result, dict) and "formatted" in result:
                return result
            return {"formatted": str(result), "structured_sources": []}
        except Exception as e:
            logger.error(f"combine_ssb_befs_data failed: {e}")
            return {
                "formatted": f"Feil ved kombinasjon av SSB og BEFS: {str(e)}",
                "structured_sources": [],
            }

    @tool
    async def get_leie_gap_analyse(year: int = 2025) -> str:
        """Analyser leie-gap: sammenlign kontraktsfestet husleie (contracts) mot GL-bokført husleie (gl_transactions) per eiendom.
Bruk for spørsmål om 'leie-gap', 'husleiedifferanse', 'kontraktsfestet vs GL-bokført husleie', 'hvorfor er det stor forskjell i husleie'.
Returnerer topp 20 eiendommer med størst avvik, sortert etter gap-beløp."""
        try:
            result = await db.execute(text("""
                WITH gl_rent AS (
                    SELECT gt.property_id, SUM(gt.amount) AS gl_rent
                    FROM gl_transactions gt
                    WHERE gt.year = :year AND gt.amount > 0
                      AND gt.account_name ILIKE '%leie%'
                    GROUP BY gt.property_id
                ),
                contracted AS (
                    SELECT u.property_id,
                           SUM((c.amount->>'amount_per_year')::float) AS contracted_rent
                    FROM contracts c
                    JOIN units u ON c.unit_id = u.unit_id
                    WHERE c.status = 'active'
                      AND c.amount IS NOT NULL
                      AND c.amount->>'amount_per_year' IS NOT NULL
                    GROUP BY u.property_id
                )
                SELECT p.name, p.address,
                       COALESCE(co.contracted_rent, 0) AS contracted_rent,
                       COALESCE(gr.gl_rent, 0) AS gl_rent,
                       COALESCE(co.contracted_rent, 0) - COALESCE(gr.gl_rent, 0) AS gap
                FROM properties p
                LEFT JOIN gl_rent gr ON gr.property_id = p.property_id
                LEFT JOIN contracted co ON co.property_id = p.property_id
                WHERE gr.property_id IS NOT NULL OR co.property_id IS NOT NULL
                ORDER BY ABS(COALESCE(co.contracted_rent, 0) - COALESCE(gr.gl_rent, 0)) DESC
                LIMIT 20
            """), {"year": year})
            rows = result.fetchall()
            if not rows:
                return f"Ingen leie-gap data funnet for {year}."
            lines = [f"LEIE-GAP ANALYSE {year} (topp 20 eiendommer etter avvik):"]
            total_contracted = 0
            total_gl = 0
            for r in rows:
                name = (r.name or r.address or "Ukjent")[:38]
                contracted = r.contracted_rent or 0
                gl_rent = r.gl_rent or 0
                gap = r.gap or 0
                total_contracted += contracted
                total_gl += gl_rent
                gap_pct = (gap / contracted * 100) if contracted > 0 else 0
                lines.append(
                    f"  {name:<38} kontraktsfestet={contracted:>14,.0f} NOK  GL-bokført={gl_rent:>14,.0f} NOK  gap={gap:>14,.0f} NOK ({gap_pct:.1f}%)"
                )
            total_gap = total_contracted - total_gl
            lines.append(f"\nSUM (topp 20): kontraktsfestet={total_contracted:,.0f} NOK, GL-bokført={total_gl:,.0f} NOK, total gap={total_gap:,.0f} NOK")
            return "\n".join(lines)
        except Exception as e:
            logger.error(f"get_leie_gap_analyse failed: {e}")
            return f"Feil ved leie-gap analyse: {str(e)}"

    @tool
    async def get_yoy_cost_analysis() -> str:
        """Analyser kostnadsvekst år-over-år: sammenlign GL-kostnader 2024 mot 2025 per kostnadskategori (account_name).
Bruk for 'kostnadsvekst', 'år-over-år endring', 'hva har økt i kostnad', 'sammenlign 2024 og 2025'.
Returnerer alle kategorier sortert etter absolutt endring (størst endring øverst)."""
        try:
            result = await db.execute(text("""
                SELECT
                    account_name,
                    SUM(CASE WHEN year = 2024 THEN amount ELSE 0 END) AS amt_2024,
                    SUM(CASE WHEN year = 2025 THEN amount ELSE 0 END) AS amt_2025,
                    SUM(CASE WHEN year = 2025 THEN amount ELSE 0 END)
                        - SUM(CASE WHEN year = 2024 THEN amount ELSE 0 END) AS change_nok
                FROM gl_transactions
                WHERE year IN (2024, 2025) AND amount > 0
                  AND account_name IS NOT NULL AND account_name != ''
                GROUP BY account_name
                ORDER BY ABS(
                    SUM(CASE WHEN year = 2025 THEN amount ELSE 0 END)
                    - SUM(CASE WHEN year = 2024 THEN amount ELSE 0 END)
                ) DESC
            """))
            rows = result.fetchall()
            if not rows:
                return "Ingen GL-data funnet for 2024/2025."
            lines = ["KOSTNADSVEKST 2024 → 2025 PER KATEGORI (sortert etter absolutt endring):"]
            total_2024 = 0
            total_2025 = 0
            for r in rows:
                cat = (r.account_name or "Ukjent")[:45]
                a2024 = r.amt_2024 or 0
                a2025 = r.amt_2025 or 0
                chg = r.change_nok or 0
                chg_pct = (chg / a2024 * 100) if a2024 > 0 else 0
                total_2024 += a2024
                total_2025 += a2025
                sign = "+" if chg >= 0 else ""
                lines.append(
                    f"  {cat:<45} 2024={a2024:>14,.0f}  2025={a2025:>14,.0f}  endring={sign}{chg:>14,.0f} ({sign}{chg_pct:.1f}%)"
                )
            total_chg = total_2025 - total_2024
            total_chg_pct = (total_chg / total_2024 * 100) if total_2024 > 0 else 0
            sign = "+" if total_chg >= 0 else ""
            lines.append(f"\nTOTAL: 2024={total_2024:,.0f} NOK, 2025={total_2025:,.0f} NOK, endring={sign}{total_chg:,.0f} NOK ({sign}{total_chg_pct:.1f}%)")
            return "\n".join(lines)
        except Exception as e:
            logger.error(f"get_yoy_cost_analysis failed: {e}")
            return f"Feil ved år-over-år analyse: {str(e)}"

    @tool
    async def get_gl_property_data(property_name: str, year: int = 2025) -> str:
        """Hent GL-regnskapsdata (faktiske kostnader) for en spesifikk eiendom fra gl_transactions.
Bruk for 'hva koster [eiendom]', 'regnskapet for [eiendom]', 'faktiske kostnader [eiendom] [år]'.
property_name kan være delvis navn eller adresse (f.eks. 'Tærudgata', 'Oslo sentrum').
Returnerer totalkostnader per kostnadskategori sortert etter beløp."""
        try:
            result = await db.execute(text("""
                SELECT p.name, p.address, gt.account_name,
                       SUM(gt.amount) AS total
                FROM gl_transactions gt
                JOIN properties p ON gt.property_id = p.property_id
                WHERE gt.year = :year
                  AND gt.amount > 0
                  AND (p.name ILIKE :q OR p.address ILIKE :q)
                GROUP BY p.name, p.address, gt.account_name
                ORDER BY total DESC
            """), {"year": year, "q": f"%{property_name}%"})
            rows = result.fetchall()
            if not rows:
                return f"Ingen GL-data funnet for eiendom «{property_name}» i {year}."
            prop_display = f"{rows[0].name or rows[0].address}"
            lines = [f"GL-REGNSKAPSDATA FOR {prop_display.upper()} ({year}):"]
            grand_total = 0
            for r in rows:
                cat = (r.account_name or "Ukjent")[:45]
                total = r.total or 0
                grand_total += total
                lines.append(f"  {cat:<45} {total:>14,.0f} NOK")
            lines.append(f"\nTOTAL: {grand_total:,.0f} NOK")
            return "\n".join(lines)
        except Exception as e:
            logger.error(f"get_gl_property_data failed: {e}")
            return f"Feil ved GL-oppslag for eiendom: {str(e)}"

    @tool
    async def get_budget_variance_report(property_name: str, year: int = 2025, period_type: str = "year") -> str:
        """Budsjett vs faktisk for en eiendom. Bruk for 'avvik for [eiendom]', 'budsjett vs faktisk [eiendom]', 'variansrapport [eiendom]'.
property_name: delvis navn eller adresse (f.eks. 'Tærudgata', 'Helsfyr').
period_type: 'year' (hele året), 'month' (krever period_value 1-12), 'quarter' (1-4), 'ytd' (1-12).
Returnerer budsjett, faktisk og varians per kategori."""
        try:
            from uuid import UUID
            # Resolve property_id from name
            prop_id = None
            if len(property_name) >= 36 and property_name.replace("-", "").replace(" ", "").isalnum():
                try:
                    prop_id = str(UUID(property_name))
                except ValueError:
                    pass
            if not prop_id:
                res = await db.execute(
                    text("SELECT property_id, name FROM properties WHERE name ILIKE :q OR address ILIKE :q LIMIT 1"),
                    {"q": f"%{property_name}%"},
                )
                row = res.fetchone()
                if not row:
                    return f"Ingen eiendom funnet for «{property_name}»."
                prop_id = str(row.property_id)
            report = await VarianceService.get_variance_report(db, prop_id, year, period_type, None)
            s = report.get("summary", {})
            lines = [
                f"BUDSJETT VS FAKTISK – {report.get('property_id', '')} ({year}):",
                f"  Budsjett: {s.get('total_budget', 0):,.0f} NOK",
                f"  Faktisk:  {s.get('total_actual', 0):,.0f} NOK",
                f"  Varians:  {s.get('total_variance', 0):,.0f} NOK ({s.get('total_variance_pct', 0):+.1f}%)",
                "",
                "Per kategori:",
            ]
            for it in report.get("items", [])[:15]:
                lines.append(
                    f"  {it['category']:<30} budsjett={it['budget']:>12,.0f}  faktisk={it['actual']:>12,.0f}  varians={it['variance']:>12,.0f} ({it['variance_pct']:+.1f}%)"
                )
            return "\n".join(lines)
        except Exception as e:
            err = str(e)
            if "mangler budsjettdata" in err.lower() or "400" in err:
                return f"Eiendommen mangler budsjettdata. Kjør budsjettgenerering fra Admin → Finans."
            logger.error(f"get_budget_variance_report failed: {e}")
            return f"Feil ved budsjett-varians: {err}"

    @tool
    async def barnevern_cost_simulation(year: int = 2026, usage_pct: float = 0.85) -> str:
        """Simuler kostnad for brukte og ubrukte barnevernsplasser.
Bruk for spørsmål om 'barnevern kostnad', 'kostnadssimulering barnevern', 'brukte/ubrukte plasser', 'egenandel barnevern'.
year: 2024–2030. usage_pct: bruksgrad 0–1 (f.eks. 0.85 = 85%). Returnerer total kostnad per region og brukte/ubrukte plasser."""
        try:
            result = await run_simulation(db=db, year=year, usage_pct=usage_pct, include_ssb=True)
            lines = [
                f"BARNEVERN KOSTNADSSIMULERING {year} (bruksgrad {(usage_pct * 100):.0f}%):",
                f"  Egenandel: {result.egenandel_maaned:,.0f} NOK/mnd ({result.egenandel_aar:,.0f} NOK/år)",
                f"  Godkjente plasser totalt: {result.total_approved_places}",
                f"  Brukte plasser: {result.total_brukte}",
                f"  Ubrukte plasser: {result.total_ubrukte}",
                f"  Kostnad brukte: {result.total_kost_brukte:,.0f} NOK",
                f"  Kostnad ubrukte: {result.total_kost_ubrukte:,.0f} NOK",
                f"  Total kostnad: {result.total_kostnad:,.0f} NOK",
                "",
                "Per region:",
            ]
            for r in result.by_region:
                lines.append(
                    f"  {r.region:<12} brukte={r.brukte_plasser:>4} ubrukte={r.ubrukte_plasser:>4} kostnad={r.total_kostnad:>14,.0f} NOK"
                )
            return "\n".join(lines)
        except Exception as e:
            logger.error(f"barnevern_cost_simulation failed: {e}")
            return f"Feil ved barnevern-kostnadssimulering: {str(e)}"

    @tool
    async def get_monthly_budget_actual(year: int = 2025) -> str:
        """Månedlig budsjett vs faktisk for hele porteføljen. Bruk for 'budsjett vs faktisk per måned', 'månedlig avvik 2025', 'rapport budsjett faktisk'.
Returnerer 12 måneder med budsjett, faktisk og varians."""
        try:
            budget_res = await db.execute(
                text("SELECT month, SUM(amount) AS budget FROM budget WHERE year = :y GROUP BY month"),
                {"y": year},
            )
            budget_by_month = {r.month: float(r.budget) for r in budget_res.fetchall()}
            actual_res = await db.execute(
                text("SELECT month, SUM(amount) AS actual FROM gl_transactions WHERE year = :y AND amount > 0 GROUP BY month"),
                {"y": year},
            )
            actual_by_month = {r.month: float(r.actual) for r in actual_res.fetchall()}
            lines = [f"MÅNEDLIG BUDSJETT VS FAKTISK {year} (hele porteføljen):"]
            total_b = total_a = 0
            for m in range(1, 13):
                b = budget_by_month.get(m, 0.0)
                a = actual_by_month.get(m, 0.0)
                var = b - a
                var_pct = round((var / b * 100), 1) if b > 0 else 0
                total_b += b
                total_a += a
                lines.append(
                    f"  {MONTH_NAMES[m-1]:>3}  budsjett={b:>14,.0f}  faktisk={a:>14,.0f}  varians={var:>14,.0f} ({var_pct:+.1f}%)"
                )
            tot_var = total_b - total_a
            tot_pct = round((tot_var / total_b * 100), 1) if total_b > 0 else 0
            lines.append(f"\nTOTAL: budsjett={total_b:,.0f} NOK, faktisk={total_a:,.0f} NOK, varians={tot_var:,.0f} NOK ({tot_pct:+.1f}%)")
            return "\n".join(lines)
        except Exception as e:
            logger.error(f"get_monthly_budget_actual failed: {e}")
            return f"Feil ved månedlig budsjett-faktisk: {str(e)}"

    @tool
    async def get_barnevern_reference_context(
        section: str = "all",
        max_items: int = 12,
    ) -> str:
        """Hent Bufdir/Storting/SSB referansedata som BEFS lagrer lokalt (JSON).
        Bruk når brukeren spør om St.prp., proposisjoner, statsbudsjett-saker for BFD,
        Bufdir årsrapport (PDF/side), eller hvilke SSB-tabeller som er relevante for
        barnevern/fosterhjem/familievern. Dette er ikke fulltekst i proposisjoner, men
        tittel, referanse og lenker – kombiner gjerne med fetch_ssb_statistics for tall.
        section: all (alt) | stprp | annual | ssb."""
        try:
            sec = (section or "all").strip().lower()
            if sec not in ("all", "stprp", "annual", "ssb"):
                sec = "all"
            return _format_barnevern_reference_context(section=sec, max_items=max_items)
        except Exception as e:
            logger.error(f"get_barnevern_reference_context failed: {e}")
            return f"Feil ved henting av barnevern-referansekontekst: {str(e)}"

    @tool
    async def get_finans_prediksjon_excel_summary(
        filename: str = "",
        file_type: str = "okonomi",
        max_rows_per_sheet: int = 28,
    ) -> str:
        """Les sammendrag av prediksjon-Excel i mappen finans/ (Prediksjon_*_Økonomi.xlsx / Lønn).
        Inneholder typisk ark som Antagelser, regioner og nøkkeltall – egnet til å forklare
        forutsetninger og sammenligne med spørsmål om budsjett/prediksjon. Dette er ikke
        sanntids-DB; for faktiske 2025-tall og budsjett i databasen bruk run_sql_query.
        La filename være tom for nyeste fil av valgt type. file_type: okonomi | lonn."""
        try:
            ft = (file_type or "okonomi").lower()
            if filename and str(filename).strip():
                path = _safe_prediksjon_path(str(filename).strip())
                if not path:
                    return (
                        "Ugyldig eller ukjent filnavn. Bruk Prediksjon_<år>_Økonomi.xlsx eller "
                        "Prediksjon_<år>_Lønn.xlsx i finans/, eller la filename være tom."
                    )
            else:
                kind = "lonn" if ft in ("lonn", "lønn", "salary") else "okonomi"
                candidates = _list_prediksjon_excel_paths(kind)
                if not candidates:
                    return (
                        "Ingen Prediksjon_*_Økonomi.xlsx / Lønn.xlsx funnet i finans/. "
                        "Generer med finans/lag_prediksjon_excel.py eller lag_lonn_excel.py (se finans/README.md)."
                    )
                path = candidates[0]
            mr = max(8, min(int(max_rows_per_sheet or 28), 60))
            return _read_prediksjon_excel_preview(path, max_rows_per_sheet=mr)
        except Exception as e:
            logger.error(f"get_finans_prediksjon_excel_summary failed: {e}")
            return f"Feil ved lesing av prediksjon-Excel: {str(e)}"

    # ─── FDVU / COMPLIANCE TOOLS ────────────────────────────────────────────────

    @tool
    async def get_fdvu_compliance_summary(property_name: str) -> str:
        """Hent FDVU compliance-oversikt for en spesifikk eiendom. Viser antall krav,
        compliance-rate, avvik (non_compliant), delvis oppfylt, ikke vurdert og forfalte
        revisjoner. Bruk dette for spørsmål som 'hva er compliance-status for X',
        'har [eiendom] avvik i FDVU', 'hvor mange krav er oppfylt på Y'.
        Krever eiendomsnavn eller del av navn (f.eks. 'Tærudgata', 'Østfold')."""
        try:
            # Finn property_id fra navn
            result = await db.execute(
                text("SELECT property_id, name, region, unit_type_derived FROM properties WHERE name ILIKE :q LIMIT 1"),
                {"q": f"%{property_name.strip()}%"},
            )
            row = result.fetchone()
            if not row:
                return f"Ingen eiendom funnet med navn «{property_name}». Prøv et annet søkeord."

            pid = str(row.property_id)
            name = row.name or property_name
            region = row.region or ""
            unit_type = row.unit_type_derived or ""

            # Hent compliance-data via SQL (unngå HTTP-kall til backend)
            res = await db.execute(text("""
                SELECT
                    COUNT(ra.assignment_id)                                          AS total,
                    SUM(CASE WHEN ca.status = 'compliant'      THEN 1 ELSE 0 END)   AS compliant,
                    SUM(CASE WHEN ca.status = 'non_compliant'  THEN 1 ELSE 0 END)   AS non_compliant,
                    SUM(CASE WHEN ca.status = 'partial'        THEN 1 ELSE 0 END)   AS partial,
                    SUM(CASE WHEN ca.status = 'not_applicable' THEN 1 ELSE 0 END)   AS not_applicable,
                    SUM(CASE WHEN ca.status IS NULL OR ca.status = 'not_assessed'
                             THEN 1 ELSE 0 END)                                      AS not_assessed,
                    SUM(CASE WHEN ca.next_review_date IS NOT NULL
                                  AND ca.next_review_date < CURRENT_DATE
                                  AND ca.status NOT IN ('not_applicable','compliant')
                             THEN 1 ELSE 0 END)                                      AS overdue
                FROM requirement_assignments ra
                LEFT JOIN compliance_assessments ca ON ca.assignment_id = ra.assignment_id
                WHERE ra.property_id = CAST(:pid AS uuid)
            """), {"pid": pid})
            s = res.fetchone()
            if not s or not s.total:
                return (
                    f"**{name}** ({region}) — FDVU ikke startet\n"
                    "Ingen krav-tildelinger er registrert for denne eiendommen ennå.\n"
                    f"Gå til /fdvu/{pid}/rapport for å starte vurdering."
                )

            total = int(s.total or 0)
            compliant = int(s.compliant or 0)
            non_compliant = int(s.non_compliant or 0)
            partial = int(s.partial or 0)
            not_applicable = int(s.not_applicable or 0)
            not_assessed = int(s.not_assessed or 0)
            overdue = int(s.overdue or 0)
            denom = total - not_applicable - not_assessed
            rate = round(compliant / denom * 100, 1) if denom > 0 else 0.0

            # Hent topp-avvik (non_compliant-krav med tittel)
            avvik_res = await db.execute(text("""
                SELECT r.code, r.title, r.regulation_set, r.severity_if_breached
                FROM requirement_assignments ra
                JOIN requirements r ON r.requirement_id = ra.requirement_id
                JOIN compliance_assessments ca ON ca.assignment_id = ra.assignment_id
                WHERE ra.property_id = CAST(:pid AS uuid)
                  AND ca.status = 'non_compliant'
                ORDER BY
                    CASE r.severity_if_breached WHEN 'critical' THEN 0 WHEN 'high' THEN 1 WHEN 'medium' THEN 2 ELSE 3 END
                LIMIT 8
            """), {"pid": pid})
            avvik_rows = avvik_res.fetchall()

            status_emoji = "✅" if rate >= 90 else "⚠️" if rate >= 60 else "🔴"
            lines = [
                f"**FDVU Compliance — {name}**",
                f"Region: {region} · Type: {unit_type}",
                f"",
                f"{status_emoji} **Compliance-rate: {rate}%** ({compliant}/{denom} vurderte krav oppfylt)",
                f"",
                f"| Status | Antall |",
                f"|--------|--------|",
                f"| ✅ Oppfylt | {compliant} |",
                f"| ❌ Avvik | {non_compliant} |",
                f"| △ Delvis | {partial} |",
                f"| — Ikke aktuelt | {not_applicable} |",
                f"| ○ Ikke vurdert | {not_assessed} |",
            ]
            if overdue:
                lines.append(f"| ⏰ Forfalt revisjon | {overdue} |")

            if avvik_rows:
                lines.append("")
                lines.append("**Registrerte avvik:**")
                for r in avvik_rows:
                    sev = r.severity_if_breached or "ukjent"
                    icon = "🔴" if sev == "critical" else "🟠" if sev == "high" else "🟡"
                    lines.append(f"- {icon} [{r.regulation_set}] {r.code} – {r.title}")

            lines.append("")
            lines.append(f"📋 Fullstendig rapport: /fdvu/{pid}/rapport")
            return "\n".join(lines)

        except Exception as e:
            logger.error("get_fdvu_compliance_summary failed: %s", e)
            return f"Feil ved henting av FDVU-status: {str(e)}"

    @tool
    async def get_fdvu_portfolio_status(region: str = "") -> str:
        """Hent FDVU-porteføljesummary for hele Bufetat eller én region.
        Viser compliance-rate, antall avvik, antall eiendommer vurdert.
        Bruk for spørsmål som 'hva er FDVU-status for porteføljen', 'compliance i Region Øst',
        'hvilke eiendommer har mest avvik', 'FDVU-oversikt Bufetat'.
        La region være tom for alle Bufetat, eller oppgi f.eks. 'Region Øst', 'Vest'."""
        try:
            region_filter = ""
            params: dict = {}
            if region and region.strip():
                region_filter = "AND p.region ILIKE :reg"
                params["reg"] = f"%{region.strip()}%"

            # Portfolio-nivå aggregering
            portfolio_res = await db.execute(text(f"""
                SELECT
                    COUNT(DISTINCT ra.property_id)                                   AS props_with_assignments,
                    COUNT(ra.assignment_id)                                           AS total,
                    SUM(CASE WHEN ca.status = 'compliant'      THEN 1 ELSE 0 END)   AS compliant,
                    SUM(CASE WHEN ca.status = 'non_compliant'  THEN 1 ELSE 0 END)   AS non_compliant,
                    SUM(CASE WHEN ca.status = 'partial'        THEN 1 ELSE 0 END)   AS partial,
                    SUM(CASE WHEN ca.status IS NULL OR ca.status = 'not_assessed' THEN 1 ELSE 0 END) AS not_assessed,
                    SUM(CASE WHEN ca.status = 'not_applicable' THEN 1 ELSE 0 END)   AS not_applicable,
                    SUM(CASE WHEN ca.next_review_date < CURRENT_DATE AND ca.status NOT IN ('not_applicable','compliant') THEN 1 ELSE 0 END) AS overdue
                FROM requirement_assignments ra
                JOIN properties p ON p.property_id = ra.property_id
                LEFT JOIN compliance_assessments ca ON ca.assignment_id = ra.assignment_id
                WHERE 1=1 {region_filter}
            """), params)
            ps = portfolio_res.fetchone()

            # Top eiendommer med mest avvik
            top_avvik_res = await db.execute(text(f"""
                SELECT p.name, p.region,
                       COUNT(CASE WHEN ca.status='non_compliant' THEN 1 END) AS avvik,
                       COUNT(CASE WHEN ca.status='partial' THEN 1 END) AS partial,
                       COUNT(ra.assignment_id) AS total
                FROM requirement_assignments ra
                JOIN properties p ON p.property_id = ra.property_id
                LEFT JOIN compliance_assessments ca ON ca.assignment_id = ra.assignment_id
                WHERE 1=1 {region_filter}
                GROUP BY p.property_id, p.name, p.region
                HAVING COUNT(CASE WHEN ca.status='non_compliant' THEN 1 END) > 0
                ORDER BY avvik DESC
                LIMIT 10
            """), params)
            top_avvik = top_avvik_res.fetchall()

            if not ps or not ps.total:
                header = f"Region {region}" if region else "Hele Bufetat"
                return f"Ingen FDVU-data funnet for {header}. Vurderingsarbeidet er ikke startet."

            total = int(ps.total or 0)
            compliant = int(ps.compliant or 0)
            non_compliant = int(ps.non_compliant or 0)
            partial = int(ps.partial or 0)
            not_assessed = int(ps.not_assessed or 0)
            not_applicable = int(ps.not_applicable or 0)
            overdue = int(ps.overdue or 0)
            props_n = int(ps.props_with_assignments or 0)
            denom = total - not_applicable - not_assessed
            rate = round(compliant / denom * 100, 1) if denom > 0 else 0.0

            header = f"Region {region.strip()}" if region else "Hele Bufetat"
            status_emoji = "✅" if rate >= 90 else "⚠️" if rate >= 60 else "🔴"
            lines = [
                f"**FDVU Porteføljestatus — {header}**",
                f"",
                f"{status_emoji} **Compliance-rate: {rate}%** ({compliant}/{denom} vurderte krav oppfylt)",
                f"{props_n} eiendommer med tildelinger · {total} totale kravtildelinger",
                f"",
                f"| Status | Antall |",
                f"|--------|--------|",
                f"| ✅ Oppfylt | {compliant} |",
                f"| ❌ Avvik | {non_compliant} |",
                f"| △ Delvis | {partial} |",
                f"| ○ Ikke vurdert | {not_assessed} |",
                f"| ⏰ Forfalt | {overdue} |",
            ]

            if top_avvik:
                lines.append("")
                lines.append("**Eiendommer med flest avvik:**")
                for r in top_avvik:
                    lines.append(f"- {r.name} ({r.region or '—'}): {r.avvik} avvik, {r.partial} delvis")

            lines.append("")
            lines.append("📋 Fullstendig rapport med filter: /fdvu/rapport")
            return "\n".join(lines)

        except Exception as e:
            logger.error("get_fdvu_portfolio_status failed: %s", e)
            return f"Feil ved henting av FDVU-porteføljestatus: {str(e)}"

    @tool
    async def generate_fdvu_rapport(property_name: str) -> str:
        """Generer en fullstendig FDVU tilsynsrapport i tekstformat for én eiendom.
        Rapporten inkluderer compliance-status per regelverk, alle avvik med alvorlighetsgrad,
        og anbefalinger. Bruk for spørsmål som 'lag tilsynsrapport for X', 'generer FDVU-rapport
        for Y', 'vis compliance-detaljer for Z', 'hva sier kravene for eiendom X'.
        Merk: For en utskriftsklar PDF-rapport, se /fdvu/[eiendom-id]/rapport i systemet."""
        try:
            # Finn eiendom
            result = await db.execute(
                text("SELECT property_id, name, address, city, region, unit_type_derived, approved_places FROM properties WHERE name ILIKE :q LIMIT 1"),
                {"q": f"%{property_name.strip()}%"},
            )
            prop = result.fetchone()
            if not prop:
                return f"Ingen eiendom funnet med navn «{property_name}»."

            pid = str(prop.property_id)

            # Hent alle assignments med requirement + assessment
            asgn_res = await db.execute(text("""
                SELECT
                    r.code, r.title, r.regulation_set, r.severity_if_breached,
                    COALESCE(ca.status, 'not_assessed') AS status,
                    ca.evidence_notes,
                    ca.next_review_date
                FROM requirement_assignments ra
                JOIN requirements r ON r.requirement_id = ra.requirement_id
                LEFT JOIN compliance_assessments ca ON ca.assignment_id = ra.assignment_id
                WHERE ra.property_id = CAST(:pid AS uuid)
                ORDER BY r.regulation_set, CASE r.severity_if_breached WHEN 'critical' THEN 0 WHEN 'high' THEN 1 WHEN 'medium' THEN 2 ELSE 3 END
            """), {"pid": pid})
            assignments = asgn_res.fetchall()

            from datetime import date as _date
            today = _date.today().strftime("%d.%m.%Y")

            lines = [
                f"# FDVU Tilsynsrapport — {prop.name}",
                f"**Adresse:** {prop.address or '—'}, {prop.city or ''}",
                f"**Region:** {prop.region or '—'} · **Type:** {prop.unit_type_derived or '—'}{f' · {prop.approved_places} plasser' if prop.approved_places else ''}",
                f"**Dato:** {today}",
                f"",
            ]

            if not assignments:
                lines.append("Ingen krav-tildelinger registrert for denne eiendommen.")
                return "\n".join(lines)

            # Grupper per regulation_set
            from collections import defaultdict
            grouped: dict = defaultdict(list)
            for a in assignments:
                grouped[a.regulation_set].append(a)

            total = len(assignments)
            compliant = sum(1 for a in assignments if a.status == "compliant")
            non_compliant = sum(1 for a in assignments if a.status == "non_compliant")
            partial = sum(1 for a in assignments if a.status == "partial")
            not_assessed = sum(1 for a in assignments if a.status in ("not_assessed", None))
            denom = total - sum(1 for a in assignments if a.status == "not_applicable") - not_assessed
            rate = round(compliant / denom * 100, 1) if denom > 0 else 0.0

            status_emoji = "✅" if rate >= 90 else "⚠️" if rate >= 60 else "🔴"
            lines += [
                f"## Sammendrag",
                f"{status_emoji} **Compliance-rate: {rate}%** | Oppfylt: {compliant} | Avvik: {non_compliant} | Delvis: {partial} | Ikke vurdert: {not_assessed}",
                f"",
            ]

            REG_LABELS = {
                "RKL6": "Risikoklasse 6 – Brann", "BVL": "Barnevernloven",
                "KVALITETSFORSKRIFTEN": "Kvalitetsforskriften", "TEK17": "TEK17",
                "HMS": "HMS / AML", "DRIFTSLEDELSE": "Driftsledelse",
                "ENOK": "Energieffektivisering", "UU": "Universell utforming",
                "SIKKERHET": "Sikkerhet", "MILJØ": "Miljø og farlige stoffer",
                "BYGG": "Bygg – NS3451", "INTERN": "Interne krav",
            }
            STATUS_LABEL = {
                "compliant": "✅ Oppfylt", "non_compliant": "❌ Avvik",
                "partial": "△ Delvis", "not_applicable": "— N/A", "not_assessed": "○ Ikke vurdert",
            }

            for reg, items in grouped.items():
                reg_label = REG_LABELS.get(reg, reg)
                reg_avvik = sum(1 for a in items if a.status == "non_compliant")
                reg_done = sum(1 for a in items if a.status not in ("not_assessed", None))
                lines.append(f"## {reg_label} ({reg_done}/{len(items)} vurdert{', ' + str(reg_avvik) + ' avvik' if reg_avvik else ''})")
                for a in items:
                    sev = a.severity_if_breached or ""
                    sev_txt = f" [{sev.upper()}]" if sev in ("critical", "high") else ""
                    status_txt = STATUS_LABEL.get(a.status, a.status)
                    lines.append(f"- **{a.code}**{sev_txt} {a.title} — {status_txt}")
                    if a.evidence_notes:
                        lines.append(f"  _Notater: {a.evidence_notes}_")
                lines.append("")

            # Anbefalinger
            critical_avvik = [a for a in assignments if a.status == "non_compliant" and a.severity_if_breached in ("critical", "high")]
            if critical_avvik:
                lines.append("## Prioriterte tiltak (kritisk/høy alvorlighet)")
                for a in critical_avvik:
                    lines.append(f"1. **{a.code}** – {a.title} → Krever umiddelbar oppfølging")

            lines.append("")
            lines.append(f"_Generer utskriftsklar PDF via /fdvu/{pid}/rapport_")
            return "\n".join(lines)

        except Exception as e:
            logger.error("generate_fdvu_rapport failed: %s", e)
            return f"Feil ved generering av FDVU-rapport: {str(e)}"

    @tool
    async def get_fdvu_kravkatalog(regulation_set: str = "") -> str:
        """Hent oversikt over FDVU-kravkatalogen (122 krav) – alle krav BEFS-eiendommer
        skal oppfylle. Viser krav gruppert per regelverk med kode, tittel og alvorlighet.
        Bruk for spørsmål som 'hvilke krav finnes for brann', 'list alle RKL6-krav',
        'hva er kravene innen HMS', 'hvilke ENOK-krav har vi'.
        regulation_set: RKL6 | BVL | KVALITETSFORSKRIFTEN | TEK17 | HMS | DRIFTSLEDELSE |
                        ENOK | UU | SIKKERHET | MILJØ | BYGG | INTERN (eller tom for alle)."""
        try:
            params: dict = {}
            where = ""
            if regulation_set and regulation_set.strip():
                where = "WHERE regulation_set = :reg"
                params["reg"] = regulation_set.strip().upper()

            res = await db.execute(text(f"""
                SELECT code, title, regulation_set, severity_if_breached, description, is_mandatory
                FROM requirements
                {where}
                ORDER BY regulation_set, CASE severity_if_breached WHEN 'critical' THEN 0 WHEN 'high' THEN 1 WHEN 'medium' THEN 2 ELSE 3 END
            """), params)
            rows = res.fetchall()

            if not rows:
                return f"Ingen krav funnet{f' for regelverk {regulation_set}' if regulation_set else ''}."

            # Grupper
            from collections import defaultdict
            grouped: dict = defaultdict(list)
            for r in rows:
                grouped[r.regulation_set].append(r)

            label_map = {
                "RKL6": "Risikoklasse 6 – Brann", "BVL": "Barnevernloven",
                "KVALITETSFORSKRIFTEN": "Kvalitetsforskriften", "TEK17": "TEK17 – Teknisk forskrift",
                "HMS": "HMS / Arbeidsmiljøloven", "DRIFTSLEDELSE": "Driftsledelse",
                "ENOK": "Energieffektivisering (ENOK)", "UU": "Universell utforming",
                "SIKKERHET": "Sikkerhet og personvern", "MILJØ": "Miljø og farlige stoffer",
                "BYGG": "Bygg – NS3451", "INTERN": "Interne krav",
            }
            sev_icon = {"critical": "🔴", "high": "🟠", "medium": "🟡", "low": "🟢"}

            lines = [f"**FDVU Kravkatalog** — {len(rows)} krav totalt", ""]
            for reg, items in grouped.items():
                lines.append(f"### {label_map.get(reg, reg)} ({len(items)} krav)")
                for r in items:
                    icon = sev_icon.get(r.severity_if_breached or "", "⚪")
                    mandatory = "" if r.is_mandatory else " *(valgfritt)*"
                    lines.append(f"- {icon} **{r.code}** — {r.title}{mandatory}")
                lines.append("")

            return "\n".join(lines)

        except Exception as e:
            logger.error("get_fdvu_kravkatalog failed: %s", e)
            return f"Feil ved henting av kravkatalog: {str(e)}"

    # --- Dynamic MCP Tool Integration ---

    # 1. Start with manual overrides (these take precedence)
    tools = [
        run_sql_query,
        lookup_properties,
        lookup_parties,
        lookup_familievernkontor_bufdir,
        search_documents,
        search_lovdata,
        assess_property_risk,
        create_jira_issue,
        fetch_ssb_statistics,
        combine_ssb_befs_data,
        get_barnevern_reference_context,
        get_finans_prediksjon_excel_summary,
        barnevern_cost_simulation,
        get_leie_gap_analyse,
        get_yoy_cost_analysis,
        get_gl_property_data,
        get_budget_variance_report,
        get_monthly_budget_actual,
        get_fdvu_compliance_summary,
        get_fdvu_portfolio_status,
        generate_fdvu_rapport,
        get_fdvu_kravkatalog,
    ]
    
    # 2. Add all tools from MCP Handler that are NOT already defined above
    existing_tool_names = {t.name for t in tools}
    
    for tool_def in mcp_handler.get_tools():
        if tool_def.name in existing_tool_names:
            continue
            
        # Create a dynamic Pydantic model for arguments
        fields = {}
        for param_name, param_schema in tool_def.parameters.get("properties", {}).items():
            param_type = str
            if param_schema.get("type") == "integer":
                param_type = int
            elif param_schema.get("type") == "number":
                param_type = float
            elif param_schema.get("type") == "boolean":
                param_type = bool
            elif param_schema.get("type") == "array":
                param_type = list
                
            # Check required
            is_required = param_name in tool_def.parameters.get("required", [])
            default = ... if is_required else None
            
            fields[param_name] = (param_type, default)
            
        args_schema = create_model(f"{tool_def.name}Args", **fields)

        async def _wrapper(**kwargs):
            # Inject db session if needed (MCPHandler manages session context internally via decorators mostly)
            # But execute_tool allows passing explicit db
            return await mcp_handler.execute_tool(tool_def.name, kwargs, db=db)
            
        dynamo_tool = StructuredTool.from_function(
            func=None,
            coroutine=_wrapper,
            name=tool_def.name,
            description=tool_def.description,
            args_schema=args_schema
        )
        
        tools.append(dynamo_tool)
        logger.info(f"Dynamically registered MCP tool: {tool_def.name}")

    return tools
