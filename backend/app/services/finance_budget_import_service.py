"""
Import av vedtatt budsjett fra økonomi-avdelingen (Bufdir-økonomi).

Leser fanen 'Budsjett 2025_2026' fra Excel-uttrekk og lagrer i
finance_budget-tabellen (IKKE budget-tabellen).

Idempotent: sletter alle rader med (year, data_source) før insert,
slik at re-import gir nøyaktig samme sluttilstand.

Beløpsstrategi:
  Bruker 'Beløp DA' (disponeringsvedtak) som primærkilde – dekker alle 12 måneder.
  'Kontantbeløp' er kun fylt for måneder som er passert (frem til uttrekksdato).
  For budsjettformål er Beløp DA korrekt kilde.

Kolonner i kildefila:
  Konto | Konto(T) | Koststed | Koststed(T) | Prosjekt | Prosjekt(T)
  | Finansiering | Finansiering(T) | Periode (YYYYMM) | Beløp DA | Kontantbeløp
"""
import logging
import uuid
from dataclasses import dataclass, field
from datetime import timezone, datetime
from typing import Optional

from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.financial_models import FINANCE_KONTO_CATEGORY

logger = logging.getLogger(__name__)

SHEET_NAME = "Budsjett 2025_2026"
KONTANT_SHEET_NAME = "Kontant 2025"

# Hvilke år vi importerer fra fila
IMPORT_YEARS = {2025, 2026}


@dataclass
class ImportReport:
    total_rows: int = 0
    inserted: int = 0
    skipped_no_periode: int = 0
    skipped_wrong_year: int = 0
    skipped_zero_amount: int = 0
    skipped_unknown_konto: int = 0
    unmatched_koststeder: list[str] = field(default_factory=list)
    matched_properties: int = 0
    direktorat_rows: int = 0
    total_2025_nok: float = 0.0
    total_2026_nok: float = 0.0
    errors: list[str] = field(default_factory=list)


def _parse_number(val) -> Optional[float]:
    """Norsk tallformat: håndterer \xa0, mellomrom, komma-desimal, None."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace("\xa0", "").replace(" ", "").replace(" ", "").strip()
    if not s or s in ("-", "–"):
        return 0.0
    try:
        if s.count(",") == 1 and s.count(".") == 0:
            s = s.replace(",", ".")
        elif s.count(",") == 1 and s.count(".") > 0:
            s = s.replace(".", "").replace(",", ".")
        return float(s)
    except ValueError:
        return None


async def import_finance_budget(
    db: AsyncSession,
    file_content: bytes,
    filename: str,
    imported_by: str,
) -> ImportReport:
    """
    Leser Excel-fila, mapper koststed → property via KoststedMapping,
    og skriver til finance_budget-tabellen for år 2025 og 2026.

    Idempotent: sletter eksisterende rader for (year, data_source) per år
    før insert.
    """
    import openpyxl
    import io

    report = ImportReport()
    batch_id = f"{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%S')}_{filename[:40]}"
    imported_at = datetime.now(timezone.utc)

    # --- Last Excel ---
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True, read_only=True)
    except Exception as exc:
        report.errors.append(f"Kunne ikke lese Excel-fil: {exc}")
        return report

    if SHEET_NAME not in wb.sheetnames:
        report.errors.append(
            f"Fant ikke fane '{SHEET_NAME}'. Tilgjengelige faner: {wb.sheetnames}"
        )
        return report

    ws = wb[SHEET_NAME]

    # --- Bygg koststed → property_id-map fra KoststedMapping ---
    try:
        ks_rows = await db.execute(text("""
            SELECT koststed_kode, property_id
            FROM koststed_mapping
            WHERE property_id IS NOT NULL
        """))
        koststed_to_property: dict[str, str] = {
            str(r.koststed_kode): str(r.property_id)
            for r in ks_rows.all()
        }
    except Exception as exc:
        logger.debug("KoststedMapping ikke tilgjengelig: %s", exc)
        koststed_to_property = {}

    # --- Slett eksisterende rader per år (idempotens) ---
    for yr in IMPORT_YEARS:
        ds = f"finance_dept_{yr}"
        try:
            await db.execute(
                text("DELETE FROM finance_budget WHERE year = :year AND data_source = :ds"),
                {"year": yr, "ds": ds},
            )
        except Exception as exc:
            logger.debug("DELETE finance_budget year=%s feilet (tabell mangler?): %s", yr, exc)

    # --- Iterér over rader ---
    rows_to_insert: list[dict] = []
    unmatched_set: set[str] = set()

    header = None
    for row_values in ws.iter_rows(values_only=True):
        if header is None:
            header = row_values
            continue

        report.total_rows += 1

        konto_raw, konto_navn, koststed_raw, koststed_navn, prosjekt, prosjekt_navn, \
            finansiering, finansiering_navn, periode_raw, belop_da, kontant_raw = (
                row_values[:11] if len(row_values) >= 11 else (row_values + (None,) * 11)[:11]
            )

        # --- Periode ---
        if periode_raw is None:
            report.skipped_no_periode += 1
            continue
        periode_str = str(int(periode_raw)) if isinstance(periode_raw, (int, float)) else str(periode_raw).strip()
        if len(periode_str) < 6:
            report.skipped_no_periode += 1
            continue
        try:
            year = int(periode_str[:4])
            month = int(periode_str[4:6])
        except ValueError:
            report.skipped_no_periode += 1
            continue

        if year not in IMPORT_YEARS:
            report.skipped_wrong_year += 1
            continue

        # --- Beløp ---
        # Beløp DA = vedtatt budsjett for hele året (brukes for begge år)
        # Kontantbeløp = kun tilgjengelig for perioder som er passert (tom for fremtid)
        # → bruk Beløp DA som primær, fall back til Kontantbeløp hvis DA er 0
        amount = _parse_number(belop_da)
        if amount is None or amount == 0.0:
            amount = _parse_number(kontant_raw)
        if amount is None or amount == 0.0:
            report.skipped_zero_amount += 1
            continue

        # --- Konto → kategori ---
        konto_str = str(int(konto_raw)) if isinstance(konto_raw, (int, float)) else str(konto_raw or "").strip()
        category = FINANCE_KONTO_CATEGORY.get(konto_str)
        if category is None:
            report.skipped_unknown_konto += 1
            continue

        # --- Koststed → property ---
        koststed_str = str(int(koststed_raw)) if isinstance(koststed_raw, (int, float)) else str(koststed_raw or "").strip()
        property_id_str = koststed_to_property.get(koststed_str)
        is_direktorat = property_id_str is None

        if is_direktorat:
            unmatched_set.add(f"{koststed_str} ({koststed_navn})")
            report.direktorat_rows += 1
        else:
            report.matched_properties += 1

        prosjekt_str = str(int(prosjekt)) if isinstance(prosjekt, (int, float)) and prosjekt else (str(prosjekt) if prosjekt else None)
        finansiering_str = str(int(finansiering)) if isinstance(finansiering, (int, float)) and finansiering else (str(finansiering) if finansiering else None)

        rows_to_insert.append({
            "finance_budget_id": str(uuid.uuid4()),
            "property_id": property_id_str,
            "koststed_kode": koststed_str,
            "koststed_navn": str(koststed_navn or "")[:200] if koststed_navn else None,
            "year": year,
            "month": month,
            "konto": konto_str,
            "konto_navn": str(konto_navn or "")[:200] if konto_navn else None,
            "category": category,
            "amount": amount,
            "finansiering_kode": finansiering_str,
            "prosjekt_kode": prosjekt_str,
            "is_direktorat_level": is_direktorat,
            "import_batch_id": batch_id,
            "imported_at": imported_at,
            "data_source": f"finance_dept_{year}",
        })

        if year == 2025:
            report.total_2025_nok += amount
        else:
            report.total_2026_nok += amount

    report.unmatched_koststeder = sorted(unmatched_set)

    # --- Batch-insert i bolker på 500 --- (budsjett)
    CHUNK = 500
    try:
        for i in range(0, len(rows_to_insert), CHUNK):
            chunk = rows_to_insert[i : i + CHUNK]
            await db.execute(text("""
                INSERT INTO finance_budget (
                    finance_budget_id, property_id, koststed_kode, koststed_navn,
                    year, month, konto, konto_navn, category, amount,
                    finansiering_kode, prosjekt_kode, is_direktorat_level,
                    import_batch_id, imported_at, data_source
                ) VALUES (
                    :finance_budget_id, :property_id, :koststed_kode, :koststed_navn,
                    :year, :month, :konto, :konto_navn, :category, :amount,
                    :finansiering_kode, :prosjekt_kode, :is_direktorat_level,
                    :import_batch_id, :imported_at, :data_source
                )
            """), chunk)
        await db.commit()
        report.inserted = len(rows_to_insert)
    except Exception as exc:
        await db.rollback()
        report.errors.append(f"DB-insert feilet: {exc}")
        logger.debug("finance_budget insert feilet: %s", exc)

    return report


async def import_kontant_actuals(
    db: AsyncSession,
    file_content: bytes,
    filename: str,
    imported_by: str,
    year: int = 2025,
) -> ImportReport:
    """
    Importer faktiske kostnader (regnskap) fra økonomi-avdelingens Excel-uttrekk.

    Leser fanen 'Kontant 2025' (eller tilsvarende for andre år).
    Kolonner: Konto | Avdeling | Kontantbeløp | Kont.periode (YYYYMM)

    Idempotent: sletter alle rader med (year, data_source='kontant_{year}') før insert.
    """
    import openpyxl
    import io

    report = ImportReport()
    data_source = f"kontant_{year}"
    batch_id = f"{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%S')}_{filename[:40]}"
    imported_at = datetime.now(timezone.utc)

    # --- Last Excel ---
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True, read_only=True)
    except Exception as exc:
        report.errors.append(f"Kunne ikke lese Excel-fil: {exc}")
        return report

    # Finn riktig fane (prøv KONTANT_SHEET_NAME, og dynamisk variant)
    sheet_candidates = [KONTANT_SHEET_NAME, f"Kontant {year}", "Kontant"]
    ws = None
    for candidate in sheet_candidates:
        if candidate in wb.sheetnames:
            ws = wb[candidate]
            break
    if ws is None:
        report.errors.append(
            f"Fant ikke kontant-fane. Prøvde: {sheet_candidates}. Tilgjengelige: {wb.sheetnames}"
        )
        return report

    # --- Bygg koststed → property_id-map ---
    try:
        ks_rows = await db.execute(text("""
            SELECT koststed_kode, property_id
            FROM koststed_mapping
            WHERE property_id IS NOT NULL
        """))
        koststed_to_property: dict[str, str] = {
            str(r.koststed_kode): str(r.property_id)
            for r in ks_rows.all()
        }
    except Exception as exc:
        logger.debug("KoststedMapping ikke tilgjengelig: %s", exc)
        koststed_to_property = {}

    # --- Idempotens: slett eksisterende rader ---
    try:
        await db.execute(
            text("DELETE FROM finance_budget WHERE year = :year AND data_source = :ds"),
            {"year": year, "ds": data_source},
        )
    except Exception as exc:
        logger.debug("DELETE finance_budget (kontant) feilet: %s", exc)

    # --- Iterér over rader ---
    # Kontant-sheet kolonner (0-indeksert):
    # 0: Konto | 1: Konto(T) | 2: Avdeling | 3: Avdeling(T) | 4: Statskonto | 5: Statskonto(T)
    # 6: Dim 2 | 7: Dim 2(T) | 8: Målgruppe | 9: Målgruppe(T) | 10: BA | 11: BA(T)
    # 12: Bilagsnr | 13: Resk.nr | 14: Resk.nr(T) | 15: Tekst | 16: Formål | 17: Formål(T)
    # 18: Kontantbeløp | 19: Kont.periode

    rows_to_insert: list[dict] = []
    unmatched_set: set[str] = set()

    header = None
    for row_values in ws.iter_rows(values_only=True):
        if header is None:
            header = row_values
            # Detect column positions dynamically
            header_list = [str(h or "").strip() for h in row_values]
            try:
                idx_konto = header_list.index("Konto")
                idx_avd = next((i for i, h in enumerate(header_list) if h in ("Avdeling", "Koststed")), 2)
                idx_avd_t = idx_avd + 1
                idx_kontant = next((i for i, h in enumerate(header_list) if "Kontantbeløp" in h), 18)
                idx_periode = next((i for i, h in enumerate(header_list) if "periode" in h.lower() or h == "Kont.periode"), 19)
            except (ValueError, StopIteration):
                # Fallback til kjente posisjoner
                idx_konto, idx_avd, idx_avd_t, idx_kontant, idx_periode = 0, 2, 3, 18, 19
            continue

        report.total_rows += 1

        def _get(idx):
            return row_values[idx] if idx < len(row_values) else None

        konto_raw = _get(idx_konto)
        avdeling_raw = _get(idx_avd)
        avdeling_navn = _get(idx_avd_t)
        kontant_raw = _get(idx_kontant)
        periode_raw = _get(idx_periode)

        # --- Periode ---
        if periode_raw is None:
            report.skipped_no_periode += 1
            continue
        periode_str = str(int(periode_raw)) if isinstance(periode_raw, (int, float)) else str(periode_raw).strip()
        if len(periode_str) < 6:
            report.skipped_no_periode += 1
            continue
        try:
            row_year = int(periode_str[:4])
            month = int(periode_str[4:6])
        except ValueError:
            report.skipped_no_periode += 1
            continue

        if row_year != year:
            report.skipped_wrong_year += 1
            continue

        # --- Beløp ---
        amount = _parse_number(kontant_raw)
        if amount is None or amount == 0.0:
            report.skipped_zero_amount += 1
            continue

        # --- Konto → kategori ---
        konto_str = str(int(konto_raw)) if isinstance(konto_raw, (int, float)) else str(konto_raw or "").strip()
        category = FINANCE_KONTO_CATEGORY.get(konto_str)
        if category is None:
            report.skipped_unknown_konto += 1
            continue

        # --- Avdeling (= koststed) → property ---
        avd_str = str(int(avdeling_raw)) if isinstance(avdeling_raw, (int, float)) else str(avdeling_raw or "").strip()
        property_id_str = koststed_to_property.get(avd_str)
        is_direktorat = property_id_str is None

        if is_direktorat:
            unmatched_set.add(f"{avd_str} ({avdeling_navn})")
            report.direktorat_rows += 1
        else:
            report.matched_properties += 1

        rows_to_insert.append({
            "finance_budget_id": str(uuid.uuid4()),
            "property_id": property_id_str,
            "koststed_kode": avd_str,
            "koststed_navn": str(avdeling_navn or "")[:200] if avdeling_navn else None,
            "year": row_year,
            "month": month,
            "konto": konto_str,
            "konto_navn": None,
            "category": category,
            "amount": amount,
            "finansiering_kode": None,
            "prosjekt_kode": None,
            "is_direktorat_level": is_direktorat,
            "import_batch_id": batch_id,
            "imported_at": imported_at,
            "data_source": data_source,
        })
        report.total_2025_nok += amount

    report.unmatched_koststeder = sorted(unmatched_set)

    # --- Batch-insert ---
    CHUNK = 500
    try:
        for i in range(0, len(rows_to_insert), CHUNK):
            chunk = rows_to_insert[i : i + CHUNK]
            await db.execute(text("""
                INSERT INTO finance_budget (
                    finance_budget_id, property_id, koststed_kode, koststed_navn,
                    year, month, konto, konto_navn, category, amount,
                    finansiering_kode, prosjekt_kode, is_direktorat_level,
                    import_batch_id, imported_at, data_source
                ) VALUES (
                    :finance_budget_id, :property_id, :koststed_kode, :koststed_navn,
                    :year, :month, :konto, :konto_navn, :category, :amount,
                    :finansiering_kode, :prosjekt_kode, :is_direktorat_level,
                    :import_batch_id, :imported_at, :data_source
                )
            """), chunk)
        await db.commit()
        report.inserted = len(rows_to_insert)
    except Exception as exc:
        await db.rollback()
        report.errors.append(f"DB-insert (kontant) feilet: {exc}")
        logger.debug("finance_budget kontant insert feilet: %s", exc)

    return report


async def import_kontant_2026_from_budget_sheet(
    db: AsyncSession,
    file_content: bytes,
    filename: str,
    imported_by: str,
) -> ImportReport:
    """
    Importer faktisk brukt (YTD) for 2026 fra Kontantbeløp-kolonnen i 'Budsjett 2025_2026'.

    Kolonner i kildefila (Budsjett 2025_2026):
      Konto | Konto(T) | Koststed | Koststed(T) | Prosjekt | Prosjekt(T)
      | Finansiering | Finansiering(T) | Periode (YYYYMM) | Beløp DA | Kontantbeløp

    Lagrer kun rader der Kontantbeløp > 0 og år = 2026 → data_source='kontant_2026'.
    Idempotent: sletter eksisterende kontant_2026-rader før insert.
    """
    import openpyxl
    import io

    TARGET_YEAR = 2026
    DATA_SOURCE = "kontant_2026"

    report = ImportReport()
    batch_id = f"{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%S')}_{filename[:40]}"
    imported_at = datetime.now(timezone.utc)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True, read_only=True)
    except Exception as exc:
        report.errors.append(f"Kunne ikke lese Excel-fil: {exc}")
        return report

    if SHEET_NAME not in wb.sheetnames:
        report.errors.append(f"Fant ikke fane '{SHEET_NAME}'. Tilgjengelige: {wb.sheetnames}")
        return report

    ws = wb[SHEET_NAME]

    # --- Koststed → property ---
    try:
        ks_rows = await db.execute(text("""
            SELECT koststed_kode, property_id FROM koststed_mapping WHERE property_id IS NOT NULL
        """))
        koststed_to_property: dict[str, str] = {
            str(r.koststed_kode): str(r.property_id) for r in ks_rows.all()
        }
    except Exception as exc:
        logger.debug("KoststedMapping ikke tilgjengelig: %s", exc)
        koststed_to_property = {}

    # --- Idempotens ---
    try:
        await db.execute(
            text("DELETE FROM finance_budget WHERE year = :y AND data_source = :ds"),
            {"y": TARGET_YEAR, "ds": DATA_SOURCE},
        )
    except Exception as exc:
        logger.debug("DELETE kontant_2026 feilet: %s", exc)

    rows_to_insert: list[dict] = []
    unmatched_set: set[str] = set()

    header = None
    for row_values in ws.iter_rows(values_only=True):
        if header is None:
            header = row_values
            continue

        report.total_rows += 1

        konto_raw, konto_navn, koststed_raw, koststed_navn, prosjekt, prosjekt_navn, \
            finansiering, finansiering_navn, periode_raw, _belop_da, kontant_raw = (
                row_values[:11] if len(row_values) >= 11 else (row_values + (None,) * 11)[:11]
            )

        # --- Periode: kun 2026 ---
        if periode_raw is None:
            report.skipped_no_periode += 1
            continue
        periode_str = str(int(periode_raw)) if isinstance(periode_raw, (int, float)) else str(periode_raw).strip()
        if len(periode_str) < 6:
            report.skipped_no_periode += 1
            continue
        try:
            year = int(periode_str[:4])
            month = int(periode_str[4:6])
        except ValueError:
            report.skipped_no_periode += 1
            continue

        if year != TARGET_YEAR:
            report.skipped_wrong_year += 1
            continue

        # --- Beløp: kun Kontantbeløp (faktisk betalt) ---
        amount = _parse_number(kontant_raw)
        if amount is None or amount == 0.0:
            report.skipped_zero_amount += 1
            continue

        # --- Konto → kategori ---
        konto_str = str(int(konto_raw)) if isinstance(konto_raw, (int, float)) else str(konto_raw or "").strip()
        category = FINANCE_KONTO_CATEGORY.get(konto_str)
        if category is None:
            report.skipped_unknown_konto += 1
            continue

        # --- Koststed → property ---
        koststed_str = str(int(koststed_raw)) if isinstance(koststed_raw, (int, float)) else str(koststed_raw or "").strip()
        property_id_str = koststed_to_property.get(koststed_str)
        is_direktorat = property_id_str is None

        if is_direktorat:
            unmatched_set.add(f"{koststed_str} ({koststed_navn})")
            report.direktorat_rows += 1
        else:
            report.matched_properties += 1

        rows_to_insert.append({
            "finance_budget_id": str(uuid.uuid4()),
            "property_id": property_id_str,
            "koststed_kode": koststed_str,
            "koststed_navn": str(koststed_navn or "")[:200] if koststed_navn else None,
            "year": year,
            "month": month,
            "konto": konto_str,
            "konto_navn": str(konto_navn or "")[:200] if konto_navn else None,
            "category": category,
            "amount": amount,
            "finansiering_kode": None,
            "prosjekt_kode": None,
            "is_direktorat_level": is_direktorat,
            "import_batch_id": batch_id,
            "imported_at": imported_at,
            "data_source": DATA_SOURCE,
        })
        report.total_2026_nok += amount

    report.unmatched_koststeder = sorted(unmatched_set)

    CHUNK = 500
    try:
        for i in range(0, len(rows_to_insert), CHUNK):
            chunk = rows_to_insert[i: i + CHUNK]
            await db.execute(text("""
                INSERT INTO finance_budget (
                    finance_budget_id, property_id, koststed_kode, koststed_navn,
                    year, month, konto, konto_navn, category, amount,
                    finansiering_kode, prosjekt_kode, is_direktorat_level,
                    import_batch_id, imported_at, data_source
                ) VALUES (
                    :finance_budget_id, :property_id, :koststed_kode, :koststed_navn,
                    :year, :month, :konto, :konto_navn, :category, :amount,
                    :finansiering_kode, :prosjekt_kode, :is_direktorat_level,
                    :import_batch_id, :imported_at, :data_source
                )
            """), chunk)
        await db.commit()
        report.inserted = len(rows_to_insert)
    except Exception as exc:
        await db.rollback()
        report.errors.append(f"DB-insert (kontant_2026) feilet: {exc}")

    return report
