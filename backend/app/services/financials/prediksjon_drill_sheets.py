"""
Ekstra Excel-ark for prediksjon 2027: eiendom × kategori, GL konto og GL-linjer.

Brukes av GET /financials/prediksjon-2027/excel og /export.xlsx.
GL-filtre matcher øvrig prediksjon-API (kun poster med beløp > 0).
"""
from __future__ import annotations

from typing import Any

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from sqlalchemy import desc, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.financial_models import Budget, GLTransaction
from app.services.financials.pred2027 import append_pred2027_basis_sheets, apply_pred2027_workbook_links

BUDGET_KAT_MAP = {
    "operations": "Drift",
    "investment": "Investering",
    "property": "Gjennomstrømning",
    "other": "Annet",
}

# Maks rader på bilagsark (unngå gigantiske filer)
GL_BILAG_MAX_ROWS = 80_000


def _hdr_row(ws, row: int, headers: list[str], fill: PatternFill) -> None:
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="left", vertical="center")


async def append_prediksjon_drill_sheets(
    db: AsyncSession,
    wb: Workbook,
    *,
    prop_info: dict[str, dict[str, Any]],
    scenario_tags: tuple[str, str] = ("xgb70", "xgb50"),
    gl_year: int = 2025,
) -> None:
    """Legger til drill-ark for prediksjon 2027, inkl. historikk og detaljert kostnadsgrunnlag."""
    fill_teal = PatternFill("solid", fgColor="0F766E")

    bud: dict[str, dict[tuple[str, str], float]] = {t: {} for t in scenario_tags}
    for tag in scenario_tags:
        rows = (
            await db.execute(
                select(
                    Budget.property_id,
                    Budget.category,
                    func.sum(Budget.amount).label("belop"),
                )
                .where(
                    Budget.year == 2027,
                    Budget.is_synthetic.is_(True),
                    Budget.data_source == f"holt_winters_2027_{tag}",
                )
                .group_by(Budget.property_id, Budget.category)
            )
        ).fetchall()
        for r in rows:
            pid = str(r.property_id)
            key = (pid, BUDGET_KAT_MAP.get(r.category or "", r.category or "Ukjent"))
            bud[tag][key] = float(r.belop or 0)

    gl_kat_rows = (
        await db.execute(
            select(
                GLTransaction.property_id,
                GLTransaction.srs_kategori,
                func.sum(GLTransaction.belop).label("belop"),
            )
            .where(
                GLTransaction.ar == gl_year,
                GLTransaction.belop > 0,
                GLTransaction.property_id.isnot(None),
            )
            .group_by(GLTransaction.property_id, GLTransaction.srs_kategori)
        )
    ).fetchall()
    gl_kat: dict[tuple[str, str], float] = {}
    for r in gl_kat_rows:
        pid = str(r.property_id)
        kat = r.srs_kategori or "Ukjent"
        gl_kat[(pid, kat)] = float(r.belop or 0)

    keys = set(bud["xgb70"].keys()) | set(bud["xgb50"].keys()) | set(gl_kat.keys())
    ws_ek = wb.create_sheet("Eiendom_kategori")
    h1 = [
        "property_id",
        "eiendom",
        "region",
        "srs_kategori",
        f"gl_{gl_year}",
        "pred_2027_xgb70",
        "pred_2027_xgb50",
        "endring_pct_70_vs_gl",
    ]
    _hdr_row(ws_ek, 1, h1, fill_teal)
    r = 2
    for pid, kat in sorted(keys):
        info = prop_info.get(pid, {})
        name = info.get("name", pid)
        region = info.get("region", "Ukjent")
        g = gl_kat.get((pid, kat), 0.0)
        p70 = bud["xgb70"].get((pid, kat), 0.0)
        p50 = bud["xgb50"].get((pid, kat), 0.0)
        endr = round((p70 - g) / g * 100, 1) if g > 0 else None
        row_vals: list[Any] = [pid, name, region, kat, round(g, 0), round(p70, 0), round(p50, 0), endr]
        for c, val in enumerate(row_vals, 1):
            cell = ws_ek.cell(row=r, column=c, value=val)
            if c >= 5 and isinstance(val, (int, float)) and c != 8:
                cell.number_format = "#,##0"
            if c == 8 and val is not None:
                cell.number_format = '0.0"%"'
        r += 1
    for col, w in enumerate([38, 36, 14, 18, 16, 16, 16, 18], 1):
        ws_ek.column_dimensions[get_column_letter(col)].width = w

    ws_k = wb.create_sheet("GL_konto")
    kh = [
        "property_id",
        "eiendom",
        "region",
        "ar",
        "srs_kategori",
        "konto",
        "konto_navn",
        "belop",
        "antall_transaksjoner",
    ]
    _hdr_row(ws_k, 1, kh, fill_teal)
    sum_belop = func.sum(GLTransaction.belop).label("sum_belop")
    konto_rows = (
        await db.execute(
            select(
                GLTransaction.property_id,
                GLTransaction.srs_kategori,
                GLTransaction.konto,
                GLTransaction.konto_navn,
                sum_belop,
                func.count(GLTransaction.transaction_id).label("n"),
            )
            .where(
                GLTransaction.ar == gl_year,
                GLTransaction.belop > 0,
                GLTransaction.property_id.isnot(None),
            )
            .group_by(
                GLTransaction.property_id,
                GLTransaction.srs_kategori,
                GLTransaction.konto,
                GLTransaction.konto_navn,
            )
            .order_by(GLTransaction.property_id, desc(sum_belop))
        )
    ).fetchall()
    rk = 2
    for row in konto_rows:
        pid = str(row[0])
        srs_kat = row[1] or ""
        konto = row[2] or ""
        konto_navn = row[3] or ""
        sbelop = float(row[4] or 0)
        n_tx = int(row[5] or 0)
        info = prop_info.get(pid, {})
        ws_k.cell(row=rk, column=1, value=pid)
        ws_k.cell(row=rk, column=2, value=info.get("name", pid))
        ws_k.cell(row=rk, column=3, value=info.get("region", "Ukjent"))
        ws_k.cell(row=rk, column=4, value=gl_year)
        ws_k.cell(row=rk, column=5, value=srs_kat)
        ws_k.cell(row=rk, column=6, value=konto)
        ws_k.cell(row=rk, column=7, value=konto_navn)
        c8 = ws_k.cell(row=rk, column=8, value=round(sbelop, 0))
        c8.number_format = "#,##0"
        ws_k.cell(row=rk, column=9, value=n_tx)
        rk += 1
    for col, w in enumerate([38, 34, 12, 6, 16, 10, 28, 14, 8], 1):
        ws_k.column_dimensions[get_column_letter(col)].width = min(w, 50)

    ws_b = wb.create_sheet("GL_bilag")
    ws_b.cell(row=1, column=1, value=f"Merk: maks {GL_BILAG_MAX_ROWS} linjer (år {gl_year}, beløp > 0). Sortert etter eiendom og beløp.")
    ws_b.merge_cells("A1:J1")
    ws_b["A1"].font = Font(italic=True, size=10)

    bh = [
        "property_id",
        "eiendom",
        "region",
        "bilagsnr",
        "bilagsdato",
        "konto",
        "konto_navn",
        "srs_kategori",
        "tekst",
        "belop",
    ]
    _hdr_row(ws_b, 2, bh, fill_teal)

    bilag_stmt = (
        select(
            GLTransaction.property_id,
            GLTransaction.bilagsnr,
            GLTransaction.bilagsdato,
            GLTransaction.konto,
            GLTransaction.konto_navn,
            GLTransaction.srs_kategori,
            GLTransaction.tekst,
            GLTransaction.belop,
        )
        .where(
            GLTransaction.ar == gl_year,
            GLTransaction.belop > 0,
            GLTransaction.property_id.isnot(None),
        )
        .order_by(GLTransaction.property_id, desc(GLTransaction.belop))
        .limit(GL_BILAG_MAX_ROWS)
    )
    bilag_rows = (await db.execute(bilag_stmt)).fetchall()
    rb = 3
    for row in bilag_rows:
        pid = str(row[0])
        bilagsnr = row[1] or ""
        bilagsdato = row[2]
        konto = row[3] or ""
        konto_navn = row[4] or ""
        srs_kat = row[5] or ""
        tekst = row[6] or ""
        belop = float(row[7] or 0)
        info = prop_info.get(pid, {})
        ws_b.cell(row=rb, column=1, value=pid)
        ws_b.cell(row=rb, column=2, value=info.get("name", pid))
        ws_b.cell(row=rb, column=3, value=info.get("region", "Ukjent"))
        ws_b.cell(row=rb, column=4, value=bilagsnr)
        ws_b.cell(row=rb, column=5, value=bilagsdato.isoformat() if bilagsdato else "")
        ws_b.cell(row=rb, column=6, value=konto)
        ws_b.cell(row=rb, column=7, value=konto_navn)
        ws_b.cell(row=rb, column=8, value=srs_kat)
        ws_b.cell(row=rb, column=9, value=str(tekst)[:500])
        c10 = ws_b.cell(row=rb, column=10, value=belop)
        c10.number_format = "#,##0.00"
        rb += 1
    for col, w in enumerate([38, 32, 12, 14, 12, 10, 24, 14, 40, 14], 1):
        ws_b.column_dimensions[get_column_letter(col)].width = min(w, 50)

    basis_refs = await append_pred2027_basis_sheets(db, wb, prop_info=prop_info)
    apply_pred2027_workbook_links(
        wb,
        history_rows=basis_refs.get("history_rows"),
        basis_rows=basis_refs.get("basis_rows"),
    )
