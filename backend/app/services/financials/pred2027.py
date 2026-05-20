from __future__ import annotations

from collections import defaultdict
from typing import Any

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from sqlalchemy import desc, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.financial_models import GLTransaction

HISTORY_YEARS: tuple[int, ...] = (2021, 2022, 2023, 2024, 2025)


def _hdr_row(ws, row: int, headers: list[str], fill: PatternFill) -> None:
    for col, value in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _set_widths(ws, widths: list[int]) -> None:
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = min(width, 50)


def _enable_filters(ws, freeze_panes: str) -> None:
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = freeze_panes


async def append_pred2027_basis_sheets(
    db: AsyncSession,
    wb: Workbook,
    *,
    prop_info: dict[str, dict[str, Any]],
    history_years: tuple[int, ...] = HISTORY_YEARS,
) -> dict[str, dict[str, int]]:
    """Legger til komplett historikkgrunnlag og detaljert kostnadsgrunnlag for Prediksjon 2027."""
    fill_navy = PatternFill("solid", fgColor="1E3A8A")

    hist_stmt = (
        select(
            GLTransaction.property_id,
            GLTransaction.srs_kategori,
            GLTransaction.ar,
            func.sum(GLTransaction.belop).label("belop"),
        )
        .where(
            GLTransaction.ar.in_(history_years),
            GLTransaction.belop > 0,
            GLTransaction.property_id.isnot(None),
        )
        .group_by(GLTransaction.property_id, GLTransaction.srs_kategori, GLTransaction.ar)
        .order_by(GLTransaction.property_id, GLTransaction.srs_kategori, GLTransaction.ar)
    )
    hist_rows = (await db.execute(hist_stmt)).fetchall()

    hist_map: dict[tuple[str, str], dict[int, float]] = defaultdict(dict)
    for row in hist_rows:
        pid = str(row.property_id)
        kat = row.srs_kategori or "Ukjent"
        hist_map[(pid, kat)][int(row.ar)] = float(row.belop or 0)

    ws_hist = wb.create_sheet("Historikk_grunnlag")
    hist_headers = [
        "property_id",
        "eiendom",
        "region",
        "srs_kategori",
        *[str(year) for year in history_years],
        "sum_grunnlag",
        "snitt_per_ar",
        "til_detaljer",
    ]
    _hdr_row(ws_hist, 1, hist_headers, fill_navy)

    history_rows_by_property: dict[str, int] = {}
    row_no = 2
    for (pid, kat), values in sorted(hist_map.items()):
        info = prop_info.get(pid, {})
        annual_values = [round(values.get(year, 0.0), 0) for year in history_years]
        total = round(sum(annual_values), 0)
        avg = round(total / len(history_years), 0) if history_years else 0
        row_values = [
            pid,
            info.get("name", pid),
            info.get("region", "Ukjent"),
            kat,
            *annual_values,
            total,
            avg,
            "Åpne detaljlinjer",
        ]
        for col, value in enumerate(row_values, 1):
            cell = ws_hist.cell(row=row_no, column=col, value=value)
            if col >= 5 and col <= (4 + len(history_years) + 2):
                cell.number_format = "#,##0"
        history_rows_by_property.setdefault(pid, row_no)
        row_no += 1

    _set_widths(ws_hist, [38, 34, 16, 18, 14, 14, 14, 14, 14, 16, 16, 18])
    _enable_filters(ws_hist, "A2")

    detail_stmt = (
        select(
            GLTransaction.transaction_id,
            GLTransaction.property_id,
            GLTransaction.ar,
            GLTransaction.periode,
            GLTransaction.bilagsnr,
            GLTransaction.bilagsdato,
            GLTransaction.konto,
            GLTransaction.konto_navn,
            GLTransaction.srs_kategori,
            GLTransaction.ba_kode,
            GLTransaction.innkjopskategori_navn,
            GLTransaction.underkategori_navn,
            GLTransaction.leverandor_navn,
            GLTransaction.tekst,
            GLTransaction.belop,
            GLTransaction.dim1_kode,
            GLTransaction.dim1_navn,
            GLTransaction.dim3_kode,
            GLTransaction.dim6_anlegg_id,
            GLTransaction.is_statsbygg,
        )
        .where(
            GLTransaction.ar.in_(history_years),
            GLTransaction.belop > 0,
            GLTransaction.property_id.isnot(None),
        )
        .order_by(GLTransaction.property_id, GLTransaction.ar, desc(GLTransaction.belop))
    )
    detail_rows = (await db.execute(detail_stmt)).fetchall()

    ws_basis = wb.create_sheet("Kostnadsgrunnlag_GL")
    title = f"Komplett kostnadsgrunnlag brukt som basis for prediksjon 2027 ({min(history_years)}–{max(history_years)})"
    ws_basis.cell(row=1, column=1, value=title)
    ws_basis.merge_cells("A1:T1")
    ws_basis["A1"].font = Font(bold=True, size=11)

    basis_headers = [
        "transaksjon_id",
        "property_id",
        "eiendom",
        "region",
        "år",
        "periode",
        "bilagsnr",
        "bilagsdato",
        "konto",
        "konto_navn",
        "srs_kategori",
        "bilagsart",
        "innkjøpskategori",
        "underkategori",
        "leverandor",
        "tekst",
        "beløp",
        "koststed",
        "koststed_navn",
        "formål",
        "anlegg_id",
        "statsbygg",
    ]
    _hdr_row(ws_basis, 2, basis_headers, fill_navy)

    basis_rows_by_property: dict[str, int] = {}
    basis_row_no = 3
    for row in detail_rows:
        pid = str(row.property_id)
        info = prop_info.get(pid, {})
        row_values = [
            str(row.transaction_id),
            pid,
            info.get("name", pid),
            info.get("region", "Ukjent"),
            int(row.ar) if row.ar is not None else None,
            row.periode or "",
            row.bilagsnr or "",
            row.bilagsdato.isoformat() if row.bilagsdato else "",
            row.konto or "",
            row.konto_navn or "",
            row.srs_kategori or "",
            row.ba_kode or "",
            row.innkjopskategori_navn or "",
            row.underkategori_navn or "",
            row.leverandor_navn or "",
            str(row.tekst or "")[:500],
            float(row.belop or 0),
            row.dim1_kode or "",
            row.dim1_navn or "",
            row.dim3_kode or "",
            row.dim6_anlegg_id or "",
            "Ja" if row.is_statsbygg else "Nei",
        ]
        for col, value in enumerate(row_values, 1):
            cell = ws_basis.cell(row=basis_row_no, column=col, value=value)
            if col == 17 and isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"
        basis_rows_by_property.setdefault(pid, basis_row_no)
        basis_row_no += 1

    _set_widths(ws_basis, [40, 38, 32, 14, 8, 10, 14, 12, 10, 24, 16, 10, 18, 18, 24, 40, 14, 12, 22, 10, 12, 10])
    _enable_filters(ws_basis, "A3")

    last_col = len(hist_headers)
    for hist_row in range(2, ws_hist.max_row + 1):
        pid = str(ws_hist.cell(row=hist_row, column=1).value or "")
        target_row = basis_rows_by_property.get(pid)
        if target_row:
            link_cell = ws_hist.cell(row=hist_row, column=last_col)
            link_cell.hyperlink = f"#Kostnadsgrunnlag_GL!A{target_row}"
            link_cell.style = "Hyperlink"

    return {
        "history_rows": history_rows_by_property,
        "basis_rows": basis_rows_by_property,
    }


def apply_pred2027_workbook_links(
    wb: Workbook,
    *,
    history_rows: dict[str, int] | None = None,
    basis_rows: dict[str, int] | None = None,
) -> None:
    """Gjør workbook enklere å bruke med filtre, freeze panes og interne hopp mellom ark."""
    history_rows = history_rows or {}
    basis_rows = basis_rows or {}

    for sheet_name, freeze_cell in {
        "Sammendrag": "A3",
        "Alle eiendommer": "A2",
        "Eiendom_kategori": "A2",
        "GL_konto": "A2",
        "GL_bilag": "A3",
        "Historikk_grunnlag": "A2",
        "Kostnadsgrunnlag_GL": "A3",
    }.items():
        if sheet_name in wb.sheetnames:
            _enable_filters(wb[sheet_name], freeze_cell)

    if "Eiendom_kategori" in wb.sheetnames:
        ws = wb["Eiendom_kategori"]
        for row in range(2, ws.max_row + 1):
            pid = str(ws.cell(row=row, column=1).value or "")
            hist_target = history_rows.get(pid)
            basis_target = basis_rows.get(pid)
            if hist_target:
                cell = ws.cell(row=row, column=1)
                cell.hyperlink = f"#Historikk_grunnlag!A{hist_target}"
                cell.style = "Hyperlink"
            if basis_target:
                cell = ws.cell(row=row, column=2)
                cell.hyperlink = f"#Kostnadsgrunnlag_GL!A{basis_target}"
                cell.style = "Hyperlink"

    if "GL_konto" in wb.sheetnames:
        ws = wb["GL_konto"]
        for row in range(2, ws.max_row + 1):
            pid = str(ws.cell(row=row, column=1).value or "")
            hist_target = history_rows.get(pid)
            basis_target = basis_rows.get(pid)
            if hist_target:
                cell = ws.cell(row=row, column=1)
                cell.hyperlink = f"#Historikk_grunnlag!A{hist_target}"
                cell.style = "Hyperlink"
            if basis_target:
                cell = ws.cell(row=row, column=2)
                cell.hyperlink = f"#Kostnadsgrunnlag_GL!A{basis_target}"
                cell.style = "Hyperlink"
