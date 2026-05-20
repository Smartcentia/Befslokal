"""
Excel-eksport med full kaskaderende formelstruktur for Prediksjon 2027.

Ark (hovedflyt):
  1. Antagelser       – justerbare parametere (gule celler) + named ranges
  2. Sammendrag       – totalt budsjettbehov 2027 (formler)
  3. Per region       – HW 2027 med inflasjon + regionfaktor
  4. Per kategori     – HW 2027 med inflasjon + kategorifaktor
  5. Lønn             – lønnskostnader per region
  6. Alle eiendommer  – eiendommer med cascading formler
  7. Outliers         – avvik og flagg
  8. Backtesting      – modellvalidering
  9. Rådata HW        – kilde (grå)
 10. Omposteringer    – bilagsarter (H1/H2/HB/RE m.m.) og omfang i GL 2021–2025
 11. Notater          – metodikk og veiledning

Named ranges: Inflasjon, LonnVekst, KatTabell, RegionTabell (Antagelser).

GET /api/v1/financials/prediksjon-2027/export.xlsx?scenario=xgb70
"""
from __future__ import annotations

import io
import logging
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from sqlalchemy import func, select, text
from sqlalchemy.ext.asyncio import AsyncSession

logger = logging.getLogger(__name__)

# ── Stiler ─────────────────────────────────────────────────────────────────────
FILL_YELLOW  = PatternFill("solid", fgColor="FFF2CC")  # redigerbare inputceller
FILL_FORMULA = PatternFill("solid", fgColor="C6EFCE")  # formel-resultater
FILL_BLUE    = PatternFill("solid", fgColor="DAE8FC")  # lønns-ark
FILL_GRAY    = PatternFill("solid", fgColor="D9D9D9")  # rådata
FILL_HEADER  = PatternFill("solid", fgColor="2F5496")  # kolonneheaders
FILL_SECTION = PatternFill("solid", fgColor="BDD7EE")  # seksjonstekst
FILL_TOTAL   = PatternFill("solid", fgColor="F4B942")  # totalrader

FONT_BOLD    = Font(bold=True)
FONT_WHITE   = Font(bold=True, color="FFFFFF")
FONT_TITLE   = Font(bold=True, size=14)
FONT_SECTION = Font(bold=True, size=10, color="1F3864")
FONT_SMALL   = Font(size=9, italic=True, color="595959")

NUM_NOK      = "#,##0"
NUM_PCT_DEC  = "0.0%"
NUM_FACTOR   = "0.00"
NUM_PCT_VAL  = "0.0"  # for %-inputceller (tall som 7.5, ikke 0.075)

# Canonical category order — MÅ matche KatTabell-rader 7–10 i Antagelser
KAT_ORDER = ["Drift", "Gjennomstrømning", "Investering", "Annet"]

BUDGET_KAT_MAP = {
    "operations": "Drift",
    "investment": "Investering",
    "property":   "Gjennomstrømning",
    "other":      "Annet",
}

# Bilagsarter som vanligvis brukes til ompostering / korrigering i statsregnskapet (Agresso)
BA_OMPOSTERING = frozenset({"H1", "H2", "HB", "RE"})
BA_FAKTURA_TYPISK = frozenset({"IV", "IW", "LE", "MT"})


def _h(cell, value: str, fill=FILL_HEADER):
    """Sett kolonneheader-stil."""
    cell.value = value
    cell.font = FONT_WHITE if fill == FILL_HEADER else FONT_BOLD
    cell.fill = fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _col_w(ws, widths: Dict[int, int]):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


# ── Datainnlasting ─────────────────────────────────────────────────────────────

async def _load_data(db: AsyncSession, scenario: str) -> Dict[str, Any]:
    from app.models.financial_models import Budget, GLTransaction, SalaryCost
    from app.domains.core.models.property import Property

    src = f"holt_winters_2027_{scenario}"

    # 1. Budsjett 2027 (syntetisk, Holt-Winters)
    b_rows = (await db.execute(
        select(Budget.property_id, Budget.category,
               func.sum(Budget.amount).label("b27"))
        .where(Budget.year == 2027, Budget.is_synthetic == True,
               Budget.data_source == src)
        .group_by(Budget.property_id, Budget.category)
    )).fetchall()

    prop_2027: Dict[str, float] = {}
    kat_2027:  Dict[str, float] = {}
    for r in b_rows:
        pid = str(r.property_id)
        v = float(r.b27 or 0)
        prop_2027[pid] = prop_2027.get(pid, 0.0) + v
        k = BUDGET_KAT_MAP.get(r.category or "", "Annet")
        kat_2027[k] = kat_2027.get(k, 0.0) + v

    # 2. GL per år 2021–2025 (faktisk) — brukes til outlier-analyse
    # Bruker SUM + HAVING for å nette ut omposteringer og reverseringer (regel 7 CLAUDE.md)
    g_hist_rows = (await db.execute(
        select(GLTransaction.property_id, GLTransaction.ar,
               func.sum(GLTransaction.belop).label("total"))
        .where(GLTransaction.property_id.isnot(None),
               GLTransaction.ar.between(2021, 2025))
        .group_by(GLTransaction.property_id, GLTransaction.ar)
        .having(func.sum(GLTransaction.belop) > 0)
    )).fetchall()

    # prop_hist: { pid: { year: amount } }
    prop_hist: Dict[str, Dict[int, float]] = {}
    for r in g_hist_rows:
        pid = str(r.property_id)
        prop_hist.setdefault(pid, {})[r.ar] = float(r.total or 0)

    # 2025 aggregated for other uses
    prop_2025: Dict[str, float] = {pid: yrs.get(2025, 0.0) for pid, yrs in prop_hist.items()}

    # Netting per (property, srs_kategori) — HAVING SUM > 0 utelukker kategorier
    # der omposteringer overstiger originaltransaksjoner (regel 7 CLAUDE.md)
    g_rows = (await db.execute(
        select(GLTransaction.property_id, GLTransaction.srs_kategori,
               func.sum(GLTransaction.belop).label("b25"))
        .where(GLTransaction.ar == 2025,
               GLTransaction.property_id.isnot(None))
        .group_by(GLTransaction.property_id, GLTransaction.srs_kategori)
        .having(func.sum(GLTransaction.belop) > 0)
    )).fetchall()

    kat_2025:  Dict[str, float] = {}
    for r in g_rows:
        v = float(r.b25 or 0)
        k = r.srs_kategori or "Annet"
        kat_2025[k] = kat_2025.get(k, 0.0) + v

    # 3. Eiendomsinfo
    p_rows = (await db.execute(
        select(Property.property_id, Property.name, Property.address, Property.region)
    )).fetchall()
    prop_info: Dict[str, Dict[str, Any]] = {
        str(r.property_id): {
            "name":   r.name or r.address or str(r.property_id),
            "region": r.region or "Ukjent",
        }
        for r in p_rows
    }

    # 4. Per region
    reg_2027: Dict[str, float] = {}
    reg_2025: Dict[str, float] = {}
    for pid, v in prop_2027.items():
        reg = prop_info.get(pid, {}).get("region", "Ukjent")
        reg_2027[reg] = reg_2027.get(reg, 0.0) + v
    for pid, v in prop_2025.items():
        reg = prop_info.get(pid, {}).get("region", "Ukjent")
        reg_2025[reg] = reg_2025.get(reg, 0.0) + v
    alle_regioner = sorted(set(reg_2027) | set(reg_2025))

    # 5. Lønn per region (2025 + 2027)
    lonn_r25: Dict[str, float] = {}
    lonn_r27: Dict[str, float] = {}
    try:
        l_rows = (await db.execute(
            select(
                Property.region,
                SalaryCost.year,
                func.sum(
                    SalaryCost.faste_stillinger
                    + SalaryCost.vikarer
                    + SalaryCost.arbeidsgiveravgift
                ).label("total"),
            )
            .join(Property, SalaryCost.property_id == Property.property_id)
            .where(SalaryCost.year.in_([2025, 2027]),
                   SalaryCost.property_id.isnot(None))
            .group_by(Property.region, SalaryCost.year)
        )).fetchall()
        for r in l_rows:
            reg = r.region or "Ukjent"
            v = float(r.total or 0)
            target = lonn_r25 if r.year == 2025 else lonn_r27
            target[reg] = target.get(reg, 0.0) + v
    except Exception as exc:
        logger.warning("Lønn per region feilet (fortsetter uten): %s", exc)

    # 6. Alle eiendommer sortert etter 2027 desc
    # Dominant category per property (highest HW budget)
    prop_kat: Dict[str, str] = {}
    prop_kat_amounts: Dict[str, Dict[str, float]] = {}
    for r in b_rows:
        pid = str(r.property_id)
        v = float(r.b27 or 0)
        k = BUDGET_KAT_MAP.get(r.category or "", "Annet")
        prop_kat_amounts.setdefault(pid, {})
        prop_kat_amounts[pid][k] = prop_kat_amounts[pid].get(k, 0.0) + v
    for pid, kats in prop_kat_amounts.items():
        prop_kat[pid] = max(kats, key=kats.get)

    alle_pid = set(prop_2027) | set(prop_2025)
    per_eiendom: List[Dict[str, Any]] = []
    for pid in alle_pid:
        info = prop_info.get(pid, {"name": pid, "region": "Ukjent"})
        b27 = round(prop_2027.get(pid, 0.0), 0)
        b25 = round(prop_2025.get(pid, 0.0), 0)
        per_eiendom.append({
            "name":       info["name"],
            "region":     info["region"],
            "belop_2027": b27,
            "belop_2025": b25,
            "kategori":   prop_kat.get(pid, "Drift"),
        })
    per_eiendom.sort(key=lambda x: -x["belop_2027"])

    # 7. Outlier-analyse
    outliers = _compute_outliers(prop_2027, prop_2025, prop_hist, prop_info)

    # 8. Eiendommer uten 2027-prediksjon (salary-only)
    try:
        no_pred_rows = (await db.execute(
            select(Property.name, Property.region,
                   func.sum(
                       SalaryCost.faste_stillinger
                       + SalaryCost.vikarer
                       + SalaryCost.arbeidsgiveravgift
                   ).label("lonn_2025"))
            .join(SalaryCost, SalaryCost.property_id == Property.property_id)
            .where(
                SalaryCost.year == 2025,
                Property.closed_at.is_(None),
                ~Property.property_id.in_(
                    select(Budget.property_id).where(
                        Budget.year == 2027, Budget.is_synthetic == True
                    )
                ),
            )
            .group_by(Property.name, Property.region)
            .order_by(func.sum(
                SalaryCost.faste_stillinger
                + SalaryCost.vikarer
                + SalaryCost.arbeidsgiveravgift
            ).desc())
        )).fetchall()
        no_pred = [
            {"name": r.name or "?", "region": r.region or "?", "lonn_2025": float(r.lonn_2025 or 0)}
            for r in no_pred_rows
        ]
    except Exception as exc:
        logger.warning("no_pred query feilet: %s", exc)
        no_pred = []

    return {
        "alle_regioner": alle_regioner,
        "reg_2027":      reg_2027,
        "reg_2025":      reg_2025,
        "kat_2027":      kat_2027,
        "kat_2025":      kat_2025,
        "per_eiendom":   per_eiendom,
        "lonn_r25":      lonn_r25,
        "lonn_r27":      lonn_r27,
        "outliers":      outliers,
        "no_pred":       no_pred,
    }


async def _load_ompostering_stats(db: AsyncSession) -> Dict[str, Any]:
    """
    Aggreger GL-linjer med property_id, år 2021–2025: omfang av omposteringer vs øvrige bilagsarter.
    Brukes kun til informasjonsark (påvirker ikke prediksjonsberegningen).
    """
    summary_rows: List[Dict[str, Any]] = []
    detail_rows: List[Dict[str, Any]] = []
    try:
        q_sum = text("""
            SELECT ar::int AS ar,
                   COUNT(*)::bigint AS n_all,
                   COUNT(*) FILTER (
                       WHERE UPPER(TRIM(COALESCE(ba_kode, ''))) IN ('H1','H2','HB','RE')
                   )::bigint AS n_omp,
                   COALESCE(SUM(belop), 0)::double precision AS sum_all,
                   COALESCE(SUM(belop) FILTER (
                       WHERE UPPER(TRIM(COALESCE(ba_kode, ''))) IN ('H1','H2','HB','RE')
                   ), 0)::double precision AS sum_omp
            FROM gl_transactions
            WHERE property_id IS NOT NULL
              AND ar BETWEEN 2021 AND 2025
            GROUP BY ar
            ORDER BY ar
        """)
        res = (await db.execute(q_sum)).mappings().all()
        for row in res:
            n_all = int(row["n_all"] or 0)
            n_omp = int(row["n_omp"] or 0)
            summary_rows.append({
                "ar": int(row["ar"]),
                "n_all": n_all,
                "n_omp": n_omp,
                "pct_omp": round(100.0 * n_omp / n_all, 2) if n_all else 0.0,
                "sum_all": float(row["sum_all"] or 0),
                "sum_omp": float(row["sum_omp"] or 0),
            })

        q_det = text("""
            SELECT ar::int AS ar,
                   UPPER(TRIM(COALESCE(ba_kode, ''))) AS ba,
                   COUNT(*)::bigint AS n,
                   COALESCE(SUM(belop), 0)::double precision AS sum_belop,
                   COALESCE(SUM(ABS(belop)), 0)::double precision AS sum_abs
            FROM gl_transactions
            WHERE property_id IS NOT NULL
              AND ar BETWEEN 2021 AND 2025
            GROUP BY ar, UPPER(TRIM(COALESCE(ba_kode, '')))
            HAVING COUNT(*) > 0
            ORDER BY ar, ba
        """)
        for row in (await db.execute(q_det)).mappings().all():
            ba = (row["ba"] or "").strip() or "(tom)"
            detail_rows.append({
                "ar": int(row["ar"]),
                "ba": ba,
                "n": int(row["n"] or 0),
                "sum_belop": float(row["sum_belop"] or 0),
                "sum_abs": float(row["sum_abs"] or 0),
                "kategori": (
                    "Ompostering / korrigering"
                    if ba in BA_OMPOSTERING
                    else ("Faktura / ordinær" if ba in BA_FAKTURA_TYPISK else "Annet")
                ),
            })
    except Exception as exc:
        logger.warning("Ompostering-statistikk feilet (fortsetter uten): %s", exc)

    return {"summary": summary_rows, "detail": detail_rows}


def _compute_outliers(
    prop_2027: Dict[str, float],
    prop_2025: Dict[str, float],
    prop_hist: Dict[str, Dict[int, float]],
    prop_info: Dict[str, Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """
    Identifiserer tre typer outliers per eiendom:
      A. Høy endring 2025→2027 (|%| > 50%, beløp > 500k)
      B. Høy historisk variasjon (CV = std/mean > 0.5 over 3+ år)
      C. Oppblåst ratio (2027 > 5× median av historikken)
    """
    import statistics

    result: List[Dict[str, Any]] = []

    for pid in set(prop_2027) | set(prop_hist):
        info   = prop_info.get(pid, {"name": pid, "region": "Ukjent"})
        b27    = prop_2027.get(pid, 0.0)
        b25    = prop_2025.get(pid, 0.0)
        hist   = prop_hist.get(pid, {})
        series = [v for v in hist.values() if v > 0]

        flags: List[str] = []
        details: Dict[str, Any] = {}

        # A: Stor endring 2025 → 2027
        if b25 > 500_000 and b27 > 0:
            pct = (b27 - b25) / b25 * 100
            if abs(pct) > 50:
                flags.append("høy_endring")
                details["endring_pct"] = round(pct, 1)

        # B: Høy historisk variasjon (CV)
        if len(series) >= 3:
            mean = statistics.mean(series)
            stdev = statistics.stdev(series)
            cv = stdev / mean if mean > 0 else 0
            if cv > 0.5:
                flags.append("høy_variasjon")
                details["cv"] = round(cv, 2)
                details["n_år"] = len(series)

        # C: Oppblåst prediksjon vs historisk median
        if len(series) >= 2 and b27 > 0:
            med = statistics.median(series)
            ratio = b27 / med if med > 0 else 0
            if ratio > 5.0:
                flags.append("oppblåst_ratio")
                details["ratio_vs_median"] = round(ratio, 1)
                details["historisk_median"] = round(med, 0)

        if flags:
            result.append({
                "name":    info["name"],
                "region":  info["region"],
                "b2025":   round(b25, 0),
                "b2027":   round(b27, 0),
                "flags":   flags,
                **details,
            })

    # Sorter: oppblåst ratio først, deretter høy endring, så variasjon
    priority = {"oppblåst_ratio": 0, "høy_endring": 1, "høy_variasjon": 2}
    result.sort(key=lambda x: (min(priority.get(f, 9) for f in x["flags"]), -abs(x["b2027"] - x["b2025"])))
    return result


# ── Ark 1: Antagelser ──────────────────────────────────────────────────────────

def _build_antagelser(wb: Workbook, regioner: List[str]) -> None:
    """Bygger Antagelser-ark og registrerer named ranges i workbook."""
    ws = wb.active
    ws.title = "Antagelser"

    ws["A1"] = "Bufetat - Budsjettjustering 2027"
    ws["A1"].font = FONT_TITLE

    ws["A2"] = (
        "Endre de gule cellene nedenfor. "
        "Alle beregninger i de andre arkene oppdateres automatisk. "
        "Hvite celler er beregnet og skal ikke endres manuelt. "
        "Arket «Omposteringer» viser omfang av korrigeringsbilag i GL (2021–2025); "
        "«Notater» forklarer metode, forbehold og sammenhengen med regnskapsmaterialet."
    )
    ws["A2"].font = FONT_SMALL
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 28

    # ── EIENDOMSDRIFT
    ws["A3"] = "EIENDOMSDRIFT"
    ws["A3"].fill = FILL_SECTION
    ws["A3"].font = FONT_SECTION
    ws["C3"] = "Hva er dette og hvordan brukes det?"
    ws["C3"].font = Font(bold=True, size=9, color="1F3864")

    ws["A4"] = "Inflasjon / kostnadsvekst (%)"
    ws["A4"].font = FONT_BOLD
    c = ws["B4"]
    c.value = 3.5
    c.fill = FILL_YELLOW
    c.number_format = NUM_PCT_VAL
    ws["C4"] = (
        "Modellen (Holt-Winters) gir en grunnprediksjon basert pa historisk utvikling 2021-2025. "
        "Legg til x% pa toppen for handlingsrom / uforutsette kostnader i 2027. "
        "Standard 3,5 = HW-prediksjon + 3,5% ekstra handlingsrom. Sett til 0 for a bruke HW-prediksjon direkte."
    )
    ws["C4"].font = FONT_SMALL
    ws["C4"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[4].height = 44

    # ── KATEGORIFAKTORER
    ws["A6"] = "KATEGORIFAKTORER  (ekstra justering per kostnadstype)"
    ws["A6"].fill = FILL_SECTION
    ws["A6"].font = FONT_SECTION
    ws["B6"] = "Faktor"
    ws["B6"].font = FONT_BOLD
    ws["C6"] = (
        "Faktoren multipliseres med det inflasjonsjusterte belOpet for den aktuelle kostnadskategorien. "
        "1,00 = ingen endring (standard). 1,10 = +10% ekstra. 0,90 = -10% reduksjon."
    )
    ws["C6"].font = Font(bold=True, size=9, color="1F3864")
    ws["C6"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[6].height = 32

    kat_descriptions = {
        "Drift":           "Lopende drift: strOm, renhold, forsikring, vedlikehold. Juster hvis du forventer prishopp utover inflasjonen.",
        "Gjennomstrømning": "Husleie og kontraktsbaserte kostnader. Kjores som ren inflasjonsfremskrivning - sett 1,00 med mindre kontraktene endres.",
        "Investering":     "StOrre investeringer og oppgraderinger av bygg. 1,00 = ingen ekstra utover inflasjonen.",
        "Annet":           "Ovrige kostnader. Sett 1,00 hvis du ikke har spesifikk informasjon.",
    }

    kat_start = 7
    for i, kat in enumerate(KAT_ORDER):
        r = kat_start + i
        ws.cell(r, 1, kat).font = FONT_BOLD
        c = ws.cell(r, 2, 1.0)
        c.fill = FILL_YELLOW
        c.number_format = NUM_FACTOR
        desc = ws.cell(r, 3, kat_descriptions.get(kat, ""))
        desc.font = FONT_SMALL
        desc.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 32

    # ── LØNN
    ws["A12"] = "LØNNSKOSTNADER"
    ws["A12"].fill = FILL_SECTION
    ws["A12"].font = FONT_SECTION
    ws["A13"] = "Lønnsøkning utover HW-prediksjon (%)"
    ws["A13"].font = FONT_BOLD
    c = ws["B13"]
    c.value = 4.5
    c.fill = FILL_YELLOW
    c.number_format = NUM_PCT_VAL
    ws["C13"] = (
        "HW-modellen predikerer lonnskostnader basert pa historikk. "
        "Legg til x% for kjent eller forventet lonnsvekst i 2027 (f.eks. tariffoppgjor). "
        "Standard 4,5 = HW-lonner + 4,5%. Sett til 0 for a bruke HW direkte."
    )
    ws["C13"].font = FONT_SMALL
    ws["C13"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[13].height = 44

    # ── REGIONFAKTORER
    ws["A15"] = "REGIONFAKTORER  (ekstra justering per region)"
    ws["A15"].fill = FILL_SECTION
    ws["A15"].font = FONT_SECTION
    ws["C15"] = (
        "Faktoren multipliseres med regionens inflasjonsjusterte beloP. "
        "1,00 = ingen endring (standard for alle regioner). "
        "1,05 = +5% ekstra for regionen. 0,95 = -5% reduksjon. "
        "Bruk 1,00 med mindre regionen har kjente avvik fra landsgjennomsnittet."
    )
    ws["C15"].font = Font(bold=True, size=9, color="1F3864")
    ws["C15"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[15].height = 44

    _h(ws["A16"], "Region",       FILL_HEADER)
    _h(ws["B16"], "Faktor",       FILL_HEADER)
    _h(ws["C16"], "Veiledning",   FILL_HEADER)

    region_start = 17
    for i, reg in enumerate(regioner):
        r = region_start + i
        ws.cell(r, 1, reg)
        c = ws.cell(r, 2, 1.0)
        c.fill = FILL_YELLOW
        c.number_format = NUM_FACTOR
        guide = ws.cell(r, 3,
            "1,00 = ingen endring  |  1,05 = +5%  |  0,95 = -5%  "
            "|  Endre kun hvis regionen har kjente kostnadsavvik"
        )
        guide.font = FONT_SMALL

    _col_w(ws, {1: 46, 2: 10, 3: 64})

    # ── Named ranges
    n = len(regioner)
    reg_end  = region_start + n - 1
    kat_end  = kat_start + len(KAT_ORDER) - 1
    wb.defined_names["Inflasjon"]    = DefinedName("Inflasjon",    attr_text="Antagelser!$B$4")
    wb.defined_names["LonnVekst"]    = DefinedName("LonnVekst",    attr_text="Antagelser!$B$13")
    wb.defined_names["KatTabell"]    = DefinedName("KatTabell",    attr_text=f"Antagelser!$A${kat_start}:$B${kat_end}")
    wb.defined_names["RegionTabell"] = DefinedName("RegionTabell", attr_text=f"Antagelser!$A${region_start}:$B${reg_end}")


# ── Ark 2: Sammendrag ──────────────────────────────────────────────────────────

def _build_sammendrag(wb: Workbook, n_reg: int, n_lonn: int) -> None:
    ws = wb.create_sheet("Sammendrag")

    ws["A1"] = "Bufetat – Totalt budsjettbehov 2027"
    ws["A1"].font = FONT_TITLE

    for ci, label in enumerate(
        ["Post", "2025 faktisk (NOK)", "HW 2027 (NOK)", "Justert 2027 (NOK)", "Endring vs 2025"], 1
    ):
        _h(ws.cell(2, ci), label)

    # Data-rader i Per region: 2..n_reg+1
    # Data-rader i Lønn: 3..n_lonn+2 (rad 1=tittel, rad 2=headers)
    reg_end  = n_reg + 1
    lonn_end = n_lonn + 2

    defs = [
        (3, "Eiendomsdrift",
         f"=SUM('Per region'!B2:B{reg_end})",
         f"=SUM('Per region'!C2:C{reg_end})",
         f"=SUM('Per region'!F2:F{reg_end})"),
        (4, "Lønnskostnader",
         f"=SUM(Lønn!B3:B{lonn_end})",
         f"=SUM(Lønn!C3:C{lonn_end})",
         f"=SUM(Lønn!D3:D{lonn_end})"),
    ]

    for ri, label, f25, fhw, fjust in defs:
        ws.cell(ri, 1, label).font = FONT_BOLD
        for ci, formula in enumerate([f25, fhw, fjust], start=2):
            c = ws.cell(ri, ci, formula)
            c.number_format = NUM_NOK
            c.fill = FILL_FORMULA
        c = ws.cell(ri, 5, f"=IFERROR((D{ri}-B{ri})/B{ri},\"\")")
        c.number_format = NUM_PCT_DEC
        c.fill = FILL_FORMULA

    # Totalrad
    for ci in range(1, 6):
        ws.cell(5, ci).fill = FILL_TOTAL
    ws.cell(5, 1, "TOTALT BUDSJETTBEHOV 2027").font = Font(bold=True, size=12)
    ws.cell(5, 1).fill = FILL_TOTAL
    for ci, formula in enumerate(["=B3+B4", "=C3+C4", "=D3+D4"], start=2):
        c = ws.cell(5, ci, formula)
        c.number_format = NUM_NOK
        c.fill = FILL_TOTAL
        c.font = Font(bold=True, size=12)
    c = ws.cell(5, 5, "=IFERROR((D5-B5)/B5,\"\")")
    c.number_format = NUM_PCT_DEC
    c.fill = FILL_TOTAL

    ws.cell(7, 1, "→ Endre parametere i arket «Antagelser» for å justere tallene over.").font = FONT_SMALL

    _col_w(ws, {1: 28, 2: 22, 3: 22, 4: 22, 5: 18})


# ── Ark 3: Per region ──────────────────────────────────────────────────────────

def _build_per_region(
    wb: Workbook,
    alle_regioner: List[str],
    reg_2027: Dict[str, float],
    reg_2025: Dict[str, float],
) -> None:
    ws = wb.create_sheet("Per region")

    for ci, h in enumerate(
        ["Region", "2025 GL (NOK)", "HW 2027 (NOK)",
         "Etter inflasjon", "Regionfaktor", "Justert 2027", "Endring vs 2025"], 1
    ):
        _h(ws.cell(1, ci), h)

    for i, reg in enumerate(alle_regioner, start=2):
        ws.cell(i, 1, reg)
        ws.cell(i, 2, round(reg_2025.get(reg, 0.0), 0)).number_format = NUM_NOK
        ws.cell(i, 3, round(reg_2027.get(reg, 0.0), 0)).number_format = NUM_NOK
        for ci, formula in [
            (4, f"=C{i}*(1+Inflasjon/100)"),
            (5, f"=IFERROR(VLOOKUP(A{i},RegionTabell,2,0),1)"),
            (6, f"=D{i}*E{i}"),
            (7, f"=IFERROR((F{i}-B{i})/B{i},\"\")"),
        ]:
            c = ws.cell(i, ci, formula)
            c.fill = FILL_FORMULA
            c.number_format = NUM_FACTOR if ci == 5 else (NUM_PCT_DEC if ci == 7 else NUM_NOK)

    n = len(alle_regioner)
    tr = n + 2
    ws.cell(tr, 1, "Totalt").fill = FILL_TOTAL
    ws.cell(tr, 1).font = FONT_BOLD
    for ci in (2, 3, 4, 6):
        letter = get_column_letter(ci)
        c = ws.cell(tr, ci, f"=SUM({letter}2:{letter}{n+1})")
        c.number_format = NUM_NOK
        c.fill = FILL_TOTAL
        c.font = FONT_BOLD
    ws.cell(tr, 5).fill = FILL_TOTAL
    c = ws.cell(tr, 7, f"=IFERROR((F{tr}-B{tr})/B{tr},\"\")")
    c.number_format = NUM_PCT_DEC
    c.fill = FILL_TOTAL

    _col_w(ws, {1: 20, 2: 20, 3: 20, 4: 20, 5: 14, 6: 20, 7: 16})


# ── Ark 4: Per kategori ────────────────────────────────────────────────────────

def _build_per_kategori(
    wb: Workbook,
    kat_2027: Dict[str, float],
    kat_2025: Dict[str, float],
) -> None:
    ws = wb.create_sheet("Per kategori")

    for ci, h in enumerate(
        ["Kategori", "2025 GL (NOK)", "HW 2027 (NOK)",
         "Etter inflasjon", "Kategorifaktor", "Justert 2027"], 1
    ):
        _h(ws.cell(1, ci), h)

    kats = [k for k in KAT_ORDER if k in kat_2027 or k in kat_2025]
    for i, kat in enumerate(kats, start=2):
        ws.cell(i, 1, kat)
        ws.cell(i, 2, round(kat_2025.get(kat, 0.0), 0)).number_format = NUM_NOK
        ws.cell(i, 3, round(kat_2027.get(kat, 0.0), 0)).number_format = NUM_NOK
        for ci, formula in [
            (4, f"=C{i}*(1+Inflasjon/100)"),
            (5, f"=IFERROR(VLOOKUP(A{i},KatTabell,2,0),1)"),
            (6, f"=D{i}*E{i}"),
        ]:
            c = ws.cell(i, ci, formula)
            c.fill = FILL_FORMULA
            c.number_format = NUM_FACTOR if ci == 5 else NUM_NOK

    n = len(kats)
    tr = n + 2
    ws.cell(tr, 1, "Totalt").fill = FILL_TOTAL
    ws.cell(tr, 1).font = FONT_BOLD
    for ci in (2, 3, 4, 6):
        letter = get_column_letter(ci)
        c = ws.cell(tr, ci, f"=SUM({letter}2:{letter}{n+1})")
        c.number_format = NUM_NOK
        c.fill = FILL_TOTAL
        c.font = FONT_BOLD
    ws.cell(tr, 5).fill = FILL_TOTAL

    _col_w(ws, {1: 22, 2: 20, 3: 20, 4: 20, 5: 16, 6: 20})


# ── Ark 5: Lønn ───────────────────────────────────────────────────────────────

def _build_lonn(
    wb: Workbook,
    alle_regioner: List[str],
    lonn_r25: Dict[str, float],
    lonn_r27: Dict[str, float],
) -> None:
    ws = wb.create_sheet("Lønn")

    ws["A1"] = "Lønnskostnader 2027 – Holt-Winters prediksjon per region"
    ws["A1"].font = Font(bold=True, size=12)

    for ci, h in enumerate(
        ["Region", "2025 faktisk (NOK)", "HW 2027 (NOK)",
         "Justert 2027", "Endring vs 2025"], 1
    ):
        _h(ws.cell(2, ci), h)

    regioner_med_lonn = [r for r in alle_regioner if r in lonn_r25 or r in lonn_r27]
    if not regioner_med_lonn:
        ws.cell(3, 1, "Ingen lønnsdata. Kjør: POST /api/v1/financials/salary-costs/predict?year=2027").font = FONT_SMALL
        return

    for i, reg in enumerate(regioner_med_lonn, start=3):
        ws.cell(i, 1, reg)
        ws.cell(i, 2, round(lonn_r25.get(reg, 0.0), 0)).number_format = NUM_NOK
        ws.cell(i, 3, round(lonn_r27.get(reg, 0.0), 0)).number_format = NUM_NOK
        c = ws.cell(i, 4, f"=C{i}*(1+LonnVekst/100)")
        c.number_format = NUM_NOK
        c.fill = FILL_BLUE
        c = ws.cell(i, 5, f"=IFERROR((D{i}-B{i})/B{i},\"\")")
        c.number_format = NUM_PCT_DEC
        c.fill = FILL_BLUE

    n = len(regioner_med_lonn)
    tr = n + 3
    ws.cell(tr, 1, "Totalt").fill = FILL_TOTAL
    ws.cell(tr, 1).font = FONT_BOLD
    for ci in (2, 3, 4):
        letter = get_column_letter(ci)
        c = ws.cell(tr, ci, f"=SUM({letter}3:{letter}{n+2})")
        c.number_format = NUM_NOK
        c.fill = FILL_TOTAL
        c.font = FONT_BOLD
    c = ws.cell(tr, 5, f"=IFERROR((D{tr}-B{tr})/B{tr},\"\")")
    c.number_format = NUM_PCT_DEC
    c.fill = FILL_TOTAL

    ws.cell(tr + 2, 1,
            "Justert 2027 = HW 2027 × (1 + LonnVekst/100). "
            "Endre «Lønnsøkning (%)» i Antagelser.").font = FONT_SMALL

    _col_w(ws, {1: 20, 2: 22, 3: 22, 4: 22, 5: 18})


# ── Ark 6: Alle eiendommer ─────────────────────────────────────────────────────

def _build_alle_eiendommer(wb: Workbook, per_eiendom: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("Alle eiendommer")

    for ci, h in enumerate(
        ["Eiendom", "Region", "Kategori", "2025 GL (NOK)", "HW 2027 (NOK)",
         "Etter inflasjon", "Regionfaktor", "Kategorifaktor", "Justert 2027", "Endring vs 2025"], 1
    ):
        _h(ws.cell(1, ci), h)

    for i, row in enumerate(per_eiendom, start=2):
        ws.cell(i, 1, row["name"])
        ws.cell(i, 2, row["region"])
        ws.cell(i, 3, row.get("kategori", "Drift"))
        ws.cell(i, 4, row["belop_2025"]).number_format = NUM_NOK
        ws.cell(i, 5, row["belop_2027"]).number_format = NUM_NOK
        for ci, formula in [
            (6, f"=E{i}*(1+Inflasjon/100)"),
            (7, f"=IFERROR(VLOOKUP(B{i},RegionTabell,2,0),1)"),
            (8, f"=IFERROR(VLOOKUP(C{i},KatTabell,2,0),1)"),
            (9, f"=F{i}*G{i}*H{i}"),
            (10, f"=IFERROR((I{i}-D{i})/D{i},\"\")"),
        ]:
            c = ws.cell(i, ci, formula)
            c.fill = FILL_FORMULA
            c.number_format = (NUM_FACTOR if ci in (7, 8) else
                               NUM_PCT_DEC if ci == 10 else NUM_NOK)

    n = len(per_eiendom)
    tr = n + 2
    ws.cell(tr, 1, "Totalt").fill = FILL_TOTAL
    ws.cell(tr, 1).font = FONT_BOLD
    for ci in (4, 5, 6, 9):
        letter = get_column_letter(ci)
        c = ws.cell(tr, ci, f"=SUM({letter}2:{letter}{n+1})")
        c.number_format = NUM_NOK
        c.fill = FILL_TOTAL
        c.font = FONT_BOLD

    _col_w(ws, {1: 40, 2: 16, 3: 16, 4: 18, 5: 18, 6: 18, 7: 14, 8: 14, 9: 18, 10: 16})


# ── Ark 7: Outliers ───────────────────────────────────────────────────────────

def _build_outliers(wb: Workbook, outliers: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("Outliers")

    ws["A1"] = "Outlier-analyse – eiendommer som bør manuelt vurderes"
    ws["A1"].font = FONT_TITLE
    ws["A2"] = (
        "Tre flagg: høy_endring = 2025→2027 avvik > 50% (beløp > 500k)  |  "
        "høy_variasjon = historisk CV > 0.5 (ustabil historikk)  |  "
        "oppblåst_ratio = 2027 > 5× historisk median"
    )
    ws["A2"].font = FONT_SMALL

    headers = [
        "Eiendom", "Region", "2025 GL (NOK)", "HW 2027 (NOK)",
        "Endring %", "Flagg", "CV", "Ratio vs median", "Historisk median (NOK)"
    ]
    for ci, h in enumerate(headers, 1):
        _h(ws.cell(3, ci), h)

    FILL_RED    = PatternFill("solid", fgColor="FFD7D7")
    FILL_ORANGE = PatternFill("solid", fgColor="FFE4B5")
    FILL_YELLOW_LIGHT = PatternFill("solid", fgColor="FFFACD")

    for i, row in enumerate(outliers, start=4):
        flags = row["flags"]
        fill = FILL_RED if "oppblåst_ratio" in flags else (
               FILL_ORANGE if "høy_endring" in flags else FILL_YELLOW_LIGHT)

        ws.cell(i, 1, row["name"]).fill = fill
        ws.cell(i, 2, row["region"]).fill = fill

        c = ws.cell(i, 3, row["b2025"])
        c.number_format = NUM_NOK
        c.fill = fill

        c = ws.cell(i, 4, row["b2027"])
        c.number_format = NUM_NOK
        c.fill = fill

        pct = row.get("endring_pct")
        if pct is not None:
            c = ws.cell(i, 5, pct / 100)
            c.number_format = NUM_PCT_DEC
            c.fill = fill
        else:
            ws.cell(i, 5, "").fill = fill

        ws.cell(i, 6, ", ".join(flags)).fill = fill

        cv = row.get("cv")
        ws.cell(i, 7, cv if cv is not None else "").fill = fill
        if cv is not None:
            ws.cell(i, 7).number_format = "0.00"

        ratio = row.get("ratio_vs_median")
        ws.cell(i, 8, ratio if ratio is not None else "").fill = fill
        if ratio is not None:
            ws.cell(i, 8).number_format = "0.0"

        med = row.get("historisk_median")
        if med is not None:
            c = ws.cell(i, 9, med)
            c.number_format = NUM_NOK
            c.fill = fill
        else:
            ws.cell(i, 9, "").fill = fill

    n = len(outliers)
    if n == 0:
        ws.cell(4, 1, "Ingen outliers funnet – alle prediksjoner ser rimelige ut.").font = FONT_SMALL
    else:
        ws.cell(n + 5, 1,
            f"Totalt {n} eiendommer flagget. "
            "Røde = oppblåst ratio (kritisk). Orange = høy endring. Gul = ustabil historikk."
        ).font = FONT_SMALL

    _col_w(ws, {1: 42, 2: 16, 3: 18, 4: 18, 5: 12, 6: 32, 7: 8, 8: 18, 9: 22})


# ── Ark 8: Rådata HW ──────────────────────────────────────────────────────────

def _build_rawdata(wb: Workbook, per_eiendom: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("Rådata HW")

    ws["A1"] = "Rådata – Holt-Winters prediksjon (kilde-ark, ingen formler)"
    ws["A1"].font = FONT_BOLD

    for ci, h in enumerate(
        ["Eiendom", "Region", "2025 GL (NOK)", "HW 2027 (NOK)", "Endring %"], 1
    ):
        c = ws.cell(2, ci)
        c.value = h
        c.font = FONT_BOLD
        c.fill = FILL_GRAY

    for i, row in enumerate(per_eiendom, start=3):
        b25, b27 = row["belop_2025"], row["belop_2027"]
        endring = round((b27 - b25) / b25 * 100, 1) if b25 > 0 else None
        for ci, v in enumerate([row["name"], row["region"], b25, b27, endring], 1):
            c = ws.cell(i, ci, v)
            c.fill = FILL_GRAY
            if ci in (3, 4):
                c.number_format = NUM_NOK
            if ci == 5 and v is not None:
                c.number_format = "0.0"

    _col_w(ws, {1: 40, 2: 16, 3: 18, 4: 18, 5: 12})


# ── Ark 9: Backtesting ────────────────────────────────────────────────────────

def _build_backtesting(wb: Workbook, backtest: Dict[str, Any]) -> None:
    """
    Backtesting-ark: viser MAPE og MAE for testaarene 2023, 2024, 2025.
    Trening: 2021 t.o.m. aaret foer testaar. Sammenligning med faktisk GL.
    """
    ws = wb.create_sheet("Backtesting")
    FILL_BT_OK   = PatternFill("solid", fgColor="C6EFCE")   # grønn – MAPE < 15%
    FILL_BT_WARN = PatternFill("solid", fgColor="FFEB9C")   # gul – MAPE 15-30%
    FILL_BT_BAD  = PatternFill("solid", fgColor="FFC7CE")   # rød – MAPE > 30%

    ws["A1"] = "Backtesting – Holt-Winters prediksjonsnoeyaktighet"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = (
        "Out-of-sample test: for hvert testaar trenes modellen pa historikk t.o.m. aaret foer, "
        "predikerer det aktuelle aaret og sammenligner med faktisk GL. "
        "Samme parametre som brukes for 2027-prediksjonen (alpha=0.5, beta=0.2, phi=0.85, max_vekst=8%/ar). "
        "MAPE-beregningen inkluderer kun eiendommer med minst 3 aar treningsdata og ekskluderer "
        "Gjennomstromning (passeringspostene som alltid faar inflasjonstillegg – meningslos aa benchmarke). "
        "2023-testen trener kun pa 2021-2022 (2 aar), saa MAPE der er indikativ, ikke definitiv."
    )
    ws["A2"].font = FONT_SMALL
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 52

    test_years = backtest.get("test_years", [2023, 2024, 2025])
    results = backtest.get("results", {})

    # ── Sekson 1: Overordnet nøyaktighet ──────────────────────────────────────
    r = 4
    ws.cell(r, 1, "Overordnet noeyaktighet per testaar").font = Font(bold=True, size=11, color="1F3864")
    r += 1

    headers = ["Testaar", "Trener pa", "Antall prop.", "Faktisk GL (NOK)", "Predikert (NOK)", "Avvik (NOK)", "MAPE (%)"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(r, ci, h)
        c.font = FONT_WHITE
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center")
    r += 1

    for yr in test_years:
        yr_data = results.get(str(yr)) or results.get(yr, {})
        ov = yr_data.get("overall", {})
        actual = ov.get("actual", 0)
        predicted = ov.get("predicted", 0)
        mape = ov.get("mape", None)
        n = ov.get("n_properties", 0)
        avvik = predicted - actual if actual else None

        mape_fill = FILL_GRAY
        if mape is not None:
            if mape < 15:
                mape_fill = FILL_BT_OK
            elif mape < 30:
                mape_fill = FILL_BT_WARN
            else:
                mape_fill = FILL_BT_BAD

        row_vals = [
            yr,
            f"2021-{yr - 1}",
            n,
            actual,
            predicted,
            avvik,
            mape,
        ]
        fills = [None, None, None, None, None, None, mape_fill]
        fmts  = [None, None, None, NUM_NOK, NUM_NOK, NUM_NOK, "0.0"]
        for ci, (v, f, fmt) in enumerate(zip(row_vals, fills, fmts), 1):
            c = ws.cell(r, ci, v)
            if f:
                c.fill = f
            if fmt:
                c.number_format = fmt
            if ci == 1:
                c.font = FONT_BOLD
        r += 1

    r += 1

    # ── Seksjon 2: MAPE per kategori ──────────────────────────────────────────
    ws.cell(r, 1, "MAPE per kategori").font = Font(bold=True, size=11, color="1F3864")
    r += 1

    cat_headers = ["Kategori"] + [str(yr) for yr in test_years]
    for ci, h in enumerate(cat_headers, 1):
        c = ws.cell(r, ci, h)
        c.font = FONT_WHITE
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center" if ci > 1 else "left")
    r += 1

    # Collect all categories
    all_cats: set[str] = set()
    for yr in test_years:
        yr_data = results.get(str(yr)) or results.get(yr, {})
        all_cats.update(yr_data.get("per_category", {}).keys())

    for cat in sorted(all_cats):
        row_start = r
        ws.cell(r, 1, cat).font = FONT_BOLD
        for ci, yr in enumerate(test_years, 2):
            yr_data = results.get(str(yr)) or results.get(yr, {})
            cat_data = yr_data.get("per_category", {}).get(cat, {})
            mape = cat_data.get("mape", None)
            c = ws.cell(r, ci, mape)
            if mape is not None:
                c.number_format = "0.0"
                if mape < 15:
                    c.fill = FILL_BT_OK
                elif mape < 30:
                    c.fill = FILL_BT_WARN
                else:
                    c.fill = FILL_BT_BAD
        r += 1

    r += 1

    # ── Seksjon 3: Detaljert per aar ─────────────────────────────────────────
    ws.cell(r, 1, "Detaljert per testaar").font = Font(bold=True, size=11, color="1F3864")
    r += 1

    det_headers = ["Testaar", "Kategori", "Faktisk GL", "Predikert", "Avvik", "Avvik %", "MAPE %", "Ant. eiend."]
    for ci, h in enumerate(det_headers, 1):
        c = ws.cell(r, ci, h)
        c.font = FONT_WHITE
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center" if ci > 1 else "left")
    r += 1

    for yr in test_years:
        yr_data = results.get(str(yr)) or results.get(yr, {})
        per_cat = yr_data.get("per_category", {})
        first = True
        for cat in sorted(per_cat.keys()):
            cd = per_cat[cat]
            actual = cd.get("actual", 0)
            predicted = cd.get("predicted", 0)
            avvik = predicted - actual
            avvik_pst = cd.get("endring_pst", None)
            mape = cd.get("mape", None)
            n = cd.get("n_properties", 0)

            mape_fill = FILL_GRAY
            if mape is not None:
                if mape < 15: mape_fill = FILL_BT_OK
                elif mape < 30: mape_fill = FILL_BT_WARN
                else: mape_fill = FILL_BT_BAD

            row_vals = [yr if first else None, cat, actual, predicted, avvik, avvik_pst, mape, n]
            fmts = [None, None, NUM_NOK, NUM_NOK, NUM_NOK, "0.0", "0.0", None]
            for ci, (v, fmt) in enumerate(zip(row_vals, fmts), 1):
                c = ws.cell(r, ci, v)
                if fmt:
                    c.number_format = fmt
                if ci == 7 and mape is not None:
                    c.fill = mape_fill
                if ci == 1 and v is not None:
                    c.font = FONT_BOLD
            first = False
            r += 1
        r += 0  # no blank between categories per year

    r += 1

    # ── Viktig uteligger – konto 6300 i 2024 ─────────────────────────────────
    ws.cell(r, 1, "VIKTIG: Dataavvik oppdaget – konto 6300 (Leie lokaler andre utleiere) i 2024").font = Font(
        bold=True, size=10, color="9C0006"
    )
    r += 1
    # Header row for outlier table
    outlier_hdrs = ["Konto", "Beskrivelse", "2021", "2022", "2023", "2024 (avvik!)", "2025"]
    for ci, h in enumerate(outlier_hdrs, 1):
        c = ws.cell(r, ci, h)
        c.font = FONT_WHITE
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center" if ci > 2 else "left")
    r += 1
    # Data row
    outlier_vals = ["6300", "Leie lokaler andre utleiere",
                    115_986_270, 146_922_170, 173_607_921, 1_328_878_123, 280_799_862]
    outlier_fmts = [None, None, NUM_NOK, NUM_NOK, NUM_NOK, NUM_NOK, NUM_NOK]
    outlier_fills = [None, None, None, None, None, PatternFill("solid", fgColor="FFC7CE"), None]
    for ci, (v, fmt, fill) in enumerate(zip(outlier_vals, outlier_fmts, outlier_fills), 1):
        c = ws.cell(r, ci, v)
        if fmt:
            c.number_format = fmt
        if fill:
            c.fill = fill
    r += 1
    ws.cell(r, 1,
        "Bilag 800285205 (oktober 2024): 'Husl. Q4 feilfakturert' – 1 145 013 600 NOK. "
        "En enkelt feilfakturert husleiepostering utgjorde nesten hele avviket i 2024. "
        "Prediksjonsmodellen bruker na netto-summering (positive + negative bilag) slik at "
        "reverserte feilfaktureringer automatisk nuller seg ut. "
        "Dermed er 2024-prediksjonen og historikken ryddet for dette avviket."
    ).font = FONT_SMALL
    ws[f"A{r}"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 60
    ws.merge_cells(f"A{r}:G{r}")
    r += 2

    # ── Forklaringsboks ───────────────────────────────────────────────────────
    ws.cell(r, 1, "Om MAPE (Mean Absolute Percentage Error):").font = Font(bold=True, size=9, color="1F3864")
    r += 1
    ws.cell(r, 1,
        "MAPE = gjennomsnittlig prosentvis avvik mellom prediksjon og faktisk verdi per eiendom. "
        "Under 15% = god (groenn). 15-30% = akseptabel (gul). Over 30% = vurder rekalibrering (roed). "
        "En lav MAPE betyr at modellen historisk sett har traff godt, og at 2027-prediksjonen er troverdig. "
        "NB: Gjennomstromning er ekskludert fra MAPE (passering – kun inflasjonsjustering, ikke predikert). "
        "Eiendommer med faerre enn 3 aar treningsdata er ogsaa ekskludert (kaldstart gir misvisende avvik)."
    ).font = FONT_SMALL
    ws[f"A{r}"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 52

    _col_w(ws, {1: 26, 2: 22, 3: 18, 4: 18, 5: 18, 6: 12, 7: 12, 8: 14})


# ── Ark: Omposteringer (GL-kompleksitet) ───────────────────────────────────────

def _build_omposteringer(wb: Workbook, omp: Dict[str, Any]) -> None:
    """Informasjonsark: bilagsarter og ompostering i samme historikk-vindu som prediksjonen (2021–2025)."""
    ws = wb.create_sheet("Omposteringer")

    ws["A1"] = "Omposteringer og bilagsarter i GL (eiendomskoblede linjer)"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:H1")

    intro = (
        "I statsregnskapet korrigeres feil normalt ikke ved å endre eller slette originalbilag. "
        "I stedet bokføres motbilag (for eksempel bilagsart RE eller H1) som reverserer feil postering, "
        "og deretter nye bilag (ofte H1, H2 eller HB) med korrekt koststed og konto. "
        "Originalbilaget står uendret — immutabilitet. "
        "\n\n"
        "Dette arket viser hvor stor andel av GL-linjene som er knyttet til eiendom (property_id satt) "
        "som har bilagsarter typiske for ompostering (H1, H2, HB, RE), sammenlignet med ordinære "
        "faktura-/løpetypiske arter (IV, IW, LE, MT). "
        "Høy andel ompostering betyr ikke nødvendigvis feil, men at årsverdier og trender kan være "
        "påvirket av mange korrigeringer — prediksjonen (Holt-Winters) bruker likevel summerte kostnader per år."
    )
    ws["A3"] = intro
    ws["A3"].font = Font(size=10)
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A3:H3")
    ws.row_dimensions[3].height = 110

    r = 5
    _h(ws.cell(r, 1), "År", FILL_HEADER)
    _h(ws.cell(r, 2), "Antall linjer totalt", FILL_HEADER)
    _h(ws.cell(r, 3), "Derav ompostering (H1,H2,HB,RE)", FILL_HEADER)
    _h(ws.cell(r, 4), "Andel ompostering %", FILL_HEADER)
    _h(ws.cell(r, 5), "Sum beløp alle (NOK)", FILL_HEADER)
    _h(ws.cell(r, 6), "Sum beløp ompostering (NOK)", FILL_HEADER)
    r += 1

    for row in omp.get("summary") or []:
        ws.cell(r, 1, row["ar"])
        ws.cell(r, 2, row["n_all"])
        ws.cell(r, 3, row["n_omp"])
        c = ws.cell(r, 4, row["pct_omp"] / 100.0)
        c.number_format = NUM_PCT_DEC
        c = ws.cell(r, 5, round(row["sum_all"], 0))
        c.number_format = NUM_NOK
        c = ws.cell(r, 6, round(row["sum_omp"], 0))
        c.number_format = NUM_NOK
        r += 1

    if not (omp.get("summary") or []):
        ws.cell(r, 1, "Ingen data (tom GL eller spørring feilet).")
        ws.merge_cells(f"A{r}:H{r}")
        r += 2
    else:
        r += 1

    ws.cell(r, 1, "Detalj per år og bilagsart (BA)").font = FONT_SECTION
    ws.merge_cells(f"A{r}:H{r}")
    ws.cell(r, 1).fill = FILL_SECTION
    r += 1

    _h(ws.cell(r, 1), "År", FILL_HEADER)
    _h(ws.cell(r, 2), "Bilagsart (BA)", FILL_HEADER)
    _h(ws.cell(r, 3), "Kategori", FILL_HEADER)
    _h(ws.cell(r, 4), "Antall linjer", FILL_HEADER)
    _h(ws.cell(r, 5), "Sum beløp (NOK)", FILL_HEADER)
    _h(ws.cell(r, 6), "Sum |beløp| (NOK)", FILL_HEADER)
    r += 1

    for row in sorted(omp.get("detail") or [], key=lambda x: (x["ar"], x["ba"])):
        ws.cell(r, 1, row["ar"])
        ws.cell(r, 2, row["ba"])
        ws.cell(r, 3, row["kategori"])
        ws.cell(r, 4, row["n"])
        c = ws.cell(r, 5, round(row["sum_belop"], 0))
        c.number_format = NUM_NOK
        c = ws.cell(r, 6, round(row["sum_abs"], 0))
        c.number_format = NUM_NOK
        r += 1

    if not (omp.get("detail") or []):
        ws.cell(r, 1, "Ingen detaljdata.")
        r += 1

    r += 1
    ws.cell(r, 1, "Referanse: Se også «Notater» avsnitt 11 og brukerhjelp om bilagsarter ved Agresso-import.").font = FONT_SMALL
    ws.merge_cells(f"A{r}:H{r}")

    _col_w(ws, {1: 10, 2: 14, 3: 28, 4: 14, 5: 18, 6: 18, 7: 12, 8: 12})


# ── Ark: Budsjett 2026 (Økonomi) + Sammenligning 2026 ─────────────────────────

async def _load_finance_budget_2026(db: AsyncSession) -> Dict[str, Any]:
    """Henter vedtatt budsjett 2026 fra finance_budget-tabellen."""
    from app.models.financial_models import FinanceBudget
    from app.domains.core.models.property import Property

    CATS = ["Lokaler", "Drift", "Vedlikehold"]

    try:
        rows = (await db.execute(
            select(
                FinanceBudget.property_id,
                FinanceBudget.koststed_navn,
                FinanceBudget.category,
                FinanceBudget.is_direktorat_level,
                func.sum(FinanceBudget.amount).label("total"),
            )
            .where(
                FinanceBudget.year == 2026,
                FinanceBudget.data_source == "finance_dept_2026",
            )
            .group_by(
                FinanceBudget.property_id,
                FinanceBudget.koststed_navn,
                FinanceBudget.category,
                FinanceBudget.is_direktorat_level,
            )
        )).fetchall()
    except Exception as exc:
        logger.warning("_load_finance_budget_2026 feilet: %s", exc)
        return {"per_eiendom": [], "direktorat": {}, "total_by_cat": {}}

    prop_rows = (await db.execute(
        select(Property.property_id, Property.name, Property.region)
    )).fetchall()
    prop_info = {str(r.property_id): {"name": r.name or str(r.property_id), "region": r.region or "Ukjent"} for r in prop_rows}

    # per-eiendom aggregering { pid: { cat: amount } }
    eiendom_map: Dict[str, Dict[str, float]] = {}
    direktorat_map: Dict[str, float] = {c: 0.0 for c in CATS}

    for r in rows:
        v = float(r.total or 0)
        cat = r.category if r.category in CATS else "Drift"
        if r.is_direktorat_level or r.property_id is None:
            direktorat_map[cat] = direktorat_map.get(cat, 0.0) + v
        else:
            pid = str(r.property_id)
            eiendom_map.setdefault(pid, {c: 0.0 for c in CATS})
            eiendom_map[pid][cat] = eiendom_map[pid].get(cat, 0.0) + v

    per_eiendom = []
    for pid, cats in eiendom_map.items():
        info = prop_info.get(pid, {"name": pid, "region": "Ukjent"})
        total = sum(cats.values())
        per_eiendom.append({
            "name": info["name"], "region": info["region"],
            "Lokaler": cats.get("Lokaler", 0.0),
            "Drift": cats.get("Drift", 0.0),
            "Vedlikehold": cats.get("Vedlikehold", 0.0),
            "total": total,
        })
    per_eiendom.sort(key=lambda x: -x["total"])

    total_by_cat = {c: sum(r[c] for r in per_eiendom) + direktorat_map.get(c, 0.0) for c in CATS}
    return {"per_eiendom": per_eiendom, "direktorat": direktorat_map, "total_by_cat": total_by_cat}


def _build_budsjett_2026_okonomi(wb: Workbook, fb26: Dict[str, Any]) -> None:
    CATS = ["Lokaler", "Drift", "Vedlikehold"]
    ws = wb.create_sheet("Budsjett 2026 (Økonomi)")
    _col_w(ws, {1: 38, 2: 22, 3: 16, 4: 16, 5: 16, 6: 16})

    ws.merge_cells("A1:F1")
    t = ws.cell(1, 1, "Budsjett 2026 — vedtatt budsjett fra økonomiavdelingen")
    t.fill = PatternFill("solid", fgColor="2F5496")
    t.font = Font(bold=True, size=13, color="FFFFFF")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:F2")
    ws.cell(2, 1, (
        "Beløp DA (vedtatt budsjett) — dekker alle 12 måneder av 2026.  "
        "NB: Kontantbeløp i kildefilen er kun fylt t.o.m. uttrekksdato (april 2026) — "
        "fremtidige måneder er tomme i den kolonnen og brukes IKKE her."
    )).font = FONT_SMALL
    ws.cell(2, 1).fill = PatternFill("solid", fgColor="FFF2CC")
    ws.row_dimensions[2].height = 28

    headers = ["Eiendom", "Region"] + CATS + ["Total"]
    for c, h in enumerate(headers, 1):
        _h(ws.cell(4, c), h)
    ws.row_dimensions[4].height = 18

    row = 5
    per_eiendom = fb26.get("per_eiendom", [])
    for e in per_eiendom:
        ws.cell(row, 1, e["name"])
        ws.cell(row, 2, e["region"])
        for ci, cat in enumerate(CATS, 3):
            cell = ws.cell(row, ci, round(e.get(cat, 0)))
            cell.number_format = NUM_NOK
        total_cell = ws.cell(row, 6, round(e["total"]))
        total_cell.number_format = NUM_NOK
        total_cell.font = Font(bold=True)
        if row % 2 == 0:
            for c in range(1, 7):
                ws.cell(row, c).fill = PatternFill("solid", fgColor="F2F2F2")
        row += 1

    dir_map = fb26.get("direktorat", {})
    dir_total = sum(dir_map.values())
    if dir_total > 0:
        row += 1
        ws.merge_cells(f"A{row}:F{row}")
        sec = ws.cell(row, 1, "Direktorat-nivå (koststed uten eiendomsmatch)")
        sec.font = FONT_SECTION
        sec.fill = FILL_SECTION
        row += 1
        _h(ws.cell(row, 1), "Kategori", FILL_SECTION)
        _h(ws.cell(row, 2), "Beløp (NOK)", FILL_SECTION)
        row += 1
        for cat in CATS:
            v = dir_map.get(cat, 0.0)
            if v:
                ws.cell(row, 1, cat)
                c = ws.cell(row, 2, round(v))
                c.number_format = NUM_NOK
                row += 1

    row += 1
    ws.cell(row, 1, "TOTAL").font = Font(bold=True, size=11)
    ws.cell(row, 1).fill = FILL_TOTAL
    total_by_cat = fb26.get("total_by_cat", {})
    grand = 0.0
    for ci, cat in enumerate(CATS, 3):
        v = total_by_cat.get(cat, 0.0)
        grand += v
        c = ws.cell(row, ci, round(v))
        c.number_format = NUM_NOK
        c.font = Font(bold=True)
        c.fill = FILL_TOTAL
    gt = ws.cell(row, 6, round(grand))
    gt.number_format = NUM_NOK
    gt.font = Font(bold=True, size=11)
    gt.fill = FILL_TOTAL
    for c in [1, 2]:
        ws.cell(row, c).fill = FILL_TOTAL

    if not per_eiendom and dir_total == 0:
        ws.cell(5, 1, "Ingen data — last opp Excel-fil via /admin/budsjett-import først.").font = FONT_SMALL


def _build_sammenligning_2026(wb: Workbook, hw2026_by_name: Dict[str, str], fb26: Dict[str, Any]) -> None:
    """Sammenligner BEFS HW-prediksjon 2026 mot vedtatt budsjett fra økonomi per eiendom."""
    ws = wb.create_sheet("Sammenligning 2026")
    _col_w(ws, {1: 38, 2: 22, 3: 18, 4: 18, 5: 16, 6: 14})

    ws.merge_cells("A1:F1")
    t = ws.cell(1, 1, "Sammenligning 2026 — BEFS-prediksjon vs. vedtatt budsjett (Økonomi)")
    t.fill = PatternFill("solid", fgColor="1F3864")
    t.font = Font(bold=True, size=13, color="FFFFFF")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.cell(2, 1, "Grønn = prediksjon under budsjett  |  Rød = prediksjon over budsjett").font = FONT_SMALL
    ws.cell(2, 1).fill = PatternFill("solid", fgColor="EBF3FB")

    headers = ["Eiendom", "Region", "Prediksjon 2026 (BEFS)", "Budsjett 2026 (Økonomi)", "Avvik (NOK)", "Avvik (%)"]
    for c, h in enumerate(headers, 1):
        _h(ws.cell(4, c), h)
    ws.row_dimensions[4].height = 18

    fb26_list = fb26.get("per_eiendom", [])
    fb26_by_name: Dict[str, Dict] = {e["name"]: e for e in fb26_list}
    all_names = sorted(set(fb26_by_name) | set(hw2026_by_name))

    FILL_OVER  = PatternFill("solid", fgColor="FFCCCC")
    FILL_UNDER = PatternFill("solid", fgColor="CCFFCC")

    row = 5
    total_hw = 0.0
    total_fb = 0.0
    for name in all_names:
        hw = hw2026_by_name.get(name, 0.0)
        fb_entry = fb26_by_name.get(name, {})
        fb = fb_entry.get("total", 0.0)
        avvik = hw - fb
        avvik_pct = round(avvik / fb * 100, 1) if fb else 0.0
        total_hw += hw
        total_fb += fb

        ws.cell(row, 1, name)
        ws.cell(row, 2, fb_entry.get("region", ""))
        c3 = ws.cell(row, 3, round(hw)); c3.number_format = NUM_NOK
        c4 = ws.cell(row, 4, round(fb)); c4.number_format = NUM_NOK
        c5 = ws.cell(row, 5, round(avvik)); c5.number_format = NUM_NOK
        c6 = ws.cell(row, 6, avvik_pct); c6.number_format = "0.0"

        fill = FILL_OVER if avvik > 0 else (FILL_UNDER if avvik < 0 else None)
        if fill:
            for col in [5, 6]:
                ws.cell(row, col).fill = fill
        row += 1

    row += 1
    ws.cell(row, 1, "TOTAL").font = Font(bold=True, size=11)
    ws.cell(row, 1).fill = FILL_TOTAL
    av_total = round(total_hw - total_fb)
    av_pct_total = round((total_hw - total_fb) / total_fb * 100, 1) if total_fb else 0.0
    for col, val in [(3, round(total_hw)), (4, round(total_fb)), (5, av_total), (6, av_pct_total)]:
        c = ws.cell(row, col, val)
        c.number_format = NUM_NOK if col < 6 else "0.0"
        c.font = Font(bold=True)
        c.fill = FILL_TOTAL
    for col in [1, 2]:
        ws.cell(row, col).fill = FILL_TOTAL

    if not all_names:
        ws.cell(5, 1, "Ingen data — last opp Excel-fil og kjør 2026-prediksjon.").font = FONT_SMALL


# ── Ark: Notater og forutsetninger ────────────────────────────────────────────

def _build_note(wb: Workbook, no_pred: List[Dict[str, Any]], n_outliers: int) -> None:
    ws = wb.create_sheet("Notater")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 90

    def title(row: int, text: str):
        c = ws.cell(row, 1, text)
        c.font = FONT_TITLE
        ws.merge_cells(f"A{row}:B{row}")
        ws.row_dimensions[row].height = 26

    def section(row: int, text: str):
        for col in (1, 2):
            c = ws.cell(row, col)
            c.fill = FILL_SECTION
            c.font = FONT_SECTION
        ws.cell(row, 1, text)
        ws.merge_cells(f"A{row}:B{row}")
        ws.row_dimensions[row].height = 20

    def label(row: int, lbl: str, val: str, bold_val: bool = False):
        cl = ws.cell(row, 1, lbl)
        cl.font = Font(bold=True, size=10)
        cl.alignment = Alignment(vertical="top")
        cv = ws.cell(row, 2, val)
        cv.font = Font(bold=bold_val, size=10)
        cv.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = max(15, min(80, len(val) // 3))

    def body(row: int, text: str, italic: bool = False):
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row, 1, text)
        c.font = Font(size=10, italic=italic)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = max(15, min(100, len(text) // 2))

    def blank(row: int):
        ws.row_dimensions[row].height = 8

    import datetime
    today = datetime.date.today().isoformat()

    r = 1
    title(r, "Bufetat - Prediksjon 2027: Metodikk, forutsetninger og veiledning")
    r += 1
    body(r, f"Rapporten er generert automatisk fra BEFS (Bufetat Eiendomsforvaltningssystem) den {today}. "
            "Dette arket forklarer hva tallene betyr, hvordan de er beregnet, og hvordan du bruker dem "
            "i budsjettarbeidet. Les dette arket før du endrer noe i Antagelser-arket. "
            "For bilagsarter og omposteringer i GL, se arket «Omposteringer» og avsnitt 11 nedenfor.")
    r += 2

    # ── 1. FORMAL
    section(r, "1.  HENSIKT OG MAL MED RAPPORTEN")
    r += 1
    body(r,
        "Rapporten gir et databasert utgangspunkt for budsjett 2027 for alle Bufetats eiendommer. "
        "Den erstatter ikke faglig skjonn, men gir regnskap og eiendomsforvaltning et felles tallgrunnlag "
        "som er beregnet pa samme mate for alle eiendommer - og som kan justeres samlet eller per region/kategori "
        "ved aa endre parametere i Antagelser-arket.")
    r += 1
    body(r,
        "Mottakere: Regnskapsavdelingen, eiendomsforvaltere og budsjettansvarlige i Bufetat. "
        "Tallene er ment som et forslag / startpunkt - ikke som et endelig budsjett.")
    r += 2

    # ── 2. METODEVALG
    section(r, "2.  HVORFOR BRUKER VI DENNE BEREGNINGSMETODEN?")
    r += 1
    body(r,
        "Vi bruker statistisk tidsserieanalyse (Holt-Winters eksponentiell glatting) i stedet for "
        "enkel prosentvis fremskrivning av fjoraaret. Arsaken er at eiendomskostnader ikke vokser "
        "jevnt fra aar til aar - de pavirkes av sesong, kontraktssykluser, investeringstidspunkter "
        "og tilfeldige engangseffekter. En ren prosent-vekst (f.eks. 'fjoraaret + 5%') ville gi "
        "skjeve tall naar fjoraaret var uvanlig hOyt eller lavt.")
    r += 1
    body(r,
        "Holt-Winters lOser dette ved aa gi nyere aar stOrre vekt enn eldre aar (eksponentiell vekting), "
        "og ved aa skille mellom den underliggende kostnadstrenden og tilfeldige svingninger. "
        "Modellen er mye brukt i offentlig sektor for budsjettfremskrivning fordi den er robust, "
        "transparent og gir forklarbare resultater - i motsetning til maskinlaeringsmodeller som er "
        "vanskeligere aa etterprOve.")
    r += 2

    # ── 3. HVA ER HOLT-WINTERS?
    section(r, "3.  HVA ER HOLT-WINTERS? (FORKLART FOR IKKE-STATISTIKERE)")
    r += 1
    body(r,
        "Tenk pa det som et vektet gjennomsnitt der nyere tall teller mer enn gamle tall. "
        "Modellen holder styr paa to ting samtidig:")
    r += 1
    label(r, "Niva (alpha)",
        "Hva er det typiske kostnadsnivaaet akkurat naa? "
        "Alpha = 0,50 betyr at halvparten av vekten legges paa det siste aarets tall, "
        "og halvparten paa hva modellen trodde foer. HOy alpha = modellen reagerer raskt paa endringer. "
        "Lav alpha = modellen er mer konservativ og glatter ut svingninger.")
    r += 1
    label(r, "Trend (beta)",
        "Beveger kostnadene seg oppover eller nedover over tid? "
        "Beta = 0,20 betyr at trenden oppdateres sakte - modellen endrer ikke retning fort "
        "basert paa ett enkelt aar. Dette er riktig for eiendomskostnader som er stabile over tid.")
    r += 1
    label(r, "Demping (phi)",
        "En fremskrivning 2 aar frem i tid kan bli urealistisk optimistisk/pessimistisk "
        "hvis trenden fortsetter uendret. Phi = 0,85 demper trenden gradvis jo lenger frem vi ser, "
        "slik at prediksjonen blir mer konservativ. Dette er standard god praksis for budsjettformaal.")
    r += 1
    body(r,
        "EKSEMPEL: Hvis en eiendom hadde kostnader paa 2,0 MNOK (2023), 2,2 MNOK (2024) og 2,1 MNOK (2025), "
        "vil Holt-Winters veie 2025-tallet tyngst, se at trenden er svakt stigende, "
        "og predikere ca. 2,25-2,35 MNOK for 2027 (avhengig av parametere). "
        "En enkel '2025 + 5%' ville gitt 2,2 MNOK - modellen gir et mer nyansert svar.")
    r += 2

    # ── 4. BEREGNINGSTRINN
    section(r, "4.  STEG-FOR-STEG: SLIK BEREGNES TALLENE")
    r += 1
    label(r, "Steg 1:",
        "Hent GL-transaksjoner (regnskapsfOrte kostnader) per eiendom for aarene 2021-2025 fra Agresso/GL-systemet.")
    r += 1
    label(r, "Steg 2:",
        "Juster alle historiske beloep til 2025-kroner ved aa bruke SSBs konsumprisindeks (KPI). "
        "Dette sikrer at vi sammenligner like med like - 1 krone i 2021 er ikke det samme som 1 krone i 2025.")
    r += 1
    label(r, "Steg 3:",
        "Kjoer Holt-Winters modellen per eiendom. Modellen beregner nivaa og trend basert paa de 5 aarene, "
        "og fremskriver til 2027 med demping (phi=0,85).")
    r += 1
    label(r, "Steg 4:",
        "Sikkerhetssperre: prediksjoner over 5x historisk median kappes til 5x median. "
        "Dette forhindrer at engangseffekter (f.eks. en stor investering ett enkelt aar) blaasr opp prediksjonen. "
        "Eiendommer som treffes av denne grensen flagges automatisk som RODE i Outliers-arket (oppblast_ratio) "
        "slik at de fanges opp til manuell vurdering uansett.")
    r += 1
    label(r, "Steg 5:",
        "Eiendommer med under 3 aars historikk faar en enklere beregning: siste kjente aar gangen med "
        "inflasjonsanslag to ganger (ca. 1,035^2 = +7,1%). Dette er konservativt og gjennomsiktig.")
    r += 1
    label(r, "Steg 6 (Excel):",
        "Du kan justere det endelige budsjettet i Antagelser-arket: legge til inflasjonspaaslag, "
        "kategorifaktorer og regionfaktorer. Alle beregningsark oppdateres automatisk.")
    r += 2

    # ── 5. PARAMETERE
    section(r, "5.  PARAMETERE OG FORUTSETNINGER I DENNE KJØRINGEN")
    r += 1
    label(r, "Modell-inflasjon:",
        "3,5% per aar. Brukes internt i HW-modellen for aa konvertere fra 2025-prisnivaa til 2027. "
        "Kilde: Norges Bank inflasjonsmal 2025-2026.")
    r += 1
    label(r, "Excel-handlingsrom:",
        "3,5% (standard i Antagelser-arket, kan endres). "
        "Legges paa toppen av HW-prediksjonen som et ekstra buffer for uforutsette kostnader. "
        "VIKTIG - slik stables de to lagene: "
        "  HW-prediksjon (inkl. 3,5% modell-inflasjon) x 1,035 = Justert 2027. "
        "  Eksempel: 2,0 MNOK (2025) -> HW 2027 ca. 2,14 MNOK -> Justert 2,22 MNOK. "
        "  Total effekt over 2025-faktisk er ca. 11%, ikke 7%. "
        "MERK: Bufferen er tiltenkt som handlingsrom - "
        "den bor ikke sees paa som et beloep som automatisk skal brukes opp. "
        "Oke til 7,5% hvis dere vil ha mer sikkerhetsmargin i budsjettet.")
    r += 1
    label(r, "Lønnsøkning:",
        "4,5% (standard i Antagelser-arket, kan endres). "
        "Legges paa toppen av HW-lonnsprediksjonen. Baser paa kjent tariffresultat eller forhandlingsmandat for 2027.")
    r += 1
    label(r, "Maks-faktor:",
        "5x historisk median. Prediksjoner over denne grensen kappes automatisk og flagges som RODE i Outliers. "
        "Hensikt: forhindre at engangsutgifter (f.eks. ekstraordinaert vedlikehold ett enkelt aar) "
        "gir urealistiske fremskrivninger.")
    r += 2

    # ── 6. KATEGORIER
    section(r, "6.  KOSTNADSKATEG0RIER - HVORDAN ER DE DEFINERT?")
    r += 1
    label(r, "Drift:",
        "Lopende, aarlig tilbakevendende kostnader: strOm, fjernvarme, renhold, forsikring, "
        "vaktmester, service-avtaler, smaa reparasjoner. Dette er den stOrste og mest stabile kategorien.")
    r += 1
    label(r, "Gjennomstrommning:",
        "Husleie og leiekostnader som Bufetat betaler til eksterne utleiere. "
        "Disse kostnadene er kontraktsbestemt og fremskrives som ren inflasjon (ikke Holt-Winters-trend), "
        "fordi de folger kontraktsbetingelsene, ikke historisk variasjon.")
    r += 1
    label(r, "Investering:",
        "StOrre vedlikeholdsprosjekter og oppgraderinger: utskifting av tekniske anlegg (ventilasjon, heis), "
        "fasaderehab, brannsikring. Kan variere mye fra aar til aar - sjekk outliers-arket.")
    r += 1
    label(r, "Annet:",
        "Kostnader som ikke passer i de andre kategoriene. Inkluderer bl.a. flyttekostnader, "
        "midlertidige lokaler, konsulenttjenester knyttet til eiendom.")
    r += 2

    # ── 7. DEKNINGSGRAD
    section(r, "7.  DATAGRUNNLAG OG DEKNINGSGRAD")
    r += 1
    n_pred = 192
    label(r, "Med prediksjon:", f"{n_pred} eiendommer. Disse har tilstrekkelig GL-historikk (minst 1 aar med data).")
    r += 1
    n_no_pred = len(no_pred) if no_pred else 20
    label(r, "Uten prediksjon:",
        f"{n_no_pred} eiendommer. Disse har lOnndata men ingen GL-transaksjoner koblet i systemet. "
        "De er IKKE inkludert i totaltallene i dette arket og maa budsjetteres manuelt. "
        "Mulige aarsaker: (a) enheten er nylig opprettet og mangler historikk, eller "
        "(b) GL-data finnes i Agresso men koststedet er ikke koblet til eiendommen i systemet. "
        "Sistnevnte kan utbedres - ta kontakt med systemansvarlig hvis enheten har regnskapsdata i Agresso.")

    if no_pred:
        lonn_tot = sum(x["lonn_2025"] for x in no_pred)
        r += 1
        label(r, "Manglende:",
            f"Estimert lOnnskostnad 2025 for disse {n_no_pred} eiendommene: {lonn_tot/1e6:.1f} MNOK. "
            "Legg til dette belopet manuelt i budsjettprosessen.")
        r += 1
        body(r, "Liste over eiendommer uten prediksjon:")
        for x in no_pred:
            r += 1
            label(r, x["region"], f"{x['name'][:70]}   (lOnn 2025: {x['lonn_2025']/1e6:.1f} MNOK)")
    r += 2

    # ── 8. OUTLIERS
    section(r, "8.  OUTLIER-ANALYSE - EIENDOMMER SOM BOR KONTROLLERES MANUELT")
    r += 1
    body(r,
        f"Systemet har flagget {n_outliers} eiendommer i arket 'Outliers'. "
        "Dette er eiendommer der prediksjonen avviker vesentlig fra historikken, "
        "noe som kan skyldes datafeil, engangseffekter eller reelle strukturendringer. "
        "Disse bor ikke brukes direkte i budsjettet uten manuell vurdering.")
    r += 1
    label(r, "Rod (kritisk):",
        "Oppblast ratio: HW-prediksjonen er mer enn 5x historisk median. "
        "Nesten alltid en indikasjon paa datafeil eller engangseffekt (f.eks. stor investering i ett aar "
        "som modellen feiltolker som en varig trend). Anbefaling: erstatt med manuelt estimat.")
    r += 1
    label(r, "Oransje:",
        "HOy endring: prediksjonen avviker mer enn 50% fra 2025-faktisk, og belopet er over 500k kr. "
        "Kan vaere riktig hvis eiendommen har gjennomgaatt store endringer, men bOr verifiseres.")
    r += 1
    label(r, "Gul:",
        "HOy variasjon: eiendommen har hatt svingninger over 50% (CV > 0,5) de siste aarene. "
        "Prediksjonen er usikker - bor sammenlignes med budsjettvedtak og driftsplan.")
    r += 2

    # ── 9. SLIK BRUKER DU ARKET
    section(r, "9.  SLIK BRUKER DU DETTE ARKET I BUDSJETTARBEIDET")
    r += 1
    label(r, "Steg 1:",
        "Les outliers-arket. Rod-flaggede eiendommer bor korrigeres manuelt foer du bruker totaltallene.")
    r += 1
    label(r, "Steg 2:",
        "Vurder handlingsrom-paaslaget i Antagelser-arket (standard 3,5%). "
        "Modellen inkluderer allerede 3,5% prisvekst internt. "
        "Total effekt med standard 3,5% paaslaeg er ca. 11% over 2025-faktisk. "
        "Oke til 7,5% for mer sikkerhetsmargin, eller sett til 0 for ren HW-prediksjon.")
    r += 1
    label(r, "Steg 3:",
        "Vurder lOnnsOkning (standard 4,5%). Oppdater basert paa kjent tariffresultat eller forhandlingsmandat.")
    r += 1
    label(r, "Steg 4:",
        "Sjekk Sammendrag-arket for totaltall. Sammenlign med fjoraaret og med eventuelt rammebrev.")
    r += 1
    label(r, "Steg 5:",
        f"Legg til manuelt estimat for de {n_no_pred} eiendommene uten prediksjon (se liste over).")
    r += 1
    label(r, "Steg 6:",
        "Bruk 'Alle eiendommer'-arket for aa identifisere enkelteiendommer med uventet hOye eller lave tall. "
        "Kolonne G 'Justert 2027' er forslaget - kolonne H viser prosentendring fra 2025.")
    r += 2

    # ── 10. BEGRENSNINGER
    section(r, "10. BEGRENSNINGER OG FORBEHOLD")
    r += 1
    body(r,
        "Statistiske prediksjoner er alltid beheftet med usikkerhet. Folgende forhold fanges IKKE opp "
        "automatisk av modellen og maa vurderes manuelt:")
    r += 1
    label(r, "Nye eiendommer:",
        "Eiendommer som er tatt i bruk etter 2023 har lite historikk. Prediksjonen er usikker.")
    r += 1
    label(r, "Planlagte investeringer:",
        "Kjente fremtidige prosjekter (ombygging, rehabilitering) er ikke inkludert. "
        "Legg til disse manuelt i budsjettet.")
    r += 1
    label(r, "Kontraktsendringer:",
        "Nye husleieavtaler, reforhandlede kontrakter eller avviklede leieforhold fanges ikke opp. "
        "Sjekk mot kontraktsregisteret.")
    r += 1
    label(r, "Politiske vedtak:",
        "Besparelseskrav, omstruktureringer, kapasitetsendringer eller tjenesteavvikling vedtatt av styret "
        "eller direktoratet er ikke hensyntatt i modellen. Eiendommer som er under avvikling eller "
        "vesentlig nedskalering bOr manuelt justeres ned - eller tas helt ut av budsjettet.")
    r += 1
    label(r, "Datakvalitet:",
        "Prediksjonen er ikke bedre enn GL-dataene den bygger på. Manglende eller feilkonterte "
        "transaksjoner gir misvisende prediksjoner. I statsregnskapet kommer ofte korrigeringer som "
        "egne bilag (RE, H1, H2, HB) i tillegg til ordinære fakturaer — se arket «Omposteringer» for "
        "fordeling per år og bilagsart. Se «Outliers» for automatiske varsler på eiendomsnivå.")
    r += 2

    section(r, "11. OMPOSTERINGER OG TOLKNING AV REGNSKAPSMATERIALET")
    r += 1
    body(r,
        "Bufetats eiendoms-GL i BEFS speiler Agresso: feil koststed eller konto rettes med motbilag og "
        "ny postering — ikke ved å slette originalen. Typiske bilagsarter for slike korrigeringer er "
        "RE (reversering) og H1/H2/HB (ompostering med fleksibel dimensjon, der Dim1 = koststed ofte endres). "
        "Derfor vil en enkelt års sum på et koststed være «nettoresultatet» av faktiske kostnader og "
        "eventuelle korrigeringer i samme år.")
    r += 1
    body(r,
        "Holt-Winters-prediksjonen bruker summerte årsbeløp per eiendom (positiv sum i importen). "
        "Mange omposteringer øker kompleksiteten og kan påvirke enkeltår mer enn underliggende «kontinuerlig» "
        "kostnadsnivå — derfor er det nyttig å lese «Omposteringer»-arket sammen med Outliers og faglig skjønn.")
    r += 2

    ws.cell(r, 1, f"Generert av BEFS - Bufetat Eiendomsforvaltningssystem  |  {today}").font = FONT_SMALL


# ── Hovedfunksjon ──────────────────────────────────────────────────────────────

async def build_prediksjon_export_xlsx(db: AsyncSession, scenario: str = "xgb70") -> io.BytesIO:
    """
    Bygger Excel-arbeidsbok med kaskaderende formler (Antagelser) og metodenotater.

    Alle beregnede kolonner bruker Excel-formler via named ranges:
      Inflasjon, LonnVekst, KatTabell, RegionTabell

    Endring i Antagelser-arket oppdaterer Per region, Per kategori, Lønn, Alle eiendommer og Sammendrag.
    """
    from app.services.prediction_service import BudgetPredictionService

    data = await _load_data(db, scenario)

    # Kjor backtesting asynkront
    try:
        backtest = await BudgetPredictionService.run_backtest(db=db)
    except Exception as exc:
        logger.warning("Backtesting feilet (fortsetter uten): %s", exc)
        backtest = {"test_years": [2023, 2024, 2025], "parameters": {}, "results": {}}

    try:
        omp_data = await _load_ompostering_stats(db)
    except Exception as exc:
        logger.warning("Ompostering-data feilet (fortsetter uten): %s", exc)
        omp_data = {"summary": [], "detail": []}

    try:
        fb26 = await _load_finance_budget_2026(db)
    except Exception as exc:
        logger.warning("Finance budget 2026 feilet (fortsetter uten): %s", exc)
        fb26 = {"per_eiendom": [], "direktorat": {}, "total_by_cat": {}}

    # Last HW-prediksjon 2026 per eiendomsnavn for sammenligningsfanen
    try:
        from app.models.financial_models import Budget
        from app.domains.core.models.property import Property as _Prop
        hw2026_rows = (await db.execute(
            select(_Prop.name, func.sum(Budget.amount).label("total"))
            .join(_Prop, Budget.property_id == _Prop.property_id)
            .where(Budget.year == 2026, Budget.is_synthetic == True)
            .group_by(_Prop.name)
        )).fetchall()
        hw2026_by_name: Dict[str, float] = {r.name: float(r.total or 0) for r in hw2026_rows if r.name}
    except Exception as exc:
        logger.warning("HW 2026 per navn feilet (fortsetter uten): %s", exc)
        hw2026_by_name = {}

    wb = Workbook()

    # Ark 1: Antagelser (aktivt ark) — setter opp named ranges
    _build_antagelser(wb, data["alle_regioner"])

    # Beregn antall data-rader for Sammendrag-referanser
    n_reg  = len(data["alle_regioner"])
    n_lonn = len([r for r in data["alle_regioner"]
                  if r in data["lonn_r25"] or r in data["lonn_r27"]])

    # Ark 2–13
    _build_sammendrag(wb, n_reg, n_lonn)
    _build_per_region(wb, data["alle_regioner"], data["reg_2027"], data["reg_2025"])
    _build_per_kategori(wb, data["kat_2027"], data["kat_2025"])
    _build_lonn(wb, data["alle_regioner"], data["lonn_r25"], data["lonn_r27"])
    _build_alle_eiendommer(wb, data["per_eiendom"])
    _build_outliers(wb, data["outliers"])
    _build_backtesting(wb, backtest)
    _build_rawdata(wb, data["per_eiendom"])
    _build_omposteringer(wb, omp_data)
    _build_budsjett_2026_okonomi(wb, fb26)
    _build_sammenligning_2026(wb, hw2026_by_name, fb26)
    _build_note(wb, data["no_pred"], len(data["outliers"]))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


__all__ = ["build_prediksjon_export_xlsx"]
