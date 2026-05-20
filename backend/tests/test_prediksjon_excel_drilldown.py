from collections import namedtuple

import pytest
from openpyxl import Workbook

from app.services.financials.prediksjon_drill_sheets import append_prediksjon_drill_sheets


class FakeResult:
    def __init__(self, rows):
        self._rows = rows

    def fetchall(self):
        return self._rows


class FakeSession:
    def __init__(self, results):
        self._results = list(results)

    async def execute(self, _stmt):
        if not self._results:
            raise AssertionError("Unexpected extra execute() call")
        return FakeResult(self._results.pop(0))


@pytest.mark.asyncio
async def test_append_prediksjon_drill_sheets_adds_history_and_full_basis_sheets():
    BudgetRow = namedtuple("BudgetRow", ["property_id", "category", "belop"])
    GlKatRow = namedtuple("GlKatRow", ["property_id", "srs_kategori", "belop"])
    HistRow = namedtuple("HistRow", ["property_id", "srs_kategori", "ar", "belop"])
    DetailRow = namedtuple(
        "DetailRow",
        [
            "transaction_id",
            "property_id",
            "ar",
            "periode",
            "bilagsnr",
            "bilagsdato",
            "konto",
            "konto_navn",
            "srs_kategori",
            "ba_kode",
            "innkjopskategori_navn",
            "underkategori_navn",
            "leverandor_navn",
            "tekst",
            "belop",
            "dim1_kode",
            "dim1_navn",
            "dim3_kode",
            "dim6_anlegg_id",
            "is_statsbygg",
        ],
    )

    db = FakeSession(
        [
            [BudgetRow("P1", "operations", 1200.0)],
            [BudgetRow("P1", "operations", 1000.0)],
            [GlKatRow("P1", "Drift", 900.0)],
            [HistRow("P1", "Drift", 2024, 850.0), HistRow("P1", "Drift", 2025, 900.0)],
            [("P1", "Drift", "6300", "Leie", 900.0, 2)],
            [("P1", "B-100", None, "6300", "Leie", "Drift", "Fakturalinje", 450.0)],
            [
                DetailRow(
                    1,
                    "P1",
                    2025,
                    202501,
                    "B-100",
                    None,
                    "6300",
                    "Leie",
                    "Drift",
                    "IV",
                    "Lokaler",
                    "Leie lokaler",
                    "Utleier AS",
                    "Fakturalinje",
                    450.0,
                    "100",
                    "Koststed",
                    "200",
                    "A-1",
                    True,
                )
            ],
        ]
    )

    wb = Workbook()
    wb.remove(wb.active)

    await append_prediksjon_drill_sheets(
        db,
        wb,
        prop_info={"P1": {"name": "Testeiendom", "region": "Øst"}},
    )

    assert "Eiendom_kategori" in wb.sheetnames
    assert "Historikk_grunnlag" in wb.sheetnames
    assert "Kostnadsgrunnlag_GL" in wb.sheetnames

    ws_hist = wb["Historikk_grunnlag"]
    headers = [ws_hist.cell(row=1, column=i).value for i in range(1, 10)]
    assert "2025" in headers

    ws_basis = wb["Kostnadsgrunnlag_GL"]
    basis_headers = [ws_basis.cell(row=2, column=i).value for i in range(1, 16)]
    assert "leverandor" in basis_headers
    assert "bilagsart" in basis_headers
